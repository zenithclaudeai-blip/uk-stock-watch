#!/usr/bin/env python3
"""
UK Stock Watch — cloud poller.

Runs on a GitHub Actions schedule instead of in a browser. Pulls the same free
sources as the browser extension (Google News, Yahoo News, Yahoo Finance quotes,
Investing.com UK's analyst-ratings feed), classifies items, pushes new
upgrade/downgrade alerts to a webhook (WhatsApp via CallMeBot, ntfy, etc.), and
writes a static dashboard page for GitHub Pages.

Stdlib only — no pip installs needed.
"""

import http.cookiejar
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from zoneinfo import ZoneInfo

STATE_DIR = os.environ.get("STATE_DIR", "state")
DOCS_DIR = os.environ.get("DOCS_DIR", "docs")
DOCS_FILENAME = os.environ.get("DOCS_FILENAME", "index.html")
WATCHLIST_FILE = os.environ.get("WATCHLIST_FILE", "watchlist.json")
SEEN_FILE = os.path.join(STATE_DIR, "seen.json")
DATA_FILE = os.path.join(STATE_DIR, "data.json")
# When true, this run skips the market-wide broker search, LSE screener, and heat map —
# used by the large-watchlist hourly job so it doesn't duplicate what the fast 5-minute
# job already covers (which would mean duplicate WhatsApp alerts from two separate runs).
SKIP_MARKET_WIDE = os.environ.get("SKIP_MARKET_WIDE", "false").lower() == "true"
MAX_ITEMS_PER_TICKER = 60
MAX_SEEN = 3000

NEWS_MAX_AGE_DAYS = 21  # news/broker items older than this are filtered from the live feed
# When true, additionally restricts to items published "today" in the real London
# calendar (handles the GMT/BST switch correctly) — set False if this proves too
# strict and cuts out genuinely relevant items from yesterday evening.
NEWS_SAME_LONDON_DAY_ONLY = True

UPGRADE_WORDS = [
    "upgrade", "raises rating", "buy rating", "outperform", "overweight",
    "raised to buy", "initiates.*buy",
    r"upgrades?\s+\S+.{0,25}\s+to\s+(buy|overweight|outperform|add|accumulate)",
    r"raises?\s+\S+.{0,25}\s+to\s+(buy|overweight|outperform)",
    r"moves?\s+\S+.{0,25}\s+to\s+(buy|overweight|outperform)",
]
DOWNGRADE_WORDS = [
    "downgrade", "cuts rating", "sell rating", "underperform", "underweight",
    "cut to sell", "initiates.*sell",
    r"cuts?\s+\S+.{0,25}\s+to\s+(sell|underweight|underperform|reduce|hold)",
    r"downgrades?\s+\S+.{0,25}\s+to\s+(sell|hold|underweight|underperform)",
    r"moves?\s+\S+.{0,25}\s+to\s+(sell|underweight|underperform)",
]
TARGET_WORDS = ["price target", "target price", "pt raised", "pt cut"]
# Director/PDMR (Persons Discharging Managerial Responsibilities) dealings — a standard
# RNS announcement type disclosing when a company director has bought or sold shares.
# Real, published, factual disclosure — not something this tool infers or predicts.
DIRECTOR_DEALING_WORDS = [
    "pdmr", "director/pdmr shareholding", "director shareholding",
    "notification of transactions by persons discharging managerial responsibilities",
    "director dealing", "directors' dealing", "holding(s) in company",
]
EVENT_WORDS = [
    "trading update", "results", "interim report", "final results", "agm",
    "dividend", "profit warning", "acquisition", "placing", "earnings", "guidance",
    "merger", "merges", "merging", "takeover", "bid for", "bids for", "acquire",
    "acquires", "acquiring", "buyout", "to be bought", "in talks to", "offer for",
]
BROKER_NAMES = [
    "Barclays", "Deutsche Bank", "JPMorgan", "JP Morgan", "HSBC", "UBS", "Citi",
    "Citigroup", "Goldman Sachs", "Morgan Stanley", "RBC", "RBC Capital Markets",
    "Jefferies", "Peel Hunt", "Numis", "Berenberg", "Panmure", "Panmure Liberum",
    "Shore Capital", "Investec", "Canaccord", "Canaccord Genuity", "Liberum", "BofA",
    "Bank of America", "Societe Generale", "BNP Paribas", "Credit Suisse", "Stifel",
    "Redburn", "Zeus Capital", "Cenkos", "finnCap", "Singer Capital Markets",
    "N+1 Singer", "WH Ireland", "Edison", "Equity Development", "Cantor Fitzgerald",
    "Exane BNP Paribas", "Kepler Cheuvreux", "Mediobanca", "Davy", "Goodbody",
    "Third Bridge", "Marex", "Turner Pope", "Arden Partners", "Charles Stanley",
    "AJ Bell", "Hargreaves Lansdown", "Killik", "Interactive Investor", "Shard Capital",
    "Oberon Capital", "Whitman Howard", "Progressive Equity Research", "Hardman & Co",
]
ANALYST_RATINGS_FEED_URL = "https://uk.investing.com/rss/news_1061.rss"
GENERAL_MARKET_NEWS_FEED_URL = "https://uk.investing.com/rss/news_25.rss"  # confirmed via their own official RSS listing page
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
}

# Yahoo's screener and quoteSummary endpoints reject anonymous requests from cloud/
# datacenter IPs (GitHub Actions runners included) with a 401 unless a session cookie
# + "crumb" token — Yahoo's own web frontend's real auth handshake — are attached.
# The basic chart/quote endpoint is more lenient and doesn't need this. This uses a
# shared cookie jar across all Yahoo requests so the session persists once established.
_yahoo_cookiejar = http.cookiejar.CookieJar()
_yahoo_opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(_yahoo_cookiejar))
_yahoo_crumb = None


def _yahoo_opener_get(url, timeout=15):
    req = urllib.request.Request(url, headers=HEADERS)
    with _yahoo_opener.open(req, timeout=timeout) as resp:
        return resp.read()


def get_yahoo_crumb():
    """Fetch (and cache) a Yahoo auth crumb, establishing the session cookie first."""
    global _yahoo_crumb
    if _yahoo_crumb:
        return _yahoo_crumb
    try:
        _yahoo_opener_get("https://fc.yahoo.com")  # sets the initial session cookie
    except Exception as e:
        print(f"  ! yahoo cookie handshake failed: {e}", file=sys.stderr)
    try:
        crumb = _yahoo_opener_get("https://query2.finance.yahoo.com/v1/test/getcrumb").decode("utf-8").strip()
        if crumb and "<html" not in crumb.lower():
            _yahoo_crumb = crumb
            return _yahoo_crumb
    except Exception as e:
        print(f"  ! yahoo crumb fetch failed: {e}", file=sys.stderr)
    return None


def http_get(url, timeout=15):
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.read()


def classify(title):
    t = title.lower()
    if any(re.search(w, t) for w in UPGRADE_WORDS):
        return "upgrade"
    if any(re.search(w, t) for w in DOWNGRADE_WORDS):
        return "downgrade"
    if any(w in t for w in TARGET_WORDS):
        return "target"
    if any(w in t for w in DIRECTOR_DEALING_WORDS):
        return "director_dealing"
    if any(w in t for w in EVENT_WORDS):
        return "event"
    return "news"


def detect_broker(title):
    t = title.lower()
    for b in BROKER_NAMES:
        if b.lower() in t:
            return b
    return ""


def parse_rss(xml_bytes):
    items = []
    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError:
        return items
    for item in root.iter("item"):
        title = (item.findtext("title") or "").strip()
        link = (item.findtext("link") or "").strip()
        pub_date = (item.findtext("pubDate") or "").strip()
        source_el = item.find("source")
        source = source_el.text.strip() if source_el is not None and source_el.text else ""
        if not source and link:
            try:
                source = urllib.parse.urlparse(link).hostname.replace("www.", "")
            except Exception:
                source = ""
        if title and link:
            items.append({"title": title, "link": link, "pubDate": pub_date, "source": source})
    return items


def fetch_yahoo_analyst(ticker):
    """
    Structured analyst ratings history via Yahoo Finance's quoteSummary endpoint —
    real firm name, exact action (up/down/init/main/reit), and date, straight from
    Yahoo's own aggregation of broker calls. Requires the crumb-authenticated opener;
    falls back to trying without a crumb if one couldn't be obtained.
    Also pulls calendarEvents (next earnings date, next ex-dividend date) in the SAME
    request — quoteSummary accepts multiple modules at once, so this adds real data
    without adding a new network call.
    """
    symbol = yahoo_symbol(ticker)
    crumb = get_yahoo_crumb()
    url = (
        f"https://query1.finance.yahoo.com/v10/finance/quoteSummary/{urllib.parse.quote(symbol)}"
        "?modules=upgradeDowngradeHistory,recommendationTrend,financialData,calendarEvents,summaryDetail,defaultKeyStatistics,price,assetProfile"
    )
    if crumb:
        url += f"&crumb={urllib.parse.quote(crumb)}"
    try:
        data = json.loads(_yahoo_opener_get(url))
        result = (((data.get("quoteSummary") or {}).get("result")) or [None])[0]
        if not result:
            return None
        history = ((result.get("upgradeDowngradeHistory") or {}).get("history")) or []
        fin = result.get("financialData") or {}
        cal = result.get("calendarEvents") or {}
        earnings_dates = ((cal.get("earnings") or {}).get("earningsDate")) or []
        next_earnings = earnings_dates[0].get("raw") if earnings_dates else None
        ex_div = (cal.get("exDividendDate") or {}).get("raw")
        summary = result.get("summaryDetail") or {}
        div_rate = (summary.get("dividendRate") or {}).get("raw")  # currency amount per share per year
        div_yield_raw = (summary.get("dividendYield") or {}).get("raw")  # Yahoo returns this as a fraction, e.g. 0.045 = 4.5%
        div_yield_pct = div_yield_raw * 100 if div_yield_raw is not None else None
        stats = result.get("defaultKeyStatistics") or {}
        price_mod = result.get("price") or {}
        trailing_pe = (summary.get("trailingPE") or {}).get("raw")
        eps = (stats.get("trailingEps") or {}).get("raw")
        market_cap = (price_mod.get("marketCap") or {}).get("raw")
        fifty_two_low = (summary.get("fiftyTwoWeekLow") or {}).get("raw")
        fifty_two_high = (summary.get("fiftyTwoWeekHigh") or {}).get("raw")
        held_insiders_raw = (stats.get("heldPercentInsiders") or {}).get("raw")  # fraction, e.g. 0.12 = 12%
        held_insiders_pct = held_insiders_raw * 100 if held_insiders_raw is not None else None
        profile = result.get("assetProfile") or {}
        sector = profile.get("sector")
        industry = profile.get("industry")
        business_summary = profile.get("longBusinessSummary")
        if business_summary and len(business_summary) > 220:
            business_summary = business_summary[:217].rsplit(" ", 1)[0] + "..."  # trim at a word boundary
        return {
            "history": history,
            "targetMeanPrice": (fin.get("targetMeanPrice") or {}).get("raw"),
            "recommendationKey": fin.get("recommendationKey"),
            "nextEarningsDate": next_earnings,  # unix epoch seconds, or None
            "exDividendDate": ex_div,  # unix epoch seconds, or None
            "dividendRate": div_rate,  # per-share currency amount, or None
            "dividendYieldPct": div_yield_pct,  # already converted to a percentage, or None
            "trailingPE": trailing_pe,
            "trailingEps": eps,
            "marketCap": market_cap,
            "fiftyTwoWeekLow": fifty_two_low,
            "fiftyTwoWeekHigh": fifty_two_high,
            "heldPercentInsidersPct": held_insiders_pct,
            "sector": sector,
            "industry": industry,
            "businessSummary": business_summary,
        }
    except Exception as e:
        print(f"  ! yahoo analyst history failed: {ticker} ({e})", file=sys.stderr)
        return None


FCA_SHORT_INTEREST_URL = "https://www.fca.org.uk/publication/data/short-positions-daily-update.xlsx"


def fetch_short_interest():
    """
    FCA's daily Aggregate Net Short Position file — free, official, no auth.
    T+2 lag (today's figures reflect positions from two working days ago).
    Returns dict: COMPANY NAME (upper, stripped) -> {"pct": float, "position_date": str}
    Fails soft (returns {}) — same pattern as every other fetch_* here.
    """
    try:
        import openpyxl
    except ImportError:
        print("  ! short interest skipped: openpyxl not installed (add to requirements.txt)", file=sys.stderr)
        return {}
    try:
        import io
        raw = http_get(FCA_SHORT_INTEREST_URL, timeout=20)
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active

        header_row_idx, headers = None, {}
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
            for cell in row:
                if cell and "issuer" in str(cell).lower():
                    header_row_idx = i + 1
                    headers = {str(h).strip().lower(): k for k, h in enumerate(row) if h}
                    break
            if header_row_idx:
                break
        if header_row_idx is None:
            print("  ! short interest: header row not found — FCA file format may have changed", file=sys.stderr)
            return {}

        name_col = next((v for k, v in headers.items() if "issuer" in k), None)
        pct_col = next((v for k, v in headers.items() if "%" in k or "percent" in k), None)
        date_col = next((v for k, v in headers.items() if "date" in k), None)
        if name_col is None or pct_col is None:
            print(f"  ! short interest: required columns not found (saw: {list(headers.keys())})", file=sys.stderr)
            return {}

        result = {}
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not row or row[name_col] is None:
                continue
            name = str(row[name_col]).strip().upper()
            try:
                pct = float(row[pct_col])
            except (TypeError, ValueError):
                continue
            result[name] = {
                "pct": pct,
                "position_date": str(row[date_col]) if date_col is not None else None,
            }
        print(f"  short interest: loaded {len(result)} companies from FCA")
        return result
    except Exception as e:
        print(f"  ! short interest fetch failed: {e}", file=sys.stderr)
        return {}


def match_short_interest(company_name, short_interest_map):
    """Loose name match — FCA issuer names rarely match Yahoo's/watchlist's naming exactly."""
    if not company_name or not short_interest_map:
        return None
    clean = company_name.upper().replace(" PLC", "").replace(" ORD", "").strip()
    for fca_name, data in short_interest_map.items():
        clean_fca = fca_name.replace(" PLC", "").replace(" ORD", "").strip()
        if clean in clean_fca or clean_fca in clean:
            return data
    return None


ACTION_CATEGORY = {"up": "upgrade", "down": "downgrade", "init": "event", "main": "news", "reit": "news"}
RECENT_WINDOW_SECONDS = 2 * 24 * 3600  # only alert on ratings from the last 48h, not backfilled history


def analyst_history_to_items(ticker, analyst):
    items = []
    if not analyst:
        return items
    symbol = yahoo_symbol(ticker)
    now = time.time()
    for h in analyst.get("history", [])[:15]:
        firm = h.get("firm", "")
        to_grade = h.get("toGrade", "")
        from_grade = h.get("fromGrade", "")
        action = h.get("action", "")
        epoch = h.get("epochGradeDate")
        if not firm or not to_grade or not epoch:
            continue
        pub_date = datetime.fromtimestamp(epoch, tz=timezone.utc).strftime("%a, %d %b %Y %H:%M:%S GMT")
        title = f"{firm} {'moves' if action=='main' else 'rates'} {ticker} to {to_grade}" + (
            f" (from {from_grade})" if from_grade and from_grade != to_grade else ""
        )
        items.append({
            "title": title,
            "link": f"https://finance.yahoo.com/quote/{symbol}/analysis#{firm}-{epoch}".replace(" ", "-"),
            "pubDate": pub_date,
            "source": "Yahoo Finance Analyst History",
            "category": ACTION_CATEGORY.get(action, "news"),
            "broker": firm,
            "_recent": (now - epoch) <= RECENT_WINDOW_SECONDS,
        })
    return items


import email.utils as _email_utils


LONDON_TZ = ZoneInfo("Europe/London")


def _parse_pub_date(pub_date_str):
    """
    Shared date parser — tries RFC 822 (standard RSS: "Wed, 31 Jul 2024 07:00:00 GMT")
    then Investing.com's plain "YYYY-MM-DD HH:MM:SS" format. Returns a tz-aware
    datetime, or None if genuinely unparseable.
    """
    if not pub_date_str:
        return None
    dt = None
    try:
        dt = _email_utils.parsedate_to_datetime(pub_date_str)
    except Exception:
        pass
    if dt is None:
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
            try:
                dt = datetime.strptime(pub_date_str.strip(), fmt)
                break
            except Exception:
                continue
    if dt is None:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt


def is_recent_enough(pub_date_str, max_age_days=NEWS_MAX_AGE_DAYS):
    """
    Returns True if pub_date_str parses to within max_age_days of now, OR if it can't
    be parsed at all (fail open — an unparseable date shouldn't silently hide an item,
    that's a display bug, not a staleness signal).
    """
    dt = _parse_pub_date(pub_date_str)
    if dt is None:
        return True
    age = datetime.now(timezone.utc) - dt
    return age.days <= max_age_days


def is_today_in_london(pub_date_str):
    """
    Stricter same-day filter, using the real Europe/London calendar date (correctly
    handles the GMT/BST switch via zoneinfo, not just a fixed UTC offset) — 'today'
    means today in London, not just 'within the last N days'. Fails open on an
    unparseable date, same reasoning as is_recent_enough.
    """
    dt = _parse_pub_date(pub_date_str)
    if dt is None:
        return True
    return dt.astimezone(LONDON_TZ).date() == datetime.now(timezone.utc).astimezone(LONDON_TZ).date()


def passes_news_filters(pub_date_str):
    """Combined recency check applied to every news/broker item before it can enter
    the feed, alerts, or dashboard."""
    if not is_recent_enough(pub_date_str):
        return False
    if NEWS_SAME_LONDON_DAY_ONLY and not is_today_in_london(pub_date_str):
        return False
    return True


def item_sort_key(it):
    """Sort key for ordering news/alert items latest-first. NEVER sort on the raw
    pubDate string directly — RFC 822 dates start with the day name ("Thu, 27 Aug...",
    "Fri, 28 Aug..."), so a plain string sort groups by day-of-week alphabetically
    ("Fri" < "Thu") rather than by actual chronological time, silently scrambling the
    real order. Parse to a real datetime first; fall back to detectedAt (stored as
    Python's .isoformat(), a different format _parse_pub_date doesn't cover — handled
    separately here), then epoch 0 (sorts last) for anything genuinely unparseable."""
    dt = _parse_pub_date(it.get("pubDate"))
    if dt is None and it.get("detectedAt"):
        try:
            dt = datetime.fromisoformat(it["detectedAt"])
        except Exception:
            dt = None
    if dt is None:
        return datetime.fromtimestamp(0, tz=timezone.utc)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt


def fetch_feed(url):
    try:
        return parse_rss(http_get(url)), False
    except Exception as e:
        print(f"  ! feed fetch failed: {url} ({e})", file=sys.stderr)
        return [], True


def google_news_url(company_name):
    q = f'{company_name} (LSE OR "London Stock Exchange") (upgrade OR downgrade OR "price target" OR rating)'
    return "https://news.google.com/rss/search?" + urllib.parse.urlencode(
        {"q": q, "hl": "en-GB", "gl": "GB", "ceid": "GB:en"}
    )


def general_news_url(company_name):
    """Broader than google_news_url — not restricted to rating/broker keywords, so it
    catches general company news (earnings, M&A, trading updates) for stocks that show
    up in the screener but aren't on the watchlist."""
    q = f'{company_name} (LSE OR "London Stock Exchange")'
    return "https://news.google.com/rss/search?" + urllib.parse.urlencode(
        {"q": q, "hl": "en-GB", "gl": "GB", "ceid": "GB:en"}
    )


def reuters_bloomberg_url(company_name):
    # Reuters dropped public RSS years ago and Bloomberg is subscription-walled, so
    # neither offers a free feed directly. This scopes the same free Google News search
    # to just those two publishers (site: filter) — legitimate, no scraping either site,
    # same mechanism already used for the general news search.
    q = f'{company_name} (LSE OR "London Stock Exchange") (site:reuters.com OR site:bloomberg.com)'
    return "https://news.google.com/rss/search?" + urllib.parse.urlencode(
        {"q": q, "hl": "en-GB", "gl": "GB", "ceid": "GB:en"}
    )


def market_wide_broker_news_url():
    # NOT scoped to a company name — this is what makes it market-wide rather than
    # watchlist-only. One request per poll cycle covers every LSE-listed company's
    # broker rating news, instead of the (infeasible) approach of adding all ~1,900
    # LSE companies to the watchlist and polling each individually.
    q = '(LSE OR "London Stock Exchange") (upgrade OR downgrade OR "price target" OR "rating") (broker OR analyst OR bank)'
    return "https://news.google.com/rss/search?" + urllib.parse.urlencode(
        {"q": q, "hl": "en-GB", "gl": "GB", "ceid": "GB:en"}
    )


def yahoo_symbol(ticker):
    t = ticker.strip().upper()
    return t if t.endswith(".L") else f"{t.rstrip('.')}.L"


def yahoo_news_url(ticker):
    return "https://feeds.finance.yahoo.com/rss/2.0/headline?" + urllib.parse.urlencode(
        {"s": yahoo_symbol(ticker), "region": "UK", "lang": "en-GB"}
    )


def fetch_ftse100():
    """The overall market headline every major finance site leads with — a single call
    once per cycle (not per-stock), for the ^FTSE index itself."""
    url = "https://query1.finance.yahoo.com/v8/finance/chart/%5EFTSE?interval=1d&range=5d"
    try:
        data = json.loads(http_get(url))
        result = (data.get("chart") or {}).get("result") or [None]
        if not result[0]:
            return None
        meta = result[0].get("meta", {})
        price = meta.get("regularMarketPrice")
        prev_close = meta.get("chartPreviousClose") or meta.get("previousClose")
        if price is None or not prev_close:
            return None
        change = price - prev_close
        change_pct = change / prev_close * 100
        return {"price": price, "change": change, "changePct": change_pct}
    except Exception as e:
        print(f"  ! FTSE 100 fetch failed: {e}", file=sys.stderr)
        return None


def fetch_yahoo_quote(ticker):
    symbol = yahoo_symbol(ticker)
    url = f"https://query1.finance.yahoo.com/v8/finance/chart/{urllib.parse.quote(symbol)}?interval=1m&range=1d"
    try:
        data = json.loads(http_get(url))
        result = (data.get("chart") or {}).get("result") or [None]
        meta = (result[0] or {}).get("meta") if result[0] else None
        if not meta or "regularMarketPrice" not in meta:
            return None
        price = meta["regularMarketPrice"]
        prev_close = meta.get("previousClose") or meta.get("chartPreviousClose")
        change = price - prev_close if prev_close else None
        change_pct = (change / prev_close * 100) if (prev_close and change is not None) else None
        day_high = meta.get("regularMarketDayHigh")
        day_low = meta.get("regularMarketDayLow")
        # Intraday range as % of previous close — a purely descriptive volatility
        # measure (how much the stock has already swung today), not a prediction.
        range_pct = ((day_high - day_low) / prev_close * 100) if (day_high and day_low and prev_close) else None
        return {
            "price": price,
            "currency": meta.get("currency", "GBp"),
            "change": change,
            "changePct": change_pct,
            "dayHigh": day_high,
            "dayLow": day_low,
            "rangePct": range_pct,
            "asOf": meta.get("regularMarketTime", int(time.time())) * 1000,
        }
    except Exception as e:
        print(f"  ! yahoo quote failed: {ticker} ({e})", file=sys.stderr)
        return None


def compute_rsi(closes, period=14):
    """Standard RSI (relative strength index) — a raw computed statistic from real
    closing prices, shown as a number, not labeled 'overbought'/'oversold' or turned
    into a recommendation. That labeling is exactly the kind of implied-advice framing
    this tool avoids; the number itself is just arithmetic on public price history."""
    if len(closes) < period + 1:
        return None
    deltas = [closes[i] - closes[i - 1] for i in range(1, len(closes))]
    recent = deltas[-period:]
    gains = [d for d in recent if d > 0]
    losses = [-d for d in recent if d < 0]
    avg_gain = sum(gains) / period
    avg_loss = sum(losses) / period
    if avg_loss == 0:
        return 100.0
    rs = avg_gain / avg_loss
    return 100 - (100 / (1 + rs))


def fetch_price_technicals(ticker):
    """Real, already-happened price history — 5-day % change, RSI(14), and 20-day
    moving average, all computed from the same single chart fetch (a longer range than
    strictly needed for the 5-day figure, so RSI/MA come along for free rather than
    requiring a second network call). Facts about the past, not predictions."""
    symbol = yahoo_symbol(ticker)
    url = f"https://query1.finance.yahoo.com/v8/finance/chart/{urllib.parse.quote(symbol)}?interval=1d&range=3mo"
    try:
        data = json.loads(http_get(url))
        result = (data.get("chart") or {}).get("result") or [None]
        if not result[0]:
            return None
        closes = (result[0].get("indicators", {}).get("quote", [{}])[0] or {}).get("close") or []
        closes = [c for c in closes if c is not None]  # some days can come back null (holidays etc.)
        if len(closes) < 6:
            return None  # not enough real trading days to compute even the 5-day change
        latest = closes[-1]
        five_days_ago = closes[-6]
        change_pct = (latest - five_days_ago) / five_days_ago * 100 if five_days_ago else None
        rsi14 = compute_rsi(closes, 14)
        ma20 = sum(closes[-20:]) / len(closes[-20:]) if len(closes) >= 20 else None
        return {
            "changePct5d": change_pct,
            "price": latest,
            "rsi14": rsi14,
            "ma20": ma20,
            "aboveMA20": (latest > ma20) if ma20 else None,
        }
    except Exception as e:
        print(f"  ! price technicals fetch failed: {ticker} ({e})", file=sys.stderr)
        return None


BIG_MOVER_THRESHOLD_PCT = 5.0  # purely descriptive flag: "this has already moved a lot today"
UPTREND_5DAY_THRESHOLD_PCT = 5.0  # flagged as "5-day uptrend" once risen at least this much
# Gainers/losers liquidity filter — without this, illiquid penny stocks with meaningless
# tiny-absolute-value swings (0.1p -> 1p = "900%") dominate the list.
MIN_VOLUME_FOR_MOVERS = 500_000
MIN_PRICE_PENCE_FOR_MOVERS = 20  # excludes sub-20p penny stocks specifically from gainers/losers
# News/broker items older than this are considered stale and filtered from the live feed —
# Google News search returns "most relevant," not "most recent," so old syndicated pieces
# (sometimes years old) can otherwise resurface in results.


def fetch_gb_screener(sort_field, sort_type="DESC", count=10):
    """
    Generic LSE-scoped screener via Yahoo Finance's free public endpoint — same source
    already used for top-volume, generalized to any sortable field (dayvolume,
    percentchange, etc). Requires the crumb-authenticated opener; Yahoo 401s this
    endpoint for anonymous requests from cloud/datacenter IPs.

    Yahoo's "region: gb" tag catches thousands of obscure GDRs/depositary receipts
    (symbols ending ".IL" etc) alongside genuine LSE-listed shares — these are thinly
    traded, so their %-change numbers are noise (a move from 0.1p to 1p shows as
    "900%"). Filtered here to only ".L"-suffixed symbols, the standard LSE ticker
    format, and over-fetched (4x) since a chunk gets filtered out.
    """
    crumb = get_yahoo_crumb()
    url = "https://query1.finance.yahoo.com/v1/finance/screener"
    if crumb:
        url += f"?crumb={urllib.parse.quote(crumb)}"
    # Illiquid penny stocks dominate the top of a raw %-change sort by definition —
    # a small over-fetch (e.g. 4x) means genuinely liquid mid/large-caps can be buried
    # past the fetch window entirely, leaving nothing after the liquidity filter runs.
    # Gainers/losers need a much bigger pool to filter from than Volume does.
    fetch_size = count * 50 if sort_field == "percentchange" else count * 4
    fetch_size = min(fetch_size, 250)  # stay within what the endpoint reliably accepts
    body = json.dumps({
        "size": fetch_size,
        "offset": 0,
        "sortField": sort_field,
        "sortType": sort_type,
        "quoteType": "EQUITY",
        "query": {"operator": "eq", "operands": ["region", "gb"]},
    }).encode("utf-8")
    req = urllib.request.Request(
        url, data=body, method="POST",
        headers={**HEADERS, "Content-Type": "application/json"},
    )
    try:
        data = json.loads(_yahoo_opener.open(req, timeout=15).read())
        quotes = (((data.get("finance") or {}).get("result") or [{}])[0]).get("quotes", [])
        out = []
        for q in quotes:
            symbol = q.get("symbol", "")
            if not symbol.upper().endswith(".L"):
                continue  # skip GDRs/IOB instruments etc — keep genuine LSE-listed shares only
            # LSE International Order Book tickers (depositary receipts for FOREIGN
            # companies, not genuine UK-listed businesses) conventionally start with a
            # digit — e.g. "0R9U.L" is PayPal Holdings, confirmed by its news being
            # entirely US-market coverage. Excluding these keeps results to genuine LSE
            # primary listings, matching what "LSE-listed stocks only" actually promises.
            ticker_part = symbol.upper().rsplit(".L", 1)[0]
            if ticker_part[:1].isdigit():
                continue
            volume = q.get("regularMarketVolume")
            price = q.get("regularMarketPrice")
            if sort_field == "dayvolume" and not volume:
                continue  # skip zero/missing-volume noise from the volume ranking specifically
            if sort_field == "percentchange":
                # Gainers/losers dominated by illiquid penny stocks: a move from 0.1p to
                # 1p shows as "900%" and is meaningless. Require real liquidity and a
                # non-penny price so the list reflects stocks actually worth noticing.
                if not volume or volume < MIN_VOLUME_FOR_MOVERS:
                    continue
                if not price or price < MIN_PRICE_PENCE_FOR_MOVERS:
                    continue
            out.append({
                "symbol": symbol,
                "name": q.get("shortName", symbol),
                "volume": volume,
                "price": price,
                "changePct": q.get("regularMarketChangePercent"),
            })
            if len(out) >= count:
                break
        return out
    except Exception as e:
        print(f"  ! screener failed (sortField={sort_field}): {e}", file=sys.stderr)
        return []


def fetch_lse_screener():
    return {
        "volume": fetch_gb_screener("dayvolume", "DESC", 10),
        "gainers": fetch_gb_screener("percentchange", "DESC", 10),
        "losers": fetch_gb_screener("percentchange", "ASC", 10),
    }


def now_stamp():
    """Human-readable London-time stamp for every outgoing message and the dashboard —
    'this is live and this is exactly when' rather than leaving it implicit. Shows
    London time first (what matters for a UK tool) with the London/BST or GMT label
    zoneinfo resolves automatically, plus UTC alongside for technical clarity."""
    now_utc = datetime.now(timezone.utc)
    now_london = now_utc.astimezone(LONDON_TZ)
    return f'{now_london.strftime("%a %d %b %Y, %H:%M")} {now_london.strftime("%Z")} ({now_utc.strftime("%H:%M")} UTC)'


def format_london_and_utc(dt_utc):
    """Same dual London/UTC format as now_stamp(), for a specific stored UTC datetime
    (e.g. lastPoll) rather than 'right now'."""
    if dt_utc.tzinfo is None:
        dt_utc = dt_utc.replace(tzinfo=timezone.utc)
    dt_london = dt_utc.astimezone(LONDON_TZ)
    return f'{dt_london.strftime("%a %d %b %Y, %H:%M")} {dt_london.strftime("%Z")} ({dt_utc.strftime("%H:%M")} UTC)'


def format_screener_sections(screener):
    """
    Returns a list of separate messages (one per section: Volume, Gainers, Losers)
    instead of one combined string. CallMeBot/WhatsApp truncates long messages, and
    all three sections combined easily exceeds that — sending separately means each
    one arrives complete rather than the later sections getting cut off silently.
    """
    def section(title, rows, show_pct=True):
        lines = [f"{title}  🕐 {now_stamp()}"]
        for i, q in enumerate(rows, 1):
            chg = q.get("changePct")
            chg_str = f'{"▲" if (chg or 0) >= 0 else "▼"}{abs(chg or 0):.2f}%' if chg is not None else ""
            vol = q.get("volume")
            vol_str = f"{vol:,}" if isinstance(vol, (int, float)) else "?"
            bit = f"vol {vol_str}" if not show_pct else chg_str
            name = q.get("name") or q.get("symbol", "")
            lines.append(f'{i}. {q.get("symbol","")} ({name}) — {bit}')
        return "\n".join(lines)

    messages = []
    if screener.get("volume"):
        messages.append(section("📊 TOP 10 BY VOLUME (LSE):", screener["volume"], show_pct=False))
    if screener.get("gainers"):
        messages.append(section("📈 TOP 10 GAINERS (LSE):", screener["gainers"]))
    else:
        messages.append(f"📈 TOP 10 GAINERS (LSE):  🕐 {now_stamp()}\nNo liquid (500k+ vol, 20p+) gainers found this run.")
    if screener.get("losers"):
        messages.append(section("📉 TOP 10 LOSERS (LSE):", screener["losers"]))
    else:
        messages.append(f"📉 TOP 10 LOSERS (LSE):  🕐 {now_stamp()}\nNo liquid (500k+ vol, 20p+) losers found this run.")
    return messages


_RATE_LIMIT_WINDOW_SECONDS = 240 * 60  # matches CallMeBot's free-tier window
_RATE_LIMIT_MAX_SENDS = 14  # stay under CallMeBot's 16/240min cap with a safety margin
_recent_send_times = []  # populated from seen_state at the start of each run, see main()


def _current_template():
    return (os.environ.get("WEBHOOK_TEMPLATE", "").strip() or "callmebot")


def _rate_limit_ok():
    """Global send budget shared across every message type this run (and recent prior
    runs) — the per-category throttles above help, but this is the actual backstop that
    guarantees the WhatsApp free-tier cap is never exceeded regardless of which
    combination of screener/alert/heartbeat messages happens to fire in a given hour.
    Only applies to CallMeBot specifically — ntfy and generic webhooks don't share that
    limit, so they're not artificially held back by a cap that doesn't apply to them."""
    if _current_template() != "callmebot":
        return True
    now = time.time()
    while _recent_send_times and now - _recent_send_times[0] > _RATE_LIMIT_WINDOW_SECONDS:
        _recent_send_times.pop(0)
    return len(_recent_send_times) < _RATE_LIMIT_MAX_SENDS


def send_webhook(message):
    if not _rate_limit_ok():
        print(f"  ! send skipped: WhatsApp rate-limit budget exhausted ({_RATE_LIMIT_MAX_SENDS}/{_RATE_LIMIT_WINDOW_SECONDS//60}min) — still on dashboard.", file=sys.stderr)
        return
    webhook_url = os.environ.get("WEBHOOK_URL", "").strip()
    # GitHub passes a missing secret through as an EMPTY STRING env var, not an absent
    # one — so os.environ.get(..., "callmebot") never actually triggers its default in
    # that case. Explicitly fall back after stripping, regardless of why it's empty.
    template = (os.environ.get("WEBHOOK_TEMPLATE", "").strip() or "callmebot")
    if not webhook_url:
        print("  ! WEBHOOK_URL is empty — check the secret exists and is named exactly WEBHOOK_URL", file=sys.stderr)
        return
    print(f"  > sending webhook via template='{template}'")
    _recent_send_times.append(time.time())  # count the attempt itself, regardless of outcome —
                                              # matches CallMeBot's own "message queued" behavior
    try:
        if template == "callmebot":
            sep = "&" if "?" in webhook_url else "?"
            url = f"{webhook_url}{sep}text={urllib.parse.quote(message)}"
            resp = urllib.request.urlopen(urllib.request.Request(url, headers=HEADERS), timeout=15)
            body = resp.read().decode("utf-8", errors="replace")
            print(f"  > callmebot response ({resp.status}): {body[:200]}")
        elif template == "ntfy":
            req = urllib.request.Request(webhook_url, data=message.encode("utf-8"), method="POST", headers=HEADERS)
            resp = urllib.request.urlopen(req, timeout=15)
            print(f"  > ntfy response: {resp.status}")
        else:
            body = json.dumps({"text": message, "message": message}).encode("utf-8")
            req = urllib.request.Request(
                webhook_url, data=body, method="POST",
                headers={**HEADERS, "Content-Type": "application/json"},
            )
            resp = urllib.request.urlopen(req, timeout=15)
            print(f"  > generic webhook response: {resp.status}")
    except Exception as e:
        print(f"  ! webhook send failed: {e}", file=sys.stderr)


def load_json(path, default):
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return default


def save_json(path, data):
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def format_market_cap(value):
    """Readable market cap — e.g. 1_234_000_000 -> '£1.23bn'. Yahoo returns this in the
    stock's own currency (GBp for most LSE stocks, so this is already pence-scale — but
    marketCap itself is typically reported in full currency units, not pence, by Yahoo)."""
    if value is None:
        return None
    if value >= 1_000_000_000:
        return f"£{value / 1_000_000_000:.2f}bn"
    if value >= 1_000_000:
        return f"£{value / 1_000_000:.1f}m"
    return f"£{value:,.0f}"


def scale_bar_html(label, value_display, position_pct, lo_label, hi_label, color_lo, color_hi, zone_lo=33, zone_hi=66):
    """
    A compact positional scale bar: shows WHERE a number sits on a labelled range,
    never whether that's "good" or "bad" — no buy/sell framing, no green-means-go
    styling. The marker's position is the only signal; colour is purely to distinguish
    the two ends of the scale (e.g. small-cap vs large-cap), not a verdict.
    position_pct: 0-100, already clamped by the caller.
    zone_lo/zone_hi: where the colour bands split (default even thirds, middle band neutral grey).
    """
    pos = max(0, min(100, position_pct))
    return (
        f'<div style="margin:5px 0 2px;max-width:260px;">'
        f'<div style="display:flex;justify-content:space-between;font-size:11px;color:#9aa0a6;margin-bottom:2px;">'
        f'<span>{esc_safe(label)}</span><span style="color:#e8eaed;font-weight:700;">{esc_safe(value_display)}</span></div>'
        f'<div style="position:relative;height:5px;border-radius:3px;'
        f'background:linear-gradient(to right, {color_lo} 0%, {color_lo} {zone_lo}%, '
        f'#2a2e37 {zone_lo}%, #2a2e37 {zone_hi}%, {color_hi} {zone_hi}%, {color_hi} 100%);">'
        f'<div style="position:absolute;left:{pos:.0f}%;top:-2px;width:2px;height:9px;'
        f'background:#e8eaed;border-radius:1px;"></div></div>'
        f'<div style="display:flex;justify-content:space-between;font-size:9px;color:#6b7078;margin-top:1px;">'
        f'<span>{esc_safe(lo_label)}</span><span>{esc_safe(hi_label)}</span></div></div>'
    )


def esc_safe(s):
    """Small standalone escaper — scale_bar_html is called from inside render_dashboard's
    nested functions where the main esc() closure isn't in scope, so this avoids relying
    on closure capture across function boundaries."""
    return (str(s) if s is not None else "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def pe_scale(pe):
    """P/E scale: 0-30 range, no 'good/bad' framing — a P/E can be low because a stock
    is cheap OR because something's wrong; can be high because of growth optimism OR
    overpricing. The bar shows POSITION only, never a verdict."""
    if pe is None or pe < 0:
        return ""
    pos = min(pe / 30 * 100, 100)
    return scale_bar_html("P/E", f"{pe:.1f}", pos, "low", "high", "#7fb3ff", "#e0d267")


def rsi_scale(rsi):
    """RSI is already 0-100 natively — no transform needed. Labels describe the fact
    (risen/fallen fast) never a recommendation (never 'overbought'/'oversold')."""
    if rsi is None:
        return ""
    return scale_bar_html("RSI (14)", f"{rsi:.1f}", rsi, "0 (fallen fast)", "100 (risen fast)", "#7fb3ff", "#e0d267")


def mktcap_scale(mcap_value):
    """Log-scale position: £1m -> 0%, £10bn -> 100%. Purely a size classification
    (micro/mid/large-cap), not a quality judgement — small isn't 'bad', it's smaller."""
    if not mcap_value or mcap_value <= 0:
        return ""
    import math
    log_val = math.log10(max(mcap_value, 1))
    pos = (log_val - 6) / (10 - 6) * 100  # 10^6 (£1m) to 10^10 (£10bn)
    pos = max(0, min(100, pos))
    display = format_market_cap(mcap_value) or ""
    return scale_bar_html("Mkt cap", display, pos, "micro-cap", "large-cap", "#f0997b", "#5dcaa5")


def eps_scale(eps):
    """EPS scale: -£0.50 to £2.00 — a company can be a legitimate early-stage business
    with negative EPS, so 'loss-making' is a factual label, not a red flag icon."""
    if eps is None:
        return ""
    pos = (eps - (-0.5)) / (2.0 - (-0.5)) * 100
    pos = max(0, min(100, pos))
    return scale_bar_html("EPS", f"£{eps:.2f}", pos, "loss-making", "strong profit/share", "#f0997b", "#5dcaa5")


def format_epoch_date(epoch_seconds):
    """Formats a Yahoo epoch timestamp (earnings/ex-dividend dates) as a plain London
    date. IMPORTANT: only returns a value if the date is today or in the future — Yahoo
    sometimes returns a stale PAST date for "next earnings"/"ex-dividend" when it has no
    genuine upcoming one (confirmed against real data: a "next earnings" date that had
    already passed, an "ex-dividend" date from 1996). Showing a past date labeled "next"
    would be actively misleading, so it's treated the same as missing data — omitted."""
    if not epoch_seconds:
        return None
    try:
        dt = datetime.fromtimestamp(epoch_seconds, tz=timezone.utc).astimezone(LONDON_TZ)
        if dt.date() < datetime.now(timezone.utc).astimezone(LONDON_TZ).date():
            return None  # stale/past date — not genuinely "next" anything
        return dt.strftime("%d %b %Y")
    except Exception:
        return None


def render_dashboard(data, watchlist):
    items_by_ticker = data.get("items", {})
    quotes = data.get("quotes", {})
    screener = data.get("screener", {})
    ftse100 = data.get("ftse100")
    screener_news = data.get("screenerNews", {})
    uptrend_stocks = data.get("uptrendStocks", [])
    big_movers = data.get("bigMovers", [])
    market_wide = data.get("marketWide", [])
    last_poll_raw = data.get("lastPoll")
    if last_poll_raw:
        try:
            last_poll = format_london_and_utc(datetime.strptime(last_poll_raw, "%Y-%m-%d %H:%M:%S"))
        except Exception:
            last_poll = last_poll_raw  # fall back to raw string if format ever changes
    else:
        last_poll = "never"
    all_items = []
    for ticker, its in items_by_ticker.items():
        all_items.extend(its)
    all_items.sort(key=item_sort_key, reverse=True)

    def esc(s):
        return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    ftse_html = ""
    if ftse100:
        chg = ftse100.get("changePct") or 0
        cls = "up" if chg >= 0 else "down"
        arrow = "▲" if chg >= 0 else "▼"
        ftse_html = (
            f'<p style="font-size:14px;margin:0 0 10px;">FTSE 100: '
            f'<span class="{cls}">{ftse100.get("price",""):,.2f} {arrow}{abs(chg):.2f}%</span></p>'
        )

    def quote_div(t, q):
        chg_cls = "up" if (q.get("change") or 0) >= 0 else "down"
        chg_arrow = "▲" if (q.get("change") or 0) >= 0 else "▼"
        target = q.get("targetMeanPrice")
        rec = q.get("recommendationKey")
        extra = ""
        if target or rec:
            extra = f'<div class="meta">consensus: {esc(rec or "?")} · target {target if target else "?"}</div>'
        return (
            f'<div class="q"><b>{esc(t)}</b> '
            f'<span class="{chg_cls}">{q.get("price", "?")}{"p" if q.get("currency") == "GBp" else ""} '
            f'{chg_arrow}{abs(q.get("changePct") or 0):.2f}%</span>{extra}</div>'
        )

    quote_rows = "".join(quote_div(t, q) for t, q in quotes.items())

    def item_div(it):
        # Defense in depth: only render a broker badge for genuine rating/target items,
        # even if something upstream ever slipped through — a "news" item mentioning a
        # bank's name is not the same as that bank issuing a rating.
        show_broker = it.get("broker") and it.get("category") in ("upgrade", "downgrade", "target")
        broker_html = f'<span class="broker">{esc(it["broker"])}</span>' if show_broker else ""
        ticker_label = "LSE" if it.get("ticker") == "MARKET" else esc(it.get("ticker", ""))
        return (
            f'<div class="item"><span class="badge {it.get("category","news")}">{it.get("category","news").upper()}</span> '
            f'<b>{ticker_label}</b> '
            f'{broker_html} '
            f'<span class="meta">{esc(it.get("source",""))} · {esc(it.get("pubDate",""))}</span><br/>'
            f'<a href="{esc(it.get("link","#"))}" target="_blank">{esc(it.get("title",""))}</a></div>'
        )

    item_rows = "".join(item_div(it) for it in all_items[:150])
    market_wide_rows = "".join(item_div(it) for it in market_wide[:60])

    def screener_table(rows, show_pct=True):
        def row(i, q):
            vol = q.get("volume")
            vol_str = f"{vol:,}" if isinstance(vol, (int, float)) else "?"
            chg = q.get("changePct") or 0
            chg_cls = "up" if chg >= 0 else "down"
            chg_str = f'{"▲" if chg >= 0 else "▼"}{abs(chg):.2f}%'
            last_col = chg_str if show_pct else vol_str
            last_cls = f' class="{chg_cls}"' if show_pct else ""
            symbol = q.get("symbol", "")
            name = esc(q.get("name") or symbol)
            # Inline news link — the same lookup already built for the "News on Today's
            # Top Movers" section, shown right where you'd actually look for it: next to
            # the stock itself, not just in a separate section further down the page.
            news_for_symbol = screener_news.get(symbol) or []
            if news_for_symbol:
                top = news_for_symbol[0]
                extra = f" (+{len(news_for_symbol)-1} more)" if len(news_for_symbol) > 1 else ""
                news_html = (
                    f'<br/><a href="{esc(top.get("link","#"))}" target="_blank" '
                    f'style="color:#7fb3ff;font-size:12px;font-weight:600;">📰 {esc(top.get("title",""))}</a>'
                    f'<span class="meta">{extra}</span>'
                )
            else:
                news_html = ""
            target = q.get("targetMeanPrice")
            rec = q.get("recommendationKey")
            target_html = (
                f'<br/><span class="meta">🎯 target <span class="val">{target:.2f}</span>{f" · {esc(rec)}" if rec else ""}</span>'
                if target else ""
            )
            earnings_date = format_epoch_date(q.get("nextEarningsDate"))
            ex_div_date = format_epoch_date(q.get("exDividendDate"))
            div_rate = q.get("dividendRate")
            div_yield = q.get("dividendYieldPct")
            calendar_html = ""
            if earnings_date:
                calendar_html += f'<br/><span class="meta">📅 next earnings: <span class="val">{earnings_date}</span></span>'
            if ex_div_date:
                calendar_html += f'<br/><span class="meta">💰 ex-dividend: <span class="val">{ex_div_date}</span></span>'
            if div_rate is not None or div_yield is not None:
                rate_str = f'<span class="val">{div_rate:.2f}</span>/share' if div_rate is not None else ''
                yield_str = f'<span class="val">{div_yield:.2f}%</span> yield' if div_yield is not None else ''
                joined = " · ".join(s for s in (rate_str, yield_str) if s)
                calendar_html += f'<br/><span class="meta">💷 dividend: {joined}</span>'
            pe = q.get("trailingPE")
            eps = q.get("trailingEps")
            mcap = format_market_cap(q.get("marketCap"))
            fundamentals_parts = []
            if pe is not None:
                fundamentals_parts.append(f'P/E <span class="val">{pe:.1f}</span>')
            if eps is not None:
                fundamentals_parts.append(f'EPS <span class="val">{eps:.2f}</span>')
            if mcap:
                fundamentals_parts.append(f'mkt cap <span class="val">{mcap}</span>')
            if fundamentals_parts:
                calendar_html += f'<br/><span class="meta">📊 {" · ".join(fundamentals_parts)}</span>'
            scales_html = pe_scale(pe) + mktcap_scale(q.get("marketCap")) + eps_scale(eps)
            if scales_html:
                calendar_html += f'<div style="margin-top:4px;">{scales_html}</div>'
            wk_low = q.get("fiftyTwoWeekLow")
            wk_high = q.get("fiftyTwoWeekHigh")
            if wk_low is not None and wk_high is not None:
                calendar_html += f'<br/><span class="meta">📏 52-wk range: <span class="val">{wk_low:.2f}</span> – <span class="val">{wk_high:.2f}</span></span>'
            insiders = q.get("heldPercentInsidersPct")
            if insiders is not None:
                calendar_html += f'<br/><span class="meta">🧑‍💼 insider ownership: <span class="val">{insiders:.1f}%</span></span>'
            short_pct = q.get("shortInterestPct")
            if short_pct is not None:
                calendar_html += f'<br/><span class="meta">📉 short interest: <span class="val">{short_pct:.2f}%</span> (FCA)</span>'
            sector = q.get("sector")
            industry = q.get("industry")
            if sector or industry:
                sector_bit = " · ".join(s for s in (sector, industry) if s)
                calendar_html += f'<br/><span class="meta">🏷️ <span class="val">{esc(sector_bit)}</span></span>'
            biz_summary = q.get("businessSummary")
            if biz_summary:
                calendar_html += f'<br/><span class="meta" style="display:block;max-width:520px;">ℹ️ {esc(biz_summary)}</span>'
            rsi14 = q.get("rsi14")
            above_ma = q.get("aboveMA20")
            technicals_html = ""
            if rsi14 is not None:
                technicals_html += f'<br/><span class="meta">RSI(14): <span class="val">{rsi14:.1f}</span></span>'
                technicals_html += f'<div style="margin-top:4px;">{rsi_scale(rsi14)}</div>'
            if above_ma is not None:
                technicals_html += f'<br/><span class="meta">Price is <span class="val">{"above" if above_ma else "below"}</span> its 20-day average</span>'
            return (
                f'<tr><td>{i+1}</td><td><b style="font-size:14px;">{esc(symbol)}</b><br/>'
                f'<span class="meta">{name}</span>{news_html}{target_html}{calendar_html}{technicals_html}</td><td{last_cls}>{last_col}</td></tr>'
            )
        return "".join(row(i, q) for i, q in enumerate(rows)) or '<tr><td colspan="3" class="meta">No data yet</td></tr>'

    vol_rows = screener_table(screener.get("volume", []), show_pct=False)
    gain_rows = screener_table(screener.get("gainers", []))
    lose_rows = screener_table(screener.get("losers", []))

    def screener_news_item(symbol, it):
        broker_html = f'<span class="broker">{esc(it["broker"])}</span>' if it.get("broker") and it.get("category") in ("upgrade", "downgrade", "target") else ""
        return (
            f'<div class="item"><span class="badge {it.get("category","news")}">{it.get("category","news").upper()}</span> '
            f'<b>{esc(symbol)}</b> {broker_html} '
            f'<span class="meta">{esc(it.get("source",""))} · {esc(it.get("pubDate",""))}</span><br/>'
            f'<a href="{esc(it.get("link","#"))}" target="_blank">{esc(it.get("title",""))}</a></div>'
        )

    screener_news_rows = "".join(
        screener_news_item(symbol, it)
        for symbol, items in screener_news.items()
        for it in items
    )

    uptrend_rows = "".join(
        f'<div class="quote-row"><b>{esc(s["symbol"])}</b> ({esc(s["name"])}) — '
        f'<span style="color:#2fbf71">▲{s["changePct5d"]:.1f}%</span> over 5 sessions</div>'
        for s in sorted(uptrend_stocks, key=lambda x: -x["changePct5d"])
    )

    # Every screener-ranked stock (Volume/Gainers/Losers) that has a real broker target
    # price attached — pulled together here as its own scannable list, in addition to
    # already showing inline under each screener row.
    all_screener_rows = (
        screener.get("volume", []) + screener.get("gainers", []) + screener.get("losers", [])
    )
    seen_target_symbols = set()
    target_price_rows = ""
    for q in all_screener_rows:
        symbol = q.get("symbol", "")
        target = q.get("targetMeanPrice")
        if not target or symbol in seen_target_symbols:
            continue
        seen_target_symbols.add(symbol)
        rec = q.get("recommendationKey")
        rec_html = f' · <span class="meta">{esc(rec)}</span>' if rec else ""
        target_price_rows += (
            f'<div class="quote-row"><b>{esc(symbol)}</b> ({esc(q.get("name") or symbol)}) — '
            f'🎯 target {target:.2f}{rec_html}</div>'
        )

    def heatmap_cell(q):
        chg = q.get("changePct") or 0
        symbol = esc(q.get("symbol", ""))
        name = esc(q.get("name") or symbol)
        # Colour by direction, intensity (lightness) by magnitude of the move —
        # bigger swings render darker/more saturated, capped at a 10% move.
        magnitude = min(abs(chg), 10) / 10
        lightness = 55 - (magnitude * 30)  # 55% (small move) down to 25% (big move)
        hue = 142 if chg >= 0 else 4  # green / red
        bg = f"hsl({hue}, 55%, {lightness:.0f}%)"
        return (
            f'<div class="heat-cell" style="background:{bg}" title="{name}">'
            f'<div class="heat-symbol">{symbol}</div>'
            f'<div class="heat-pct">{"▲" if chg >= 0 else "▼"}{abs(chg):.1f}%</div>'
            f"</div>"
        )

    heatmap_pool = (screener.get("gainers", []) + screener.get("losers", []))
    heatmap_pool.sort(key=lambda q: abs(q.get("changePct") or 0), reverse=True)
    heatmap_cells = "".join(heatmap_cell(q) for q in heatmap_pool[:20])

    mover_rows = "".join(
        f'<div class="q"><b>{esc(m["ticker"])}</b> '
        f'<span class="{"up" if (m.get("changePct") or 0) >= 0 else "down"}">'
        f'{"▲" if (m.get("changePct") or 0) >= 0 else "▼"}{abs(m.get("changePct") or 0):.2f}%</span></div>'
        for m in big_movers
    )

    html = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"/><meta name="viewport" content="width=device-width, initial-scale=1"/>
<meta http-equiv="refresh" content="90">
<title>UK Stock Watch</title>
<style>
body{{background:#0f1115;color:#e8eaed;font-family:-apple-system,sans-serif;margin:0;padding:12px;font-size:14px}}
h1{{font-size:20px;margin:4px 0;font-weight:800}}
h2{{font-size:17px;margin:22px 0 8px;font-weight:800;border-left:4px solid #7fb3ff;padding-left:10px}}
h3{{font-size:13px;margin:0 0 6px;color:#c2c7d0;font-weight:700}}
.screener-grid{{display:grid;grid-template-columns:1fr;gap:10px}}
@media(min-width:600px){{.screener-grid{{grid-template-columns:1fr 1fr 1fr}}}}
.heatmap-grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:4px;margin-bottom:16px}}
@media(min-width:600px){{.heatmap-grid{{grid-template-columns:repeat(8,1fr)}}}}
.heat-cell{{border-radius:4px;padding:8px 4px;text-align:center;color:#fff}}
.heat-symbol{{font-size:11px;font-weight:700}}
.heat-pct{{font-size:10px;opacity:0.9}}
.disclaimer{{background:#1c2b25;border:1px solid #274235;color:#9aa0a6;border-radius:6px;padding:8px;font-size:11px;margin-bottom:10px}}
.quotes{{display:flex;flex-wrap:wrap;gap:8px;margin-bottom:14px}}
.q{{background:#171a21;border:1px solid #2a2e37;border-radius:6px;padding:6px 10px;font-size:12px}}
.up{{color:#50dc96;font-weight:800;font-size:14px}} .down{{color:#ff6b6b;font-weight:800;font-size:14px}}
table{{width:100%;border-collapse:collapse;margin-bottom:16px;font-size:13px}}
table td, table th{{padding:7px 8px;border-bottom:1px solid #2a2e37;text-align:left}}
.item{{background:#171a21;border:1px solid #2a2e37;border-radius:8px;padding:10px;margin-bottom:8px}}
.item a{{color:#e8eaed;text-decoration:none;font-size:14px;font-weight:600}}
.item a:hover{{text-decoration:underline}}
.meta{{color:#9aa0a6;font-size:12px;line-height:1.9}}
.val{{color:#e8eaed;font-weight:700}}
.badge{{border-radius:4px;padding:1px 6px;font-size:10px;font-weight:700;margin-right:4px}}
.badge.upgrade{{background:#163a2a;color:#50dc96}}
.badge.downgrade{{background:#3a1919;color:#ff6b6b}}
.badge.target{{background:#2a2a17;color:#e0d267}}
.badge.director_dealing{{background:#1a2a3a;color:#7fb3ff}}
.badge.event{{background:#1c2a3a;color:#6ab6ff}}
.badge.news{{background:#22262f;color:#9aa0a6}}
.broker{{background:#2a1c3a;color:#c69bf0;border-radius:4px;padding:1px 6px;font-size:10px;font-weight:700;margin-right:4px}}
.lastpoll{{color:#9aa0a6;font-size:11px;text-align:right}}
</style></head>
<body>
<h1>UK Stock Watch — Live Feed</h1>
<p class="lastpoll" style="text-align:left;font-size:13px;color:#50dc96;margin:0 0 10px;">🕐 Live as of {esc(str(last_poll))} — this page auto-refreshes every 90s (data itself updates every ~5 min)</p>
{ftse_html}
<p class="disclaimer">LSE-listed stocks only. Informational only — not investment advice, not a guarantee of any outcome.</p>

<nav style="margin:14px 0;padding:10px;background:#161920;border-radius:6px;font-size:13px;line-height:2.2;">
<b style="color:#9aa0a6;margin-right:8px;">Jump to:</b>
<a href="#heatmap" style="color:#7fb3ff;margin-right:14px;text-decoration:none;">🗺️ Heat Map</a>
<a href="#screener" style="color:#7fb3ff;margin-right:14px;text-decoration:none;">📊 Screener</a>
<a href="#mover-news" style="color:#7fb3ff;margin-right:14px;text-decoration:none;">📰 Mover News</a>
<a href="#uptrend" style="color:#7fb3ff;margin-right:14px;text-decoration:none;">📈 5-Day Uptrend</a>
<a href="#targets" style="color:#7fb3ff;margin-right:14px;text-decoration:none;">🎯 Target Prices</a>
<a href="#movers-today" style="color:#7fb3ff;margin-right:14px;text-decoration:none;">🔥 Moving Today</a>
<a href="#watchlist" style="color:#7fb3ff;margin-right:14px;text-decoration:none;">👀 Watchlist</a>
<a href="#broker-alerts" style="color:#7fb3ff;margin-right:14px;text-decoration:none;">⬆⬇🎯 Broker Alerts</a>
<a href="#news-feed" style="color:#7fb3ff;text-decoration:none;">📰 News Feed</a>
</nav>

<h2 id="heatmap">🗺️ Heat Map (top movers, by size of move)</h2>
<div class="heatmap-grid">{heatmap_cells or '<span class="meta">No data yet</span>'}</div>

<h2 id="screener">📊 LSE Screener (Volume / Gainers / Losers)</h2>
<div class="screener-grid">
  <div><h3>Top Volume</h3><table><tr><th>#</th><th>Symbol</th><th>Volume</th></tr>{vol_rows}</table></div>
  <div><h3>Top Gainers</h3><table><tr><th>#</th><th>Symbol</th><th>Chg%</th></tr>{gain_rows}</table></div>
  <div><h3>Top Losers</h3><table><tr><th>#</th><th>Symbol</th><th>Chg%</th></tr>{lose_rows}</table></div>
</div>

<h2 id="mover-news">📰 News on Today's Top Movers</h2>
<p class="meta">Real, dated-today news for any stock currently in Volume/Gainers/Losers above — not limited to your watchlist.</p>
<div>{screener_news_rows or '<span class="meta">No same-day news found for today&#39;s ranked stocks yet.</span>'}</div>

<h2 id="uptrend">📈 5-Day Uptrend ({UPTREND_5DAY_THRESHOLD_PCT:.0f}%+, screener + watchlist)</h2>
<p class="meta">Real closing-price history over the last 5 trading days — a fact about the past, not a forecast of what happens next.</p>
<div class="quotes">{uptrend_rows or '<span class="meta">Nothing has met the 5-day threshold right now</span>'}</div>

<h2 id="targets">🎯 Broker Target Prices (screener stocks)</h2>
<p class="meta">Real, already-published broker consensus targets from Yahoo's aggregation — not generated by this tool.</p>
<div class="quotes">{target_price_rows or '<span class="meta">No target price data available for today&#39;s ranked stocks yet.</span>'}</div>

<h2 id="movers-today">🔥 Already Moving Today (watchlist, ±{BIG_MOVER_THRESHOLD_PCT:.0f}%+)</h2>
<p class="meta">A fact about what already happened today — not a forecast of what happens next.</p>
<div class="quotes">{mover_rows or '<span class="meta">Nothing past the threshold right now</span>'}</div>

<h2 id="watchlist">👀 Your Watchlist</h2>
<div class="quotes">{quote_rows or '<span class="meta">No quotes yet</span>'}</div>

<h2 id="broker-alerts">⬆⬇🎯 Market-wide Broker Alerts (all LSE, not just watchlist)</h2>
<p class="meta">Upgrades/downgrades from anywhere on the LSE, not limited to your watchlist below.</p>
{market_wide_rows or '<p class="meta">No market-wide alerts yet.</p>'}

<h2 id="news-feed">📰 News &amp; Broker Feed (watchlist)</h2>
{item_rows or '<p class="meta">No items yet — first run may still be in progress.</p>'}
<p class="lastpoll">Last checked: {esc(str(last_poll))}</p>
</body></html>"""
    os.makedirs(DOCS_DIR, exist_ok=True)
    with open(os.path.join(DOCS_DIR, DOCS_FILENAME), "w", encoding="utf-8") as f:
        f.write(html)


HEARTBEAT_INTERVAL_SECONDS = 6 * 3600  # confirmation ping every 6h, not every 5 min — a signal, not spam


AI_DIGEST_INTERVAL_SECONDS = 6 * 3600  # throttled like the heartbeat — a paid API call, not free

# CallMeBot's free WhatsApp tier allows ~16 messages per 240 minutes (~4/hour sustained).
# Sending the screener report every 5-minute cycle would blow through that in half an hour
# regardless of anything else — throttle the WhatsApp SEND specifically (the dashboard/
# data itself still refreshes every cycle; only the push notification is rationed).
SCREENER_MESSAGE_INTERVAL_SECONDS = 60 * 60  # once per hour
MAX_ALERTS_PER_RUN = 5  # cap a burst of alerts in one run; rest are still on the dashboard
# A stock is flagged as "getting attention today" purely on mention COUNT — how many
# already-published, real items exist about it today. This is not a signal about where
# the price might go; it's a fact about today's coverage volume, same category as the
# screener's volume ranking.
ATTENTION_MENTION_THRESHOLD = 3


def format_uptrend_message(uptrend_stocks):
    """Purely descriptive: stocks that have genuinely risen UPTREND_5DAY_THRESHOLD_PCT%+
    over the last 5 real trading days, computed from actual closing prices. This is a
    fact about the past — explicitly not a prediction the rise continues."""
    if not uptrend_stocks:
        return None
    flagged = sorted(uptrend_stocks, key=lambda x: -x["changePct5d"])
    lines = [
        f"📈 5-DAY UPTREND ({UPTREND_5DAY_THRESHOLD_PCT:.0f}%+, screener + watchlist)  🕐 {now_stamp()}",
        "Real closing-price history over the last 5 trading days — not a forecast of what happens next.",
    ]
    for s in flagged[:15]:
        lines.append(f'{s["symbol"]} ({s["name"]}) — ▲{s["changePct5d"]:.1f}% over 5 sessions')
    return "\n".join(lines)


def format_screener_news_message(screener_news):
    """Real, dated-today news for stocks currently ranked in Volume/Gainers/Losers —
    not limited to the watchlist. Capped per message to stay a reasonable length."""
    lines = []
    for symbol, items in screener_news.items():
        for it in items[:2]:  # at most 2 headlines per stock in the message (full list is on the dashboard)
            lines.append(f'{symbol}: {it["title"]}')
    if not lines:
        return None
    header = f"📰 NEWS ON TODAY'S TOP MOVERS  🕐 {now_stamp()}"
    return header + "\n" + "\n".join(lines[:15])  # overall cap so one message can't run away


def format_attention_message(mention_counts):
    """Purely descriptive: which watchlist stocks have unusually high real news/broker
    mention counts today. Never says why that might matter for price — just the count."""
    flagged = sorted(
        ((t, d) for t, d in mention_counts.items() if d["count"] >= ATTENTION_MENTION_THRESHOLD),
        key=lambda x: -x[1]["count"],
    )
    if not flagged:
        return None
    lines = [f"📢 GETTING ATTENTION TODAY (watchlist, {ATTENTION_MENTION_THRESHOLD}+ mentions)  🕐 {now_stamp()}"]
    lines.append("A count of today's coverage — not a signal about where price might go.")
    for ticker, d in flagged:
        lines.append(f"{ticker} ({d['name']}) — {d['count']} mentions today")
    return "\n".join(lines)
AI_DIGEST_MODEL = "claude-haiku-4-5-20251001"  # cheap/fast model, appropriate for a short summary

AI_DIGEST_SYSTEM_PROMPT = """You are a strictly factual summarizer for a UK stock market news digest sent to one person's WhatsApp.

CRITICAL RULES — violating any of these makes your output unusable:
- NEVER recommend buying, selling, or holding any stock, in any form or phrasing.
- NEVER predict future price movements or say what "might" or "could" happen next.
- NEVER use directive/advisory language: "should", "consider", "opportunity", "worth watching for a play", "good time to", etc.
- ONLY describe what has already happened: which broker rated which stock how, what news broke, what moved and by how much, per the data given to you.
- Attribute every claim to its source (the broker name, or "news reports") — never state something as your own conclusion.
- Do not add analysis, speculation, or connect-the-dots reasoning between separate facts.
- Keep it under 120 words, plain prose, no markdown headers.

If you cannot summarize something factually without it reading like advice, state the raw fact more plainly instead of omitting it or softening it into a suggestion."""

# Safety net beyond the system prompt — if the model drifts into advice-shaped language
# despite the instructions, this blocks the message from ever being sent rather than
# trusting the prompt alone.
FORBIDDEN_DIGEST_PATTERNS = [
    r"\byou should\b", r"\bconsider (buying|selling|holding)\b", r"\brecommend(s|ed|ing)?\b",
    r"\bworth (a )?(buy|look|watch)\b", r"\bgood (time|opportunity)\b", r"\bi (suggest|advise)\b",
    r"\bmight (rise|fall|climb|drop)\b", r"\bcould (rise|fall|climb|drop|see)\b",
    r"\bexpect(ed)? to (rise|fall|climb|drop)\b",
]


def generate_ai_digest(context_text):
    """
    Optional, opt-in, paid: summarizes already-gathered facts into a short WhatsApp
    digest. Returns None (silently, no error) if no API key is configured — the rest
    of the tool works identically either way. Never asked to judge or predict, only to
    restate what was already found; output is scanned before sending as a second layer
    of protection against advice-shaped language slipping through.
    """
    api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        return None
    body = json.dumps({
        "model": AI_DIGEST_MODEL,
        "max_tokens": 300,
        "system": AI_DIGEST_SYSTEM_PROMPT,
        "messages": [{"role": "user", "content": f"Today's gathered data:\n{context_text}\n\nWrite the factual summary."}],
    }).encode("utf-8")
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=body,
        headers={"x-api-key": api_key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
        method="POST",
    )
    try:
        resp_data = json.loads(urllib.request.urlopen(req, timeout=30).read())
        text = "".join(b.get("text", "") for b in resp_data.get("content", []) if b.get("type") == "text").strip()
        if not text:
            return None
        if any(re.search(pat, text, re.IGNORECASE) for pat in FORBIDDEN_DIGEST_PATTERNS):
            print("  ! AI digest blocked: output matched an advice-shaped pattern, not sent.", file=sys.stderr)
            return None
        return text
    except Exception as e:
        print(f"  ! AI digest generation failed: {e}", file=sys.stderr)
        return None


def build_digest_context(screener, market_wide_enriched, big_movers, watchlist_alerts):
    """Plain-text summary of this run's facts, handed to the AI as its only input —
    it never sees anything beyond what's already been gathered and shown elsewhere."""
    lines = []
    if screener.get("gainers"):
        lines.append("Top gainers: " + ", ".join(f"{q['symbol']} +{q.get('changePct',0):.1f}%" for q in screener["gainers"][:5]))
    if screener.get("losers"):
        lines.append("Top losers: " + ", ".join(f"{q['symbol']} {q.get('changePct',0):.1f}%" for q in screener["losers"][:5]))
    if market_wide_enriched:
        lines.append("Market-wide broker calls: " + "; ".join(
            f'{it.get("broker","?")} {it["category"]} on {it["title"][:80]}' for it in market_wide_enriched[:8]
        ))
    if big_movers:
        lines.append("Watchlist stocks already moving today: " + ", ".join(
            f"{m['ticker']} {m.get('changePct',0):+.1f}%" for m in big_movers
        ))
    if watchlist_alerts:
        lines.append("Watchlist news/broker alerts: " + "; ".join(
            f'{a.get("ticker","")} {a["category"]}: {a["title"][:80]}' for a in watchlist_alerts[:8]
        ))
    return "\n".join(lines) if lines else "No notable items this cycle."


def maybe_send_ai_digest(seen_state, screener, market_wide_enriched, big_movers, watchlist_alerts):
    if not os.environ.get("ANTHROPIC_API_KEY", "").strip():
        return False  # feature is off — no key configured, no cost, no message
    last_digest = seen_state.get("lastAiDigest", 0)
    if time.time() - last_digest < AI_DIGEST_INTERVAL_SECONDS:
        return False
    context = build_digest_context(screener, market_wide_enriched, big_movers, watchlist_alerts)
    digest = generate_ai_digest(context)
    if digest:
        msg = f"🤖 AI DAILY DIGEST (factual summary only, not advice)  🕐 {now_stamp()}\n\n{digest}"
        print(msg)
        send_webhook(msg)
    seen_state["lastAiDigest"] = time.time()
    return True


def maybe_send_heartbeat(seen_state, watchlist, alert_count, mover_count):
    last_heartbeat = seen_state.get("lastHeartbeat", 0)
    now = time.time()
    if now - last_heartbeat < HEARTBEAT_INTERVAL_SECONDS:
        return False
    ts = datetime.now(timezone.utc).strftime("%H:%M UTC")
    msg = (
        f"✅ UK Stock Watch — still checking. {ts}, watching {len(watchlist)} stocks. "
        f"{alert_count} broker alert(s) and {mover_count} big mover(s) since last heartbeat."
    )
    print(msg)
    send_webhook(msg)
    seen_state["lastHeartbeat"] = now
    return True


def main():
    watchlist = load_json(WATCHLIST_FILE, [])
    if not watchlist:
        print("watchlist.json is empty — nothing to do.")
        return

    seen_state = load_json(SEEN_FILE, {"seen": []})
    seen = set(seen_state.get("seen", []))
    global _recent_send_times
    _recent_send_times = list(seen_state.get("recentSendTimes", []))
    data = load_json(DATA_FILE, {"items": {}, "quotes": {}, "lastPoll": None})
    items_by_ticker = data.get("items", {})
    quotes = data.get("quotes", {})

    yahoo_crumb = get_yahoo_crumb()
    print(f"Yahoo auth crumb: {'obtained' if yahoo_crumb else 'FAILED — screener/analyst-history will likely 401'}")

    new_alerts = []
    mention_counts = {}  # ticker -> {"name": ..., "count": ...} — today's mention volume, purely descriptive
    big_movers = []
    market_wide_enriched = []
    screener = {}
    ratings_items = []
    screener_news = {}
    uptrend_stocks = []
    screener_targets = {}
    ftse100 = fetch_ftse100()

    if not SKIP_MARKET_WIDE:
        ratings_items, _ = fetch_feed(ANALYST_RATINGS_FEED_URL)
        market_wide_items, _ = fetch_feed(market_wide_broker_news_url())
        screener = fetch_lse_screener()

        # News for every stock ranked in Volume/Gainers/Losers, not just the watchlist —
        # deduped by symbol (a stock can appear in more than one list), one query each,
        # staggered to avoid the burst-triggered 503s seen earlier in this project.
        screener_news = {}
        ranked_stocks = {}
        for section in ("volume", "gainers", "losers"):
            for row in screener.get(section, []):
                ranked_stocks[row["symbol"]] = row.get("name", row["symbol"])
        print(f"Fetching news for {len(ranked_stocks)} screener-ranked stocks (volume/gainers/losers)...")
        for symbol, name in ranked_stocks.items():
            items, _ = fetch_feed(general_news_url(name))
            items = [it for it in items if passes_news_filters(it.get("pubDate"))]
            if items:
                now_iso_sc = datetime.now(timezone.utc).isoformat()
                enriched_items = []
                for it in items[:5]:  # cap per-stock to keep dashboard/message size sane
                    category = classify(it["title"])
                    broker = detect_broker(it["title"]) if category in ("upgrade", "downgrade", "target") else None
                    enriched_items.append({**it, "ticker": symbol, "company": name, "category": category, "broker": broker, "detectedAt": now_iso_sc})
                screener_news[symbol] = enriched_items
            time.sleep(1)  # stagger — this is the change most likely to trip Google's rate limiting if rushed

        # 5-day uptrend: real closing-price history for the same deduped stock set (no
        # extra tickers beyond what's already being fetched news for) plus the watchlist.
        uptrend_stocks = []
        uptrend_targets = dict(ranked_stocks)
        for stock in watchlist:
            uptrend_targets.setdefault(stock["ticker"], stock["name"])
        print(f"Checking price technicals, targets, earnings/dividend dates for {len(uptrend_targets)} stocks...")
        short_interest_map = fetch_short_interest()  # one fetch per run, not per ticker
        screener_targets = {}  # symbol -> {targetMeanPrice, recommendationKey, nextEarningsDate, exDividendDate, rsi14, ma20, aboveMA20}
        for symbol, name in uptrend_targets.items():
            hist = fetch_price_technicals(symbol)
            if hist and hist.get("changePct5d") is not None and hist["changePct5d"] >= UPTREND_5DAY_THRESHOLD_PCT:
                uptrend_stocks.append({"symbol": symbol, "name": name, **hist})
            time.sleep(0.3)  # lighter stagger — this hits Yahoo, not Google, different rate-limit budget

            # Real broker price targets + earnings/dividend calendar dates — all from
            # Yahoo's own published data, in the SAME request as before (calendarEvents
            # was added to the existing modules list, not a new call).
            analyst = fetch_yahoo_analyst(symbol)
            if analyst:
                entry = {}
                if analyst.get("targetMeanPrice"):
                    entry["targetMeanPrice"] = analyst["targetMeanPrice"]
                    entry["recommendationKey"] = analyst.get("recommendationKey")
                if analyst.get("nextEarningsDate"):
                    entry["nextEarningsDate"] = analyst["nextEarningsDate"]
                if analyst.get("exDividendDate"):
                    entry["exDividendDate"] = analyst["exDividendDate"]
                if analyst.get("dividendRate") is not None:
                    entry["dividendRate"] = analyst["dividendRate"]
                if analyst.get("dividendYieldPct") is not None:
                    entry["dividendYieldPct"] = analyst["dividendYieldPct"]
                if analyst.get("trailingPE") is not None:
                    entry["trailingPE"] = analyst["trailingPE"]
                if analyst.get("trailingEps") is not None:
                    entry["trailingEps"] = analyst["trailingEps"]
                if analyst.get("marketCap") is not None:
                    entry["marketCap"] = analyst["marketCap"]
                if analyst.get("fiftyTwoWeekLow") is not None:
                    entry["fiftyTwoWeekLow"] = analyst["fiftyTwoWeekLow"]
                if analyst.get("fiftyTwoWeekHigh") is not None:
                    entry["fiftyTwoWeekHigh"] = analyst["fiftyTwoWeekHigh"]
                if analyst.get("heldPercentInsidersPct") is not None:
                    entry["heldPercentInsidersPct"] = analyst["heldPercentInsidersPct"]
                if analyst.get("sector"):
                    entry["sector"] = analyst["sector"]
                if analyst.get("industry"):
                    entry["industry"] = analyst["industry"]
                if analyst.get("businessSummary"):
                    entry["businessSummary"] = analyst["businessSummary"]
                if hist:
                    entry["rsi14"] = hist.get("rsi14")
                    entry["ma20"] = hist.get("ma20")
                    entry["aboveMA20"] = hist.get("aboveMA20")
                si = match_short_interest(name, short_interest_map)
                if si:
                    entry["shortInterestPct"] = si["pct"]
                    entry["shortInterestDate"] = si["position_date"]
                if entry:
                    screener_targets[symbol] = entry
            time.sleep(0.3)

        # Attach the target-price data directly onto each screener row so it displays
        # with the row itself — no separate lookup needed on the dashboard/message side.
        for section in ("volume", "gainers", "losers"):
            for row in screener.get(section, []):
                extra = screener_targets.get(row["symbol"])
                if extra:
                    row.update(extra)  # attach whichever fields were found: target, recommendation, earnings/ex-div dates, RSI/MA

        # Market-wide broker alerts: covers ALL LSE companies for upgrade/downgrade news,
        # not just the watchlist — this is what "all LSE companies" actually needs, without
        # trying to poll ~1,900 individual tickers (which would take hours per cycle and
        # get rate-limited). Combines the unfiltered ratings feed + the market-wide search.
        # general_market_items (investing.com's general "Stock Market News" feed) was
        # removed from this pool entirely, and its fetch removed too — despite being on
        # the uk.investing.com subdomain, real data proved it's not actually UK-scoped
        # (surfaced US broker actions on Workday, Ulta Beauty, Elastic, Autodesk as if
        # they were "all LSE" alerts). Not reused elsewhere either, so keeping the fetch
        # around would just be a wasted request for data that's no longer used anywhere.
        market_wide_pool = ratings_items + market_wide_items
        market_wide_pool = [it for it in market_wide_pool if passes_news_filters(it.get("pubDate"))]
        now_iso_mw = datetime.now(timezone.utc).isoformat()
        for it in market_wide_pool:
            category = classify(it["title"])
            # "target" = a broker raising/cutting their price target — a genuine, already-
            # published broker action, same category of fact as an upgrade/downgrade, so it
            # belongs in the same alert stream rather than being silently dropped.
            if category not in ("upgrade", "downgrade", "target"):
                continue
            market_wide_enriched.append({
                **it,
                "ticker": "MARKET",
                "company": "",
                "category": category,
                "broker": detect_broker(it["title"]),
                "detectedAt": now_iso_mw,
            })
        market_wide_dedup = {}
        for it in market_wide_enriched:
            market_wide_dedup[it["link"]] = it  # de-dupe within this run (same story can appear in both feeds)
        market_wide_enriched = list(market_wide_dedup.values())

        for it in market_wide_enriched:
            if it["link"] in seen:
                continue
            seen.add(it["link"])
            new_alerts.append(it)
    else:
        print("SKIP_MARKET_WIDE set — skipping market-wide search/screener/heatmap (covered by the other job).")

    for stock in watchlist:
        ticker, name = stock["ticker"], stock["name"]
        print(f"Polling {ticker} ({name})...")

        g_items, _ = fetch_feed(google_news_url(name))
        y_items, _ = fetch_feed(yahoo_news_url(ticker))
        rb_items, _ = fetch_feed(reuters_bloomberg_url(name))
        matched_ratings = [it for it in ratings_items if name.lower() in it["title"].lower()]
        combined = g_items + y_items + rb_items + matched_ratings
        combined = [it for it in combined if passes_news_filters(it.get("pubDate"))]
        # Purely a count of real, already-published items mentioning this stock today
        # (deduped by link) — a fact about today's coverage volume, not a prediction of
        # anything. NEWS_SAME_LONDON_DAY_ONLY already restricts `combined` to today.
        mention_links = {it["link"] for it in combined}
        mention_counts[ticker] = {"name": name, "count": len(mention_links)}

        quote = fetch_yahoo_quote(ticker)
        if quote:
            quotes[ticker] = quote
            # Purely descriptive: this stock has already moved a lot today.
            # Not a prediction of further movement, just a factual flag.
            if abs(quote.get("changePct") or 0) >= BIG_MOVER_THRESHOLD_PCT:
                big_movers.append({**quote, "ticker": ticker})

        analyst = fetch_yahoo_analyst(ticker)
        analyst_items = analyst_history_to_items(ticker, analyst)
        if analyst and analyst.get("targetMeanPrice") and ticker in quotes:
            quotes[ticker]["targetMeanPrice"] = analyst["targetMeanPrice"]
            quotes[ticker]["recommendationKey"] = analyst.get("recommendationKey")

        now_iso = datetime.now(timezone.utc).isoformat()
        enriched = []
        for it in combined:
            category = classify(it["title"])
            # Only tag a broker when the item is actually a rating/target call — otherwise
            # a story that merely mentions a bank's name (e.g. a personnel/legal story
            # about "Barclays") gets mislabeled as if that bank issued the rating.
            broker = detect_broker(it["title"]) if category in ("upgrade", "downgrade", "target") else None
            enriched.append({
                **it,
                "ticker": ticker,
                "company": name,
                "category": category,
                "broker": broker,
                "detectedAt": now_iso,
            })
        # Analyst history items already carry structured category/broker/pubDate —
        # merge as-is rather than re-running keyword classification on them.
        for it in analyst_items:
            enriched.append({**it, "ticker": ticker, "company": name, "detectedAt": now_iso})

        existing = items_by_ticker.get(ticker, [])
        merged = enriched + existing
        seen_links = set()
        deduped = []
        for it in merged:
            if it["link"] in seen_links:
                continue
            seen_links.add(it["link"])
            deduped.append(it)
        deduped.sort(key=item_sort_key, reverse=True)
        items_by_ticker[ticker] = deduped[:MAX_ITEMS_PER_TICKER]

        for it in enriched:
            if it["link"] in seen:
                continue
            seen.add(it["link"])
            # Analyst-history items are backfilled (up to 15 past ratings) so first-run
            # dashboard context isn't empty — but only alert on genuinely recent ones,
            # not old history that happens to be new to our "seen" set. For your
            # watchlist specifically, event-category news (mergers, trading updates,
            # profit warnings, results) is genuinely price-moving and worth pushing —
            # unlike the market-wide stream, this is scoped to stocks you actually track.
            if it["category"] in ("upgrade", "downgrade", "target", "director_dealing", "event") and it.get("_recent", True):
                new_alerts.append(it)

        time.sleep(0.5)  # be polite to free sources

    # Merge this run's market-wide items with previously stored ones (same pattern as
    # the per-ticker feed) so the dashboard shows recent history, not just this cycle.
    existing_market_wide = data.get("marketWide", [])
    merged_market_wide = market_wide_enriched + existing_market_wide
    seen_links_mw = set()
    deduped_market_wide = []
    for it in merged_market_wide:
        if it["link"] in seen_links_mw:
            continue
        seen_links_mw.add(it["link"])
        deduped_market_wide.append(it)
    deduped_market_wide.sort(key=item_sort_key, reverse=True)
    deduped_market_wide = deduped_market_wide[:MAX_ITEMS_PER_TICKER]

    data = {
        "items": items_by_ticker,
        "quotes": quotes,
        "screener": screener,
        "ftse100": ftse100,
        "screenerNews": screener_news,
        "uptrendStocks": uptrend_stocks,
        "bigMovers": big_movers,
        "marketWide": deduped_market_wide,
        "lastPoll": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
    }
    save_json(DATA_FILE, data)
    render_dashboard(data, watchlist)

    screener_last_sent = seen_state.get("lastScreenerMessage", 0)
    screener_interval = SCREENER_MESSAGE_INTERVAL_SECONDS if _current_template() == "callmebot" else 0
    if time.time() - screener_last_sent >= screener_interval:
        screener_messages = format_screener_sections(screener)
        for i, msg in enumerate(screener_messages):
            print(msg)
            send_webhook(msg)
            if i < len(screener_messages) - 1:
                time.sleep(2)  # small gap so rapid-fire messages don't collapse/drop
        if screener_messages:
            seen_state["lastScreenerMessage"] = time.time()
    else:
        print("Screener WhatsApp send throttled (sent within the last hour) — dashboard/data still updated above.")

    screener_news_msg = format_screener_news_message(screener_news)
    if screener_news_msg:
        print(screener_news_msg)
        send_webhook(screener_news_msg)

    uptrend_msg = format_uptrend_message(uptrend_stocks)
    if uptrend_msg:
        print(uptrend_msg)
        send_webhook(uptrend_msg)

    attention_msg = format_attention_message(mention_counts)
    if attention_msg:
        print(attention_msg)
        send_webhook(attention_msg)

    if big_movers:
        lines = [f"🔥 ALREADY MOVING TODAY (±{BIG_MOVER_THRESHOLD_PCT:.0f}%+, watchlist)  🕐 {now_stamp()} — facts about today, not a forecast:"]
        for m in big_movers:
            arrow = "▲" if (m.get("changePct") or 0) >= 0 else "▼"
            lines.append(f'{m["ticker"]} {arrow}{abs(m.get("changePct") or 0):.2f}%')
        movers_msg = "\n".join(lines)
        print(movers_msg)
        send_webhook(movers_msg)

    ALERT_LABELS = {"upgrade": "⬆ UPGRADE", "downgrade": "⬇ DOWNGRADE", "target": "🎯 PRICE TARGET", "director_dealing": "🧑‍💼 DIRECTOR DEALING", "event": "📰 NEWS"}
    alert_cap = MAX_ALERTS_PER_RUN if _current_template() == "callmebot" else len(new_alerts)
    alerts_to_send = new_alerts[:alert_cap]
    skipped_count = len(new_alerts) - len(alerts_to_send)
    for alert in alerts_to_send:
        label = ALERT_LABELS.get(alert["category"], "📰 NEWS")
        broker_tag = f' ({alert["broker"]})' if alert["broker"] else ""
        ticker_bit = "LSE (market-wide)" if alert["ticker"] == "MARKET" else alert["ticker"]
        article_date = alert.get("pubDate") or ""
        msg = f'{label}: {ticker_bit}{broker_tag}\n{alert["title"]}\n{alert["link"]}\n📅 {article_date}  🕐 seen {now_stamp()}'
        print(f"Alert: {msg}")
        send_webhook(msg)
        time.sleep(2)  # CallMeBot can silently drop messages sent in rapid succession
                        # even while still returning 200 — space every send out.
    if skipped_count > 0:
        skip_msg = f"ℹ️ {skipped_count} more alert(s) this run — see the full list on the dashboard (capped to avoid WhatsApp rate limits)."
        print(skip_msg)
        send_webhook(skip_msg)

    print(f"Done. {len(new_alerts)} new alert(s) found, {len(alerts_to_send)} sent to WhatsApp.")

    if new_alerts:
        time.sleep(2)  # gap before heartbeat/digest, same reasoning as between alerts

    heartbeat_sent = maybe_send_heartbeat(seen_state, watchlist, len(new_alerts), len(big_movers))
    digest_sent = maybe_send_ai_digest(seen_state, screener, market_wide_enriched, big_movers, new_alerts)

    # Always persist here (not just when heartbeat/digest fired) — this is the only place
    # lastScreenerMessage's updated value actually gets written to disk.
    save_json(SEEN_FILE, {
        "seen": list(seen)[-MAX_SEEN:],
        "lastHeartbeat": seen_state.get("lastHeartbeat", 0),
        "lastAiDigest": seen_state.get("lastAiDigest", 0),
        "lastScreenerMessage": seen_state.get("lastScreenerMessage", 0),
        "recentSendTimes": _recent_send_times,  # carries the rate-limit budget forward to the next run
    })


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        # A totally unhandled crash (vs the per-source try/excepts already inside main)
        # should still tell you something's wrong, not just go silent.
        err_msg = f"⚠️ UK Stock Watch poller crashed: {type(e).__name__}: {e}\nWill retry next scheduled run."
        print(err_msg, file=sys.stderr)
        try:
            send_webhook(err_msg)
        except Exception:
            pass  # if even the failure notification fails, there's nothing more we can do here
        raise  # re-raise so the GitHub Actions run shows red/failed in the Actions tab too
