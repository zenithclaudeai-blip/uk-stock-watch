#!/usr/bin/env python3
"""
UK Stock Watch — cloud poller.

Runs on a GitHub Actions schedule instead of in a browser. Pulls the same free
sources as the browser extension (Google News, Yahoo News, Yahoo Finance quotes,
Investing.com UK's analyst-ratings feed), classifies items, pushes new
upgrade/downgrade alerts to a webhook (WhatsApp via CallMeBot, ntfy, etc.), and
writes a static dashboard page for GitHub Pages.

Stdlib only — no pip installs needed.
"""

import http.cookiejar
import json
import os
import re
import difflib
import sys
import time
import uuid
import urllib.error
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
from html.parser import HTMLParser
from datetime import datetime, timezone, date, timedelta
from zoneinfo import ZoneInfo

STATE_DIR = os.environ.get("STATE_DIR", "state")
DOCS_DIR = os.environ.get("DOCS_DIR", "docs")
DOCS_FILENAME = os.environ.get("DOCS_FILENAME", "index.html")

# Shared CSS - extracted once so the main dashboard and every dedicated
# standalone page (screener.html, heatmap.html, etc.) use the EXACT SAME
# styling from one single source, rather than independent copies that
# could silently drift apart over future edits.
DASHBOARD_CSS = """<style>
body{background:#0f1115;color:#e8eaed;font-family:-apple-system,sans-serif;margin:0 auto;padding:12px;font-size:17px;line-height:1.6;max-width:1200px;word-wrap:break-word;overflow-wrap:break-word}
h1{font-size:26px;margin:4px 0;font-weight:800}
h2{font-size:21px;margin:26px 0 10px;font-weight:800;border-left:4px solid #7fb3ff;padding-left:10px}
h3{font-size:16px;margin:0 0 8px;color:#c2c7d0;font-weight:700}
a:focus-visible,summary:focus-visible,button:focus-visible{outline:2px solid #7fb3ff;outline-offset:2px;border-radius:2px}
.screener-grid{display:grid;grid-template-columns:1fr;gap:10px}
@media(min-width:600px){.screener-grid{grid-template-columns:1fr 1fr 1fr}}
.heatmap-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:4px;margin-bottom:16px}
@media(max-width:380px){.heatmap-grid{grid-template-columns:repeat(3,1fr)}}
@media(min-width:600px){.heatmap-grid{grid-template-columns:repeat(8,1fr)}}
.heat-cell{border-radius:4px;padding:8px 4px;text-align:center;color:#fff}
.heat-symbol{font-size:14px;font-weight:700}
.heat-pct{font-size:13px;opacity:0.9}
.disclaimer{background:#1c2b25;border:1px solid #274235;color:#9aa0a6;border-radius:6px;padding:10px;font-size:14px;margin-bottom:10px}
.quotes{display:flex;flex-wrap:wrap;gap:8px;margin-bottom:14px}
.q{background:#171a21;border:1px solid #2a2e37;border-radius:6px;padding:8px 12px;font-size:15px}
.up{color:#50dc96;font-weight:800;font-size:17px} .down{color:#ff6b6b;font-weight:800;font-size:17px}
.radar-card{background:#171a21;border:1px solid #2a2e37;border-radius:8px;padding:14px 16px;margin-bottom:14px}
.radar-compact-wrap{margin-bottom:10px}
.radar-compact{background:#141821;border-left:3px solid #ffb454;border-radius:6px;padding:10px 14px}
.radar-compact .meta{line-height:1.7;margin:2px 0}
.radar-header{display:flex;flex-wrap:wrap;align-items:baseline;gap:10px;margin-bottom:6px}
.radar-ticker{font-size:20px;font-weight:800;color:#e8eaed}
.radar-freshness{font-size:13px;padding:2px 8px;border-radius:10px;background:#20242d;display:inline-block}
.radar-multi{font-size:12px;font-weight:700;color:#0f1115;background:#7fb3ff;padding:2px 8px;border-radius:10px;display:inline-block}
.radar-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin:10px 0}
.radar-grid-2{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin:10px 0}
.radar-col{background:#12141a;border:1px solid #22262f;border-radius:6px;padding:10px 12px}
.radar-col-label{font-size:13px;font-weight:800;letter-spacing:0.6px;color:#7fb3ff;margin-bottom:6px;text-transform:uppercase}
.radar-col-warn .radar-col-label{color:#e8918a}
.radar-col-target{border-left:3px solid #7fb3ff}
@media (max-width:800px){
  .radar-grid, .radar-grid-2{grid-template-columns:1fr}
  .radar-ticker{font-size:18px}
  nav[aria-label="Section navigation"] a{font-size:14px !important}
}
table{display:block;overflow-x:auto;-webkit-overflow-scrolling:touch;width:100%;border-collapse:collapse;margin-bottom:16px;font-size:16px}
table td, table th{padding:9px 10px;border-bottom:1px solid #2a2e37;text-align:left}
.radar-table-wrap{overflow-x:auto;-webkit-overflow-scrolling:touch;border-radius:8px;border:1px solid #2a2e37;margin-bottom:16px}
nav[aria-label="Section navigation"] a{white-space:nowrap}
.radar-table{display:table;border-collapse:collapse;width:100%;min-width:1100px;margin-bottom:0}
.radar-table thead th{background:#171a21;color:#7fb3ff;text-align:left;padding:12px 14px;font-size:13px;letter-spacing:0.5px;text-transform:uppercase;border-bottom:2px solid #2a2e37;white-space:nowrap;position:sticky;top:0;z-index:1}
.radar-table tbody td{padding:14px;border-bottom:1px solid #22262f;font-size:15px;vertical-align:top}
.radar-table tbody tr:hover{background:#161920}
.radar-table th:first-child, .radar-table td:first-child{position:sticky;left:0;background:#0f1115;z-index:2;border-right:1px solid #2a2e37}
.radar-table thead th:first-child{background:#171a21;z-index:3}
.source-pill{display:inline-block;background:#20242d;color:#9aa0a6;font-size:12px;padding:2px 8px;border-radius:10px;margin:1px 3px 1px 0}
.signal{font-weight:800;font-size:15px;letter-spacing:0.3px}
.signal-strong{color:#50dc96} .signal-mixed{color:#e0b84a} .signal-weak{color:#9aa0a6}
.confidence{font-size:14px;font-weight:700;color:#c7cad1}
.evidence-list{margin:0;padding-left:16px;font-size:13.5px;color:#8fd6b4}
.evidence-list li{margin-bottom:3px}
.dont-chase-badge{display:inline-block;background:#3a1f1f;color:#ff8f87;font-size:11px;font-weight:800;padding:2px 7px;border-radius:4px;letter-spacing:0.4px}
@media (max-width:800px){
  .radar-table{font-size:13px}
}
.item{background:#171a21;border:1px solid #2a2e37;border-radius:8px;padding:12px;margin-bottom:8px}
.item a{color:#e8eaed;text-decoration:none;font-size:17px;font-weight:600}
.item a:hover{text-decoration:underline}
.meta{color:#9aa0a6;font-size:15px;line-height:2.0}
.status-ok{color:#50dc96;font-size:15px;font-weight:600}
.status-warn{color:#f0b429;font-size:15px;font-weight:700}
.status-bad{color:#ff6b6b;font-size:15px;font-weight:700}
.val{color:#e8eaed;font-weight:700}
.badge{border-radius:4px;padding:2px 8px;font-size:12px;font-weight:700;margin-right:4px}
.badge.upgrade{background:#163a2a;color:#50dc96}
.badge.downgrade{background:#3a1919;color:#ff6b6b}
.badge.target{background:#2a2a17;color:#e0d267}
.badge.target_raise{background:#1c3a1c;color:#7bd97b}
.badge.target_cut{background:#3a2317;color:#e0977f}
.badge.initiation{background:#1f2a3a;color:#8fb8ff}
.badge.reiteration{background:#2a2532;color:#b8a0d9}
.badge.director_dealing{background:#1a2a3a;color:#7fb3ff}
.badge.event{background:#1c2a3a;color:#6ab6ff}
.badge.news{background:#22262f;color:#9aa0a6}
.broker{background:#2a1c3a;color:#c69bf0;border-radius:4px;padding:2px 8px;font-size:12px;font-weight:700;margin-right:4px}
.lastpoll{color:#9aa0a6;font-size:11px;text-align:right}
.status-neutral{color:#9aa0a6;font-size:15px;font-weight:600}
</style>"""
WATCHLIST_FILE = os.environ.get("WATCHLIST_FILE", "watchlist.json")
SEEN_FILE = os.path.join(STATE_DIR, "seen.json")
DATA_FILE = os.path.join(STATE_DIR, "data.json")
# When true, this run skips the market-wide broker search, LSE screener, and heat map —
# used by the large-watchlist hourly job so it doesn't duplicate what the fast 5-minute
# job already covers (which would mean duplicate WhatsApp alerts from two separate runs).
SKIP_MARKET_WIDE = os.environ.get("SKIP_MARKET_WIDE", "false").lower() == "true"
MAX_ITEMS_PER_TICKER = 60
MAX_SEEN = 3000

NEWS_MAX_AGE_DAYS = 21  # news/broker items older than this are filtered from the live feed
# When true, additionally restricts to items published "today" in the real London
# calendar (handles the GMT/BST switch correctly) — set False if this proves too
# strict and cuts out genuinely relevant items from yesterday evening.
NEWS_SAME_LONDON_DAY_ONLY = True

UPGRADE_WORDS = [
    "upgrade", "raises rating", "buy rating",
    "raised to buy", "initiates.*buy",
    r"upgrades?\s+\S+.{0,25}\s+to\s+['\"‘’]?(buy|overweight|outperform|add|accumulate)",
    r"raises?\s+\S+.{0,25}\s+to\s+['\"‘’]?(buy|overweight|outperform)",
    r"\bups\b\s+\S+.{0,25}\s+to\s+['\"‘’]?(buy|overweight|outperform)",
    r"moves?\s+\S+.{0,25}\s+to\s+['\"‘’]?(buy|overweight|outperform)",
]
DOWNGRADE_WORDS = [
    "downgrade", "cuts rating", "sell rating",
    "cut to sell", "initiates.*sell",
    r"cuts?\s+\S+.{0,25}\s+to\s+['\"‘’]?(sell|underweight|underperform|reduce|hold)",
    r"downgrades?\s+\S+.{0,25}\s+to\s+['\"‘’]?(sell|hold|underweight|underperform)",
    r"moves?\s+\S+.{0,25}\s+to\s+['\"‘’]?(sell|underweight|underperform)",
    # Narrow addition: "Broker cuts Company('s) [stock] rating" — no
    # destination rating stated, just the fact a rating was cut. Bounded
    # to at most ONE extra intervening word (covers a two-word company
    # name, e.g. "JD Sports") — deliberately NOT a wider allowance: an
    # adversarial test during development showed a {0,3}-word gap could
    # bridge across an unrelated clause ("cuts marketing costs while
    # credit rating agency...") and wrongly fire. {0,1} still matches
    # both real confirmed headlines ("cuts Pearson rating", "cuts
    # Beiersdorf stock rating") while resisting that false positive.
    r"\bcuts?\s+\S+(?:\s+\S+){0,1}\s+(?:stock\s+)?rating\b",
]
TARGET_WORDS = ["price target", "target price", "pt raised", "pt cut"]

# =========================================================================
# NORMALIZED ACTION VOCABULARY (per user's structured broker-monitor spec)
# =========================================================================
# Distinguishes a RATING change (upgrade/downgrade) from a TARGET-only change
# (target raised/cut with rating unchanged) — the two are different facts.
# Example: "BUY maintained, target £500->£550" = TARGET_RAISE, not UPGRADE.
# Checked in priority order inside classify(): a genuine rating change always
# wins over a target-only mention, since it's the more significant fact.
INITIATION_WORDS = [
    "initiates coverage", "initiated coverage", "starts coverage", "started coverage",
    "resumes coverage", "resumed coverage", "begins coverage", "initiate coverage",
]
REITERATION_WORDS = [
    "reiterates", "reiterated", "reaffirms", "reaffirmed", "maintains rating",
    "keeps rating", "retains rating", "sticks with",
]
TARGET_RAISE_WORDS = [
    r"hik(e|es|ed)\s+\S*.{0,30}(price target|target price)",
    r"rais(e|es|ed)\s+\S*.{0,30}(price target|target price)",
    r"lift(s|ed)?\s+\S*.{0,30}(price target|target price)",
    r"\bups\b\s+\S*.{0,30}(price target|target price)",
    r"increas(e|es|ed)\s+\S*.{0,30}(price target|target price)",
    r"boost(s|ed)?\s+\S*.{0,30}(price target|target price)",
    r"price target.{0,20}(raised|hiked|lifted|increased)",
    r"target price.{0,20}(raised|hiked|lifted|increased)",
]
TARGET_CUT_WORDS = [
    r"\bcuts?\b\s+\S*.{0,30}(price target|target price)",
    r"lower(s|ed)?\s+\S*.{0,30}(price target|target price)",
    r"trim(s|med)?\s+\S*.{0,30}(price target|target price)",
    r"slash(es|ed)?\s+\S*.{0,30}(price target|target price)",
    r"reduc(e|es|ed)\s+\S*.{0,30}(price target|target price)",
    r"price target.{0,20}(cut|lowered|trimmed|reduced|slashed)",
    r"target price.{0,20}(cut|lowered|trimmed|reduced|slashed)",
]

# Rating-text -> BULLISH/NEUTRAL/BEARISH bucket, for a plain-English "what
# does this rating mean" signal without inventing a prediction — purely a
# classification of the label itself, same as the user's spec.
BULLISH_RATING_TERMS = {
    "buy", "strong buy", "outperform", "overweight", "add", "accumulate",
    "top pick", "speculative buy", "house stock",
}
NEUTRAL_RATING_TERMS = {
    "hold", "neutral", "market perform", "equal weight", "equal-weight",
    "sector perform", "in-line", "no recommendation", "coverage pending",
}
BEARISH_RATING_TERMS = {
    "sell", "strong sell", "underperform", "underweight", "reduce", "trading sell",
}


def normalize_rating_bucket(raw_rating):
    """Maps any broker rating text to BULLISH/NEUTRAL/BEARISH, or None if
    unrecognized. Purely descriptive classification of the label itself —
    never a recommendation, never predicts anything."""
    if not raw_rating:
        return None
    r = raw_rating.strip().lower().replace("_", " ")
    if r in BULLISH_RATING_TERMS:
        return "BULLISH"
    if r in NEUTRAL_RATING_TERMS:
        return "NEUTRAL"
    if r in BEARISH_RATING_TERMS:
        return "BEARISH"
    return None


# Combined vocabulary for extract_new_rating_from_headline() below — reuses
# the SAME rating terms already defined above for bucket classification,
# rather than introducing a second/parallel rating vocabulary. Sorted
# longest-first so e.g. "strong buy" matches before the shorter "buy".
_ALL_RATING_TERMS_SORTED = sorted(
    BULLISH_RATING_TERMS | NEUTRAL_RATING_TERMS | BEARISH_RATING_TERMS,
    key=len, reverse=True,
)
_RATING_TERMS_ALTERNATION = "|".join(re.escape(t) for t in _ALL_RATING_TERMS_SORTED)
_NEW_RATING_RE = re.compile(
    rf"\b(?:upgrade[sd]?|downgrade[sd]?|cuts?|raise[sd]?)\s+to\s+['\"‘’]?({_RATING_TERMS_ALTERNATION})\b",
    re.IGNORECASE,
)


def extract_new_rating_from_headline(title):
    """
    Narrow, EXPLICIT-only extractor for a stated destination rating —
    "upgrade to Neutral", "downgrade to Buy", "cut to Hold". Only fires
    when an action word (upgrade/downgrade/cut/raise) is directly
    followed by "to <a known rating term>", using the SAME rating
    vocabulary already defined above — no new/parallel classification
    scheme. Returns the rating text exactly as it appeared in the
    headline (preserves original casing, e.g. "Neutral"), or None if no
    explicit rating is stated. NEVER inferred from the action word alone
    — "UBS downgrades Boliden" with no destination stated returns None,
    not a guessed rating.
    """
    m = _NEW_RATING_RE.search(title)
    if not m:
        return None
    return m.group(1)


def normalize_action_from_grades(from_grade, to_grade, yahoo_action_code=None):
    """
    Returns one of: UPGRADE, DOWNGRADE, REITERATION, INITIATION, RATING_CHANGE,
    NO_CHANGE. Prefers Yahoo's own action code (most reliable, since Yahoo
    already knows the real event type); falls back to comparing bullish/
    neutral/bearish buckets of from/to grade when no action code is given.
    """
    if yahoo_action_code == "init":
        return "INITIATION"
    if yahoo_action_code == "up":
        return "UPGRADE"
    if yahoo_action_code == "down":
        return "DOWNGRADE"
    if yahoo_action_code == "reit":
        return "REITERATION"
    from_bucket = normalize_rating_bucket(from_grade)
    to_bucket = normalize_rating_bucket(to_grade)
    if from_bucket and to_bucket:
        order = {"BEARISH": 0, "NEUTRAL": 1, "BULLISH": 2}
        if from_bucket == to_bucket:
            return "NO_CHANGE"
        return "UPGRADE" if order[to_bucket] > order[from_bucket] else "DOWNGRADE"
    if from_grade and to_grade and from_grade != to_grade:
        return "RATING_CHANGE"
    return "NO_CHANGE"


def normalize_headline_action(title):
    """
    Classifies a NEWS HEADLINE (not structured Yahoo data) into the same
    normalized vocabulary. Checked in priority order: a genuine rating
    change always wins over a target-only mention, since a headline saying
    "Citi upgrades X, raises target" is fundamentally an UPGRADE, not a
    TARGET_RAISE, even though it mentions the target too.
    """
    t = title.lower()
    if any(re.search(w, t) for w in UPGRADE_WORDS):
        return "UPGRADE"
    if any(re.search(w, t) for w in DOWNGRADE_WORDS):
        return "DOWNGRADE"
    if any(w in t for w in INITIATION_WORDS):
        return "INITIATION"
    if any(re.search(w, t) for w in TARGET_RAISE_WORDS):
        return "TARGET_RAISE"
    if any(re.search(w, t) for w in TARGET_CUT_WORDS):
        return "TARGET_CUT"
    if any(w in t for w in REITERATION_WORDS):
        return "REITERATION"
    return None  # not a broker-action headline at all — classify() handles the rest

# Maps classify()'s lowercase category (used for CSS badge classes) to the
# uppercase normalized vocabulary word (used in the structured data /
# alert labels) — keeps the two representations consistent everywhere.
CATEGORY_TO_NORMALIZED_ACTION = {
    "upgrade": "UPGRADE", "downgrade": "DOWNGRADE", "initiation": "INITIATION",
    "target_raise": "TARGET_RAISE", "target_cut": "TARGET_CUT", "reiteration": "REITERATION",
    "target": "RATING_CHANGE", "director_dealing": None, "event": None, "news": None,
}

# Director/PDMR (Persons Discharging Managerial Responsibilities) dealings — a standard
# RNS announcement type disclosing when a company director has bought or sold shares.
# Real, published, factual disclosure — not something this tool infers or predicts.
DIRECTOR_DEALING_WORDS = [
    "pdmr", "director/pdmr shareholding", "director shareholding",
    "notification of transactions by persons discharging managerial responsibilities",
    "director dealing", "directors' dealing", "holding(s) in company",
]
EVENT_WORDS = [
    "trading update", "results", "interim report", "final results", "agm",
    "dividend", "profit warning", "acquisition", "placing", "earnings", "guidance",
    "merger", "merges", "merging", "takeover", "bid for", "bids for", "acquire",
    "acquires", "acquiring", "buyout", "to be bought", "in talks to", "offer for",
]
BROKER_NAMES = [
    "Barclays", "Deutsche Bank", "JPMorgan", "JP Morgan", "HSBC", "UBS", "Citi",
    "Citigroup", "Goldman Sachs", "Morgan Stanley", "RBC", "RBC Capital Markets",
    "Jefferies", "Peel Hunt", "Numis", "Berenberg", "Panmure", "Panmure Liberum",
    "Shore Capital", "Investec", "Canaccord", "Canaccord Genuity", "Liberum", "BofA",
    "Bank of America", "Societe Generale", "BNP Paribas", "Credit Suisse", "Stifel",
    "Redburn", "Zeus Capital", "Cenkos", "finnCap", "Singer Capital Markets",
    "N+1 Singer", "WH Ireland", "Edison", "Equity Development", "Cantor Fitzgerald",
    "Exane BNP Paribas", "Kepler Cheuvreux", "Mediobanca", "Davy", "Goodbody",
    "Third Bridge", "Marex", "Turner Pope", "Arden Partners", "Charles Stanley",
    "AJ Bell", "Hargreaves Lansdown", "Killik", "Interactive Investor", "Shard Capital",
    "Oberon Capital", "Whitman Howard", "Progressive Equity Research", "Hardman & Co",
]
ANALYST_RATINGS_FEED_URL = "https://uk.investing.com/rss/news_1061.rss"
GENERAL_MARKET_NEWS_FEED_URL = "https://uk.investing.com/rss/news_25.rss"  # confirmed via their own official RSS listing page
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
}

# Yahoo's screener and quoteSummary endpoints reject anonymous requests from cloud/
# datacenter IPs (GitHub Actions runners included) with a 401 unless a session cookie
# + "crumb" token — Yahoo's own web frontend's real auth handshake — are attached.
# The basic chart/quote endpoint is more lenient and doesn't need this. This uses a
# shared cookie jar across all Yahoo requests so the session persists once established.
_yahoo_cookiejar = http.cookiejar.CookieJar()
_yahoo_opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(_yahoo_cookiejar))
_yahoo_crumb = None


def _yahoo_opener_get(url, timeout=15):
    req = urllib.request.Request(url, headers=HEADERS)
    with _yahoo_opener.open(req, timeout=timeout) as resp:
        return resp.read()


def get_yahoo_crumb():
    """Fetch (and cache) a Yahoo auth crumb, establishing the session cookie first."""
    global _yahoo_crumb
    if _yahoo_crumb:
        return _yahoo_crumb
    try:
        _yahoo_opener_get("https://fc.yahoo.com")  # sets the initial session cookie
    except Exception as e:
        print(f"  ! yahoo cookie handshake failed: {e}", file=sys.stderr)
    try:
        crumb = _yahoo_opener_get("https://query2.finance.yahoo.com/v1/test/getcrumb").decode("utf-8").strip()
        if crumb and "<html" not in crumb.lower():
            _yahoo_crumb = crumb
            return _yahoo_crumb
    except Exception as e:
        print(f"  ! yahoo crumb fetch failed: {e}", file=sys.stderr)
    return None


def http_get(url, timeout=15):
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.read()


def classify(title):
    t = title.lower()
    # Downgrade checked before upgrade: a headline like "downgraded X to Y from
    # outperform" would otherwise risk matching an upgrade-side pattern first —
    # a genuine downgrade always contains "downgrade"/"cuts"-style language
    # that's a stronger, less ambiguous signal than any upgrade-side term.
    if any(re.search(w, t) for w in DOWNGRADE_WORDS):
        return "downgrade"
    if any(re.search(w, t) for w in UPGRADE_WORDS):
        return "upgrade"
    if any(w in t for w in INITIATION_WORDS):
        return "initiation"
    if any(re.search(w, t) for w in TARGET_RAISE_WORDS):
        return "target_raise"
    if any(re.search(w, t) for w in TARGET_CUT_WORDS):
        return "target_cut"
    if any(w in t for w in REITERATION_WORDS):
        return "reiteration"
    if any(w in t for w in TARGET_WORDS):
        return "target"
    if any(w in t for w in DIRECTOR_DEALING_WORDS):
        return "director_dealing"
    if any(w in t for w in EVENT_WORDS):
        return "event"
    return "news"


def detect_broker(title):
    """
    Returns whichever known broker name appears EARLIEST in the actual
    headline TEXT — not earliest in the BROKER_NAMES list. The previous
    list-order-only approach meant a headline like "Berenberg downgrades
    Barclays to Hold" incorrectly returned "Barclays" as the broker,
    purely because "Barclays" happens to be listed before "Berenberg" in
    BROKER_NAMES, even though "Berenberg" is the one actually performing
    the action. Found and confirmed while testing the ticker-resolution
    fix above — a genuine, separate pre-existing bug, not introduced by
    that change, fixed here since correct broker attribution is a
    prerequisite for that fix to work correctly.
    """
    t = title.lower()
    best_name, best_pos = "", None
    for b in BROKER_NAMES:
        idx = t.find(b.lower())
        if idx != -1 and (best_pos is None or idx < best_pos):
            best_name, best_pos = b, idx
    return best_name


def parse_rss(xml_bytes):
    items = []
    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError:
        return items
    for item in root.iter("item"):
        title = (item.findtext("title") or "").strip()
        link = (item.findtext("link") or "").strip()
        pub_date = (item.findtext("pubDate") or "").strip()
        source_el = item.find("source")
        source = source_el.text.strip() if source_el is not None and source_el.text else ""
        if not source and link:
            try:
                source = urllib.parse.urlparse(link).hostname.replace("www.", "")
            except Exception:
                source = ""
        if title and link:
            items.append({"title": title, "link": link, "pubDate": pub_date, "source": source})
    return items


def fetch_yahoo_broker_target(ticker):
    """
    Lightweight, dedicated broker-consensus lookup — requests ONLY the
    financialData module (not the 8-module fetch_yahoo_analyst does for
    the full watchlist research view), since this is used for a
    genuinely separate purpose: enriching LSE-primary Screener rows
    (Gainers/Losers/Volume) with the broker-target fields LSE's own
    market-data response doesn't provide, never anything more.

    This is Yahoo ENRICHMENT of LSE market data, not a replacement
    source — callers must attach these fields alongside the existing
    LSE price/changePct/volume, never overwrite them.

    Returns {"targetMeanPrice", "targetHighPrice", "targetLowPrice",
    "numberOfAnalystOpinions", "recommendationKey"} or None if
    genuinely unavailable.
    """
    symbol = yahoo_symbol(ticker)
    crumb = get_yahoo_crumb()
    url = (
        f"https://query1.finance.yahoo.com/v10/finance/quoteSummary/{urllib.parse.quote(symbol)}"
        "?modules=financialData"
    )
    if crumb:
        url += f"&crumb={urllib.parse.quote(crumb)}"
    try:
        data = json.loads(_yahoo_opener_get(url))
        result = (((data.get("quoteSummary") or {}).get("result")) or [None])[0]
        if not result:
            return None
        fin = result.get("financialData") or {}
        target_mean = (fin.get("targetMeanPrice") or {}).get("raw")
        if target_mean is None:
            return None  # never fabricate a target that genuinely doesn't exist
        return {
            "targetMeanPrice": target_mean,
            "targetHighPrice": (fin.get("targetHighPrice") or {}).get("raw"),
            "targetLowPrice": (fin.get("targetLowPrice") or {}).get("raw"),
            "numberOfAnalystOpinions": (fin.get("numberOfAnalystOpinions") or {}).get("raw"),
            "recommendationKey": fin.get("recommendationKey"),
        }
    except Exception as e:
        print(f"  ! yahoo broker-target enrichment failed: {ticker} ({e})", file=sys.stderr)
        return None


def fetch_yahoo_analyst(ticker):
    """
    Structured analyst ratings history via Yahoo Finance's quoteSummary endpoint —
    real firm name, exact action (up/down/init/main/reit), and date, straight from
    Yahoo's own aggregation of broker calls. Requires the crumb-authenticated opener;
    falls back to trying without a crumb if one couldn't be obtained.
    Also pulls calendarEvents (next earnings date, next ex-dividend date) in the SAME
    request — quoteSummary accepts multiple modules at once, so this adds real data
    without adding a new network call.
    """
    symbol = yahoo_symbol(ticker)
    crumb = get_yahoo_crumb()
    url = (
        f"https://query1.finance.yahoo.com/v10/finance/quoteSummary/{urllib.parse.quote(symbol)}"
        "?modules=upgradeDowngradeHistory,recommendationTrend,financialData,calendarEvents,summaryDetail,defaultKeyStatistics,price,assetProfile"
    )
    if crumb:
        url += f"&crumb={urllib.parse.quote(crumb)}"
    try:
        data = json.loads(_yahoo_opener_get(url))
        result = (((data.get("quoteSummary") or {}).get("result")) or [None])[0]
        if not result:
            return None
        history = ((result.get("upgradeDowngradeHistory") or {}).get("history")) or []
        fin = result.get("financialData") or {}
        cal = result.get("calendarEvents") or {}
        earnings_dates = ((cal.get("earnings") or {}).get("earningsDate")) or []
        next_earnings = earnings_dates[0].get("raw") if earnings_dates else None
        ex_div = (cal.get("exDividendDate") or {}).get("raw")
        summary = result.get("summaryDetail") or {}
        div_rate = (summary.get("dividendRate") or {}).get("raw")  # currency amount per share per year
        div_yield_raw = (summary.get("dividendYield") or {}).get("raw")  # Yahoo returns this as a fraction, e.g. 0.045 = 4.5%
        div_yield_pct = div_yield_raw * 100 if div_yield_raw is not None else None
        stats = result.get("defaultKeyStatistics") or {}
        price_mod = result.get("price") or {}
        trailing_pe = (summary.get("trailingPE") or {}).get("raw")
        # Same summaryDetail module already being fetched above for P/E, dividend,
        # 52-week range — averageVolume sits right alongside them in the same
        # response. Verified against real Yahoo API response structure (multiple
        # independent open-source tools/wrappers scraping this same live endpoint
        # all show this exact field name in this exact module) before adding this
        # line — not assumed. Zero new network calls.
        average_volume = (summary.get("averageVolume") or {}).get("raw")
        eps = (stats.get("trailingEps") or {}).get("raw")
        market_cap = (price_mod.get("marketCap") or {}).get("raw")
        fifty_two_low = (summary.get("fiftyTwoWeekLow") or {}).get("raw")
        fifty_two_high = (summary.get("fiftyTwoWeekHigh") or {}).get("raw")
        held_insiders_raw = (stats.get("heldPercentInsiders") or {}).get("raw")  # fraction, e.g. 0.12 = 12%
        held_insiders_pct = held_insiders_raw * 100 if held_insiders_raw is not None else None
        profile = result.get("assetProfile") or {}
        sector = profile.get("sector")
        industry = profile.get("industry")
        business_summary = profile.get("longBusinessSummary")
        if business_summary and len(business_summary) > 220:
            business_summary = business_summary[:217].rsplit(" ", 1)[0] + "..."  # trim at a word boundary
        return {
            "history": history,
            "targetMeanPrice": (fin.get("targetMeanPrice") or {}).get("raw"),
            "recommendationKey": fin.get("recommendationKey"),
            "nextEarningsDate": next_earnings,  # unix epoch seconds, or None
            "exDividendDate": ex_div,  # unix epoch seconds, or None
            "dividendRate": div_rate,  # per-share currency amount, or None
            "dividendYieldPct": div_yield_pct,  # already converted to a percentage, or None
            "trailingPE": trailing_pe,
            "trailingEps": eps,
            "marketCap": market_cap,
            "averageVolume": average_volume,
            "fiftyTwoWeekLow": fifty_two_low,
            "fiftyTwoWeekHigh": fifty_two_high,
            "heldPercentInsidersPct": held_insiders_pct,
            "sector": sector,
            "industry": industry,
            "businessSummary": business_summary,
        }
    except Exception as e:
        print(f"  ! yahoo analyst history failed: {ticker} ({e})", file=sys.stderr)
        return None


FCA_SHORT_INTEREST_URL = "https://www.fca.org.uk/publication/data/short-positions-daily-update.xlsx"


def fetch_short_interest():
    """
    FCA's daily Aggregate Net Short Position file — free, official, no auth.
    T+2 lag (today's figures reflect positions from two working days ago).
    Returns dict: COMPANY NAME (upper, stripped) -> {"pct": float, "position_date": str}
    Fails soft (returns {}) — same pattern as every other fetch_* here.
    """
    try:
        import openpyxl
    except ImportError:
        print("  ! short interest skipped: openpyxl not installed (add to requirements.txt)", file=sys.stderr)
        return {}
    try:
        import io
        raw = http_get(FCA_SHORT_INTEREST_URL, timeout=20)
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active

        header_row_idx, headers = None, {}
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
            for cell in row:
                if cell and "issuer" in str(cell).lower():
                    header_row_idx = i + 1
                    headers = {str(h).strip().lower(): k for k, h in enumerate(row) if h}
                    break
            if header_row_idx:
                break
        if header_row_idx is None:
            print("  ! short interest: header row not found — FCA file format may have changed", file=sys.stderr)
            return {}

        name_col = next((v for k, v in headers.items() if "issuer" in k), None)
        pct_col = next((v for k, v in headers.items() if "%" in k or "percent" in k), None)
        date_col = next((v for k, v in headers.items() if "date" in k), None)
        if name_col is None or pct_col is None:
            print(f"  ! short interest: required columns not found (saw: {list(headers.keys())})", file=sys.stderr)
            return {}

        result = {}
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not row or row[name_col] is None:
                continue
            name = str(row[name_col]).strip().upper()
            try:
                pct = float(row[pct_col])
            except (TypeError, ValueError):
                continue
            result[name] = {
                "pct": pct,
                "position_date": str(row[date_col]) if date_col is not None else None,
            }
        print(f"  short interest: loaded {len(result)} companies from FCA")
        return result
    except Exception as e:
        print(f"  ! short interest fetch failed: {e}", file=sys.stderr)
        return {}


def match_short_interest(company_name, short_interest_map):
    """Loose name match — FCA issuer names rarely match Yahoo's/watchlist's naming exactly."""
    if not company_name or not short_interest_map:
        return None
    clean = company_name.upper().replace(" PLC", "").replace(" ORD", "").strip()
    for fca_name, data in short_interest_map.items():
        clean_fca = fca_name.replace(" PLC", "").replace(" ORD", "").strip()
        if clean in clean_fca or clean_fca in clean:
            return data
    return None


ACTION_CATEGORY = {"up": "upgrade", "down": "downgrade", "init": "initiation", "main": "news", "reit": "reiteration"}
RECENT_WINDOW_SECONDS = 2 * 24 * 3600  # only alert on ratings from the last 48h, not backfilled history


def format_normalized_at(iso_string):
    """Formats a normalizedAt ISO timestamp (when this item's classification was
    computed) as a plain London date/time, same dual-timezone convention as the
    rest of the dashboard — distinct from pubDate (the article/event's own date),
    since a backfilled Yahoo history item can be re-classified on a later run
    than when the underlying event actually happened."""
    if not iso_string:
        return None
    try:
        dt = datetime.fromisoformat(iso_string)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return format_london_and_utc(dt)
    except Exception:
        return None


def analyst_history_to_items(ticker, analyst):
    items = []
    if not analyst:
        return items
    symbol = yahoo_symbol(ticker)
    now = time.time()
    normalized_at = datetime.now(timezone.utc).isoformat()  # when THIS classification ran —
                                                              # distinct from the event's own date,
                                                              # since backfilled history can be
                                                              # re-classified on a later run than
                                                              # when the actual rating happened.
    for h in analyst.get("history", [])[:15]:
        firm = h.get("firm", "")
        to_grade = h.get("toGrade", "")
        from_grade = h.get("fromGrade", "")
        action = h.get("action", "")
        epoch = h.get("epochGradeDate")
        if not firm or not to_grade or not epoch:
            continue
        pub_date = datetime.fromtimestamp(epoch, tz=timezone.utc).strftime("%a, %d %b %Y %H:%M:%S GMT")
        title = f"{firm} {'moves' if action=='main' else 'rates'} {ticker} to {to_grade}" + (
            f" (from {from_grade})" if from_grade and from_grade != to_grade else ""
        )
        items.append({
            "title": title,
            "link": f"https://finance.yahoo.com/quote/{symbol}/analysis#{firm}-{epoch}".replace(" ", "-"),
            "pubDate": pub_date,
            "source": "Yahoo Finance Analyst History",
            "category": ACTION_CATEGORY.get(action, "news"),
            "broker": firm,
            "_recent": (now - epoch) <= RECENT_WINDOW_SECONDS,
            # Normalized structured fields (per the broker-monitor spec) — kept
            # alongside the original wording (fromGrade/toGrade preserved as-is).
            "normalizedAction": normalize_action_from_grades(from_grade, to_grade, action),
            "fromRating": from_grade or None,
            "toRating": to_grade,
            "fromRatingBucket": normalize_rating_bucket(from_grade),
            "toRatingBucket": normalize_rating_bucket(to_grade),
            "normalizedAt": normalized_at,  # when this classification was computed (ISO UTC)
        })
    return items



import email.utils as _email_utils


LONDON_TZ = ZoneInfo("Europe/London")

# =========================================================================
# Pipeline health — freshness of the poll ITSELF, deliberately separate
# from feed-level health (e.g. ftseUniverseStatus). This block never
# claims anything about Yahoo, Google News, broker feeds, or FTSE
# constituents — only "when did this pipeline last successfully write
# data, and how does that compare to expectations."
# =========================================================================

PIPELINE_GREEN_THRESHOLD_MIN = 15    # market hours: fresher than this -> green
PIPELINE_AMBER_THRESHOLD_MIN = 45    # market hours: fresher than this -> amber, else red
PIPELINE_OUT_OF_HOURS_STALE_THRESHOLD_MIN = 24 * 60  # outside market hours: beyond this -> worth flagging anyway
UK_MARKET_OPEN_MINUTES = 8 * 60       # 08:00
UK_MARKET_CLOSE_MINUTES = 16 * 60 + 30  # 16:30


def _easter_sunday(year):
    """Anonymous Gregorian algorithm for the date of Easter Sunday — the
    standard, well-tested method, not an approximation. Needed because
    Good Friday and Easter Monday (both LSE bank holidays) fall on a
    different date every year."""
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)


def uk_bank_holidays(year):
    """England & Wales bank holidays — the calendar the LSE actually
    follows — computed programmatically from the year's own calendar
    rather than a hardcoded list of dates, so this stays correct every
    year rather than needing a manual update each January. Covers the
    standard 8 holidays: New Year's Day, Good Friday, Easter Monday,
    Early May, Spring (late May), Summer (late August), Christmas Day,
    Boxing Day — including the weekend-shift rules for the fixed-date
    ones. Doesn't attempt one-off royal/coronation-style extra holidays,
    which aren't predictable from the calendar alone."""
    holidays = set()
    new_year = date(year, 1, 1)
    if new_year.weekday() == 5:
        new_year += timedelta(days=2)
    elif new_year.weekday() == 6:
        new_year += timedelta(days=1)
    holidays.add(new_year)
    easter = _easter_sunday(year)
    holidays.add(easter - timedelta(days=2))  # Good Friday
    holidays.add(easter + timedelta(days=1))  # Easter Monday
    early_may = date(year, 5, 1)
    while early_may.weekday() != 0:
        early_may += timedelta(days=1)
    holidays.add(early_may)
    spring = date(year, 5, 31)
    while spring.weekday() != 0:
        spring -= timedelta(days=1)
    holidays.add(spring)
    summer = date(year, 8, 31)
    while summer.weekday() != 0:
        summer -= timedelta(days=1)
    holidays.add(summer)
    christmas = date(year, 12, 25)
    boxing = date(year, 12, 26)
    if christmas.weekday() in (5, 6):
        christmas += timedelta(days=2)
    if boxing.weekday() in (5, 6) or boxing == christmas:
        boxing = christmas + timedelta(days=1)
        while boxing.weekday() in (5, 6):
            boxing += timedelta(days=1)
    holidays.add(christmas)
    holidays.add(boxing)
    return holidays


def _is_uk_bank_holiday(dt_utc):
    """True if this UTC instant, in London local time, falls on an LSE
    bank holiday. Computes only the current year's calendar each call —
    cheap, and avoids ever needing a stored/maintained holiday list."""
    dt_london = dt_utc.astimezone(LONDON_TZ)
    return dt_london.date() in uk_bank_holidays(dt_london.year)


def _is_uk_market_hours(dt_utc):
    """
    Mon-Fri, 08:00-16:30 Europe/London LOCAL time (standard LSE trading
    hours) — correctly handles the GMT/BST switch via zoneinfo (the same
    LONDON_TZ already used elsewhere), not a fixed UTC offset. Also
    correctly treats LSE bank holidays as closed, even on an otherwise
    ordinary weekday.
    """
    dt_london = dt_utc.astimezone(LONDON_TZ)
    if dt_london.weekday() >= 5:  # Saturday=5, Sunday=6
        return False
    if _is_uk_bank_holiday(dt_utc):
        return False
    minutes_since_midnight = dt_london.hour * 60 + dt_london.minute
    return UK_MARKET_OPEN_MINUTES <= minutes_since_midnight < UK_MARKET_CLOSE_MINUTES


def compute_pipeline_health(last_poll_str, now_utc=None):
    """
    Pure function — no I/O. Returns a dict describing how fresh the LAST
    SUCCESSFUL WRITE of state/data.json is, relative to `now_utc`.

    IMPORTANT, stated both here and in the returned dict itself: this is
    a WRITE-TIME EVALUATION. main() calls this at the same moment it sets
    `lastPoll`, so the persisted status will always show as fresh
    ("green"/"market_closed" with age~0) at the instant it's written —
    a run that stops happening can't recompute its own staleness. This
    persisted value is NOT proof the pipeline is currently healthy by
    the time someone reads state/data.json later; it only proves what
    the state was AT THAT WRITE. Determining CURRENT freshness — "how
    stale is this relative to right now" — is the dashboard's
    client-side script's job (see render_dashboard), which recomputes
    continuously from the embedded `lastPoll` value using the visitor's
    own real clock, completely independent of this persisted snapshot.

    Returns dict with: status, lastPoll, ageMinutes, marketHours,
    message, evaluatedAt, note.
    """
    now_utc = now_utc or datetime.now(timezone.utc)
    market_hours = _is_uk_market_hours(now_utc)
    note = ("Computed once, at the moment this data was written — does NOT reflect time "
            "elapsed since. Current freshness is determined by the dashboard's client-side "
            "script from lastPoll, independently of this persisted status.")

    if not last_poll_str:
        return {
            "status": "unknown", "lastPoll": None, "ageMinutes": None,
            "marketHours": market_hours,
            "message": "No successful poll has been recorded yet.",
            "evaluatedAt": now_utc.isoformat(), "note": note,
        }
    try:
        last_poll_dt = datetime.strptime(last_poll_str, "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
    except Exception:
        return {
            "status": "unknown", "lastPoll": last_poll_str, "ageMinutes": None,
            "marketHours": market_hours,
            "message": "The last-poll timestamp could not be parsed.",
            "evaluatedAt": now_utc.isoformat(), "note": note,
        }

    age_minutes = (now_utc - last_poll_dt).total_seconds() / 60

    if market_hours:
        if age_minutes <= PIPELINE_GREEN_THRESHOLD_MIN:
            status = "green"
            message = f"Last successful poll {age_minutes:.0f} min ago — up to date."
        elif age_minutes <= PIPELINE_AMBER_THRESHOLD_MIN:
            status = "amber"
            message = f"Last successful poll {age_minutes:.0f} min ago — later than usual during market hours."
        else:
            status = "red"
            message = f"Last successful poll {age_minutes:.0f} min ago — significantly overdue during market hours."
    else:
        if age_minutes <= PIPELINE_OUT_OF_HOURS_STALE_THRESHOLD_MIN:
            status = "market_closed"
            message = f"Markets are closed. Last successful poll {age_minutes / 60:.1f}h ago."
        else:
            status = "stale_out_of_hours"
            message = f"Markets are closed, but no successful poll in {age_minutes / 60:.1f}h — worth checking."

    return {
        "status": status, "lastPoll": last_poll_str, "ageMinutes": round(age_minutes, 1),
        "marketHours": market_hours, "message": message,
        "evaluatedAt": now_utc.isoformat(), "note": note,
    }


def format_news_timestamp(pub_date_str):
    """
    Converts a raw RSS pubDate string (e.g. "Mon, 31 Aug 2026 14:23:00 GMT" —
    always labeled GMT by the feeds regardless of actual season) into a
    genuinely correct, consistently-formatted display timestamp, matching
    the same BST/UTC convention used everywhere else on the dashboard.

    Confirmed real problem this fixes: showing the raw RSS string directly
    is misleading during British Summer Time (the feed always says "GMT"
    even in August, when the real local time is BST, one hour later) and
    is visually inconsistent with every other timestamp on the page.

    Returns "HH:MM BST (HH:MM UTC)" for today's items (matches the
    same-day-only guarantee this project already enforces via
    NEWS_SAME_LONDON_DAY_ONLY), or "DD Mon, HH:MM BST" for the rare
    older fallback item (never presented as if it were today's). Falls
    back to the raw string, never blank, if genuinely unparseable.
    """
    dt = _parse_pub_date(pub_date_str)
    if dt is None:
        return esc(pub_date_str or "")
    dt_london = dt.astimezone(LONDON_TZ)
    dt_utc = dt.astimezone(timezone.utc)
    is_today = dt_london.date() == datetime.now(timezone.utc).astimezone(LONDON_TZ).date()
    if is_today:
        return f'{dt_london.strftime("%H:%M")} {dt_london.strftime("%Z")} ({dt_utc.strftime("%H:%M")} UTC)'
    return f'{dt_london.strftime("%d %b, %H:%M")} {dt_london.strftime("%Z")}'


def _parse_pub_date(pub_date_str):
    """
    Shared date parser — tries RFC 822 (standard RSS: "Wed, 31 Jul 2024 07:00:00 GMT")
    then Investing.com's plain "YYYY-MM-DD HH:MM:SS" format. Returns a tz-aware
    datetime, or None if genuinely unparseable.
    """
    if not pub_date_str:
        return None
    dt = None
    try:
        dt = _email_utils.parsedate_to_datetime(pub_date_str)
    except Exception:
        pass
    if dt is None:
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
            try:
                dt = datetime.strptime(pub_date_str.strip(), fmt)
                break
            except Exception:
                continue
    if dt is None:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt


def is_recent_enough(pub_date_str, max_age_days=NEWS_MAX_AGE_DAYS):
    """
    Returns True if pub_date_str parses to within max_age_days of now, OR if it can't
    be parsed at all (fail open — an unparseable date shouldn't silently hide an item,
    that's a display bug, not a staleness signal).
    """
    dt = _parse_pub_date(pub_date_str)
    if dt is None:
        return True
    age = datetime.now(timezone.utc) - dt
    return age.days <= max_age_days


def is_today_in_london(pub_date_str):
    """
    Stricter same-day filter, using the real Europe/London calendar date (correctly
    handles the GMT/BST switch via zoneinfo, not just a fixed UTC offset) — 'today'
    means today in London, not just 'within the last N days'. Fails open on an
    unparseable date, same reasoning as is_recent_enough.
    """
    dt = _parse_pub_date(pub_date_str)
    if dt is None:
        return True
    return dt.astimezone(LONDON_TZ).date() == datetime.now(timezone.utc).astimezone(LONDON_TZ).date()


def passes_news_filters(pub_date_str):
    """Combined recency check applied to every news/broker item before it can enter
    the feed, alerts, or dashboard."""
    if not is_recent_enough(pub_date_str):
        return False
    if NEWS_SAME_LONDON_DAY_ONLY and not is_today_in_london(pub_date_str):
        return False
    return True


RECENT_NEWS_FALLBACK_MAX_AGE_DAYS = 7  # a "most recent available, not from today" fallback
# is only shown for items within this window — old enough to survive a quiet
# weekend, nowhere near NEWS_MAX_AGE_DAYS (21), which exists for a different
# purpose (the outer ceiling on what's retained at all).


def passes_recency_filter_wide(pub_date_str, max_age_days=RECENT_NEWS_FALLBACK_MAX_AGE_DAYS):
    """
    Age-only recency check, deliberately WITHOUT the same-day requirement —
    used ONLY to build a separate, display-only "most recent available"
    fallback pool, completely independent of items_by_ticker/marketWide
    (which stay strictly same-day, per passes_news_filters, and are what
    feeds classify_evidence/compute_entry_exit_evidence/the scorecard).
    That system's own displayed text makes explicit same-day claims
    ("no relevant same-day catalyst found") — feeding it anything wider
    than today would make those claims false. This function exists
    specifically so a genuinely relevant 2-day-old headline can still be
    SHOWN to the user (clearly dated, never presented as today's), without
    ever being treated as today's catalyst by the evidence system.
    """
    return is_recent_enough(pub_date_str, max_age_days=max_age_days)


def item_sort_key(it):
    """Sort key for ordering news/alert items latest-first. NEVER sort on the raw
    pubDate string directly — RFC 822 dates start with the day name ("Thu, 27 Aug...",
    "Fri, 28 Aug..."), so a plain string sort groups by day-of-week alphabetically
    ("Fri" < "Thu") rather than by actual chronological time, silently scrambling the
    real order. Parse to a real datetime first; fall back to detectedAt (stored as
    Python's .isoformat(), a different format _parse_pub_date doesn't cover — handled
    separately here), then epoch 0 (sorts last) for anything genuinely unparseable."""
    dt = _parse_pub_date(it.get("pubDate"))
    if dt is None and it.get("detectedAt"):
        try:
            dt = datetime.fromisoformat(it["detectedAt"])
        except Exception:
            dt = None
    if dt is None:
        return datetime.fromtimestamp(0, tz=timezone.utc)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt


# HTTP codes worth retrying — genuinely transient (rate-limiting, upstream
# overload), confirmed happening repeatedly against Google News in
# production. 404/DNS-failure/etc are NOT included: those won't resolve on
# retry, so retrying them only wastes time inside the 5-minute poll window.
_TRANSIENT_HTTP_CODES = {429, 500, 502, 503, 504}


def fetch_feed(url, max_retries=1, backoff_seconds=2.0):
    """
    Fetches and parses an RSS feed, with a short, capped retry specifically
    for TRANSIENT failures (503 Service Unavailable and similar upstream
    codes, plus timeouts/connection errors) — exactly the failure pattern
    observed repeatedly against Google News in production (a run showing
    dozens of consecutive 503s). A non-transient HTTP error (404, etc) is
    NOT retried, since retrying an error that will never resolve just
    wastes time. max_retries=1 (try twice total) with a short fixed
    backoff keeps the worst-case added delay per feed small (~2s) — this
    fetch happens many times per poll cycle, so an aggressive retry policy
    could itself meaningfully extend the run, especially in exactly the
    mass-rate-limiting scenario where retries matter most.
    """
    last_error = None
    for attempt in range(max_retries + 1):
        try:
            return parse_rss(http_get(url)), False
        except urllib.error.HTTPError as e:
            last_error = e
            if e.code not in _TRANSIENT_HTTP_CODES or attempt == max_retries:
                break
            time.sleep(backoff_seconds)
        except Exception as e:
            last_error = e
            if attempt == max_retries:
                break
            time.sleep(backoff_seconds)
    print(f"  ! feed fetch failed: {url} ({last_error})", file=sys.stderr)
    return [], True


def clean_company_name(name):
    """
    Yahoo's screener returns full share-listing names, e.g. "MCBRIDE PLC ORD 10P",
    "GEIGER COUNTER LIMITED ORD NPV", "SMITH & NEPHEW PLC ORD USD0.20" — the "ORD ..."
    part is share-class jargon (par value, currency) that never appears in actual news
    articles. Searching Google News for that literal string returns almost nothing,
    because real headlines just say "McBride" — confirmed against real data: McBride's
    genuine same-day news (a Vestacy partnership announcement) was being missed
    entirely because "ORD 10P" polluted the search query. Trimming everything from
    " ORD " onward gives a name that actually matches how news outlets refer to the
    company. Names that don't contain " ORD " (e.g. watchlist names you supplied
    yourself, like "Vodafone Group") pass through unchanged.
    """
    if not name:
        return name
    cleaned = re.split(r"\bORD\b", name, maxsplit=1, flags=re.IGNORECASE)[0].strip()
    return cleaned or name  # never return an empty string


# =========================================================================
# News relevance — ONE shared rule, applied identically wherever a
# headline needs to be judged relevant to a company, whether it's a
# freshly fetched item or a stored one being carried forward to another
# run. Deliberately separate from date/staleness filtering
# (passes_news_filters): an item can be perfectly fresh and still be
# about the wrong company (the confirmed real cases: an ABDN-tagged item
# about a bond coupon schedule with no mention of "abrdn" anywhere, and
# a SHEL-tagged item about an Allianz/AA takeover with no mention of
# "Shell" anywhere).
# =========================================================================

# Names identified by direct audit of the actual watchlist (not
# speculative) as carrying real collision risk with ordinary English
# vocabulary or unrelated short initialisms — word-boundary matching
# alone does not protect against these, because the risk isn't a
# company name being embedded INSIDE an unrelated word (that's what
# word-boundary matching fixes, e.g. "BP" inside "GBP"); it's the
# company name being a genuine, ordinary whole word or a very short,
# generic-looking token in its own right. "Shell" is a common English
# noun (seashell, artillery shell, "shell script", "shell company") with
# no relation to Shell plc in most ordinary usage. "BP" is short enough
# (2 characters) that even a clean whole-word match still has meaningful
# collision risk with unrelated initialisms (blood pressure, basis
# points, etc). Names like "Aviva" or "Haleon", despite also being
# short, are NOT included here — they are invented brand strings with no
# competing ordinary-English meaning, so their collision risk is much
# lower and treating them the same way would only reduce genuine recall
# for no real safety benefit. This set of NAMES lives as the keys of
# HIGH_COLLISION_DISQUALIFIER_PHRASES below — kept as a single source of
# truth rather than a separate list that could drift out of sync.

# Legal-suffix names where the STORED name (as supplied on the
# watchlist) is a real company name but is meaningfully more specific
# than how the company is normally referred to in headlines — "GSK plc"
# in a headline is rare; "GSK reports strong quarter" is how it's
# actually written. Rather than a general suffix-stripping rule applied
# to everything (which could shorten a name into something genuinely
# more ambiguous), this is scoped narrowly: only strip a trailing legal
# suffix, and only as an ADDITIONAL accepted variant alongside the full
# name — never a replacement for it.
_LEGAL_SUFFIX_RE = re.compile(r"\s+(plc|ltd|limited)\s*$", re.IGNORECASE)


def _relevance_name_variants(cleaned_name):
    """
    Returns the set of name forms to accept a match against — normally
    just the cleaned name itself, plus (only when a trailing legal
    suffix is present) the same name with that suffix stripped, so
    "GSK reports strong quarter" still matches "GSK plc" without
    requiring the suffix verbatim in the headline.
    """
    if not cleaned_name:
        return set()
    variants = {cleaned_name}
    stripped = _LEGAL_SUFFIX_RE.sub("", cleaned_name).strip()
    if stripped and stripped != cleaned_name:
        variants.add(stripped)
    return variants


def _is_relevant_to_company(title, cleaned_name):
    """
    Word-boundary-aware match of `cleaned_name` (already run through
    clean_company_name()) against `title` — not a plain substring check.
    This matters most for short names: BP is on the watchlist with
    exactly that 2-character name, and a plain substring check would
    incorrectly accept "GBP falls against the dollar" (the letters "bp"
    appear inside "GBP", but BP the company is never mentioned).
    Case-insensitive. Regex-special characters in the name (e.g. the
    "&" in "Legal & General", the "." some tickers carry) are escaped,
    since they're literal characters in a company name, not regex
    syntax. Checks every legal-suffix variant from
    _relevance_name_variants, not just the exact stored form.
    """
    if not cleaned_name or not title:
        return False
    title_lower = title.lower()
    for variant in _relevance_name_variants(cleaned_name):
        pattern = r"\b" + re.escape(variant.lower()) + r"\b"
        if re.search(pattern, title_lower):
            return True
    return False


# Small, fixed, auditable set of well-known idiomatic phrases that
# specifically signal a NON-company usage of a high-collision-risk name —
# used ONLY as the second check for names in HIGH_COLLISION_DISQUALIFIER_PHRASES, never as a
# relevance signal on its own.
#
# This inverts an earlier, rejected design (a positive "finance context
# word" whitelist): that approach tried to enumerate every possible
# LEGITIMATE company-news topic, an unbounded and inevitably incomplete
# list — confirmed directly when it rejected obviously genuine headlines
# like "Shell announces new refinery investment" and "BP faces investor
# lawsuit over pipeline leak" purely for not containing an earnings/
# ratings-style word. The set of "ways a company can legitimately be in
# the news" is unbounded; the set of "common English idioms that happen
# to use this specific word" is small and can realistically approach
# completeness. A word-boundary match on a high-collision name is now
# ACCEPTED BY DEFAULT — exactly as permissive as any ordinary company
# name — UNLESS the headline also contains one of these known
# disqualifying phrases.
#
# "Shell" and "BP" get their OWN separate phrase sets, deliberately not
# shared — they are different KINDS of collision risk. "Shell" is an
# ordinary English noun/verb with a small, genuinely closed set of
# common idiomatic uses (shell script, shell company, shell shock,
# seashell, artillery shell, nutshell — there are only so many). "BP" is
# a short initialism that collides with OTHER short initialisms and
# jargon (blood pressure, basis points) — and unlike Shell's idioms,
# bare medical "BP" usage ("High BP linked to heart disease risk") is
# structurally more open-ended than Shell's closed idiom set, so this
# list is knowingly less complete than Shell's; that's a disclosed
# property of the collision type, not an oversight.
HIGH_COLLISION_DISQUALIFIER_PHRASES = {
    "shell": {
        "shell script", "shell scripting", "bash shell", "unix shell",
        "shell command", "shell prompt", "command shell",
        "shell company", "shell corporation", "shell corporations",
        "shell shock", "shellshocked", "shell-shocked",
        "sea shell", "seashell", "seashells",
        "turtle shell", "egg shell", "eggshell",
        "artillery shell", "mortar shell", "shell fire", "shelling",
        "nutshell",
    },
    "bp": {
        "blood pressure", "bp reading", "bp level", "bp levels", "bp monitor",
        "bp measurement", "basis points", "basis point",
        "high bp", "low bp", "checking bp", "monitor bp", "monitor your bp",
        "heart disease", "hypertension", "doctors",
    },
}


def _has_disqualifying_phrase(title, cleaned_name):
    """
    Plain substring check (deliberately NOT word-boundary, unlike
    _is_relevant_to_company) against the small, explicit disqualifier
    phrase set for this collision-risk name. Substring matching is safe
    here — these are specific, multi-word strings with negligible risk
    of appearing embedded inside an unrelated legitimate headline by
    coincidence — and it's necessary: word-boundary matching on "shell
    script" would miss "shell scripting" (no boundary between "script"
    and "ing", both being word characters), incorrectly letting that
    headline through. The asymmetry (word-boundary for the positive
    company-name check, substring for the negative disqualifier check)
    is intentional, not an inconsistency: the positive check needed
    word-boundary because short names like "BP" have real embedding
    risk (GBP); these disqualifier phrases are long and specific enough
    that the same risk doesn't apply.
    """
    if not title or not cleaned_name:
        return False
    phrases = HIGH_COLLISION_DISQUALIFIER_PHRASES.get(cleaned_name.lower(), set())
    if not phrases:
        return False
    title_lower = title.lower()
    return any(phrase in title_lower for phrase in phrases)


def passes_relevance_filter(title, cleaned_name, fetch_source=None):
    """
    THE single relevance gate — used identically everywhere a title
    needs judging against a company name: every fetch source (g/rb/
    ratings/y) and carried-forward stored items being re-evaluated on a
    later run.

    This is a PURE function of (title, cleaned_name) alone — fetch_source
    is accepted only for optional diagnostic tagging and never changes
    the outcome. That's deliberate: it's what guarantees the same
    headline and the same company always produce the same relevance
    result, whether the item was just fetched (from any source) or is
    being carried forward from an earlier run.

    Two earlier designs were tried and rejected here, for reference:
    (1) a blanket rejection of every Yahoo-sourced match for
    high-collision names — simple, but too broad, discarding genuine
    headlines purely for arriving via one source; (2) a positive
    "finance context word" whitelist — closer, but an unbounded,
    inevitably incomplete list that rejected obviously genuine
    headlines like investment/legal/operational news. Replaced with a
    small, disclosed disqualifier-phrase check (see
    HIGH_COLLISION_DISQUALIFIER_PHRASES) — accept by default, reject
    only on a known non-company idiom.

    For ordinary names, a word-boundary company-name match
    (_is_relevant_to_company) is sufficient on its own. For names
    flagged in HIGH_COLLISION_DISQUALIFIER_PHRASES, that match is
    necessary but not sufficient: the headline must NOT also contain
    one of that name's own disqualifying phrases.
    """
    if not _is_relevant_to_company(title, cleaned_name):
        return False
    if cleaned_name and cleaned_name.lower() in HIGH_COLLISION_DISQUALIFIER_PHRASES:
        return not _has_disqualifying_phrase(title, cleaned_name)
    return True


def revalidate_stored_news_items(stored_items, current_name, date_filter_fn=None):
    """
    Re-validates a ticker's carried-forward news items against BOTH
    today's date filter (passes_news_filters) AND today's relevance
    rule (passes_relevance_filter) — not date alone. Previously, only
    date was re-checked on carry-forward, so an item that would fail
    relevance if fetched fresh today kept appearing indefinitely as
    long as its date stayed "today": confirmed live with an ABDN item
    about a bond coupon schedule and a SHEL item about an unrelated
    Allianz/AA takeover, neither mentioning the company at all.

    Each item is judged purely from its OWN recorded title/company/
    fetchSource — never from whether THIS run's fresh fetch of any
    source succeeded. A temporarily-unavailable source (a 503, a
    timeout) has zero effect here: it cannot cause a genuinely
    relevant, already-stored item to be discarded, because this
    function never looks at this run's fetch results at all.

    Items with no recorded fetchSource (stored before this fix existed)
    are judged on relevance alone, with no source-specific refinement —
    there's no source to apply one to. Items tagged "analyst" (Yahoo's
    structured analyst-history API, ticker-scoped by construction, not
    a text search) are exempt from relevance re-checking, same as at
    fetch time.

    date_filter_fn: which date/recency check to apply — defaults to
    passes_news_filters (same-day), preserving exact existing behaviour
    for every current caller. Pass passes_recency_filter_wide instead
    to revalidate the wider "most recent available" fallback pool
    against its own (looser) recency rule — relevance checking is
    identical either way.
    """
    date_filter_fn = date_filter_fn or passes_news_filters
    kept = []
    for it in stored_items:
        if not date_filter_fn(it.get("pubDate")):
            continue
        fetch_source = it.get("fetchSource")
        if fetch_source == "analyst":
            kept.append(it)
            continue
        item_company = it.get("company") or current_name
        cleaned = clean_company_name(item_company)
        if passes_relevance_filter(it.get("title", ""), cleaned, fetch_source):
            kept.append(it)
    return kept


def google_news_url(company_name):
    q = f'{company_name} (LSE OR "London Stock Exchange") (upgrade OR downgrade OR "price target" OR rating)'
    return "https://news.google.com/rss/search?" + urllib.parse.urlencode(
        {"q": q, "hl": "en-GB", "gl": "GB", "ceid": "GB:en"}
    )


def general_news_url(company_name):
    """Broader than google_news_url — not restricted to rating/broker keywords, so it
    catches general company news (earnings, M&A, trading updates) for stocks that show
    up in the screener but aren't on the watchlist."""
    q = f'{company_name} (LSE OR "London Stock Exchange")'
    return "https://news.google.com/rss/search?" + urllib.parse.urlencode(
        {"q": q, "hl": "en-GB", "gl": "GB", "ceid": "GB:en"}
    )


# FT's own official RSS feeds — confirmed via a real captured feed file
# (ft.com/rss/home/international), not just a third-party listing.
# Deliberately just these two GENERAL feeds, not company-specific search
# (FT doesn't offer one) and not sector feeds (which would systematically
# under-cover the watchlist's actual sector spread). Headlines + summary
# only, matching the same copyright-safe depth every other source here
# already uses — never full paywalled article text.
FT_MARKETS_URL = "https://www.ft.com/markets?format=rss"
FT_INTERNATIONAL_URL = "https://www.ft.com/rss/home/international"


def reuters_bloomberg_url(company_name):
    # Reuters dropped public RSS years ago and Bloomberg is subscription-walled, so
    # neither offers a free feed directly. This scopes the same free Google News search
    # to just those two publishers (site: filter) — legitimate, no scraping either site,
    # same mechanism already used for the general news search.
    q = f'{company_name} (LSE OR "London Stock Exchange") (site:reuters.com OR site:bloomberg.com)'
    return "https://news.google.com/rss/search?" + urllib.parse.urlencode(
        {"q": q, "hl": "en-GB", "gl": "GB", "ceid": "GB:en"}
    )


def market_wide_broker_news_url():
    # NOT scoped to a company name — this is what makes it market-wide rather than
    # watchlist-only. One request per poll cycle covers every LSE-listed company's
    # broker rating news, instead of the (infeasible) approach of adding all ~1,900
    # LSE companies to the watchlist and polling each individually.
    q = '(LSE OR "London Stock Exchange") (upgrade OR downgrade OR "price target" OR "rating") (broker OR analyst OR bank)'
    return "https://news.google.com/rss/search?" + urllib.parse.urlencode(
        {"q": q, "hl": "en-GB", "gl": "GB", "ceid": "GB:en"}
    )


def yahoo_symbol(ticker):
    t = ticker.strip().upper()
    return t if t.endswith(".L") else f"{t.rstrip('.')}.L"


def bare_ticker(symbol):
    """
    Normalizes a ticker to its bare, suffix-free form so the SAME
    underlying stock can be recognized across different key formats —
    e.g. a screener symbol "GLEN.L" and a watchlist ticker "GLEN" both
    normalize to "GLEN". Strips a trailing ".L" LSE suffix, then any
    leftover trailing dot — so "BP." (a real, literal trailing dot in
    that specific ticker, not a stripped suffix) still compares equal to
    "BP.L"'s bare form "BP", not to a literal "BP." Extracted from logic
    that was previously duplicated inline in two places (both computing
    the same bare form for the same reason) into one shared, tested
    function — the exact match this project has consistently preferred
    over risking the two copies drifting apart.
    """
    s = symbol.upper()
    if s.endswith(".L"):
        s = s.rsplit(".L", 1)[0]
    return s.rstrip(".")


def discover_radar_stocks(watchlist, big_movers, screener, latest_broker_events=None):
    """
    Scans every EXISTING discovery source — Watchlist, Heat Map
    (big_movers), Screener Volume/Gainers/Losers, and market-wide Broker
    Research (latest_broker_events, which genuinely covers tickers
    beyond Watchlist/Screener — see get_latest_broker_event_per_ticker's
    own docstring) — and returns an ordered dict, keyed by bare_ticker
    (so "GLEN" and "GLEN.L" merge into one entry), of every stock any of
    them found. Each entry is {"ticker": <the specific ticker string to
    use for lookups>, "name": ..., "sources": [(label, reason_or_None),
    ...]}. A stock appearing in multiple sources gets multiple (label,
    reason) pairs, never duplicated into multiple entries.

    Every "reason" is derived directly from data ALREADY computed
    elsewhere for that exact source (the same changePct/volume numbers
    the Heat Map or Screener sections already display) — never
    invented, never a second, independent calculation.

    NEWS and AI EVIDENCE are deliberately NOT independent discovery
    sources here: this project's news fetching is per-ticker (it
    searches for news ABOUT a ticker already known from another source),
    not a whole-market news scan that could discover an unknown ticker —
    and AI Evidence Review only ever runs on an already-identified
    stock, for the same structural reason. Treating either as a
    "discovery source" would misrepresent what's actually happening;
    both remain fully shown as EVIDENCE for stocks discovered another
    way (see render_stock_research_html), just never claimed as a
    discovery route in their own right.

    This function only DISCOVERS which stocks and why — it renders
    nothing and computes no evidence itself, keeping discovery cleanly
    separate from the (existing, reused) evidence rendering that
    happens elsewhere.
    """
    discoveries = {}

    def add(ticker, name, label, reason):
        if not ticker:
            return
        key = bare_ticker(ticker)
        if key not in discoveries:
            discoveries[key] = {"ticker": ticker, "name": name or ticker, "sources": []}
        discoveries[key]["sources"].append((label, reason))

    for stock in watchlist:
        add(stock.get("ticker"), stock.get("name"), "Watchlist", None)

    for entry in big_movers:
        ticker = entry.get("ticker")
        chg = entry.get("changePct")
        reason = f"{'▲' if (chg or 0) >= 0 else '▼'}{abs(chg):.1f}% move today" if chg is not None else None
        add(ticker, entry.get("name"), "Heat Map", reason)

    screener_source_labels = {"volume": "LSE Volume", "gainers": "LSE Gainers", "losers": "LSE Losers"}
    for section_key, label in screener_source_labels.items():
        for entry in screener.get(section_key, []):
            ticker = entry.get("symbol")
            chg = entry.get("changePct")
            vol = entry.get("volume")
            if section_key == "volume":
                reason = f"volume {vol:,}" if isinstance(vol, (int, float)) else None
            elif chg is not None:
                reason = f"{'▲' if chg >= 0 else '▼'}{abs(chg):.2f}% today"
            else:
                reason = None
            add(ticker, entry.get("name"), label, reason)

    for ticker, event in (latest_broker_events or {}).items():
        action = event.get("normalizedAction")
        broker = event.get("broker")
        if action and broker:
            reason = f"{broker} {action.lower()}"
        elif action:
            reason = action.lower()
        else:
            reason = None
        add(ticker, event.get("company") or ticker, "Broker Research", reason)

    return discoveries


# --- Live Radar: persistent cross-poll-cycle lifecycle tracking ---------
# Models the SAME first_seen/last_seen-preserved-forever discipline already
# proven in load_events_store/enrich_stored_event (see that function's own
# docstring) — never overwriting or deleting history, only ever adding to
# it or bumping last_seen when a stock is genuinely re-discovered.
RADAR_HISTORY_FILE = os.path.join(STATE_DIR, "radar_history.json")
RADAR_HISTORY_VERSION = 1
RADAR_AGING_THRESHOLD_MINUTES = 60  # not re-discovered by any source in
# this long -> AGING (still shown prominently, clearly flagged)
RADAR_STALE_THRESHOLD_MINUTES = 240  # not re-discovered in this long ->
# STALE below this, CLEARED beyond it (the historical record in
# radar_history.json is still preserved, per the explicit "don't
# immediately erase it" requirement — CLEARED just means "no longer
# shown as if it were current" in the LIVE section)
RADAR_HISTORY_RETENTION_DAYS = 90  # how long a CLEARED entry's full
# discovery record is kept before being pruned from the persisted store
# entirely, to avoid unbounded growth


class RadarHistoryCorruptError(Exception):
    """Same discipline as every other state file here — a file that
    EXISTS but fails to parse must never be silently treated as empty."""
    pass


def load_radar_history(path=None):
    """Missing file -> fresh empty store (legitimate first-run case).
    Existing-but-corrupt file -> raises, never silently treated as
    empty — same contract as load_events_store/load_daily_snapshots."""
    path = path or RADAR_HISTORY_FILE
    if not os.path.exists(path):
        return {"version": RADAR_HISTORY_VERSION, "stocks": {}}
    with open(path, "r", encoding="utf-8") as f:
        raw = f.read()
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        raise RadarHistoryCorruptError(f"{path} contains invalid JSON: {e}") from e
    if not isinstance(data, dict) or not isinstance(data.get("stocks"), dict):
        raise RadarHistoryCorruptError(f"{path} does not match the expected {{version, stocks}} shape")
    return data


def compute_radar_status(first_seen_iso, last_seen_iso, now_iso, discovered_this_run):
    """
    Pure function: NEW / ACTIVE / AGING / STALE / CLEARED — never a
    guess, always derived from the actual timestamps. NEW and ACTIVE
    both mean "discovered again THIS run" (freshly confirmed); the
    distinction is only whether this is the very first time ever
    (first_seen == this run) or a continuing detection. AGING/STALE/
    CLEARED all mean "NOT re-discovered this run", escalating purely by
    how long ago it was last genuinely re-detected by any source —
    exactly the configured thresholds above, nothing implicit.
    """
    now_dt = datetime.fromisoformat(now_iso)
    if discovered_this_run:
        return "NEW" if first_seen_iso == last_seen_iso == now_iso else "ACTIVE"
    last_seen_dt = datetime.fromisoformat(last_seen_iso)
    age_minutes = (now_dt - last_seen_dt).total_seconds() / 60
    if age_minutes < RADAR_AGING_THRESHOLD_MINUTES:
        return "AGING"
    if age_minutes < RADAR_STALE_THRESHOLD_MINUTES:
        return "STALE"
    return "CLEARED"


def format_radar_age(now_iso, since_iso):
    """Human-readable age string ('5 minutes', '2 hours', '3 days') —
    never fabricated: a negative or malformed timestamp pair returns
    None rather than a nonsensical age, so the caller can show an
    honest 'unavailable' instead of a wrong number."""
    try:
        now_dt = datetime.fromisoformat(now_iso)
        since_dt = datetime.fromisoformat(since_iso)
    except (ValueError, TypeError):
        return None
    delta_seconds = (now_dt - since_dt).total_seconds()
    if delta_seconds < 0:
        return None  # a future timestamp is invalid data, never displayed as an age
    minutes = delta_seconds / 60
    if minutes < 1:
        return "less than a minute"
    if minutes < 60:
        return f"{int(minutes)} minute{'s' if int(minutes) != 1 else ''}"
    hours = minutes / 60
    if hours < 24:
        return f"{hours:.1f} hour{'s' if hours != 1 else ''}"
    days = hours / 24
    return f"{days:.1f} day{'s' if days != 1 else ''}"


def merge_radar_history(history, radar_discovery, now_iso):
    """
    Pure function — takes the LOADED prior history and THIS run's fresh
    discovery, returns a NEW, updated history dict (never mutates the
    input) plus a per-ticker lifecycle dict for rendering. Reused
    discipline from enrich_stored_event: a ticker not yet in history is
    added with first_seen=last_seen=now; a ticker already in history
    that's discovered again gets last_seen bumped to now and its
    sourcesEverSeen UNIONED (never shrunk) with this run's sources; a
    ticker in history but NOT discovered this run is left completely
    untouched here — its last_seen stays in the past, which is exactly
    what lets its age (and therefore its AGING/STALE/CLEARED status)
    grow naturally on subsequent runs, with NO explicit "mark as
    aging" step needed anywhere.

    CLEARED entries older than RADAR_HISTORY_RETENTION_DAYS are pruned
    from the returned store entirely — the only case this function ever
    removes a record, and only after it's been long enough that the
    live "don't immediately erase it" requirement is well satisfied.
    """
    new_stocks = {k: dict(v) for k, v in history.get("stocks", {}).items()}
    lifecycle = {}

    for key, disco in radar_discovery.items():
        sources_this_run = sorted({label for label, _reason in disco["sources"]})
        if key not in new_stocks:
            new_stocks[key] = {
                "ticker": disco["ticker"], "name": disco["name"],
                "firstSeen": now_iso, "lastSeen": now_iso,
                "sourcesEverSeen": sources_this_run,
            }
        else:
            existing = new_stocks[key]
            existing["lastSeen"] = now_iso
            existing["ticker"] = disco["ticker"]
            existing["name"] = disco["name"] or existing.get("name")
            existing["sourcesEverSeen"] = sorted(set(existing.get("sourcesEverSeen", [])) | set(sources_this_run))
        record = new_stocks[key]
        status = compute_radar_status(record["firstSeen"], record["lastSeen"], now_iso, discovered_this_run=True)
        lifecycle[key] = {
            "firstSeen": record["firstSeen"], "lastSeen": record["lastSeen"], "status": status,
            "sourcesEverSeen": record["sourcesEverSeen"], "sourcesActiveThisRun": sources_this_run,
            "age": format_radar_age(now_iso, record["firstSeen"]),
        }

    # Stocks in history but NOT discovered this run — left untouched
    # (first_seen/last_seen never modified here), just given a computed
    # status/age for rendering purposes and pruned if long CLEARED.
    pruned_stocks = {}
    for key, record in new_stocks.items():
        if key in lifecycle:
            pruned_stocks[key] = record
            continue
        status = compute_radar_status(record["firstSeen"], record["lastSeen"], now_iso, discovered_this_run=False)
        if status == "CLEARED":
            try:
                last_seen_dt = datetime.fromisoformat(record["lastSeen"])
                now_dt = datetime.fromisoformat(now_iso)
                if (now_dt - last_seen_dt).days > RADAR_HISTORY_RETENTION_DAYS:
                    continue  # pruned entirely — the only case this function ever removes a record
            except (ValueError, TypeError):
                pass
        pruned_stocks[key] = record
        lifecycle[key] = {
            "firstSeen": record["firstSeen"], "lastSeen": record["lastSeen"], "status": status,
            "sourcesEverSeen": record.get("sourcesEverSeen", []), "sourcesActiveThisRun": [],
            "age": format_radar_age(now_iso, record["firstSeen"]),
        }

    return {"version": RADAR_HISTORY_VERSION, "stocks": pruned_stocks}, lifecycle


def save_radar_history(history, path=None):
    """Straightforward atomic write — the merge/pruning logic above is
    where the actual discipline lives; this just persists whatever
    merge_radar_history already computed."""
    path = path or RADAR_HISTORY_FILE
    atomic_write_json(path, history)
    return {"written": True, "stockCount": len(history.get("stocks", {}))}


def yahoo_news_url(ticker):
    return "https://feeds.finance.yahoo.com/rss/2.0/headline?" + urllib.parse.urlencode(
        {"s": yahoo_symbol(ticker), "region": "UK", "lang": "en-GB"}
    )


def fetch_ftse100():
    """The overall market headline every major finance site leads with — a single call
    once per cycle (not per-stock), for the ^FTSE index itself."""
    url = "https://query1.finance.yahoo.com/v8/finance/chart/%5EFTSE?interval=1d&range=5d"
    try:
        data = json.loads(http_get(url))
        result = (data.get("chart") or {}).get("result") or [None]
        if not result[0]:
            return None
        meta = result[0].get("meta", {})
        price = meta.get("regularMarketPrice")
        prev_close = meta.get("chartPreviousClose") or meta.get("previousClose")
        if price is None or not prev_close:
            return None
        change = price - prev_close
        change_pct = change / prev_close * 100
        return {"price": price, "change": change, "changePct": change_pct}
    except Exception as e:
        print(f"  ! FTSE 100 fetch failed: {e}", file=sys.stderr)
        return None


def fetch_yahoo_quote(ticker):
    symbol = yahoo_symbol(ticker)
    url = f"https://query1.finance.yahoo.com/v8/finance/chart/{urllib.parse.quote(symbol)}?interval=1m&range=1d"
    try:
        data = json.loads(http_get(url))
        result = (data.get("chart") or {}).get("result") or [None]
        meta = (result[0] or {}).get("meta") if result[0] else None
        if not meta or "regularMarketPrice" not in meta:
            return None
        price = meta["regularMarketPrice"]
        prev_close = meta.get("previousClose") or meta.get("chartPreviousClose")
        change = price - prev_close if prev_close else None
        change_pct = (change / prev_close * 100) if (prev_close and change is not None) else None
        day_high = meta.get("regularMarketDayHigh")
        day_low = meta.get("regularMarketDayLow")
        # Intraday range as % of previous close — a purely descriptive volatility
        # measure (how much the stock has already swung today), not a prediction.
        range_pct = ((day_high - day_low) / prev_close * 100) if (day_high and day_low and prev_close) else None
        return {
            "price": price,
            "currency": meta.get("currency", "GBp"),
            "change": change,
            "changePct": change_pct,
            "dayHigh": day_high,
            "dayLow": day_low,
            "rangePct": range_pct,
            "asOf": meta.get("regularMarketTime", int(time.time())) * 1000,
        }
    except Exception as e:
        print(f"  ! yahoo quote failed: {ticker} ({e})", file=sys.stderr)
        return None


def compute_rsi(closes, period=14):
    """Standard RSI (relative strength index) — a raw computed statistic from real
    closing prices, shown as a number, not labeled 'overbought'/'oversold' or turned
    into a recommendation. That labeling is exactly the kind of implied-advice framing
    this tool avoids; the number itself is just arithmetic on public price history."""
    if len(closes) < period + 1:
        return None
    deltas = [closes[i] - closes[i - 1] for i in range(1, len(closes))]
    recent = deltas[-period:]
    gains = [d for d in recent if d > 0]
    losses = [-d for d in recent if d < 0]
    avg_gain = sum(gains) / period
    avg_loss = sum(losses) / period
    if avg_loss == 0:
        return 100.0
    rs = avg_gain / avg_loss
    return 100 - (100 / (1 + rs))


CHART_HISTORY_DAYS = 30  # deliberately shorter than the 1y fetch range —
# a compact glance, not a full trading chart. Retained from the SAME
# fetch already happening for RSI/MA/ATR, never a second request.


def fetch_price_technicals(ticker):
    """Real, already-happened price history — 5-day % change, RSI(14), moving
    averages (20/50/200-day, each only computed when enough real history exists),
    a simple MA20-vs-MA50 crossover state, and ATR(14) — all computed from the
    SAME single chart fetch. Facts about the past, not predictions.

    Range extended from 3 months to 1 year specifically so 50/200-day MAs become
    computable when a stock has that much history — this increases the response
    SIZE (more bars in the same one HTTP call), not the number of network calls
    made; RSI/5-day-change/20-day-MA are unaffected, since they still just take
    the LAST N closes regardless of how much more history precedes them.

    Also extracts the latest day's volume, high, and low from the SAME response
    — verified against real Yahoo API response structure (documented in multiple
    independent open-source tools that scrape this exact endpoint) to contain
    `volume`/`high`/`low` arrays alongside `close`, aligned by index with the
    same `timestamp` array. All four are paired and filtered TOGETHER (not
    filtered separately) before taking values — filtering `close` alone for
    nulls (holidays etc.) would silently desynchronize which day's volume/high/
    low ends up attached to a given close if a null close and a present value
    (or vice versa) ever occurred on the same raw index.
    """
    symbol = yahoo_symbol(ticker)
    url = f"https://query1.finance.yahoo.com/v8/finance/chart/{urllib.parse.quote(symbol)}?interval=1d&range=1y"
    try:
        data = json.loads(http_get(url))
        result = (data.get("chart") or {}).get("result") or [None]
        if not result[0]:
            return None
        quote0 = (result[0].get("indicators", {}).get("quote", [{}])[0] or {})
        raw_closes = quote0.get("close") or []
        raw_volumes = quote0.get("volume") or []
        raw_highs = quote0.get("high") or []
        raw_lows = quote0.get("low") or []
        paired = [
            (c, v, h, l) for c, v, h, l in zip(raw_closes, raw_volumes, raw_highs, raw_lows)
            if c is not None
        ]
        closes = [c for c, _v, _h, _l in paired]  # some days can come back null (holidays etc.)
        if len(closes) < 6:
            return None  # not enough real trading days to compute even the 5-day change
        latest = closes[-1]
        latest_volume = paired[-1][1] if paired else None
        five_days_ago = closes[-6]
        change_pct = (latest - five_days_ago) / five_days_ago * 100 if five_days_ago else None
        rsi14 = compute_rsi(closes, 14)
        ma20 = sum(closes[-20:]) / len(closes[-20:]) if len(closes) >= 20 else None
        ma50 = sum(closes[-50:]) / 50 if len(closes) >= 50 else None
        ma200 = sum(closes[-200:]) / 200 if len(closes) >= 200 else None
        ma_crossover = None
        if ma20 is not None and ma50 is not None:
            ma_crossover = "bullish" if ma20 > ma50 else ("bearish" if ma20 < ma50 else "flat")
        atr14 = compute_atr(paired, 14)
        highs = [h for _c, _v, h, _l in paired]
        lows = [l for _c, _v, _h, l in paired]
        support_resistance = compute_support_resistance(highs, lows)
        breakout_status = compute_breakout_status(latest, highs[:-1], lows[:-1])
        # Retains the LAST 30 trading days' close+volume from the SAME
        # `paired` list already built above for RSI/MA/ATR — zero
        # additional network calls, and every existing calculation above
        # this line (rsi14/ma20/ma50/ma200/atr14/support_resistance/
        # breakout_status) is completely untouched by this addition.
        # Used only for the Phase 5 compact price/volume sparkline chart.
        price_volume_series = [{"close": c, "volume": v} for c, v, _h, _l in paired[-CHART_HISTORY_DAYS:]]
        return {
            "changePct5d": change_pct,
            "price": latest,
            "rsi14": rsi14,
            "ma20": ma20,
            "ma50": ma50,
            "ma200": ma200,
            "maCrossover": ma_crossover,
            "atr14": atr14,
            "supportResistance": support_resistance,
            "breakoutStatus": breakout_status,
            "aboveMA20": (latest > ma20) if ma20 else None,
            "latestVolume": latest_volume,
            "priceVolumeSeries": price_volume_series,
        }
    except Exception as e:
        print(f"  ! price technicals fetch failed: {ticker} ({e})", file=sys.stderr)
        return None


def compute_atr(paired_close_vol_high_low, period=14):
    """
    Average True Range over `period` days — a standard, deterministic
    volatility measure (how much a stock typically moves day-to-day),
    NOT a prediction. True range for a day = max(high-low,
    |high-prev_close|, |low-prev_close|); ATR = the average of the most
    recent `period` true-range values. Returns None if there isn't
    enough paired (close, volume, high, low) history to compute even
    one true-range value plus the period average — never fabricated
    from partial data.
    """
    if len(paired_close_vol_high_low) < period + 1:
        return None
    true_ranges = []
    for i in range(1, len(paired_close_vol_high_low)):
        prev_close = paired_close_vol_high_low[i - 1][0]
        _c, _v, high, low = paired_close_vol_high_low[i]
        if high is None or low is None or prev_close is None:
            continue
        tr = max(high - low, abs(high - prev_close), abs(low - prev_close))
        true_ranges.append(tr)
    if len(true_ranges) < period:
        return None
    return sum(true_ranges[-period:]) / period


SUPPORT_RESISTANCE_WINDOW_DAYS = 20  # standard Donchian-channel-style window


def compute_support_resistance(highs, lows, window=SUPPORT_RESISTANCE_WINDOW_DAYS):
    """
    Deterministic support/resistance via the standard N-day rolling
    high/low convention (a Donchian-channel-style definition, not a
    subjective reading of chart patterns): resistance = the highest high
    over the last `window` days (including today); support = the lowest
    low over the same window. A fixed, publicly documented, reproducible
    rule — investigated specifically because the brief asked whether an
    OBJECTIVE definition exists before implementing anything here.
    Returns None if fewer than `window` days of high/low data exist —
    never approximated from a shorter window.
    """
    recent_highs = highs[-window:]
    recent_lows = lows[-window:]
    if len(recent_highs) < window or len(recent_lows) < window:
        return None
    valid_highs = [h for h in recent_highs if h is not None]
    valid_lows = [l for l in recent_lows if l is not None]
    if len(valid_highs) < window or len(valid_lows) < window:
        return None
    return {"resistance": max(valid_highs), "support": min(valid_lows)}


def compute_breakout_status(latest_close, prior_highs, prior_lows, window=SUPPORT_RESISTANCE_WINDOW_DAYS):
    """
    Deterministic breakout/breakdown status. A "breakout" is today's
    close exceeding the highest high of the `window` days PRECEDING
    today (today itself deliberately excluded from that comparison —
    comparing a value against a window that includes itself is
    circular and would make "breakout" nearly meaningless). A
    "breakdown" is today's close falling below the lowest low of the
    same preceding window. Same standard, reproducible convention as
    compute_support_resistance, just applied as a threshold test rather
    than reporting the levels themselves. Returns None if there isn't
    enough preceding history — never guessed from a partial window.
    """
    recent_highs = prior_highs[-window:]
    recent_lows = prior_lows[-window:]
    if len(recent_highs) < window or len(recent_lows) < window or latest_close is None:
        return None
    valid_highs = [h for h in recent_highs if h is not None]
    valid_lows = [l for l in recent_lows if l is not None]
    if len(valid_highs) < window or len(valid_lows) < window:
        return None
    if latest_close > max(valid_highs):
        return "breakout"
    if latest_close < min(valid_lows):
        return "breakdown"
    return "within_range"


BIG_MOVER_THRESHOLD_PCT = 5.0  # purely descriptive flag: "this has already moved a lot today"
UPTREND_5DAY_THRESHOLD_PCT = 5.0  # flagged as "5-day uptrend" once risen at least this much
# Gainers/losers liquidity filter — without this, illiquid penny stocks with meaningless
# tiny-absolute-value swings (0.1p -> 1p = "900%") dominate the list.
MIN_VOLUME_FOR_MOVERS = 500_000
MIN_PRICE_PENCE_FOR_MOVERS = 20  # excludes sub-20p penny stocks specifically from gainers/losers
# News/broker items older than this are considered stale and filtered from the live feed —
# Google News search returns "most relevant," not "most recent," so old syndicated pieces
# (sometimes years old) can otherwise resurface in results.


def fetch_gb_screener(sort_field, sort_type="DESC", count=10):
    """
    Generic LSE-scoped screener via Yahoo Finance's free public endpoint — same source
    already used for top-volume, generalized to any sortable field (dayvolume,
    percentchange, etc). Requires the crumb-authenticated opener; Yahoo 401s this
    endpoint for anonymous requests from cloud/datacenter IPs.

    Yahoo's "region: gb" tag catches thousands of obscure GDRs/depositary receipts
    (symbols ending ".IL" etc) alongside genuine LSE-listed shares — these are thinly
    traded, so their %-change numbers are noise (a move from 0.1p to 1p shows as
    "900%"). Filtered here to only ".L"-suffixed symbols, the standard LSE ticker
    format, and over-fetched (4x) since a chunk gets filtered out.
    """
    crumb = get_yahoo_crumb()
    url = "https://query1.finance.yahoo.com/v1/finance/screener"
    if crumb:
        url += f"?crumb={urllib.parse.quote(crumb)}"
    # Illiquid penny stocks dominate the top of a raw %-change sort by definition —
    # a small over-fetch (e.g. 4x) means genuinely liquid mid/large-caps can be buried
    # past the fetch window entirely, leaving nothing after the liquidity filter runs.
    # Gainers/losers need a much bigger pool to filter from than Volume does.
    fetch_size = count * 50 if sort_field == "percentchange" else count * 4
    fetch_size = min(fetch_size, 250)  # stay within what the endpoint reliably accepts
    body = json.dumps({
        "size": fetch_size,
        "offset": 0,
        "sortField": sort_field,
        "sortType": sort_type,
        "quoteType": "EQUITY",
        "query": {"operator": "eq", "operands": ["region", "gb"]},
    }).encode("utf-8")
    req = urllib.request.Request(
        url, data=body, method="POST",
        headers={**HEADERS, "Content-Type": "application/json"},
    )
    try:
        data = json.loads(_yahoo_opener.open(req, timeout=15).read())
        quotes = (((data.get("finance") or {}).get("result") or [{}])[0]).get("quotes", [])
        out = []
        for q in quotes:
            symbol = q.get("symbol", "")
            if not symbol.upper().endswith(".L"):
                continue  # skip GDRs/IOB instruments etc — keep genuine LSE-listed shares only
            # LSE International Order Book tickers (depositary receipts for FOREIGN
            # companies, not genuine UK-listed businesses) conventionally start with a
            # digit — e.g. "0R9U.L" is PayPal Holdings, confirmed by its news being
            # entirely US-market coverage. Excluding these keeps results to genuine LSE
            # primary listings, matching what "LSE-listed stocks only" actually promises.
            ticker_part = symbol.upper().rsplit(".L", 1)[0]
            if ticker_part[:1].isdigit():
                continue
            volume = q.get("regularMarketVolume")
            price = q.get("regularMarketPrice")
            if sort_field == "dayvolume" and not volume:
                continue  # skip zero/missing-volume noise from the volume ranking specifically
            if sort_field == "percentchange":
                # Gainers/losers dominated by illiquid penny stocks: a move from 0.1p to
                # 1p shows as "900%" and is meaningless. Require real liquidity and a
                # non-penny price so the list reflects stocks actually worth noticing.
                if not volume or volume < MIN_VOLUME_FOR_MOVERS:
                    continue
                if not price or price < MIN_PRICE_PENCE_FOR_MOVERS:
                    continue
            out.append({
                "symbol": symbol,
                # Yahoo's "shortName" is known to be truncated at a fixed length for
                # longer UK share-class names (confirmed on the live dashboard: names
                # were cut off mid-word with no ellipsis, e.g. "PREMIER AFRICAN
                # MINERALS LIMITE"). "longName" is the fuller field Yahoo's screener
                # response also carries; preferring it when present is a strict
                # improvement — falls through to the exact same behavior as before
                # when a given quote doesn't happen to include it.
                "name": q.get("longName") or q.get("shortName", symbol),
                "volume": volume,
                "price": price,
                "changePct": q.get("regularMarketChangePercent"),
            })
            if len(out) >= count:
                break
        return out, "ok"
    except Exception as e:
        print(f"  ! screener failed (sortField={sort_field}): {e}", file=sys.stderr)
        return [], "failed"


def fetch_lse_screener(raw_count=10):
    """
    Returns (screener_dict, status_dict) — deliberately SEPARATE from each
    other. Before this, fetch_gb_screener returned an empty list for BOTH
    "the fetch worked and genuinely found nothing" and "the fetch broke
    entirely" — indistinguishable to any caller, and the only trace of a
    real failure was a line in the server log nobody looking at the
    dashboard would ever see. status_dict lets the rendered page show a
    real fetch failure as a real fetch failure, distinct from "nothing to
    show right now" — the person using this dashboard should never have
    to guess which one they're looking at.
    """
    volume_rows, volume_status = fetch_gb_screener("dayvolume", "DESC", raw_count)
    gainers_rows, gainers_status = fetch_gb_screener("percentchange", "DESC", raw_count)
    losers_rows, losers_status = fetch_gb_screener("percentchange", "ASC", raw_count)
    return (
        {"volume": volume_rows, "gainers": gainers_rows, "losers": losers_rows},
        {"volume": volume_status, "gainers": gainers_status, "losers": losers_status},
    )


# =========================================================================
# LSE first-party market data — api.londonstockexchange.com/api/v1/
# components/refresh, the SAME endpoint the live LSE website itself calls
# to render its Risers/Fallers/Volume Leaders and Heatmap pages.
#
# Confirmed genuinely usable by this application via a dedicated,
# deliberate standalone test (project diagnostic v5): a completely fresh
# process, zero cookies, zero Refinitiv/SAML session, POSTing the exact
# same request body the browser sends, received back byte-identical real
# market data (ISIN, lastprice, netchange, volume all present and
# correct) to what the live page itself displays. The Refinitiv SAML
# login that also fires on the live page is NOT required for this
# specific endpoint — proven, not assumed.
#
# Honesty note (important, not hidden): the full response structure was
# not exhaustively inspected field-by-field — only the fields confirmed
# present via that diagnostic (isin, lastprice, netchange, volume). A
# genuine ticker SYMBOL and company NAME were not independently
# confirmed present under any specific key; the parser below tries
# several plausible key names for those two fields and — critically —
# never fabricates a value for either. Any instrument missing a genuine
# ISIN or a genuine price is skipped, not guessed at.
# =========================================================================

LSE_COMPONENTS_REFRESH_URL = "https://api.londonstockexchange.com/api/v1/components/refresh"

# Exact parameters/component IDs captured from the real browser's own
# request during diagnostic v5 — not guessed.
LSE_TAB_CONFIG = {
    "risersFallersVolume": {
        "parameters": "indexname%3Dftse-100%26tab%3Drisers-and-fallers-and-volume-leaders%26tabId%3D94aeb1d3-fd7b-46f5-b19b-389796d96214",
        "componentId": "block_content%3Aa193314b-d46e-4ca3-ad4f-9b814df5bafe",
    },
    "heatmap": {
        "parameters": "indexname%3Dftse-100%26tab%3Dheatmap%26tabId%3Ddcd47cbd-346e-4bd0-bf77-039301c7d329",
        "componentId": "block_content%3A72d8cb8c-5ef6-41a9-9bb9-49db0a064214",
    },
}
LSE_MARKET_DATA_TIMEOUT_SECONDS = 15

# News Explorer's own path/parameters/componentId. Its "parameters"
# now follows the EXACT SAME construction pattern as LSE_TAB_CONFIG's
# entries above (tab=X&tabId=Y) — a genuine, confirmed request-shape
# bug: every earlier version omitted tabId entirely, which every other
# working endpoint always included. Confirmed directly (diagnostic
# v13, live): the identical request WITHOUT tabId returns an empty
# "[]" (HTTP 200, no error, no data); the SAME request WITH tabId
# returns a populated response (19,599 bytes, 16 real stories,
# newsexplorersearch present). The tabId itself
# (58734a12-d97c-40cb-8047-df76e660f23f) is the News Explorer tab's own
# ID, confirmed against the page's own /api/v1/pages configuration —
# not invented, not brute-forced.
LSE_NEWS_PATH = "news"
LSE_NEWS_TAB_ID = "58734a12-d97c-40cb-8047-df76e660f23f"
LSE_NEWS_PARAMETERS = f"tab%3Dnews-explorer%26tabId%3D{LSE_NEWS_TAB_ID}"
LSE_NEWS_COMPONENT_ID = "block_content%3A431d02ac-09b8-40c9-aba6-04a72a4f2e49"
LSE_NEWS_TIMEOUT_SECONDS = 15


def _parse_lse_instrument_row(row):
    """
    Shared per-row extraction — the exact same identity/price/name/change
    handling used by every LSE instrument parser in this project, so
    there's exactly one place this logic exists rather than duplicated
    across the generic parser and the risers/fallers/volume-specific one.
    Returns None if the row is genuinely missing identity or price
    (never fabricated), otherwise the project's standard instrument dict.
    """
    if not isinstance(row, dict):
        return None
    isin = row.get("isin") or row.get("ISIN")
    tidm = row.get("tidm") or row.get("TIDM")
    price = row.get("lastprice")
    if price is None:
        price = row.get("lastPrice")
    if price is None:
        price = row.get("last_price")
    # Accept EITHER genuine identity field — never require isin
    # specifically (that was a confirmed bug: it silently discarded
    # every row on endpoints, like Heatmap, that only ever provide tidm).
    if not (isin or tidm) or price is None:
        return None  # never fabricate a row missing genuine identity or price
    identity_type = "isin" if isin else "tidm"
    ticker = (row.get("ticker") or row.get("symbol") or tidm
              or row.get("epic") or isin)
    name = (row.get("name") or row.get("companyname") or row.get("issuername")
            or row.get("description") or row.get("longname"))
    change_pct = (row.get("percentchange") or row.get("changepercent")
                  or row.get("pctchange") or row.get("changePercent")
                  or row.get("percentualchange"))
    net_change = row.get("netchange") or row.get("netChange")
    volume = row.get("volume") or row.get("tradedvolume") or row.get("dayvolume")
    return {
        "symbol": ticker,
        "isin": isin,  # None if genuinely absent - never fabricated from tidm
        "tidm": tidm,  # None if genuinely absent
        "identityType": identity_type,
        # Never invent a company name — if genuinely absent from the
        # response, the ticker/identity is shown as-is rather than a
        # fabricated label.
        "name": name or ticker,
        "price": price,
        "changePct": change_pct,
        "netChange": net_change,
        "volume": volume,
    }


def _parse_lse_risers_fallers_volume(data):
    """
    Dedicated parser for the risersFallersVolume response's REAL
    structure — confirmed by direct inspection of a genuine captured
    response: it contains THREE SEPARATE lists, each explicitly labeled
    by LSE itself via a "type" field ("RISERS", "FALLERS", "VOLUME"),
    NOT one combined list that a caller should locally re-sort into
    three categories.

    This fixes a confirmed, real bug: the previous approach
    (_parse_lse_components_refresh) walked the whole structure, found
    all three 9-row lists as candidates, and picked ONE of them via
    max(..., key=len) — since all three are the same length, this
    silently kept only whichever list happened to be encountered first
    (in practice, RISERS) and discarded the other two entirely. Deriving
    "losers" by locally re-sorting a risers-only list correctly finds
    nothing — which is exactly the "no liquid losers found" symptom this
    fixes, not a genuine absence of fallers.

    Returns {"gainers": [...], "losers": [...], "volume": [...]} using
    LSE's own pre-categorized lists directly — each row parsed via the
    shared _parse_lse_instrument_row, never re-derived by local sorting.
    Returns None if the expected RISERS/FALLERS/VOLUME structure isn't
    found at all (caller should treat this as "structure not
    recognized", not silently return empty lists as if that were a
    genuine empty result).
    """
    type_to_key = {"RISERS": "gainers", "FALLERS": "losers", "VOLUME": "volume"}
    found = {}

    def walk(obj):
        if isinstance(obj, dict):
            block_type = obj.get("type")
            values = obj.get("values")
            if isinstance(block_type, str) and block_type in type_to_key and isinstance(values, list):
                key = type_to_key[block_type]
                rows = [r for r in (_parse_lse_instrument_row(row) for row in values) if r is not None]
                found[key] = rows
            for v in obj.values():
                walk(v)
        elif isinstance(obj, list):
            for item in obj:
                walk(item)

    walk(data)
    if not found:
        return None
    # Only return categories genuinely found — never fabricate an empty
    # list for a category that wasn't present in the response at all,
    # since that's a different situation from "present but empty".
    return found


def _parse_lse_components_refresh(data):
    """
    Defensive, shape-agnostic GENERIC parser — used for endpoints
    (Heatmap) whose response is genuinely one combined instrument list,
    not pre-split into labeled categories the way risersFallersVolume is
    (see _parse_lse_risers_fallers_volume for that one specifically).
    Rather than hardcode an exact path and silently break if LSE's own
    layout shifts slightly, this walks the WHOLE structure looking for
    any list whose items look like genuine instrument records (an
    isin/tidm/ticker AND a price), and uses the largest such list found
    — i.e., the real data, not a short nested fragment. Returns a list
    of dicts already shaped to match this project's EXISTING screener
    row shape so no downstream rendering code needs to change to
    consume it.
    """
    candidate_lists = []

    def walk(obj):
        if isinstance(obj, dict):
            for key in ("values", "content", "instruments", "constituents"):
                v = obj.get(key)
                if isinstance(v, list) and v and isinstance(v[0], dict) and (
                        "isin" in v[0] or "tidm" in v[0] or "lastprice" in v[0] or "lastPrice" in v[0]):
                    candidate_lists.append(v)
            for v in obj.values():
                walk(v)
        elif isinstance(obj, list):
            for item in obj:
                walk(item)

    walk(data)
    if not candidate_lists:
        return []
    raw_rows = max(candidate_lists, key=len)

    out = [r for r in (_parse_lse_instrument_row(row) for row in raw_rows) if r is not None]
    return out


def _lse_log(run_id, message):
    """
    Single logging helper for every LSE-related event this project
    logs — added specifically to fix a real, confirmed problem: a
    production run showed a ~3 minute gap between when a fetch actually
    happened (its own embedded "retrieved" timestamp) and when its log
    line was actually written to the Actions log, strongly indicating
    heavy stdout buffering that made the log's physical line order
    unreliable for reconstructing true chronology - explaining an
    apparent "duplicate" fetch and a heatmap message appearing before
    a Yahoo crumb line that should have logged earlier.

    Every call includes: an explicit wall-clock timestamp (so true
    chronology is recoverable even if buffering reorders lines), a
    run_id shared by every LSE call within ONE poll run (so a genuine
    duplicate fetch - same run_id, same event, twice - is
    distinguishable from output that merely LOOKS duplicated because
    two separate runs' buffered output interleaved), and flush=True so
    this specific line is written immediately rather than sitting in a
    buffer that might never fully flush before the process exits.
    Printed to stderr, which is unbuffered by Python's own default
    (unlike stdout) - flush=True is added on top as a second guarantee,
    not a replacement for choosing the right stream.
    """
    ts = datetime.now(timezone.utc).isoformat()
    print(f"[{ts}] [{run_id}] {message}", file=sys.stderr, flush=True)


def new_lse_run_id():
    """A short, unique identifier for one poll run's worth of LSE calls
    — generated once per main() execution and threaded through every
    LSE fetch this run makes, so log lines can be grouped by run
    unambiguously."""
    return uuid.uuid4().hex[:8]


def _fetch_lse_components_refresh_raw(path, parameters, component_id, timeout=LSE_MARKET_DATA_TIMEOUT_SECONDS,
                                       label="components/refresh", run_id=None):
    """
    The single shared request/error-handling core for EVERY LSE
    components/refresh call this project makes — market data, heatmap,
    and News Explorer all build on this one function, so the request
    construction, timeout handling, and error handling only exist in
    one place. Returns (parsed_json_or_None, retrieved_at_iso, error_or_None).
    Never raises — every failure mode (network, timeout, invalid JSON)
    is caught and returned as a clear error string, never silently
    swallowed and never partially-populated.

    HTTP error responses (4xx/5xx — e.g. rate-limiting) are reported
    with their actual status code explicitly in the error string, not
    just a generic exception message, since "which of the 4 LSE calls
    this run failed, and why" needs to be diagnosable from the log
    alone, not guessed at afterward.

    label identifies which of the three LSE fetches this is
    ("Screener", "Heatmap", "News Explorer") for logging only — never
    affects the request itself. run_id ties every log line from this
    call to the SAME poll run, so a genuine duplicate fetch (identical
    run_id, same label, twice) is distinguishable from output that
    merely looks duplicated due to buffered/reordered log lines from
    different runs.
    """
    if run_id is None:
        run_id = new_lse_run_id()
    _lse_log(run_id, f"LSE {label}: FETCH START (path={path!r}, parameters={parameters!r})")
    retrieved_at = datetime.now(timezone.utc).isoformat()
    body = json.dumps({
        "path": path,
        "parameters": parameters,
        "components": [{"componentId": component_id, "parameters": None}],
    }).encode("utf-8")
    req = urllib.request.Request(
        LSE_COMPONENTS_REFRESH_URL, data=body, method="POST",
        headers={**HEADERS, "Content-Type": "application/json", "Accept": "application/json, text/plain, */*"},
    )
    try:
        resp = urllib.request.urlopen(req, timeout=timeout)
        raw = resp.read().decode("utf-8", errors="replace")
    except urllib.error.HTTPError as e:
        error_detail = ""
        try:
            error_detail = e.read().decode("utf-8", errors="replace")[:300]
        except Exception:
            pass
        error_msg = f"HTTP {e.code} {e.reason}" + (f" — {error_detail}" if error_detail else "")
        _lse_log(run_id, f"LSE {label}: RESULT — FAILED — {error_msg}")
        return None, retrieved_at, error_msg
    except Exception as e:
        _lse_log(run_id, f"LSE {label}: RESULT — FAILED — {type(e).__name__}: {e}")
        return None, retrieved_at, str(e)
    _lse_log(run_id, f"LSE {label}: RESULT — HTTP {resp.status}, {len(raw)} bytes")
    try:
        parsed = json.loads(raw)
    except Exception as e:
        _lse_log(run_id, f"LSE {label}: PARSE FAILED — {type(e).__name__}: {e}")
        return None, retrieved_at, f"invalid JSON: {e}"
    return parsed, retrieved_at, None


def fetch_lse_ftse100_market_data(tab, timeout=LSE_MARKET_DATA_TIMEOUT_SECONDS, run_id=None):
    """
    Fetches genuine FTSE 100 market data directly from LSE's own
    first-party endpoint — the PRIMARY source for this data, not a
    fallback. tab must be "risersFallersVolume" or "heatmap".

    Returns {"status": "ok"|"failed", "source": "LSE", "retrievedAt":
    iso timestamp, "instruments": [...], "error": str|None}. Callers
    MUST check status explicitly and fall back to
    fetch_lse_screener()/fetch_gb_screener() (Yahoo) themselves when
    status is "failed" — this function never silently substitutes
    another source, and never mixes LSE and Yahoo values in one row.
    """
    if tab not in LSE_TAB_CONFIG:
        raise ValueError(f"Unknown LSE tab '{tab}' - must be one of {list(LSE_TAB_CONFIG)}")
    if run_id is None:
        run_id = new_lse_run_id()
    label = "Screener" if tab == "risersFallersVolume" else "Heatmap"
    cfg = LSE_TAB_CONFIG[tab]
    parsed, retrieved_at, error = _fetch_lse_components_refresh_raw(
        "ftse-constituents", cfg["parameters"], cfg["componentId"], timeout, label=label, run_id=run_id)

    if error:
        print(f"  ! LSE market data fetch failed for tab={tab}: {error}", file=sys.stderr)
        return {"status": "failed", "source": "LSE", "retrievedAt": retrieved_at,
                "instruments": [], "categorized": None, "error": error}

    categorized = None
    if tab == "risersFallersVolume":
        # Confirmed real structure (a genuine bug fix): this endpoint's
        # response contains THREE SEPARATE lists, each explicitly
        # labeled by LSE itself (type: RISERS/FALLERS/VOLUME) — not one
        # combined list to locally re-sort. Using LSE's own
        # categorization directly, rather than re-deriving "losers" by
        # sorting a risers-only list (which always found nothing).
        categorized = _parse_lse_risers_fallers_volume(parsed)

    if categorized:
        # Flat "instruments" list still populated (deduped by identity)
        # for callers that only need "every returned row" regardless of
        # category, e.g. the FTSE 100 coverage check.
        seen_keys = set()
        instruments = []
        for rows in categorized.values():
            for r in rows:
                key = r.get("isin") or r.get("tidm") or r.get("symbol")
                if key in seen_keys:
                    continue
                seen_keys.add(key)
                instruments.append(r)
    else:
        # Either not the risersFallersVolume tab, or the dedicated
        # RISERS/FALLERS/VOLUME structure genuinely wasn't found this
        # run (e.g. LSE's own layout changed) — fall back to the
        # generic flatten-largest-list parser rather than returning
        # nothing at all.
        instruments = _parse_lse_components_refresh(parsed)

    if not instruments:
        _lse_log(run_id, f"LSE {label}: PARSED 0 instruments — no genuine instrument rows "
                          f"found in an otherwise successful response — treating as a failure")
        print(f"  ! LSE market data for tab={tab}: response parsed but no genuine "
              f"instrument rows found — treating as a failure, not an empty-but-valid result", file=sys.stderr)
        return {"status": "failed", "source": "LSE", "retrievedAt": retrieved_at,
                "instruments": [], "categorized": None, "error": "no instrument rows found in response"}

    if categorized:
        _lse_log(run_id, f"LSE {label}: PARSED {len(instruments)} instruments "
                          f"(gainers={len(categorized.get('gainers', []))}, "
                          f"losers={len(categorized.get('losers', []))}, "
                          f"volume={len(categorized.get('volume', []))})")
    else:
        _lse_log(run_id, f"LSE {label}: PARSED {len(instruments)} instruments")
    print(f"  > LSE market data for tab={tab}: {len(instruments)} instruments, retrieved {retrieved_at}")
    return {"status": "ok", "source": "LSE", "retrievedAt": retrieved_at,
            "instruments": instruments, "categorized": categorized, "error": None}


def _find_newsexplorersearch_block(obj):
    """Recursively locates the "newsexplorersearch" DATATYPE block
    within a components/refresh response — confirmed (diagnostic v10,
    verified directly against a genuinely captured real response) to
    sit alongside the filter-configuration data in the SAME response,
    not a separate request. Returns the block's "value" dict (with
    "content"/"totalElements"/"totalPages"/etc) or None if genuinely
    absent — never fabricates a placeholder structure."""
    if isinstance(obj, dict):
        if obj.get("name") == "newsexplorersearch" and isinstance(obj.get("value"), dict):
            return obj["value"]
        for v in obj.values():
            result = _find_newsexplorersearch_block(v)
            if result is not None:
                return result
    elif isinstance(obj, list):
        for item in obj:
            result = _find_newsexplorersearch_block(item)
            if result is not None:
                return result
    return None


def _parse_lse_news_explorer(data):
    """
    Parses the confirmed newsexplorersearch block into this project's
    own clean story shape. Defensive about individual story rows —
    skips (never fabricates) any row genuinely missing a title, since a
    headline-less "story" isn't a real displayable result.
    """
    block = _find_newsexplorersearch_block(data)
    if block is None:
        return None
    raw_stories = block.get("content", [])
    stories = []
    for row in raw_stories:
        if not isinstance(row, dict) or not row.get("title"):
            continue
        stories.append({
            "id": row.get("id"),
            "headline": row.get("title"),
            "companyCode": row.get("companycode"),
            "companyName": row.get("companyname"),
            "source": row.get("source"),
            "newsSource": row.get("newssource"),
            "rnsNumber": row.get("rnsnumber"),
            "datetime": row.get("datetime"),
            "price": row.get("lastprice"),
            "percentChange": row.get("percentualchange"),
            "url": row.get("url"),
        })
    return {
        "stories": stories,
        "totalElements": block.get("totalElements"),
        "totalPages": block.get("totalPages"),
        "pageNumber": block.get("number"),
        "pageSize": block.get("size"),
        "isLastPage": block.get("last"),
    }


def fetch_lse_news_explorer(timeout=LSE_NEWS_TIMEOUT_SECONDS, run_id=None):
    """
    Fetches genuine LSE News Explorer stories directly from LSE's own
    first-party endpoint — the PRIMARY source, not a fallback. There is
    no Yahoo equivalent for this specific data (regulatory/company
    announcements), so a failure here means an honest "unavailable"
    state on the dashboard, never a substitute source silently
    presented as if it were the same thing.

    Returns {"status": "ok"|"failed", "source": "LSE", "retrievedAt":
    iso timestamp, "stories": [...], "totalElements": int|None,
    "totalPages": int|None, "error": str|None}.

    Guaranteed to log exactly one of a FETCH START line (via the
    shared _fetch_lse_components_refresh_raw core) followed by either a
    PARSED-stories line or a FAILED line — there is no path through
    this function that returns without both having happened. This was
    added specifically because a production run showed zero log
    evidence, in either direction, of this function ever having run —
    confirmed via static analysis to have no enclosing try/except that
    could have swallowed an exception here, so the most likely
    explanation was buffered output being lost, not a genuine silent
    skip. These explicit, flushed lines close that gap for good.
    """
    if run_id is None:
        run_id = new_lse_run_id()
    parsed, retrieved_at, error = _fetch_lse_components_refresh_raw(
        LSE_NEWS_PATH, LSE_NEWS_PARAMETERS, LSE_NEWS_COMPONENT_ID, timeout, label="News Explorer", run_id=run_id)

    if error:
        _lse_log(run_id, f"LSE News Explorer: FAILED — {error}")
        print(f"  ! LSE News Explorer fetch failed: {error}", file=sys.stderr)
        return {"status": "failed", "source": "LSE", "retrievedAt": retrieved_at,
                "stories": [], "totalElements": None, "totalPages": None, "error": error}

    parsed_news = _parse_lse_news_explorer(parsed)
    if parsed_news is None:
        _lse_log(run_id, "LSE News Explorer: FAILED — response parsed but no newsexplorersearch "
                          "block found")
        print(f"  ! LSE News Explorer: response parsed but no newsexplorersearch block "
              f"found — treating as a failure, not an empty-but-valid result", file=sys.stderr)
        return {"status": "failed", "source": "LSE", "retrievedAt": retrieved_at,
                "stories": [], "totalElements": None, "totalPages": None,
                "error": "no newsexplorersearch block found in response"}

    _lse_log(run_id, f"LSE News Explorer: {len(parsed_news['stories'])} stories "
                      f"(totalElements={parsed_news['totalElements']}, totalPages={parsed_news['totalPages']})")
    print(f"  > LSE News Explorer: {len(parsed_news['stories'])} stories "
          f"(totalElements={parsed_news['totalElements']}, totalPages={parsed_news['totalPages']}), "
          f"retrieved {retrieved_at}")
    return {
        "status": "ok", "source": "LSE", "retrievedAt": retrieved_at,
        "stories": parsed_news["stories"], "totalElements": parsed_news["totalElements"],
        "totalPages": parsed_news["totalPages"], "error": None,
    }


def fetch_lse_screener_primary(raw_count=10, display_count=10, run_id=None):
    """
    LSE-primary replacement for fetch_lse_screener(): tries the genuine
    LSE first-party endpoint FIRST for Volume/Gainers/Losers (all three
    derived from the SAME risersFallersVolume LSE response, sorted
    locally — matching the same "sort ourselves, don't trust an
    upstream top-N" principle already used elsewhere in this project),
    and falls back to the existing Yahoo-based fetch_lse_screener() ONLY
    for whichever of the three genuinely failed. Never mixes LSE and
    Yahoo values within the same row — each of the three lists is
    either entirely LSE or entirely Yahoo, and source_dict says exactly
    which, per section, so the dashboard can label it honestly rather
    than implying one uniform source.

    raw_count and display_count are deliberately separate: raw_count is
    ONLY meaningful for the Yahoo fallback path, which over-fetches a
    larger pool (e.g. 60) before the caller's own FTSE-universe name
    filter trims it back down — the LSE path needs no such over-fetch,
    since its response is already genuine FTSE 100 constituents by
    construction, so it's always trimmed straight to display_count.
    Passing raw_count's larger over-fetch value into the LSE path's own
    slice would silently return far more rows than intended for display.

    Returns a 4th value, lse_result: the FULL fetch_lse_ftse100_market_data()
    result dict (including "instruments" and "retrievedAt") when LSE
    succeeded, or None when the Yahoo fallback was used. Callers that
    need the raw instrument list or the exact retrieval timestamp (e.g.
    the coverage check, or displaying "Retrieved: ...") should use this
    rather than re-fetching — reusing it avoids an entirely redundant
    second POST to the same LSE endpoint within the same poll run,
    which is a real, avoidable rate-limiting risk factor.
    """
    if run_id is None:
        run_id = new_lse_run_id()
    lse_result = fetch_lse_ftse100_market_data("risersFallersVolume", run_id=run_id)
    source_dict = {"volume": "LSE", "gainers": "LSE", "losers": "LSE"}

    if lse_result["status"] == "ok":
        categorized = lse_result.get("categorized")
        if categorized:
            # Use LSE's own pre-categorized RISERS/FALLERS/VOLUME lists
            # directly — confirmed real structure, not locally re-sorted
            # from one combined list (that approach was a genuine bug:
            # re-deriving "losers" by sorting a risers-only list always
            # found nothing, which was never a genuine absence of real
            # fallers).
            by_gainers = categorized.get("gainers", [])[:display_count]
            by_losers = categorized.get("losers", [])[:display_count]
            by_volume = categorized.get("volume", [])[:display_count]
        else:
            # Defensive fallback only — the dedicated RISERS/FALLERS/
            # VOLUME structure genuinely wasn't found this run (e.g.
            # LSE's own layout changed), so fall back to locally
            # deriving categories from the flat instrument list rather
            # than returning nothing.
            instruments = lse_result["instruments"]
            by_volume = sorted([r for r in instruments if r.get("volume")],
                                key=lambda r: r["volume"], reverse=True)[:display_count]
            by_gainers = sorted([r for r in instruments if r.get("changePct") is not None],
                                 key=lambda r: r["changePct"], reverse=True)[:display_count]
            by_losers = sorted([r for r in instruments if r.get("changePct") is not None],
                                key=lambda r: r["changePct"])[:display_count]
        screener = {"volume": by_volume, "gainers": by_gainers, "losers": by_losers}
        status = {"volume": "ok", "gainers": "ok", "losers": "ok"}
        return screener, status, source_dict, lse_result

    # LSE failed entirely for this run — fall back to the existing,
    # already-working Yahoo-based screener, clearly labeled as a
    # fallback rather than silently presented as if it were LSE data.
    _lse_log(run_id, f"Yahoo fallback: START (reason: LSE market data unavailable — {lse_result['error']})")
    print(f"  ! LSE market data unavailable ({lse_result['error']}) — falling back to Yahoo "
          f"for Volume/Gainers/Losers this run", file=sys.stderr)
    screener, status = fetch_lse_screener(raw_count)
    source_dict = {"volume": "Yahoo (fallback)", "gainers": "Yahoo (fallback)", "losers": "Yahoo (fallback)"}
    _lse_log(run_id, f"Yahoo fallback: END — volume={len(screener.get('volume', []))}, "
                      f"gainers={len(screener.get('gainers', []))}, losers={len(screener.get('losers', []))}")
    return screener, status, source_dict, None


def _normalize_lse_ticker(ticker):
    """
    Normalizes genuine ticker format differences between LSE's own
    convention and the independent (Yahoo-derived) FTSE 100 constituent
    list's convention, so a real match isn't missed purely over
    formatting — never changes WHICH instrument is being compared, only
    how its ticker is spelled for comparison purposes.

    Confirmed real case this fixes: Aviva's genuine LSE ticker is "AV."
    (the trailing period is part of LSE's own disambiguation
    convention). The independent list's SOURCE data represents this
    same ticker as "AV/.L" — confirmed directly against the real
    captured fixture (fixtures/ftse100_yfiua_real_response.json) —
    i.e. Yahoo's own convention escapes an embedded period as a forward
    slash before the ".L" exchange suffix, rather than dropping it. An
    earlier version of this function only stripped a trailing period
    and was tested against invented data ("AV.L") that did not match
    this real escaping convention, so it never actually fixed the live
    case — this version is verified directly against the real fixture.
    Also strips a plain ".L" suffix for tickers that don't need
    unescaping at all.
    """
    if not ticker:
        return ""
    t = ticker.strip().upper()
    if t.endswith(".L"):
        t = t[:-2]
    # Yahoo's escaped-period convention: a trailing "/" (after the ".L"
    # suffix has already been stripped above) represents an embedded
    # period LSE's own ticker would show directly - e.g. "AV/" is "AV."
    if t.endswith("/"):
        t = t[:-1]
    if t.endswith("."):
        t = t[:-1]
    return t


def check_lse_ftse100_coverage(instruments=None):
    """
    A genuine constituent-coverage check, not an assumption: compares
    the FULL LSE risersFallersVolume instrument list against the
    independently-sourced FTSE 100 constituent list (yfiua's ticker+name
    JSON, already used elsewhere in this project), then reports exactly
    how many of each were matched, unmatched, or duplicated — by
    ticker (normalized for genuine formatting differences like LSE's
    "AV." vs the independent list's "AV" — see _normalize_lse_ticker),
    falling back to a cleaned company-name comparison for any row whose
    extracted "symbol" turned out to be an ISIN/tidm rather than a
    plain ticker (this project's own honest fallback when a genuine
    ticker field wasn't confirmed present).

    IMPORTANT — this check validates that returned rows genuinely
    belong to the FTSE 100; it does NOT assume or imply that a
    top-movers widget (risersFallersVolume) should return anywhere
    close to all 100 constituents. "lseInstrumentsReturned" is simply
    how many rows this specific widget returned this run (routinely a
    small, curated top-movers list by the endpoint's own design, not a
    constituent dump) — "matched"/"unmatched" describe how many of
    THOSE rows are genuine FTSE 100 members, which is the actual
    question this check answers.

    instruments, if provided, is the ALREADY-FETCHED instrument list
    from this run's own fetch_lse_screener_primary() call — reusing it
    here avoids a second, entirely redundant POST to the same LSE
    endpoint within the same poll run. Multiple rapid back-to-back
    requests to the same first-party endpoint is a real, avoidable risk
    factor for rate-limiting, and this coverage check has no need to
    re-fetch data this run already has. If instruments is None (e.g.
    called standalone, outside the normal poll flow), it fetches fresh,
    same as before.

    Returns a dict with the full breakdown. Logs a clear summary either
    way — this function's job is to report the truth, not to make LSE
    coverage look better or worse than it actually is.
    """
    ftse100_rows = []
    try:
        raw = http_get(FTSE100_CONSTITUENTS_URL)
        ftse100_rows = _parse_ftse100_json(raw)
    except Exception as e:
        print(f"  ! Coverage check: could not fetch the independent FTSE 100 constituent "
              f"list ({e}) — coverage cannot be verified this run", file=sys.stderr)
        return {"status": "failed", "error": str(e)}

    expected_tickers = {_normalize_lse_ticker(r["ticker"]) for r in ftse100_rows if r.get("ticker")}
    expected_names = {clean_company_name(r["name"]).lower() for r in ftse100_rows if r.get("name")}

    if instruments is None:
        lse_result = fetch_lse_ftse100_market_data("risersFallersVolume")
        if lse_result["status"] != "ok":
            print(f"  ! Coverage check: LSE fetch itself failed ({lse_result['error']}) — "
                  f"cannot verify coverage this run", file=sys.stderr)
            return {"status": "failed", "error": lse_result["error"]}
        instruments = lse_result["instruments"]

    seen_tickers = set()
    matched, unmatched, duplicates = [], [], []
    for row in instruments:
        symbol_norm = _normalize_lse_ticker(row.get("symbol") or "")
        name_clean = clean_company_name(row.get("name") or "").lower()
        is_matched = symbol_norm in expected_tickers or name_clean in expected_names
        key = symbol_norm or row.get("isin") or row.get("tidm")
        if key in seen_tickers:
            duplicates.append(key)
            continue
        seen_tickers.add(key)
        (matched if is_matched else unmatched).append(row)

    report = {
        "status": "ok",
        "ftse100Expected": len(expected_tickers),
        "lseInstrumentsReturned": len(instruments),
        "matched": len(matched),
        "unmatched": len(unmatched),
        "duplicates": len(duplicates),
        "unmatchedSamples": [{"symbol": r.get("symbol"), "name": r.get("name"), "isin": r.get("isin")}
                              for r in unmatched[:10]],
    }
    # The warning marker is about whether returned rows are genuinely
    # FTSE 100 members, NOT whether the widget returned close to the
    # full ~100-constituent universe (risersFallersVolume is a
    # top-movers list by design and routinely returns far fewer rows on
    # a quiet day — that alone is never a problem).
    returned = report["lseInstrumentsReturned"]
    match_rate_ok = returned == 0 or (report["matched"] / returned) >= 0.9
    marker = "" if match_rate_ok else "!!! "
    print(f"{marker}LSE FTSE 100 coverage check: this widget returned {report['lseInstrumentsReturned']} "
          f"row(s) (a top-movers list, not a constituent dump — {report['ftse100Expected']} is the "
          f"full FTSE 100 universe size for reference only), "
          f"{report['matched']} matched to FTSE 100, {report['unmatched']} unmatched, "
          f"{report['duplicates']} duplicate(s)")
    if unmatched:
        print(f"  Unmatched sample: {report['unmatchedSamples'][:3]}")
    return report


# =========================================================================
# FTSE 100 / FTSE 250 universe — restricts the market-wide screener (and
# everything downstream of it: News on Movers, 5-Day Uptrend, Broker
# Target Prices, and — via the shared ticker_lookup pool — Market-wide
# Broker Alerts) to genuine FTSE 350 constituents, instead of Yahoo's
# whole LSE universe (which includes AIM micro-caps with no relation to
# either index — confirmed live: Forgent, Tower Resources, Premier
# African Minerals were all appearing under "LSE Screener" despite none
# being FTSE 100/250 members).
# =========================================================================

FTSE100_CONSTITUENTS_URL = "https://yfiua.github.io/index-constituents/constituents-ftse100.json"
FTSE250_WIKI_API_URL = (
    "https://en.wikipedia.org/w/api.php?action=parse&page=FTSE_250_Index"
    "&format=json&prop=text"
)
FTSE_UNIVERSE_CACHE_FILE = os.path.join(STATE_DIR, "ftse_universe.json")
FTSE_UNIVERSE_MAX_AGE_HOURS = 24  # index membership changes quarterly at most
FTSE100_EXPECTED_RANGE = (85, 115)   # plausible row-count band; real value is ~100
FTSE250_EXPECTED_RANGE = (210, 290)  # plausible row-count band; real value is ~250
_TICKER_SHAPE_RE = re.compile(r"^[A-Z0-9]{1,5}[./]?[A-Z0-9]{0,3}$")


# --- Pure parsers: no I/O, no network, fully unit-testable against saved
# fixtures built from genuinely captured real source data. Each returns a
# list of {"name": str, "ticker": str} dicts, or raises on structurally
# unparseable input (caught by the fetch_* wrappers below). -------------

def _parse_ftse100_json(raw_text):
    """
    Parses yfiua/index-constituents' FTSE 100 JSON — confirmed by direct
    live fetch during development to be a flat `[{"Symbol":..,"Name":..}]`
    list; tested against that exact captured real response
    (fixtures/ftse100_yfiua_real_response.json).
    """
    data = json.loads(raw_text)
    rows = []
    for row in data:
        name = (row.get("Name") or "").strip()
        symbol = (row.get("Symbol") or "").strip().upper()
        ticker = symbol.rsplit(".L", 1)[0] if symbol.endswith(".L") else symbol
        rows.append({"name": name, "ticker": ticker})
    return rows


def _parse_ftse250_wiki_html(raw_html):
    """
    Parses MediaWiki's rendered HTML for the FTSE_250_Index article,
    identifying the constituents table by its HEADER CONTENT ("company"
    AND "ticker" both present) rather than by position on the page — the
    article also contains two unrelated historical annual-return tables,
    and the constituents table's position could shift with future edits.
    Never assumes a fixed table index.

    Tested against fixtures/ftse250_wikipedia_real_data.html — a
    reconstruction using genuinely captured real constituent data (all
    250 real company/ticker/sector rows, fetched and transcribed from
    the live page during development) wrapped in standard, well-
    documented MediaWiki wikitable HTML conventions. The DATA is real
    and verified; the exact markup shape has NOT been independently
    confirmed against a live call to this specific API endpoint from any
    environment available during development (no network path to
    Wikipedia's API existed there) — see fetch_ftse250_constituents()
    for how this gap is handled: aggressive validation, safe fallback to
    cached data, and an explicit first-live-run verification note.
    """
    parser = _WikiTableParser()
    parser.feed(raw_html)
    table = _find_constituents_table(parser.tables)
    if table is None:
        raise ValueError("constituents table not found (no table header contained both 'company' and 'ticker')")
    rows = []
    for row in table[1:]:
        if len(row) >= 2:
            rows.append({"name": row[0].strip(), "ticker": row[1].strip().upper()})
    return rows


class _WikiTableParser(HTMLParser):
    """
    Minimal HTML table parser (stdlib only, no new dependency) — collects
    EVERY table on a page as a list of rows, each row a list of cell
    strings. Identifying which table is the constituents table happens
    separately (_find_constituents_table), after parsing completes.
    """
    def __init__(self):
        super().__init__()
        self.tables = []
        self._cur_table = None
        self._cur_row = None
        self._cur_cell_parts = None

    def handle_starttag(self, tag, attrs):
        if tag == "table":
            self._cur_table = []
        elif tag == "tr" and self._cur_table is not None:
            self._cur_row = []
        elif tag in ("td", "th") and self._cur_row is not None:
            self._cur_cell_parts = []

    def handle_endtag(self, tag):
        if tag == "table" and self._cur_table is not None:
            self.tables.append(self._cur_table)
            self._cur_table = None
        elif tag == "tr" and self._cur_row is not None:
            if self._cur_table is not None:
                self._cur_table.append(self._cur_row)
            self._cur_row = None
        elif tag in ("td", "th") and self._cur_cell_parts is not None:
            if self._cur_row is not None:
                self._cur_row.append("".join(self._cur_cell_parts).strip())
            self._cur_cell_parts = None

    def handle_data(self, data):
        if self._cur_cell_parts is not None:
            self._cur_cell_parts.append(data)


def _find_constituents_table(tables):
    """Finds the table whose header row mentions BOTH "company" and
    "ticker" (case-insensitive) — never assumes table position. Returns
    None if no such table exists."""
    for table in tables:
        if not table:
            continue
        header = " ".join(table[0]).lower()
        if "company" in header and "ticker" in header:
            return table
    return None


def _validate_constituent_rows(rows, index_label, expected_range):
    """
    Generic validator applied identically to both FTSE 100 and FTSE 250
    parsed row lists. Checks, in order:
      - row count plausibly within expected_range
      - every row has a non-empty name and a ticker matching a permissive
        real-ticker shape (letters/digits, optional single "." or "/"
        separator — covers UK conventions like "BP.", "BT/A")
      - no duplicate tickers
    Tolerates a SMALL number of individually malformed rows (parsing
    imperfections on a handful of entries out of ~100-250 shouldn't
    invalidate an otherwise-good fetch) but rejects if more than 5% of
    rows are malformed, or if any ticker is duplicated (a duplicate
    suggests the parse genuinely went wrong, not just an isolated
    formatting quirk).

    Returns (is_valid: bool, issues: list[str]) — issues always explains
    why when is_valid is False.
    """
    issues = []
    lo, hi = expected_range
    if not (lo <= len(rows) <= hi):
        issues.append(f"{index_label}: row count {len(rows)} outside plausible range [{lo}, {hi}]")
        return False, issues

    seen_tickers = set()
    malformed = 0
    for row in rows:
        name = (row.get("name") or "").strip()
        ticker = (row.get("ticker") or "").strip().upper()
        if not name or not ticker or not _TICKER_SHAPE_RE.match(ticker):
            malformed += 1
            continue
        if ticker in seen_tickers:
            issues.append(f"{index_label}: duplicate ticker '{ticker}'")
            return False, issues
        seen_tickers.add(ticker)

    if malformed:
        issues.append(f"{index_label}: {malformed}/{len(rows)} malformed row(s) (missing name or invalid ticker shape)")
        if malformed > len(rows) * 0.05:
            return False, issues

    return True, issues


# --- I/O wrappers: fetch -> parse -> validate. Each returns a set of
# cleaned company names on success, or None on ANY failure — callers
# must treat None as "couldn't determine this run", never as "confirmed
# empty universe". ------------------------------------------------------

def fetch_ftse100_constituents():
    try:
        raw = http_get(FTSE100_CONSTITUENTS_URL)
        rows = _parse_ftse100_json(raw)
    except Exception as e:
        print(f"  ! FTSE 100 fetch/parse failed: {e}", file=sys.stderr)
        return None
    is_valid, issues = _validate_constituent_rows(rows, "FTSE100", FTSE100_EXPECTED_RANGE)
    for msg in issues:
        print(f"  {'!' if not is_valid else '~'} {msg}", file=sys.stderr)
    if not is_valid:
        return None
    return {clean_company_name(r["name"]).lower() for r in rows if r["name"]}


def fetch_ftse250_constituents():
    """
    See _parse_ftse250_wiki_html's docstring for the important caveat:
    the exact markup shape this expects has not been independently
    confirmed against a live call to Wikipedia's API from any
    development environment available. This function's job is to make
    that gap SAFE, not to pretend it doesn't exist: validation is
    aggressive (row count, per-row shape, duplicate tickers), and ANY
    failure — fetch, parse, or validation — returns None, which
    load_ftse_universe() treats as "keep whatever was already cached",
    never as license to wipe or corrupt the existing universe.
    """
    try:
        raw = http_get(FTSE250_WIKI_API_URL)
        data = json.loads(raw)
        html_content = data["parse"]["text"]["*"]
        rows = _parse_ftse250_wiki_html(html_content)
    except Exception as e:
        print(f"  ! FTSE 250 fetch/parse failed: {e}", file=sys.stderr)
        return None
    is_valid, issues = _validate_constituent_rows(rows, "FTSE250", FTSE250_EXPECTED_RANGE)
    for msg in issues:
        print(f"  {'!' if not is_valid else '~'} {msg}", file=sys.stderr)
    if not is_valid:
        return None
    return {clean_company_name(r["name"]).lower() for r in rows if r["name"]}


def load_ftse_universe(path=None, now=None):
    """
    Returns (names_set, source_description) for the combined FTSE 100 +
    FTSE 250 universe used to scope the market-wide screener to genuine
    FTSE 350 constituents. Re-fetches at most once every
    FTSE_UNIVERSE_MAX_AGE_HOURS — index membership changes quarterly at
    most, so this is deliberately kept OUT of the normal 5-minute poll
    cadence (a fresh cache short-circuits to zero network calls).

    Fail-safe contract, in priority order — bad or missing source data
    must NEVER wipe or corrupt an existing good universe:
    1. Fresh cache -> used directly.
    2. Both sources fetch AND validate successfully -> combined, compared
       against the previous cached list for turnover, cached, used.
    3. Only FTSE 100 succeeds -> FTSE-100-only (a real improvement over
       no filtering), cached as partial, logged clearly.
    4. Either/both fail validation or fetch, but ANY previous cache
       exists (even stale, even partial) -> that previous cache is kept
       and reused UNCHANGED. A failed refresh attempt never overwrites
       good data with nothing or with bad data.
    5. Both fail and no cache exists at all -> returns (None,
       "unavailable"); callers MUST skip FTSE filtering entirely this
       run, never treat this as an empty-but-valid universe.

    Every path logs: which source(s) were used, the timestamp, the row
    count, the validation outcome, and whether a fallback was activated —
    so production logs make the actual behaviour of any given run
    inspectable after the fact, not just assumed.
    """
    path = path or FTSE_UNIVERSE_CACHE_FILE
    now = now or datetime.now(timezone.utc)
    cached = load_json(path, None)

    if cached:
        try:
            fetched_at = datetime.fromisoformat(cached["fetched_at"])
            age_hours = (now - fetched_at).total_seconds() / 3600
            if age_hours < FTSE_UNIVERSE_MAX_AGE_HOURS:
                print(f"FTSE universe: using fresh cache (source={cached.get('source')}, "
                      f"{len(cached.get('names', []))} names, age={age_hours:.1f}h)")
                return set(cached["names"]), cached.get("source", "cache")
        except Exception:
            pass  # corrupt/unexpected cache shape — fall through to refetch

    ftse100_names = fetch_ftse100_constituents()
    ftse250_names = fetch_ftse250_constituents()

    if ftse100_names is not None and ftse250_names is not None:
        combined = ftse100_names | ftse250_names
        source = "ftse100+ftse250"
    elif ftse100_names is not None:
        combined = ftse100_names
        source = "ftse100_only"
        print("  ! FTSE 250 unavailable this run — scoping to FTSE 100 only", file=sys.stderr)
    elif cached:
        combined = set(cached["names"])
        source = cached.get("source", "cache") + "_stale"
        print(f"  ! Both FTSE constituent sources failed this run — keeping previous cached universe "
              f"unchanged (source={cached.get('source')}, {len(combined)} names, "
              f"cached {cached.get('fetched_at')})", file=sys.stderr)
        return combined, source
    else:
        print("  ! Both FTSE constituent sources failed and no cache exists — FTSE filtering skipped this run", file=sys.stderr)
        return None, "unavailable"

    # Turnover check against the previous known-good list, if one exists —
    # a genuine quarterly rebalancing can legitimately change a handful of
    # names, but a huge swing usually means something parsed wrong even
    # though it technically passed the row-count/shape validation above.
    # This does NOT block the update (a real large rebalance is possible
    # and shouldn't get the tool stuck on stale data forever) — it's
    # logged prominently for human review, nothing more.
    if cached and cached.get("names"):
        previous = set(cached["names"])
        if previous:
            unchanged = len(previous & combined)
            turnover_pct = 100 * (1 - unchanged / len(previous))
            level = "!" if turnover_pct > 15 else "~"
            print(f"  {level} FTSE universe turnover vs previous cache: {turnover_pct:.1f}% "
                  f"({len(previous)} -> {len(combined)} names)" +
                  (" — unusually large, worth a manual check" if turnover_pct > 15 else ""))

    print(f"FTSE universe: refreshed (source={source}, {len(combined)} names)")

    try:
        save_json(path, {
            "names": sorted(combined),
            "source": source,
            "fetched_at": now.isoformat(),
        })
    except Exception as e:
        print(f"  ! Could not cache FTSE universe: {e}", file=sys.stderr)

    return combined, source


def ftse_universe_status_label(source):
    """
    Maps load_ftse_universe()'s internal `source` string into one of four
    CLEARLY DISTINCT, human-readable states — specifically so these never
    get silently collapsed into each other, on the dashboard or anywhere
    else:

    - "healthy": both FTSE 100 and FTSE 250 are present and current
      (fresh fetch this run, or a fresh — not stale — cache).
    - "degraded_ftse100_only": FTSE 100 is CURRENT, but FTSE 250 is
      entirely missing right now (fetch/validation failed AND no
      previous FTSE 250 data exists to fall back on — e.g. the very
      first production run, or every prior attempt also failed). This is
      its own distinct state, separate from "stale_cache": it is not
      reusing old data, it genuinely has none.
    - "stale_cache": today's fetch(es) failed, but a previous
      known-good cache (of whatever quality it was) is being reused.
    - "unavailable": total failure and no cache of any kind exists;
      screener filtering is skipped entirely this run (whole-LSE,
      unrestricted).
    - "not_checked": SKIP_MARKET_WIDE runs (the hourly FTSE350 job)
      never touch FTSE universe status at all — distinct from all of
      the above, not an error state.
    """
    if source == "not_checked":
        return "not_checked"
    if source == "unavailable":
        return "unavailable"
    if source.endswith("_stale"):
        return "stale_cache"
    if source == "ftse100_only":
        return "degraded_ftse100_only"
    if source == "ftse100+ftse250":
        return "healthy"
    return "unknown"


def now_stamp():
    """Human-readable London-time stamp for every outgoing message and the dashboard —
    'this is live and this is exactly when' rather than leaving it implicit. Shows
    London time first (what matters for a UK tool) with the London/BST or GMT label
    zoneinfo resolves automatically, plus UTC alongside for technical clarity."""
    now_utc = datetime.now(timezone.utc)
    now_london = now_utc.astimezone(LONDON_TZ)
    return f'{now_london.strftime("%a %d %b %Y, %H:%M")} {now_london.strftime("%Z")} ({now_utc.strftime("%H:%M")} UTC)'


def format_london_and_utc(dt_utc):
    """Same dual London/UTC format as now_stamp(), for a specific stored UTC datetime
    (e.g. lastPoll) rather than 'right now'."""
    if dt_utc.tzinfo is None:
        dt_utc = dt_utc.replace(tzinfo=timezone.utc)
    dt_london = dt_utc.astimezone(LONDON_TZ)
    return f'{dt_london.strftime("%a %d %b %Y, %H:%M")} {dt_london.strftime("%Z")} ({dt_utc.strftime("%H:%M")} UTC)'


def format_screener_sections(screener):
    """
    Returns a list of separate messages (one per section: Volume, Gainers, Losers)
    instead of one combined string. CallMeBot/WhatsApp truncates long messages, and
    all three sections combined easily exceeds that — sending separately means each
    one arrives complete rather than the later sections getting cut off silently.
    """
    def section(title, rows, show_pct=True):
        lines = [f"{title}  🕐 {now_stamp()}"]
        for i, q in enumerate(rows, 1):
            chg = q.get("changePct")
            chg_str = f'{"▲" if (chg or 0) >= 0 else "▼"}{abs(chg or 0):.2f}%' if chg is not None else ""
            vol = q.get("volume")
            vol_str = f"{vol:,}" if isinstance(vol, (int, float)) else "?"
            bit = f"vol {vol_str}" if not show_pct else chg_str
            name = q.get("name") or q.get("symbol", "")
            lines.append(f'{i}. {q.get("symbol","")} ({name}) — {bit}')
        return "\n".join(lines)

    messages = []
    if screener.get("volume"):
        messages.append(section("📊 TOP 10 BY VOLUME (LSE):", screener["volume"], show_pct=False))
    if screener.get("gainers"):
        messages.append(section("📈 TOP 10 GAINERS (LSE):", screener["gainers"]))
    else:
        messages.append(f"📈 TOP 10 GAINERS (LSE):  🕐 {now_stamp()}\nNo liquid (500k+ vol, 20p+) gainers found this run.")
    if screener.get("losers"):
        messages.append(section("📉 TOP 10 LOSERS (LSE):", screener["losers"]))
    else:
        messages.append(f"📉 TOP 10 LOSERS (LSE):  🕐 {now_stamp()}\nNo liquid (500k+ vol, 20p+) losers found this run.")
    return messages


_RATE_LIMIT_WINDOW_SECONDS = 240 * 60  # matches CallMeBot's free-tier window
_RATE_LIMIT_MAX_SENDS = 14  # stay under CallMeBot's 16/240min cap with a safety margin
_recent_send_times = []  # populated from seen_state at the start of each run, see main()


def _current_template():
    return (os.environ.get("WEBHOOK_TEMPLATE", "").strip() or "callmebot")


def _rate_limit_ok():
    """Global send budget shared across every message type this run (and recent prior
    runs) — the per-category throttles above help, but this is the actual backstop that
    guarantees the WhatsApp free-tier cap is never exceeded regardless of which
    combination of screener/alert/heartbeat messages happens to fire in a given hour.
    Only applies to CallMeBot specifically — ntfy and generic webhooks don't share that
    limit, so they're not artificially held back by a cap that doesn't apply to them."""
    if _current_template() != "callmebot":
        return True
    now = time.time()
    while _recent_send_times and now - _recent_send_times[0] > _RATE_LIMIT_WINDOW_SECONDS:
        _recent_send_times.pop(0)
    return len(_recent_send_times) < _RATE_LIMIT_MAX_SENDS


def send_webhook(message, bypass_market_hours_gate=False):
    if not bypass_market_hours_gate and not _is_uk_market_hours(datetime.now(timezone.utc)):
        # Covers weekends AND LSE bank holidays (both handled inside
        # _is_uk_market_hours) — e.g. the recurring "still checking"
        # heartbeat, screener summaries, and mover/broker alerts have
        # nothing genuinely new to report when the market's shut, so
        # sending them is just noise. A scheduled run can still fire on
        # a bank holiday (cron has no concept of holidays), so this is
        # the actual gate that stops it becoming a notification, not the
        # workflow schedule. bypass_market_hours_gate=True is for
        # genuinely exceptional cases — e.g. "the poller itself crashed"
        # — where the reader needs to know regardless of what day it is.
        print("  ! send skipped: LSE closed (weekend or bank holiday) — still on dashboard.", file=sys.stderr)
        return
    if not _rate_limit_ok():
        print(f"  ! send skipped: WhatsApp rate-limit budget exhausted ({_RATE_LIMIT_MAX_SENDS}/{_RATE_LIMIT_WINDOW_SECONDS//60}min) — still on dashboard.", file=sys.stderr)
        return
    webhook_url = os.environ.get("WEBHOOK_URL", "").strip()
    # GitHub passes a missing secret through as an EMPTY STRING env var, not an absent
    # one — so os.environ.get(..., "callmebot") never actually triggers its default in
    # that case. Explicitly fall back after stripping, regardless of why it's empty.
    template = (os.environ.get("WEBHOOK_TEMPLATE", "").strip() or "callmebot")
    if not webhook_url:
        print("  ! WEBHOOK_URL is empty — check the secret exists and is named exactly WEBHOOK_URL", file=sys.stderr)
        return
    print(f"  > sending webhook via template='{template}'")
    _recent_send_times.append(time.time())  # count the attempt itself, regardless of outcome —
                                              # matches CallMeBot's own "message queued" behavior
    try:
        if template == "callmebot":
            sep = "&" if "?" in webhook_url else "?"
            url = f"{webhook_url}{sep}text={urllib.parse.quote(message)}"
            resp = urllib.request.urlopen(urllib.request.Request(url, headers=HEADERS), timeout=15)
            body = resp.read().decode("utf-8", errors="replace")
            print(f"  > callmebot response ({resp.status}): {body[:200]}")
        elif template == "ntfy":
            req = urllib.request.Request(webhook_url, data=message.encode("utf-8"), method="POST", headers=HEADERS)
            resp = urllib.request.urlopen(req, timeout=15)
            print(f"  > ntfy response: {resp.status}")
        else:
            body = json.dumps({"text": message, "message": message}).encode("utf-8")
            req = urllib.request.Request(
                webhook_url, data=body, method="POST",
                headers={**HEADERS, "Content-Type": "application/json"},
            )
            resp = urllib.request.urlopen(req, timeout=15)
            print(f"  > generic webhook response: {resp.status}")
    except Exception as e:
        print(f"  ! webhook send failed: {e}", file=sys.stderr)


def load_json(path, default):
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return default


def save_json(path, data):
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def format_market_cap(value):
    """Readable market cap — e.g. 1_234_000_000 -> '£1.23bn'. Yahoo returns this in the
    stock's own currency (GBp for most LSE stocks, so this is already pence-scale — but
    marketCap itself is typically reported in full currency units, not pence, by Yahoo)."""
    if value is None:
        return None
    if value >= 1_000_000_000:
        return f"£{value / 1_000_000_000:.2f}bn"
    if value >= 1_000_000:
        return f"£{value / 1_000_000:.1f}m"
    return f"£{value:,.0f}"


SPARKLINE_MIN_POINTS = 5  # fewer real data points than this isn't a meaningful chart
SPARKLINE_WIDTH = 280
SPARKLINE_PRICE_HEIGHT = 50
SPARKLINE_VOLUME_HEIGHT = 20
SPARKLINE_GAP = 4
SPARKLINE_TOTAL_HEIGHT = SPARKLINE_PRICE_HEIGHT + SPARKLINE_GAP + SPARKLINE_VOLUME_HEIGHT


def render_price_volume_sparkline(series, currency_suffix=""):
    """
    Compact server-side SVG sparkline — price line above, volume as a
    thin comb of vertical lines below, visually distinguished by both
    colour AND shape/position (never colour alone). Pure function of the
    ALREADY-RETAINED series (see fetch_price_technicals's own
    priceVolumeSeries field) — never fetches anything itself, never
    computes RSI/MA/ATR/etc, which stay entirely untouched elsewhere.

    Returns "" (never a broken/empty <svg> tag) when the series is
    missing or has fewer than SPARKLINE_MIN_POINTS genuine closes — the
    caller simply omits the chart in that case, same graceful-degradation
    pattern used everywhere else in this codebase.

    The <title> element is calculated from this SAME rendered series, so
    a screen reader never gets a fact the visual chart doesn't also show
    (and vice versa) — no information exists only visually.
    """
    if not series:
        return ""
    closes = [pt.get("close") for pt in series if pt.get("close") is not None]
    if len(closes) < SPARKLINE_MIN_POINTS:
        return ""
    min_c, max_c = min(closes), max(closes)
    c_range = (max_c - min_c) or 1  # avoid division by zero on a genuinely flat price
    volumes = [pt.get("volume") or 0 for pt in series]
    max_v = max(volumes) or 1

    n = len(series)

    def x_for(i):
        return (i / (n - 1)) * SPARKLINE_WIDTH if n > 1 else 0

    price_points = []
    for i, pt in enumerate(series):
        c = pt.get("close")
        if c is None:
            continue
        x = x_for(i)
        y = SPARKLINE_PRICE_HEIGHT - ((c - min_c) / c_range) * SPARKLINE_PRICE_HEIGHT
        price_points.append(f"{x:.1f},{y:.1f}")
    if len(price_points) < SPARKLINE_MIN_POINTS:
        return ""
    price_polyline = " ".join(price_points)

    # Volume as ONE compact <path> (a comb of vertical strokes) rather
    # than one <rect> per bar — meaningfully smaller output for the same
    # visual result.
    bar_width = max(SPARKLINE_WIDTH / n * 0.6, 1)
    vol_path_parts = []
    for i, pt in enumerate(series):
        v = pt.get("volume") or 0
        x = x_for(i)
        h = (v / max_v) * SPARKLINE_VOLUME_HEIGHT
        y_top = SPARKLINE_TOTAL_HEIGHT - h
        vol_path_parts.append(f"M{x:.1f} {SPARKLINE_TOTAL_HEIGHT} L{x:.1f} {y_top:.1f}")
    vol_path = " ".join(vol_path_parts)

    first_close, last_close = closes[0], closes[-1]
    title_text = (
        f"{n}-day price chart: opened {first_close:.2f}{currency_suffix}, "
        f"closed {last_close:.2f}{currency_suffix}, ranged {min_c:.2f}\u2013{max_c:.2f}{currency_suffix}. "
        f"Volume bars shown below the price line."
    )

    return (
        f'<svg viewBox="0 0 {SPARKLINE_WIDTH} {SPARKLINE_TOTAL_HEIGHT}" '
        f'style="width:100%;max-width:320px;height:auto;display:block;" role="img" preserveAspectRatio="xMidYMid meet">'
        f'<title>{esc(title_text)}</title>'
        f'<path d="{vol_path}" stroke="#3a4150" stroke-width="{bar_width:.1f}" fill="none"/>'
        f'<polyline points="{price_polyline}" fill="none" stroke="#7fb3ff" stroke-width="1.5"/>'
        f'</svg>'
    )


def scale_bar_html(label, value_display, position_pct, lo_label, hi_label, color_lo, color_hi, zone_lo=33, zone_hi=66):
    """
    A compact positional scale bar: shows WHERE a number sits on a labelled range,
    never whether that's "good" or "bad" — no buy/sell framing, no green-means-go
    styling. The marker's position is the only signal; colour is purely to distinguish
    the two ends of the scale (e.g. small-cap vs large-cap), not a verdict.
    position_pct: 0-100, already clamped by the caller.
    zone_lo/zone_hi: where the colour bands split (default even thirds, middle band neutral grey).
    """
    pos = max(0, min(100, position_pct))
    return (
        f'<div style="margin:6px 0 3px;max-width:320px;">'
        f'<div style="display:flex;justify-content:space-between;font-size:14px;color:#9aa0a6;margin-bottom:3px;">'
        f'<span>{esc_safe(label)}</span><span style="color:#e8eaed;font-weight:700;">{esc_safe(value_display)}</span></div>'
        f'<div style="position:relative;height:7px;border-radius:4px;'
        f'background:linear-gradient(to right, {color_lo} 0%, {color_lo} {zone_lo}%, '
        f'#2a2e37 {zone_lo}%, #2a2e37 {zone_hi}%, {color_hi} {zone_hi}%, {color_hi} 100%);">'
        f'<div style="position:absolute;left:{pos:.0f}%;top:-3px;width:3px;height:13px;'
        f'background:#e8eaed;border-radius:1px;"></div></div>'
        f'<div style="display:flex;justify-content:space-between;font-size:12px;color:#6b7078;margin-top:2px;">'
        f'<span>{esc_safe(lo_label)}</span><span>{esc_safe(hi_label)}</span></div></div>'
    )


def esc_safe(s):
    """Small standalone escaper — scale_bar_html is called from inside render_dashboard's
    nested functions where the main esc() closure isn't in scope, so this avoids relying
    on closure capture across function boundaries."""
    return (str(s) if s is not None else "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def pe_scale(pe):
    """P/E scale: 0-30 range, no 'good/bad' framing — a P/E can be low because a stock
    is cheap OR because something's wrong; can be high because of growth optimism OR
    overpricing. The bar shows POSITION only, never a verdict."""
    if pe is None or pe < 0:
        return ""
    pos = min(pe / 30 * 100, 100)
    return scale_bar_html("P/E", f"{pe:.1f}", pos, "low", "high", "#7fb3ff", "#e0d267")


def rsi_scale(rsi):
    """RSI is already 0-100 natively — no transform needed. Labels describe the fact
    (risen/fallen fast) never a recommendation (never 'overbought'/'oversold')."""
    if rsi is None:
        return ""
    return scale_bar_html("RSI (14)", f"{rsi:.1f}", rsi, "0 (fallen fast)", "100 (risen fast)", "#7fb3ff", "#e0d267")


def mktcap_scale(mcap_value):
    """Log-scale position: £1m -> 0%, £10bn -> 100%. Purely a size classification
    (micro/mid/large-cap), not a quality judgement — small isn't 'bad', it's smaller."""
    if not mcap_value or mcap_value <= 0:
        return ""
    import math
    log_val = math.log10(max(mcap_value, 1))
    pos = (log_val - 6) / (10 - 6) * 100  # 10^6 (£1m) to 10^10 (£10bn)
    pos = max(0, min(100, pos))
    display = format_market_cap(mcap_value) or ""
    return scale_bar_html("Mkt cap", display, pos, "micro-cap", "large-cap", "#f0997b", "#5dcaa5")


def eps_scale(eps):
    """EPS scale: -£0.50 to £2.00 — a company can be a legitimate early-stage business
    with negative EPS, so 'loss-making' is a factual label, not a red flag icon."""
    if eps is None:
        return ""
    pos = (eps - (-0.5)) / (2.0 - (-0.5)) * 100
    pos = max(0, min(100, pos))
    return scale_bar_html("EPS", f"£{eps:.2f}", pos, "loss-making", "strong profit/share", "#f0997b", "#5dcaa5")


# =========================================================================
# Research-view calculations — Phase 1 of the "actionable LSE research
# dashboard" work. Every function here is a PURE, deterministic
# calculation over fields already present on an enriched screener/
# watchlist row (target price, current price, volume, averageVolume,
# ma20, rsi14, changePct) — no new network calls, no new data source.
# None of these functions produce investment advice or a recommendation
# — they compute and label FACTS (a distance, a ratio, a flag describing
# an observed combination of already-known figures), never a judgement
# on what to do about them.
# =========================================================================

def compute_target_upside_pct(price, target):
    """
    (target / price - 1) * 100 — how far the ALREADY-PUBLISHED broker
    consensus target sits from the current price. This is a distance
    calculation, not a forecast: it says nothing about whether the
    target will be reached, only how far away it currently is.
    Returns None if either input is missing or price is zero/invalid.
    """
    if price is None or target is None or price == 0:
        return None
    return (target / price - 1) * 100


def compute_ma20_distance_pct(price, ma20):
    """
    (price / ma20 - 1) * 100 — exact % distance from the 20-day moving
    average, replacing the previous above/below-only binary. A fact
    about where today's price sits relative to a recent trend baseline,
    not a signal to act on by itself.
    """
    if price is None or ma20 is None or ma20 == 0:
        return None
    return (price / ma20 - 1) * 100


def compute_volume_ratio(volume, average_volume):
    """
    Today's volume as a multiple of the published average volume — e.g.
    3.1 means "3.1x average volume". Returns None if either input is
    missing or the average is zero/invalid (never fabricates a ratio
    from incomplete data).
    """
    if volume is None or average_volume is None or average_volume == 0:
        return None
    return volume / average_volume


HIGH_VOLUME_RATIO_THRESHOLD = 1.5  # today's volume at least 1.5x the published average
OVEREXTENDED_RSI_THRESHOLD = 70    # standard RSI "overbought" reference level

# DON'T CHASE warning thresholds — deliberately separate from
# UPTREND_5DAY_THRESHOLD_PCT (5.0%, used for the "5-Day Uptrend" listing,
# which just flags "notable"). This warning is about genuine chasing
# risk, so it needs a materially higher bar for the move itself, PLUS at
# least one other sign of being technically extended — reuses the
# EXISTING OVEREXTENDED_RSI_THRESHOLD/HIGH_VOLUME_RATIO_THRESHOLD above
# rather than inventing new ones. All three constants are grouped here
# specifically so they're easy to find and adjust together later.
DONT_CHASE_5DAY_MOVE_THRESHOLD_PCT = 15.0


def compute_dont_chase_warning(change_pct_5d, rsi14, volume_ratio):
    """
    Returns None, or {"reasons": [str, ...]} — a purely factual pattern
    match over data ALREADY computed elsewhere (5-day price change, RSI,
    volume ratio), never a new score and never a buy/sell instruction.
    Fires only when the 5-day move itself clears
    DONT_CHASE_5DAY_MOVE_THRESHOLD_PCT AND at least one of RSI/volume
    also independently clears its own existing threshold — a big move
    alone, with no other sign of being stretched, is not flagged; that's
    just what a normal healthy trend looks like. Every reason shown is
    a specific number from real fetched/calculated data, not a generic
    label. Missing data (any input None) is handled safely: that
    specific factor simply can't contribute, and if the move itself is
    unknown, no warning is produced at all — absence of data is never
    treated as a signal, exactly as elsewhere in this codebase.
    """
    if change_pct_5d is None or abs(change_pct_5d) < DONT_CHASE_5DAY_MOVE_THRESHOLD_PCT:
        return None
    reasons = [f"{'+' if change_pct_5d >= 0 else ''}{change_pct_5d:.1f}% over 5 days"]
    extended = False
    if rsi14 is not None and rsi14 >= OVEREXTENDED_RSI_THRESHOLD:
        reasons.append(f"RSI {rsi14:.0f}")
        extended = True
    if volume_ratio is not None and volume_ratio >= HIGH_VOLUME_RATIO_THRESHOLD:
        reasons.append(f"Volume {volume_ratio:.1f}× average")
        extended = True
    if not extended:
        return None  # a large move alone, with no overextension signal, isn't flagged
    return {"reasons": reasons}


def compute_opportunity_flags(changePct, volume_ratio, above_ma20, rsi14, has_news):
    """
    Deterministic, explainable research flags — transparent rule
    combinations over already-known figures, NEVER a prediction, and
    NEVER labelled as a buy/sell signal. Each flag returned as
    (flag_id, label, reason_string) so the dashboard can show WHY a flag
    fired, not just that it did. A row can carry more than one flag
    (e.g. a big rise on strong volume with no news yet is both
    "momentum + volume" AND, if RSI is also elevated, "overextended" —
    these are not mutually exclusive facts about the same move).

    Any input that's None is treated as "insufficient data for that
    specific check" — a flag requiring volume_ratio never fires if
    volume_ratio is None, rather than guessing.
    """
    flags = []
    chg = changePct or 0
    significant_move = abs(chg) >= BIG_MOVER_THRESHOLD_PCT
    high_volume = volume_ratio is not None and volume_ratio >= HIGH_VOLUME_RATIO_THRESHOLD

    if chg > 0 and high_volume and above_ma20:
        flags.append((
            "momentum_volume", "🚀 Momentum + volume",
            f"+{chg:.1f}% · {volume_ratio:.1f}× average volume · above 20-day MA",
        ))
    if significant_move and has_news:
        flags.append((
            "positive_catalyst" if chg >= 0 else "negative_catalyst",
            "🟢 Catalyst found" if chg >= 0 else "🔴 Negative catalyst",
            f"{'+' if chg >= 0 else ''}{chg:.1f}% · relevant same-day news found",
        ))
    if significant_move and not has_news:
        flags.append((
            "move_without_catalyst", "🟡 Move without catalyst",
            f"{'+' if chg >= 0 else ''}{chg:.1f}% · "
            f"{'high volume' if high_volume else 'no volume confirmation'} · no relevant same-day news found",
        ))
    if chg > 0 and rsi14 is not None and rsi14 >= OVEREXTENDED_RSI_THRESHOLD:
        flags.append((
            "overextended", "🟡 Overextended",
            f"+{chg:.1f}% recent move · RSI {rsi14:.0f}",
        ))
    if chg < 0 and high_volume:
        flags.append((
            "weakness_volume", "🔴 Weakness + volume",
            f"{chg:.1f}% · {volume_ratio:.1f}× average volume",
        ))
    return flags


# =========================================================================
# Phase 2 — connective intelligence: evidence-agreement classification,
# multi-event broker momentum, and market-context helpers. Same
# discipline as Phase 1: pure, deterministic, no new network calls, and
# — the important addition here — no fabricated sentiment. This system
# has no text sentiment analysis and will not simulate one: a generic
# "news"/"event"/"reiteration"/"target" category item confirms a
# catalyst EXISTS, never its direction. Direction is only ever asserted
# from a genuinely directional classification already computed elsewhere
# (classify()'s upgrade/downgrade/target_raise/target_cut categories).
# =========================================================================

# Categories with a genuinely known direction — everything else that's
# still a real catalyst (news, event, reiteration, target, initiation)
# only ever confirms presence, never a guessed direction. "initiation" is
# deliberately excluded from the positive set: an initiation's own
# direction depends on what rating it initiated AT, which classify()'s
# category alone doesn't tell us — asserting "positive" without knowing
# that would be exactly the kind of fabrication being avoided here.
EVIDENCE_POSITIVE_CATEGORIES = {"upgrade", "target_raise"}
EVIDENCE_NEGATIVE_CATEGORIES = {"downgrade", "target_cut"}


def classify_evidence(changePct, volume_ratio, news_items, has_significant_move_threshold=None, latest_broker_event=None):
    """
    Deterministic evidence-agreement classification — a fact about whether
    the AVAILABLE signals (price direction, volume, and any directionally-
    classified catalyst) agree or conflict, never a prediction and never a
    buy/sell instruction.

    Returns a dict with every component visible, not just a final label —
    "no black box": {
        "label": one of "supported" / "conflicting" / "catalyst_unclear_direction"
                  / "unexplained_move" / "no_signal",
        "hasCatalyst": bool — was ANY relevant news/event found for this stock today,
        "catalystDirection": None | "positive" | "negative" — ONLY set when a
            genuinely directional item (upgrade/downgrade/target_raise/target_cut,
            from either a news item's category OR a same-day dated broker event)
            was found; a generic news item never sets this,
        "volumeConfirms": True | False | None (None when volume_ratio itself
            is unavailable, so absence of data is never silently treated as
            "no confirmation"),
    }

    "supported": significant move + a directional catalyst whose direction
    matches the price move (e.g. price up + an upgrade/target-raise today).
    "conflicting": significant move + a directional catalyst pointing the
    OTHER way (e.g. price falling despite a same-day upgrade/target-raise —
    exactly the "conflicting evidence" example given).
    "catalyst_unclear_direction": significant move + a real catalyst exists,
    but it's a generic news/event item with no determinable direction —
    the honest answer when sentiment can't be inferred, not a guess.
    "unexplained_move": significant move, no catalyst found at all.
    "no_signal": the move itself isn't significant enough to classify.

    latest_broker_event: the SAME dated event object the Broker
    Intelligence block renders (see get_latest_broker_event_per_ticker) —
    added after a real end-to-end scenario walkthrough caught a genuine
    inconsistency: without this, a same-day broker downgrade that wasn't
    ALSO independently surfaced by the separate news-scraping pipeline
    (a realistic gap — the structured events pipeline and the news feed
    are two different sources) was invisible to this function, so a stock
    rising +6% on a same-day downgrade was labelled "unexplained_move"
    here while the Broker Intelligence block directly below correctly
    showed the downgrade — two parts of the same rendered picture
    disagreeing. Deliberately scoped to TODAY's date only (London terms),
    matching the same-day principle news items are already held to; an
    event from weeks ago being folded into today's evidence picture would
    be its own kind of misleading.
    """
    threshold = has_significant_move_threshold if has_significant_move_threshold is not None else BIG_MOVER_THRESHOLD_PCT
    chg = changePct or 0
    significant_move = abs(chg) >= threshold
    volume_confirms = (volume_ratio >= HIGH_VOLUME_RATIO_THRESHOLD) if volume_ratio is not None else None

    catalyst_direction = None
    has_catalyst = False
    for it in (news_items or []):
        category = it.get("category")
        if category in EVIDENCE_POSITIVE_CATEGORIES:
            catalyst_direction = "positive"
            has_catalyst = True
        elif category in EVIDENCE_NEGATIVE_CATEGORIES:
            catalyst_direction = "negative"
            has_catalyst = True
        elif category:
            has_catalyst = True  # a real catalyst, direction just not determinable

    if latest_broker_event and latest_broker_event.get("date"):
        try:
            today_london = datetime.now(timezone.utc).astimezone(LONDON_TZ).strftime("%Y-%m-%d")
            if latest_broker_event["date"] == today_london:
                action = latest_broker_event.get("action")
                if action == "UPGRADE":
                    catalyst_direction = "positive"
                    has_catalyst = True
                elif action == "DOWNGRADE":
                    catalyst_direction = "negative"
                    has_catalyst = True
        except Exception:
            pass  # never let a malformed date field break relevance classification

    if not significant_move:
        label = "no_signal"
    elif catalyst_direction is not None:
        agrees = (chg >= 0 and catalyst_direction == "positive") or (chg < 0 and catalyst_direction == "negative")
        label = "supported" if agrees else "conflicting"
    elif has_catalyst:
        label = "catalyst_unclear_direction"
    else:
        label = "unexplained_move"

    return {
        "label": label,
        "hasCatalyst": has_catalyst,
        "catalystDirection": catalyst_direction,
        "volumeConfirms": volume_confirms,
    }


BROKER_MOMENTUM_LOOKBACK_DAYS = 90  # a quarter-ish window — long enough to see a real
                                     # trend across multiple events, not just noise from one


def compute_broker_momentum(events_for_ticker, lookback_days=BROKER_MOMENTUM_LOOKBACK_DAYS, now_utc=None):
    """
    Deterministic broker-sentiment TREND across multiple dated events, not
    just the single latest one (see get_latest_broker_event_per_ticker for
    that). A rating UPGRADE or a target RAISE (new_target > old_target,
    compared numerically when both are present — never inferred from
    action alone) each count as one positive-direction event; DOWNGRADE
    or a target CUT each count as one negative. REITERATION/INITIATION/
    NO_CHANGE/RATING_CHANGE (an ambiguous bucket-comparison fallback)
    contribute to neither count — never guessed.

    Returns {"direction": "improving"|"stable"|"deteriorating"|"no_recent_activity",
             "netScore": int, "eventCount": int} — netScore and eventCount
    are always included so the direction label is never a black box; a
    caller can always see exactly what it was computed from.
    """
    now_utc = now_utc or datetime.now(timezone.utc)
    positive = 0
    negative = 0
    counted = 0
    for e in (events_for_ticker or []):
        if e.get("superseded_by"):
            continue
        ts = e.get("timestamp")
        if not ts:
            continue
        try:
            dt = datetime.fromisoformat(ts)
        except (ValueError, TypeError):
            continue
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        if (now_utc - dt).days > lookback_days:
            continue
        counted += 1
        action = e.get("action")
        if action == "UPGRADE":
            positive += 1
        elif action == "DOWNGRADE":
            negative += 1
        old_t, new_t = e.get("old_target"), e.get("new_target")
        if old_t is not None and new_t is not None:
            try:
                old_f, new_f = float(old_t), float(new_t)
                if new_f > old_f:
                    positive += 1
                elif new_f < old_f:
                    negative += 1
            except (TypeError, ValueError):
                pass
    if counted == 0:
        return {"direction": "no_recent_activity", "netScore": 0, "eventCount": 0}
    net = positive - negative
    direction = "improving" if net > 0 else ("deteriorating" if net < 0 else "stable")
    return {"direction": direction, "netScore": net, "eventCount": counted}


def compute_relative_to_ftse(stock_change_pct, ftse_change_pct):
    """
    (stock % move) - (FTSE 100 % move) — the SAME kind of story-changing
    context your brief describes ("Shell +4% while FTSE +0.2%" reads very
    differently from "Shell +4% while FTSE -0.5%"). Genuinely available
    now: ftse100's own changePct is already fetched every cycle, no new
    source. Returns None if either input is missing — never fabricated
    from a partial figure.
    """
    if stock_change_pct is None or ftse_change_pct is None:
        return None
    return stock_change_pct - ftse_change_pct


MIN_SECTOR_SAMPLE_SIZE = 3  # below this, a "sector average" is barely more informative
                            # than one or two other stocks dressed up as a market signal


def compute_sector_relative_context(ticker, sector, all_enriched_rows):
    """
    A DELIBERATELY CAUTIOUS sector-context approximation — explicitly NOT
    presented as an authoritative sector index, because no such data
    source exists or was found (investigated directly; see the design
    report). This averages the % move of OTHER stocks that happen to
    already be in the SAME enriched pool (watchlist + today's screener
    rankings) and share the same sector — a small, coincidental sample,
    not the real sector universe.

    Returns None if fewer than MIN_SECTOR_SAMPLE_SIZE other same-sector
    stocks exist in the pool THIS run — never manufactures a "sector
    average" from an inadequate sample and presents it as if authoritative,
    per the explicit requirement not to do that. When it DOES return a
    value, callers must show the sample size and methodology alongside it
    (see the rendering — this function only returns the number and count,
    it doesn't decide how it's labelled).
    """
    if not sector:
        return None
    same_sector_moves = [
        row.get("changePct") for row in all_enriched_rows
        if row.get("sector") == sector and row.get("symbol_or_ticker") != ticker
        and row.get("changePct") is not None
    ]
    if len(same_sector_moves) < MIN_SECTOR_SAMPLE_SIZE:
        return None
    avg = sum(same_sector_moves) / len(same_sector_moves)
    return {"avgChangePct": avg, "sampleSize": len(same_sector_moves)}


def esc(s):
    """Minimal HTML-escaping — module-level (not nested in render_dashboard) so
    the shared per-stock research rendering function can use it too, from
    outside render_dashboard's own closure."""
    return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def format_epoch_date(epoch_seconds):
    """Formats a Yahoo epoch timestamp (earnings/ex-dividend dates) as a plain London
    date. IMPORTANT: only returns a value if the date is today or in the future — Yahoo
    sometimes returns a stale PAST date for "next earnings"/"ex-dividend" when it has no
    genuine upcoming one (confirmed against real data: a "next earnings" date that had
    already passed, an "ex-dividend" date from 1996). Showing a past date labeled "next"
    would be actively misleading, so it's treated the same as missing data — omitted."""
    if not epoch_seconds:
        return None
    try:
        dt = datetime.fromtimestamp(epoch_seconds, tz=timezone.utc).astimezone(LONDON_TZ)
        if dt.date() < datetime.now(timezone.utc).astimezone(LONDON_TZ).date():
            return None  # stale/past date — not genuinely "next" anything
        return dt.strftime("%d %b %Y")
    except Exception:
        return None


# Urgency thresholds for the Upcoming Catalysts section — grouped here,
# easy to find and adjust. Deliberately icon-based, not color-only.
CATALYST_URGENT_DAYS = 2   # 🔴 today/in the next couple of days
CATALYST_SOON_DAYS = 7     # 🟡 this week
CATALYST_URGENCY_LABELS = {
    "urgent": "🔴 Imminent",
    "soon": "🟡 This week",
    "upcoming": "⚪ Upcoming",
}


def collect_catalyst_events(quotes, screener, watchlist):
    """
    Pulls together every known upcoming earnings/ex-dividend date across
    BOTH the watchlist and screener-ranked pools — same combined-pool,
    dedupe-by-bare-ticker pattern already established (see the GLEN
    merge-lookup fix and bare_ticker's own docstring), reused here rather
    than building a new aggregation approach. A stock present in both
    pools contributes each event only once.

    Purely calendar/scheduling data — never affected by whether the
    market is currently open, so this section is not wrapped in any of
    the market-hours-aware empty-state logic used for live price data
    elsewhere; a genuinely empty result here just means no known dates,
    regardless of what time it is.

    Stale/past dates are excluded (reuses the exact same "past = not
    genuinely upcoming" rule as format_epoch_date, applied independently
    here since this also needs the numeric days-until, not just a
    display string). Missing dates are simply absent, never fabricated.

    Returns a list of {"ticker", "name", "eventType" ("earnings" or
    "ex_dividend"), "date", "daysUntil", "urgency"}, sorted soonest-first.
    """
    def _event_or_none(epoch_seconds, event_type, ticker, name):
        if not epoch_seconds:
            return None
        try:
            dt_london = datetime.fromtimestamp(epoch_seconds, tz=timezone.utc).astimezone(LONDON_TZ)
        except (ValueError, OSError, OverflowError, TypeError):
            return None  # a malformed epoch never breaks the whole section
        today_london = datetime.now(timezone.utc).astimezone(LONDON_TZ).date()
        if dt_london.date() < today_london:
            return None  # stale/past — never shown as "upcoming"
        days_until = (dt_london.date() - today_london).days
        if days_until <= CATALYST_URGENT_DAYS:
            urgency = "urgent"
        elif days_until <= CATALYST_SOON_DAYS:
            urgency = "soon"
        else:
            urgency = "upcoming"
        return {
            "ticker": ticker, "name": name, "eventType": event_type,
            "date": dt_london.strftime("%d %b %Y"), "daysUntil": days_until, "urgency": urgency,
        }

    pool = {}  # bare ticker -> (display_ticker, name, data_dict) — first pool wins, matching the existing Broker Target Prices pattern
    for q in screener.get("volume", []) + screener.get("gainers", []) + screener.get("losers", []):
        symbol = q.get("symbol", "")
        if not symbol:
            continue
        bare = bare_ticker(symbol)
        if bare not in pool:
            pool[bare] = (symbol, q.get("name") or symbol, q)
    for stock in watchlist:
        ticker = stock["ticker"]
        bare = bare_ticker(ticker)
        if bare not in pool:
            pool[bare] = (ticker, stock["name"], quotes.get(ticker, {}))

    events = []
    for bare, (ticker, name, q) in pool.items():
        for epoch_key, event_type in (("nextEarningsDate", "earnings"), ("exDividendDate", "ex_dividend")):
            evt = _event_or_none(q.get(epoch_key), event_type, ticker, name)
            if evt:
                events.append(evt)

    events.sort(key=lambda e: e["daysUntil"])
    return events


EVIDENCE_LABEL_TEXT = {
    "supported": ("🟢 Supported", "price move agrees with a same-day directional broker action"),
    "conflicting": ("🔴 Conflicting evidence", "price move disagrees with a same-day directional broker action"),
    "catalyst_unclear_direction": ("🟡 Catalyst found, direction unclear", "relevant news exists but isn't broker-classified as directionally positive or negative"),
    "unexplained_move": ("🟡 Unexplained move", "no relevant same-day catalyst found for this move"),
    "no_signal": ("No signal", "move isn't large enough to classify"),
}

BROKER_MOMENTUM_LABEL_TEXT = {
    "improving": "↑ improving",
    "stable": "→ stable",
    "deteriorating": "↓ deteriorating",
    "no_recent_activity": "no recent broker activity",
}

# =========================================================================
# Data-quality taxonomy — every evidence line in the entry/exit panel is
# tagged with one of these, so provenance is always visible rather than
# blended together. Deliberately scoped to the NEW entry/exit panel for
# this phase, not retrofitted across every existing line on the
# dashboard — that would be a much larger, higher-risk pass across
# already-working, already-tested rendering, and is called out here as a
# scoping decision, not an oversight.
# =========================================================================
DATA_QUALITY_TAGS = {
    "SOURCE_FACT": "📌 fact",
    "CALCULATED": "🧮 calculated",
    "BROKER_OPINION": "🏦 broker opinion",
    "NEWS_REPORT": "📰 news",
    "SYSTEM_INTERPRETATION": "🧭 interpretation",
}


SCORECARD_DIMENSIONS = ["TREND", "MOMENTUM", "VOLUME", "NEWS", "BROKER", "TECHNICAL", "MARKET", "RISK"]


def compute_research_scorecard(
    change_pct_5d, above_ma20, rsi14, ma_crossover, volume_ratio, change_pct,
    evidence, broker_momentum, breakout_status, ftse_relative, sector_context,
    upside_pct,
):
    """
    A fully decomposable, deterministic research scorecard — explicitly
    NOT a prediction or a probability. Every dimension's score is a
    small integer (typically -2..+2) computed from a fixed, documented
    rule stated in this docstring, never a hidden weighting or anything
    resembling model output. The TOTAL is simply the sum of the visible
    parts — never separately calibrated, never converted into a
    "probability of rising" or similar framing anywhere in this system.

    Rules (each independent of the others, some inputs deliberately
    contribute to more than one dimension — that's intentional, so a
    genuine risk shows up wherever it's relevant rather than being
    hidden by only counting once):

    TREND: +1 if price above 20-day MA else -1; +1 if 5-day change
    positive else -1 (each only scored when its input is available).

    MOMENTUM: +1 if RSI in a healthy bullish zone (45-70) else -1 if
    RSI < 30 (weak) else 0; +1 for a bullish MA20/50 crossover, -1 for
    bearish.

    VOLUME: +2 if volume >=1.5x average AND price rising (confirmed
    up-move); -2 if >=1.5x average AND price falling (confirmed
    weakness); -1 if volume <0.8x average (weak participation); else 0.

    NEWS: uses classify_evidence's label directly — "supported": +2,
    "conflicting": -2, "unexplained_move": -1, else 0. Never infers a
    score from generic news sentiment (classify_evidence itself never
    does, so neither does this).

    BROKER: uses compute_broker_momentum's direction — "improving": +2,
    "deteriorating": -2, else 0.

    TECHNICAL: uses compute_breakout_status — "breakout": +1,
    "breakdown": -1, else 0. (ATR/volatility is reported elsewhere as
    context, deliberately NOT scored here — volatility itself isn't
    inherently positive or negative, so forcing it into a directional
    point would misrepresent what it actually means.)

    MARKET: +1 if outperforming FTSE 100, -1 if underperforming; +1 more
    if also outperforming its (sample-safeguarded) sector context, -1 if
    underperforming it — capped at ±2 total for this dimension.

    RISK: a caution-only accumulator, always <= 0: -1 if RSI >= 70
    (overextended), -1 if price already >5% above the broker consensus
    target (limited stated upside), -1 if evidence is "conflicting".

    Returns {"dimensions": {name: (score, [reason strings])}, "total": int,
    "confidence": "High"|"Medium"|"Low"}. Confidence is based on how many
    of the underlying inputs were actually AVAILABLE (data completeness),
    never on how confident the resulting score looks — a score built from
    mostly-missing data is explicitly labelled Low confidence even if its
    total happens to look decisive.
    """
    dims = {}

    # --- TREND ---
    trend_score, trend_reasons = 0, []
    if above_ma20 is True:
        trend_score += 1; trend_reasons.append("above 20-day MA (+1)")
    elif above_ma20 is False:
        trend_score -= 1; trend_reasons.append("below 20-day MA (-1)")
    if change_pct_5d is not None and change_pct_5d > 0:
        trend_score += 1; trend_reasons.append("positive 5-day trend (+1)")
    elif change_pct_5d is not None and change_pct_5d < 0:
        trend_score -= 1; trend_reasons.append("negative 5-day trend (-1)")
    dims["TREND"] = (trend_score, trend_reasons)

    # --- MOMENTUM ---
    mom_score, mom_reasons = 0, []
    if rsi14 is not None:
        if 45 <= rsi14 <= 70:
            mom_score += 1; mom_reasons.append(f"RSI {rsi14:.0f} in healthy bullish zone (+1)")
        elif rsi14 < 30:
            mom_score -= 1; mom_reasons.append(f"RSI {rsi14:.0f} weak (-1)")
    if ma_crossover == "bullish":
        mom_score += 1; mom_reasons.append("bullish MA20/50 crossover (+1)")
    elif ma_crossover == "bearish":
        mom_score -= 1; mom_reasons.append("bearish MA20/50 crossover (-1)")
    dims["MOMENTUM"] = (mom_score, mom_reasons)

    # --- VOLUME ---
    vol_score, vol_reasons = 0, []
    high_vol = volume_ratio is not None and volume_ratio >= HIGH_VOLUME_RATIO_THRESHOLD
    chg = change_pct or 0
    if high_vol and chg > 0:
        vol_score += 2; vol_reasons.append(f"{volume_ratio:.1f}× average volume confirms the rise (+2)")
    elif high_vol and chg < 0:
        vol_score -= 2; vol_reasons.append(f"{volume_ratio:.1f}× average volume confirms the decline (-2)")
    elif volume_ratio is not None and volume_ratio < 0.8:
        vol_score -= 1; vol_reasons.append(f"only {volume_ratio:.1f}× average — weak participation (-1)")
    dims["VOLUME"] = (vol_score, vol_reasons)

    # --- NEWS ---
    news_score, news_reasons = 0, []
    if evidence["label"] == "supported":
        news_score += 2; news_reasons.append("evidence supported by a directional broker action (+2)")
    elif evidence["label"] == "conflicting":
        news_score -= 2; news_reasons.append("evidence conflicts with a directional broker action (-2)")
    elif evidence["label"] == "unexplained_move":
        news_score -= 1; news_reasons.append("no catalyst found for this move (-1)")
    dims["NEWS"] = (news_score, news_reasons)

    # --- BROKER ---
    broker_score, broker_reasons = 0, []
    if broker_momentum and broker_momentum["direction"] == "improving":
        broker_score += 2; broker_reasons.append(f"broker momentum improving ({broker_momentum['eventCount']} action(s)) (+2)")
    elif broker_momentum and broker_momentum["direction"] == "deteriorating":
        broker_score -= 2; broker_reasons.append(f"broker momentum deteriorating ({broker_momentum['eventCount']} action(s)) (-2)")
    dims["BROKER"] = (broker_score, broker_reasons)

    # --- TECHNICAL ---
    tech_score, tech_reasons = 0, []
    if breakout_status == "breakout":
        tech_score += 1; tech_reasons.append(f"breakout above {SUPPORT_RESISTANCE_WINDOW_DAYS}-day high (+1)")
    elif breakout_status == "breakdown":
        tech_score -= 1; tech_reasons.append(f"breakdown below {SUPPORT_RESISTANCE_WINDOW_DAYS}-day low (-1)")
    dims["TECHNICAL"] = (tech_score, tech_reasons)

    # --- MARKET ---
    mkt_score, mkt_reasons = 0, []
    if ftse_relative is not None and ftse_relative > 0:
        mkt_score += 1; mkt_reasons.append(f"outperforming FTSE 100 by {ftse_relative:.1f}% (+1)")
    elif ftse_relative is not None and ftse_relative < 0:
        mkt_score -= 1; mkt_reasons.append(f"underperforming FTSE 100 by {abs(ftse_relative):.1f}% (-1)")
    if sector_context is not None:
        if sector_context["avgChangePct"] < chg:
            mkt_score += 1; mkt_reasons.append(f"outperforming its tracked sector sample, n={sector_context['sampleSize']} (+1)")
        elif sector_context["avgChangePct"] > chg:
            mkt_score -= 1; mkt_reasons.append(f"underperforming its tracked sector sample, n={sector_context['sampleSize']} (-1)")
    dims["MARKET"] = (mkt_score, mkt_reasons)

    # --- RISK (caution-only, always <= 0) ---
    risk_score, risk_reasons = 0, []
    if rsi14 is not None and rsi14 >= OVEREXTENDED_RSI_THRESHOLD:
        risk_score -= 1; risk_reasons.append(f"RSI {rsi14:.0f} overextended (-1)")
    if upside_pct is not None and upside_pct <= -5:
        risk_score -= 1; risk_reasons.append(f"price {abs(upside_pct):.1f}% above broker target (-1)")
    if evidence["label"] == "conflicting":
        risk_score -= 1; risk_reasons.append("conflicting evidence (-1)")
    dims["RISK"] = (risk_score, risk_reasons)

    total = sum(score for score, _reasons in dims.values())

    # Confidence: data-completeness based, never probability-based.
    available_inputs = [
        change_pct_5d, above_ma20, rsi14, ma_crossover, volume_ratio,
        upside_pct, ftse_relative,
        broker_momentum["eventCount"] > 0 if broker_momentum else False,
        breakout_status,
    ]
    available_count = sum(1 for v in available_inputs if v not in (None, False))
    if available_count >= 7:
        confidence = "High"
    elif available_count >= 4:
        confidence = "Medium"
    else:
        confidence = "Low"

    return {"dimensions": dims, "total": total, "confidence": confidence}


# Scorecard subtotals — purely a presentation-layer regrouping of the
# SAME 8 already-computed dimension scores above, never a new
# calculation. Every dimension is assigned to EXACTLY one bucket, so
# TECHNICAL_MARKET_DIMENSIONS + RESEARCH_EVIDENCE_DIMENSIONS +
# {"RISK"} together account for all of SCORECARD_DIMENSIONS with no
# overlap and no omission — the three groups reconcile to the existing
# total by construction, not by any additional check.
#
# "Research Evidence" (not "Fundamental") deliberately — NEWS and BROKER
# never touch P/E, earnings, revenue, or any actual financial-statement
# data (nothing in this scorecard does); they measure whether a same-day
# catalyst exists and which way recent analyst opinion is moving. Calling
# that "Fundamental" would overclaim what's actually being measured.
#
# MARKET sits under Technical/Market Evidence, not Research Evidence —
# relative-strength-vs-benchmark is a standard technical-analysis
# concept, and MARKET never draws on news/broker/research inputs at all,
# only price-vs-FTSE and price-vs-sector comparisons.
#
# RISK is deliberately excluded from BOTH subtotals — it's a caution-only
# dimension (always <=0) that itself draws on inputs from both other
# groups (RSI from Technical/Market, target-proximity and conflicting-
# evidence from Research), so folding it into either bucket would
# misrepresent it. It's shown on its own, separately.
TECHNICAL_MARKET_DIMENSIONS = ["TREND", "MOMENTUM", "VOLUME", "TECHNICAL", "MARKET"]
RESEARCH_EVIDENCE_DIMENSIONS = ["NEWS", "BROKER"]


def compute_scorecard_subtotals(dimensions):
    """
    Returns {"technicalMarket": int, "researchEvidence": int, "risk": int}
    — pure sums over the ALREADY-COMPUTED per-dimension scores passed in.
    Never recalculates anything compute_research_scorecard already
    decided; never called before that function has already run.
    Mathematically: technicalMarket + researchEvidence + risk always
    equals the existing scorecard total exactly, since every one of the
    8 dimensions is counted in precisely one of these three groups.
    """
    technical_market = sum(dimensions[d][0] for d in TECHNICAL_MARKET_DIMENSIONS)
    research_evidence = sum(dimensions[d][0] for d in RESEARCH_EVIDENCE_DIMENSIONS)
    risk = dimensions["RISK"][0]
    return {"technicalMarket": technical_market, "researchEvidence": research_evidence, "risk": risk}


# Signal Quality — deliberately separate from "confidence" above (data
# completeness) and never touches it. This is about agreement: given
# whatever the scorecard actually concluded, do the scored dimensions
# point the same way? A stock can have High confidence (lots of raw
# data available) with Mixed signal quality (the data that IS available
# disagrees), or Low confidence with Strong signal quality (little data,
# but what little exists agrees) — two genuinely independent facts, not
# two words for the same thing.
SIGNAL_QUALITY_MIN_AGREEING_DIMENSIONS = 2  # fewer non-zero dimensions than
# this can't meaningfully demonstrate "agreement" — a single dimension has
# nothing else to agree WITH, so it falls to Weak rather than Strong.


def compute_signal_quality(dimensions, contradictions):
    """
    Returns "Strong", "Mixed", or "Weak" — purely derived from data
    ALREADY computed by compute_research_scorecard (dimensions) and
    detect_contradictions (contradictions), never a new score and never
    a new fetch.

    1. Any specific, already-DETECTED contradiction means Mixed, full
       stop — a named conflict is a stronger signal than a bare tally of
       positive vs negative dimensions, so it overrides the count below.
    2. Otherwise, count dimensions with a genuinely NON-ZERO score (a
       dimension scoring exactly 0 is silence, not agreement OR
       disagreement, so it's excluded from this count entirely):
       - fewer than SIGNAL_QUALITY_MIN_AGREEING_DIMENSIONS non-zero
         dimensions -> Weak (nothing meaningful to judge agreement from)
       - all non-zero dimensions share the same sign -> Strong
       - signs are mixed -> Mixed
    """
    if contradictions:
        return "Mixed"
    positive = sum(1 for score, _reasons in dimensions.values() if score > 0)
    negative = sum(1 for score, _reasons in dimensions.values() if score < 0)
    non_zero = positive + negative
    if non_zero < SIGNAL_QUALITY_MIN_AGREEING_DIMENSIONS:
        return "Weak"
    if positive > 0 and negative > 0:
        return "Mixed"
    return "Strong"


NEWS_TYPE_KEYWORDS = {
    "earnings": {"earnings", "profit", "revenue", "quarterly", "interim results", "annual results", "trading update", "full-year results", "half-year results"},
    "legal_regulatory": {"regulator", "fine", "lawsuit", "investigation", "fca", "antitrust", "compliance", "breach", "sanction"},
    "operational": {"outage", "shutdown", "incident", "disruption", "fire", "recall", "strike", "closure"},
    "investment_capex": {"invest", "capex", "expansion", "new plant", "new facility", "to build", "construction"},
    "management": {"chief executive", "chairman", "appoints", "resigns", "steps down", "names new", "management change"},
    "mna": {"acquisition", "acquire", "merger", "takeover", "buyout", "divest"},
}


def classify_news_type(title, category=None):
    """
    Purely DESCRIPTIVE secondary categorization of what KIND of news a
    headline is — earnings / legal-regulatory / operational / investment-
    capex / management / M&A / broker action / other. Completely separate
    from relevance filtering (passes_relevance_filter never uses this)
    and NEVER asserts sentiment or direction — same discipline as the
    relevance-disqualifier work: small, explicit, auditable keyword
    lists, not an attempt at full topic modelling.

    If `category` is already one of classify()'s broker-action categories
    (upgrade/downgrade/target/target_raise/target_cut/initiation/
    reiteration), returns "broker_action" directly — reusing that
    decision rather than re-deriving it, so the two can never disagree.
    Otherwise applies the keyword lists above to the headline text; a
    headline matching none of them is honestly labelled "other" rather
    than forced into a category it doesn't fit.
    """
    if category in ("upgrade", "downgrade", "target", "target_raise", "target_cut", "initiation", "reiteration"):
        return "broker_action"
    if not title:
        return "other"
    t = title.lower()
    for type_name, keywords in NEWS_TYPE_KEYWORDS.items():
        if any(kw in t for kw in keywords):
            return type_name
    return "other"


NEWS_TYPE_LABELS = {
    "earnings": "📊 Earnings",
    "legal_regulatory": "⚖️ Legal/Regulatory",
    "operational": "🏭 Operational",
    "investment_capex": "💰 Investment/Capex",
    "management": "👔 Management",
    "mna": "🤝 M&A",
    "broker_action": "🏦 Broker Action",
    "other": "📰 Other",
}


# =========================================================================
# Source-type classification — a SEPARATE taxonomy from classify_news_type
# above (which describes WHAT KIND of story a headline is: earnings, M&A,
# etc). This describes WHERE it came from and how authoritative that
# provenance is: company/RNS fact, formal broker data, financial
# journalism, retail commentary, or genuinely unclassifiable. Same
# discipline as classify_news_type: never guesses. A source is only ever
# attributed with certainty determinable from HOW the item was fetched
# (a structured feed, not a text search) or from the item's OWN real URL
# (its actual publishing domain) — never inferred from search source,
# title wording, or any other soft signal. Anything not confidently
# determinable this way is honestly labelled "Unknown/Aggregated" rather
# than assigned a specific attribution that might be wrong.
# =========================================================================

SOURCE_TYPE_DOMAIN_MAP = {
    "reuters.com": ("financial_journalism", "Reuters"),
    "bloomberg.com": ("financial_journalism", "Bloomberg"),
    "ft.com": ("financial_journalism", "FT — Financial Times"),
    "ii.co.uk": ("retail_commentary", "interactive investor"),
}

# A genuinely distinctive, well-documented RNS announcement title format
# ("REG - RNS - Final Announcement Released", "RNS Number : 1234A") —
# confirmed via real captured examples, not assumed. Matching on this
# specific, narrow pattern (not just the word "RNS" anywhere, which could
# appear in unrelated commentary) keeps this a genuine detection, not a
# guess dressed up as one.
_RNS_TITLE_RE = re.compile(r"^\s*REG\s*[-–]\s*RNS\s*[-–]", re.IGNORECASE)


def classify_source_type(fetch_source, link=None, title=None):
    """
    Returns (source_type, source_label). source_type is one of:
    "company_announcement", "broker_data", "financial_journalism",
    "retail_commentary", "aggregated_unknown".

    Structured, non-text-search sources are known with certainty from HOW
    they were fetched (fetch_source itself, not the article content) —
    the broker-ratings feed and Yahoo's structured analyst-history API are
    never reached via a keyword search, so there's nothing to guess.
    Everything else is judged from the item's own real link (its actual
    publishing domain, via SOURCE_TYPE_DOMAIN_MAP) or a narrowly-matched
    RNS title pattern — never from which search found it, since a Google
    News search for "Barclays" can legitimately surface a Reuters
    article, a company statement, or a retail blog, and only the
    resulting item's OWN link says which.
    """
    if fetch_source in ("ratings", "analyst"):
        return "broker_data", "Broker Data"
    if fetch_source == "ft":
        return "financial_journalism", "FT — Financial Times"

    hostname = ""
    if link:
        try:
            hostname = (urllib.parse.urlparse(link).hostname or "").replace("www.", "")
        except Exception:
            hostname = ""
    for domain, (source_type, label) in SOURCE_TYPE_DOMAIN_MAP.items():
        if domain in hostname:
            return source_type, label

    if title and _RNS_TITLE_RE.match(title):
        return "company_announcement", "RNS — Company Announcement"

    return "aggregated_unknown", "Unknown/Aggregated"


# Priority order for resolving near-duplicate headlines from DIFFERENT
# sources (different links, same underlying story) — reuses the SAME
# authoritativeness ordering classify_source_type's own docstring
# already documents, rather than inventing a separate ranking.
SOURCE_TYPE_PRIORITY = {
    "company_announcement": 0,
    "broker_data": 1,
    "financial_journalism": 2,
    "retail_commentary": 3,
    "aggregated_unknown": 4,
}

NEAR_DUPLICATE_HEADLINE_THRESHOLD = 0.85  # deliberately high (85%+ similarity) -
# conservative on purpose: a missed near-duplicate is a minor cosmetic
# issue, but a FALSE positive would silently hide a genuinely different
# story just because it shares some wording, which is a real information
# loss. Never applied across different tickers or different days (see
# dedupe_near_duplicate_headlines) - only ever compares headlines already
# confirmed to be about the SAME stock, on the SAME day.


def _normalize_headline_for_comparison(title):
    """Lowercase, strip punctuation/extra whitespace - so 'Barclays Q3
    profit beats forecasts' and 'Barclays: Q3 profit beats forecasts.'
    compare as near-identical rather than differing only in punctuation
    that carries no real meaning."""
    return re.sub(r"[^\w\s]", "", (title or "").lower()).split()


def dedupe_near_duplicate_headlines(items):
    """
    Removes near-duplicate STORIES (not just exact-link duplicates,
    which the existing seen_links check already catches) — the same
    underlying event covered independently by two different outlets
    with two different URLs and slightly different headline wording.

    Only ever compares items already known to be about the SAME ticker
    on the SAME day (callers pass in a single ticker's already-same-day-
    filtered list) - never cross-ticker, never cross-day. When two
    items are near-duplicates (headline similarity >= threshold), keeps
    the one from the more authoritative source (via SOURCE_TYPE_PRIORITY,
    the same ordering classify_source_type already documents), and among
    equal-priority sources, keeps whichever was published first.

    Never removes a headline that's merely SHORT or that shares a few
    common words - the threshold is deliberately high specifically to
    avoid ever discarding a genuinely different story.
    """
    if len(items) <= 1:
        return items

    kept = []
    kept_normalized = []
    # Process in priority order first (most authoritative source wins
    # when a near-duplicate is found), then by earliest publish time
    # within equal priority.
    ordered = sorted(
        items,
        key=lambda it: (SOURCE_TYPE_PRIORITY.get(it.get("sourceType"), 4),
                         _parse_pub_date(it.get("pubDate")) or datetime.min.replace(tzinfo=timezone.utc)),
    )
    for it in ordered:
        norm = _normalize_headline_for_comparison(it.get("title", ""))
        is_near_duplicate = False
        for existing_norm in kept_normalized:
            ratio = difflib.SequenceMatcher(None, norm, existing_norm).ratio()
            if ratio >= NEAR_DUPLICATE_HEADLINE_THRESHOLD:
                is_near_duplicate = True
                break
        if not is_near_duplicate:
            kept.append(it)
            kept_normalized.append(norm)
    return kept


def detect_contradictions(
    change_pct, change_pct_5d, above_ma20, rsi14, volume_ratio,
    evidence, broker_momentum, upside_pct, ma200,
):
    """
    Deeper contradiction detection — checks a FIXED set of specific
    signal pairs for disagreement (the exact pairs given in the brief),
    each returned as {"type", "conflict", "supportingA", "supportingB",
    "missingData"} — what conflicts, what supports each side, and what
    data would help resolve it further. Distinct from classify_evidence's
    single overall label (still used here for the price/catalyst check):
    a stock can have more than one kind of tension at once, and this
    surfaces all of them rather than collapsing to one verdict. Never
    resolves a contradiction into an answer — only describes it, so the
    person can weigh it themselves.
    """
    contradictions = []
    chg = change_pct or 0
    significant_move = abs(chg) >= BIG_MOVER_THRESHOLD_PCT

    # 1. Price direction vs a same-day directional broker catalyst
    if evidence["label"] == "conflicting":
        direction_word = "rising" if chg >= 0 else "falling"
        catalyst_word = evidence["catalystDirection"] or "?"
        contradictions.append({
            "type": "price_vs_catalyst",
            "conflict": f"Price is {direction_word} ({chg:+.1f}%) but the same-day directional broker action was {catalyst_word}",
            "supportingA": [f"Price move: {chg:+.1f}%"],
            "supportingB": [f"Directional broker catalyst: {catalyst_word}"],
            "missingData": [],
        })

    # 2. Large move + weak/unconfirmed volume
    if significant_move and volume_ratio is not None and volume_ratio < 0.8:
        contradictions.append({
            "type": "move_without_volume",
            "conflict": f"A {abs(chg):.1f}% move is not backed by above-average volume",
            "supportingA": [f"Price move: {chg:+.1f}%"],
            "supportingB": [f"Volume only {volume_ratio:.1f}× average"],
            "missingData": [],
        })
    elif significant_move and volume_ratio is None:
        contradictions.append({
            "type": "move_without_volume",
            "conflict": f"A {abs(chg):.1f}% move — volume confirmation cannot be checked",
            "supportingA": [f"Price move: {chg:+.1f}%"],
            "supportingB": [],
            "missingData": ["average volume data unavailable"],
        })

    # 3. Strong rise + already-overextended RSI
    if significant_move and chg > 0 and rsi14 is not None and rsi14 >= OVEREXTENDED_RSI_THRESHOLD:
        contradictions.append({
            "type": "move_vs_overextension",
            "conflict": f"A strong rise (+{chg:.1f}%) alongside an already-overextended RSI ({rsi14:.0f})",
            "supportingA": [f"Price move: +{chg:.1f}%"],
            "supportingB": [f"RSI {rsi14:.0f} — overbought territory"],
            "missingData": [],
        })

    # 4. Positive target upside + deteriorating broker momentum
    if upside_pct is not None and upside_pct > 0 and broker_momentum and broker_momentum["direction"] == "deteriorating":
        contradictions.append({
            "type": "upside_vs_broker_deterioration",
            "conflict": f"Consensus target implies +{upside_pct:.1f}% upside, but broker momentum is deteriorating",
            "supportingA": [f"Target upside: +{upside_pct:.1f}%"],
            "supportingB": [f"Broker momentum: deteriorating ({broker_momentum['eventCount']} action(s))"],
            "missingData": [],
        })

    # 5. Improving broker momentum + weakening price trend
    if broker_momentum and broker_momentum["direction"] == "improving" and change_pct_5d is not None and change_pct_5d < 0:
        contradictions.append({
            "type": "broker_improving_vs_price_weak",
            "conflict": f"Broker momentum is improving, but the 5-day price trend is negative ({change_pct_5d:.1f}%)",
            "supportingA": [f"Broker momentum: improving ({broker_momentum['eventCount']} action(s))"],
            "supportingB": [f"5-day trend: {change_pct_5d:.1f}%"],
            "missingData": [],
        })

    # 6. Strong 5-day move + still below a longer-term moving average.
    # Uses MA20 (above_ma20) as the available longer-term-relative check;
    # when 200-day MA data isn't available (needs a full year of
    # history), that's disclosed explicitly rather than silently
    # substituted for without comment.
    if change_pct_5d is not None and change_pct_5d >= UPTREND_5DAY_THRESHOLD_PCT and above_ma20 is False:
        contradictions.append({
            "type": "short_term_strength_vs_long_term_weakness",
            "conflict": f"Strong 5-day gain (+{change_pct_5d:.1f}%) while still below the 20-day MA",
            "supportingA": [f"5-day trend: +{change_pct_5d:.1f}%"],
            "supportingB": ["Price below 20-day MA"],
            "missingData": [] if ma200 is not None else ["insufficient history for 200-day MA context"],
        })

    return contradictions


def compute_entry_exit_evidence(
    change_pct, change_pct_5d, above_ma20, ma_crossover, rsi14,
    volume_ratio, upside_pct, evidence, broker_momentum, latest_broker_event,
    ftse_relative, breakout_status=None,
):
    """
    Structures ALREADY-COMPUTED signals into the ENTRY EVIDENCE / EXIT-RISK
    EVIDENCE framework — this introduces no new signal of its own; every
    line here comes from a calculation or classification that already
    exists elsewhere (compute_ma20_distance_pct, classify_evidence,
    compute_broker_momentum, etc). Never a buy/sell instruction — purely
    an organized presentation of existing facts/calculations/opinions,
    each tagged with its DATA_QUALITY_TAGS provenance so nothing here is
    presented as more authoritative than it actually is.

    Returns {"entry": {"supporting": [(text, tag), ...], "opposing": [...]},
             "exit": {"supporting": [...], "opposing": [...]},
             "holdWait": [(text, tag), ...]}
    "exit.supporting" = evidence supporting caution/reduction (a risk
    case, matching the brief's terminology exactly); "exit.opposing" =
    evidence AGAINST exiting — i.e. the position may still be sound.
    "holdWait" fires on genuine ambiguity (conflicting evidence, mixed
    signals on both sides) or genuine data gaps — the system is
    deliberately willing to say "insufficient evidence" rather than
    forcing every stock into an entry or exit framing.

    Every rule here fires independently and only when its OWN specific
    condition is met — a missing input (None) simply means that
    particular line doesn't appear, never a fabricated placeholder.
    """
    entry_supporting, entry_opposing = [], []
    exit_supporting, exit_opposing = [], []
    chg = change_pct or 0
    high_volume = volume_ratio is not None and volume_ratio >= HIGH_VOLUME_RATIO_THRESHOLD

    # --- Trend ---
    if change_pct_5d is not None and change_pct_5d > 0:
        entry_supporting.append((f"5-day trend positive (+{change_pct_5d:.1f}%)", "CALCULATED"))
    elif change_pct_5d is not None and change_pct_5d < 0:
        exit_supporting.append((f"5-day trend negative ({change_pct_5d:.1f}%) — weakening trend", "CALCULATED"))
        entry_opposing.append((f"5-day trend negative ({change_pct_5d:.1f}%)", "CALCULATED"))
    if above_ma20 is True:
        entry_supporting.append(("Price above 20-day MA", "CALCULATED"))
    elif above_ma20 is False:
        exit_supporting.append(("Price below 20-day MA — weakening trend", "CALCULATED"))
        entry_opposing.append(("Price below 20-day MA", "CALCULATED"))

    # --- Momentum ---
    if ma_crossover == "bullish":
        entry_supporting.append(("20-day MA above 50-day MA (bullish crossover)", "CALCULATED"))
    elif ma_crossover == "bearish":
        exit_supporting.append(("20-day MA below 50-day MA (bearish crossover)", "CALCULATED"))
        entry_opposing.append(("20-day MA below 50-day MA (bearish crossover)", "CALCULATED"))
    if rsi14 is not None and rsi14 >= OVEREXTENDED_RSI_THRESHOLD:
        entry_opposing.append((f"RSI {rsi14:.0f} — overextended / excessive move risk", "CALCULATED"))
        exit_supporting.append((f"RSI {rsi14:.0f} — overextended", "CALCULATED"))
    elif rsi14 is not None and rsi14 <= 30:
        exit_opposing.append((f"RSI {rsi14:.0f} — oversold, downside momentum may already be exhausted", "CALCULATED"))

    # --- Volume ---
    if high_volume and chg > 0:
        entry_supporting.append((f"Volume {volume_ratio:.1f}× average confirms the move", "CALCULATED"))
    elif high_volume and chg < 0:
        exit_supporting.append((f"Volume {volume_ratio:.1f}× average on a decline — weakness confirmed by volume", "CALCULATED"))

    # --- News / catalyst (contradiction signals live here) ---
    if evidence["label"] == "supported":
        entry_supporting.append(("Price move agrees with a same-day directional broker action", "SYSTEM_INTERPRETATION"))
    elif evidence["label"] == "conflicting":
        entry_opposing.append(("Price move CONFLICTS with a same-day directional broker action", "SYSTEM_INTERPRETATION"))
        exit_supporting.append(("Conflicting evidence between price move and broker action", "SYSTEM_INTERPRETATION"))
    elif evidence["label"] == "unexplained_move":
        if chg > 0:
            entry_opposing.append(("No relevant same-day catalyst found for this rise", "SYSTEM_INTERPRETATION"))
        else:
            exit_supporting.append(("No relevant same-day catalyst found for this decline", "SYSTEM_INTERPRETATION"))

    # --- Broker evidence ---
    if latest_broker_event:
        action = latest_broker_event.get("action")
        broker_name = latest_broker_event.get("broker", "?")
        old_r, new_r = latest_broker_event.get("old_rating"), latest_broker_event.get("new_rating")
        if action == "UPGRADE":
            entry_supporting.append((f"Latest broker action: {broker_name} upgraded ({old_r} → {new_r})", "BROKER_OPINION"))
        elif action == "DOWNGRADE":
            exit_supporting.append((f"Latest broker action: {broker_name} downgraded ({old_r} → {new_r})", "BROKER_OPINION"))
    if broker_momentum and broker_momentum["direction"] == "improving":
        entry_supporting.append((f"Broker momentum improving ({broker_momentum['eventCount']} action(s) recently)", "BROKER_OPINION"))
        exit_opposing.append(("Broker momentum improving", "BROKER_OPINION"))
    elif broker_momentum and broker_momentum["direction"] == "deteriorating":
        entry_opposing.append((f"Broker momentum deteriorating ({broker_momentum['eventCount']} action(s) recently)", "BROKER_OPINION"))
        exit_supporting.append(("Broker momentum deteriorating", "BROKER_OPINION"))
    if upside_pct is not None:
        if upside_pct >= 10:
            entry_supporting.append((f"Broker consensus target implies +{upside_pct:.1f}% upside", "BROKER_OPINION"))
        elif upside_pct <= -5:
            exit_supporting.append((f"Price already {abs(upside_pct):.1f}% above broker consensus target", "CALCULATED"))
            entry_opposing.append((f"Price already {abs(upside_pct):.1f}% above broker consensus target — limited stated upside", "CALCULATED"))

    # --- Market context ---
    if ftse_relative is not None and ftse_relative > 0:
        entry_supporting.append((f"Outperforming FTSE 100 by {ftse_relative:.1f}%", "CALCULATED"))
    elif ftse_relative is not None and ftse_relative < 0:
        exit_supporting.append((f"Underperforming FTSE 100 by {abs(ftse_relative):.1f}%", "CALCULATED"))

    # --- Technical confirmation / resistance risk (breakout/breakdown) ---
    if breakout_status == "breakout":
        entry_supporting.append((f"Breakout above the {SUPPORT_RESISTANCE_WINDOW_DAYS}-day high — technical confirmation", "CALCULATED"))
    elif breakout_status == "breakdown":
        exit_supporting.append((f"Breakdown below the {SUPPORT_RESISTANCE_WINDOW_DAYS}-day low", "CALCULATED"))
        entry_opposing.append((f"Breakdown below the {SUPPORT_RESISTANCE_WINDOW_DAYS}-day low — technical/resistance risk", "CALCULATED"))

    # --- HOLD / WAIT evidence — the system must be comfortable saying
    # "insufficient evidence" rather than forcing every stock into an
    # entry or exit framing. Fires on genuine ambiguity or genuine data
    # gaps, never as a default filler.
    hold_wait = []
    if evidence["label"] == "conflicting":
        hold_wait.append(("Conflicting evidence between price move and broker action", "SYSTEM_INTERPRETATION"))
    if evidence["label"] in ("unexplained_move", "no_signal") and evidence["hasCatalyst"] is False:
        hold_wait.append(("No meaningful catalyst identified", "SYSTEM_INTERPRETATION"))
    if entry_supporting and entry_opposing:
        hold_wait.append(("Mixed signals: real evidence exists on both sides", "SYSTEM_INTERPRETATION"))
    missing_inputs = [
        x is None for x in (
            change_pct_5d, rsi14, ma_crossover, volume_ratio, upside_pct,
        )
    ] + [broker_momentum is None or broker_momentum.get("eventCount", 0) == 0]
    if sum(missing_inputs) >= 4:
        hold_wait.append(("Insufficient data available across trend/momentum/volume/broker signals to form a confident view", "SYSTEM_INTERPRETATION"))

    return {
        "entry": {"supporting": entry_supporting, "opposing": entry_opposing},
        "exit": {"supporting": exit_supporting, "opposing": exit_opposing},
        "holdWait": hold_wait,
    }


def render_stock_research_html(
    ticker, name, price, change_pct, currency,
    volume, average_volume, rsi14, ma20, change_pct_5d, above_ma20,
    target, recommendation, market_cap, wk_low, wk_high, sector,
    ftse_change_pct=None, sector_context=None,
    news_items=None, latest_broker_event=None, broker_momentum=None,
    anchor_id=None, css_class="", include_header=True, ma_crossover=None,
    ma50=None, ma200=None, atr14=None, support_resistance=None, breakout_status=None,
    show_stock_intelligence_label=False, progressive_disclosure=False,
    suppress_extended_market_cap=False,
    ai_evidence_confidence=None, ai_evidence_caveat=None,
    scorecard_summary_collector=None,
    price_volume_series=None,
):
    """
    THE single shared rendering function for a stock's full research
    picture — called identically by the Watchlist, Screener, and Moving
    Today sections (see their respective call sites), so all three views
    can never drift into presenting the same underlying data two
    different ways.

    Every section is OMITTED, not fabricated, when its underlying data is
    missing — this function never invents a target, a catalyst, a broker
    action, or a sector context it doesn't have real data for.

    anchor_id: if given (e.g. f"stock-{ticker}"), wraps the whole block in
    a div with that id, making it a link target from elsewhere (Screener
    rows for a stock that's also on the watchlist link to this exact
    anchor — see the "watchlist_tickers" cross-linking in screener_table).

    progressive_disclosure: Phase 5 addition. When True, the output is
    split into an always-visible CORE (price/move, volume×average,
    5-day trend, RSI, MA20 position, target+upside, broker consensus,
    latest broker action, evidence status, key flags, top news headline)
    and an EXTENDED section — MA50/MA200/crossover/ATR/support-resistance/
    breakout, FTSE/sector context, market-cap/52-week/sector fundamentals,
    broker momentum/history, contradictions, the full entry/exit/hold-wait
    evidence panel, and the scorecard — wrapped in a native HTML
    <details>/<summary> element, collapsed by default. No JavaScript, no
    new dependency: <details> is standard HTML with built-in expand/
    collapse behaviour in every modern browser. NOTHING is removed or
    computed differently — every extended-section calculation is
    identical to before, only WHERE it renders changes. When False (the
    Watchlist's existing behaviour — deliberately unchanged, "do not
    downgrade the Watchlist to match the compact Screener"), core and
    extended content are concatenated directly, exactly as before this
    phase.
    """
    news_items = news_items or []
    vol_ratio = compute_volume_ratio(volume, average_volume)
    ma_dist_pct = compute_ma20_distance_pct(price, ma20)
    upside_pct = compute_target_upside_pct(price, target)
    ftse_relative = compute_relative_to_ftse(change_pct, ftse_change_pct)
    evidence = classify_evidence(change_pct, vol_ratio, news_items, latest_broker_event=latest_broker_event)

    currency_suffix = "p" if currency == "GBp" else ""
    chg_cls = "up" if (change_pct or 0) >= 0 else "down"
    chg_arrow = "▲" if (change_pct or 0) >= 0 else "▼"

    header = ""
    if include_header:
        intel_label = '<div style="font-size:10px;color:#7fb3ff;font-weight:700;letter-spacing:0.5px;">🔬 STOCK INTELLIGENCE</div>' if show_stock_intelligence_label else ""
        header = (
            f'{intel_label}<b>{esc(ticker)}</b>{f" — {esc(name)}" if name else ""} '
            f'<span class="{chg_cls}">{price if price is not None else "?"}{currency_suffix} '
            f'{chg_arrow}{abs(change_pct or 0):.2f}%</span>'
        )

    # --- CORE: DON'T CHASE warning (always visible, shown first when present) ---------------------------------------------
    core_lines = []
    dont_chase = compute_dont_chase_warning(change_pct_5d, rsi14, vol_ratio)
    if dont_chase:
        core_lines.append(
            f'<div class="meta" style="color:#f0997b;font-weight:700;">🔥 DON\'T CHASE</div>'
            f'<div class="meta" style="font-size:13px;">{esc(" · ".join(dont_chase["reasons"]))}</div>'
            f'<div class="meta" style="font-size:12px;opacity:0.85;">Reason: an unusually large 5-day move together with at least one sign of the stock being technically extended (elevated RSI and/or well-above-average volume) — a factual observation about current conditions, not a buy/sell instruction.</div>'
        )

    # --- CORE: Volume × average (always visible) ---------------------------------------------
    if volume is not None:
        ratio_str = f' (<span class="val">{vol_ratio:.1f}×</span> average)' if vol_ratio is not None else ""
        core_lines.append(f'<div class="meta">📊 volume: <span class="val">{volume:,}</span>{ratio_str}</div>')

    # --- EXTENDED: FTSE-relative context ---------------------------------------------
    extended_lines = []
    if ftse_relative is not None:
        rel_cls = "up" if ftse_relative >= 0 else "down"
        extended_lines.append(
            f'<div class="meta">🇬🇧 vs FTSE 100: <span class="{rel_cls}">{"+" if ftse_relative >= 0 else ""}{ftse_relative:.1f}%</span></div>'
        )

    # --- CORE: 5-day trend, MA20 position, RSI (always visible) ---------------------------------------------
    trend_parts = []
    if change_pct_5d is not None:
        d5_cls = "up" if change_pct_5d >= 0 else "down"
        trend_parts.append(f'5-day: <span class="{d5_cls}">{"+" if change_pct_5d >= 0 else ""}{change_pct_5d:.1f}%</span>')
    if ma_dist_pct is not None:
        ma_cls = "up" if ma_dist_pct >= 0 else "down"
        trend_parts.append(f'20-day MA: <span class="{ma_cls}">{"+" if ma_dist_pct >= 0 else ""}{ma_dist_pct:.1f}%</span>')
    elif above_ma20 is not None:
        trend_parts.append(f'<span class="val">{"above" if above_ma20 else "below"}</span> 20-day MA')
    if rsi14 is not None:
        trend_parts.append(f'RSI: <span class="val">{rsi14:.1f}</span>')
    if trend_parts:
        core_lines.append(f'<div class="meta">📈 {" · ".join(trend_parts)}</div>')

    # --- EXTENDED: MA50/MA200/crossover/ATR/support-resistance/breakout/sector-context ---------------------------------------------
    tech_extra_parts = []
    if ma50 is not None:
        tech_extra_parts.append(f'MA50: <span class="val">{ma50:.2f}</span>')
    if ma200 is not None:
        tech_extra_parts.append(f'MA200: <span class="val">{ma200:.2f}</span>')
    elif ma50 is not None:
        tech_extra_parts.append('<span style="opacity:0.6;">MA200: insufficient history (needs ~1yr)</span>')
    if ma_crossover:
        tech_extra_parts.append(f'MA20/50: <span class="val">{esc(ma_crossover)}</span>')
    if tech_extra_parts:
        extended_lines.append(f'<div class="meta" style="font-size:11px;">{" · ".join(tech_extra_parts)}</div>')
    if atr14 is not None:
        extended_lines.append(f'<div class="meta" style="font-size:11px;">ATR(14): <span class="val">{atr14:.2f}</span> (typical daily range — volatility context, not a signal)</div>')
    if support_resistance is not None:
        extended_lines.append(
            f'<div class="meta" style="font-size:11px;">{SUPPORT_RESISTANCE_WINDOW_DAYS}-day support/resistance: '
            f'<span class="val">{support_resistance["support"]:.2f}</span> / <span class="val">{support_resistance["resistance"]:.2f}</span></div>'
        )
    if breakout_status and breakout_status != "within_range":
        extended_lines.append(f'<div class="meta" style="font-size:11px;">📐 {esc(breakout_status.capitalize())} vs {SUPPORT_RESISTANCE_WINDOW_DAYS}-day range</div>')

    if sector_context is not None:
        sec_cls = "up" if sector_context["avgChangePct"] >= 0 else "down"
        extended_lines.append(
            f'<div class="meta" style="font-size:11px;">vs {esc(sector or "sector")} '
            f'(n={sector_context["sampleSize"]} stocks tracked, not the full sector): '
            f'<span class="{sec_cls}">{"+" if sector_context["avgChangePct"] >= 0 else ""}{sector_context["avgChangePct"]:.1f}%</span></div>'
        )

    # --- EXTENDED: Fundamentals (52-wk range, sector; mkt cap optionally
    # suppressed here — see suppress_extended_market_cap) ---------------------------------------------
    fundamentals_parts = []
    mcap_str = format_market_cap(market_cap)
    # suppress_extended_market_cap: the Screener already shows market cap
    # in its OWN pre-existing, always-visible P/E/EPS/scale-bar block
    # (built from the exact same canonical `market_cap` value, not a
    # second calculation) — showing it again here, behind the click,
    # was a genuine harmless-but-real duplication caught in the prior
    # audit. Scoped ONLY to that one call site: Watchlist and Moving
    # Today have no such separate block, so they keep showing it here,
    # completely unchanged.
    if mcap_str and not suppress_extended_market_cap:
        fundamentals_parts.append(f'mkt cap <span class="val">{mcap_str}</span>')
    if wk_low is not None and wk_high is not None:
        fundamentals_parts.append(f'52-wk <span class="val">{wk_low:.2f}</span>–<span class="val">{wk_high:.2f}</span>')
    if sector:
        fundamentals_parts.append(f'<span class="val">{esc(sector)}</span>')
    if fundamentals_parts:
        extended_lines.append(f'<div class="meta">{" · ".join(fundamentals_parts)}</div>')

    # --- CORE: top news headline (always visible) ---------------------------------------------
    if news_items:
        top = news_items[0]
        extra_count = f" (+{len(news_items) - 1} more)" if len(news_items) > 1 else ""
        news_type = classify_news_type(top.get("title", ""), top.get("category"))
        news_type_label = NEWS_TYPE_LABELS.get(news_type, "📰")
        core_lines.append(
            f'<div class="meta"><span style="opacity:0.7;font-size:11px;">{esc(news_type_label)}</span><br/>'
            f'<a href="{esc(top.get("link", "#"))}" target="_blank" '
            f'style="color:#7fb3ff;">{esc(top.get("title", ""))}</a>{extra_count}</div>'
        )
    else:
        core_lines.append('<div class="meta">📰 No relevant same-day news found.</div>')

    # --- CORE: broker consensus/target/upside + latest dated action (always visible) ---------------------------------------------
    if target or recommendation:
        upside_html = ""
        if upside_pct is not None:
            up_cls = "up" if upside_pct >= 0 else "down"
            upside_html = f' · distance to target <span class="{up_cls}">{"+" if upside_pct >= 0 else ""}{upside_pct:.1f}%</span>'
        core_lines.append(f'<div class="meta">🎯 consensus: {esc(recommendation or "?")} · target {target if target else "?"}{upside_html}</div>')
    else:
        # Explicit, never a silent omission — no AI-estimated or otherwise
        # fabricated target is ever substituted in when none genuinely exists.
        core_lines.append('<div class="meta">🎯 No current broker target available</div>')
    if latest_broker_event:
        old_r = latest_broker_event.get("old_rating") or "?"
        new_r = latest_broker_event.get("new_rating") or "?"
        old_t, new_t = latest_broker_event.get("old_target"), latest_broker_event.get("new_target")
        action = latest_broker_event.get("action")
        # Explicit action-type label (UPGRADE/DOWNGRADE/TARGET RAISE/TARGET
        # CUT) rather than just showing the before/after numbers — the
        # brief specifically wants "not just Consensus: Buy" but the
        # actual nature of the change made explicit.
        action_label = ""
        if action == "UPGRADE":
            action_label = '<b style="color:#5dcaa5;">UPGRADE</b> '
        elif action == "DOWNGRADE":
            action_label = '<b style="color:#f0997b;">DOWNGRADE</b> '
        elif old_t is not None and new_t is not None:
            try:
                action_label = ('<b style="color:#5dcaa5;">TARGET RAISE</b> ' if float(new_t) > float(old_t)
                                 else '<b style="color:#f0997b;">TARGET CUT</b> ' if float(new_t) < float(old_t) else "")
            except (TypeError, ValueError):
                pass
        target_change = f' · target {old_t} → {new_t}' if (old_t is not None and new_t is not None) else ""
        core_lines.append(
            f'<div class="meta">🏦 {action_label}{esc(latest_broker_event.get("date", ""))} — {esc(latest_broker_event.get("broker", "?"))} — '
            f'{esc(str(old_r))} → {esc(str(new_r))}{esc(target_change)}</div>'
        )

    # --- EXTENDED: broker momentum/history ---------------------------------------------
    if broker_momentum and broker_momentum.get("eventCount", 0) > 0:
        label = BROKER_MOMENTUM_LABEL_TEXT.get(broker_momentum["direction"], broker_momentum["direction"])
        extended_lines.append(
            f'<div class="meta" style="font-size:11px;">Broker momentum: <span class="val">{esc(label)}</span> '
            f'({broker_momentum["eventCount"]} action(s) in the last {BROKER_MOMENTUM_LOOKBACK_DAYS} days)</div>'
        )

    # --- CORE: evidence status + key flags (always visible) ---------------------------------------------
    evidence_label, evidence_reason = EVIDENCE_LABEL_TEXT.get(evidence["label"], (evidence["label"], ""))
    core_lines.append(f'<div class="meta">🧭 Evidence: <span class="val">{esc(evidence_label)}</span> — {esc(evidence_reason)}</div>')
    # AI evidence-quality review — PURELY an additional, visible caveat line
    # underneath the deterministic Evidence: label above, which is NEVER
    # altered by this. Only rendered when a genuine, validated AI review
    # exists (ai_evidence_confidence is None whenever no API key is
    # configured, the call failed, or the response was rejected as
    # malformed/advice-shaped — in every one of those cases this block
    # simply doesn't render, and the page looks exactly as it does without
    # the feature at all).
    if ai_evidence_confidence:
        confidence_display = AI_EVIDENCE_CONFIDENCE_LABELS.get(ai_evidence_confidence, ai_evidence_confidence)
        caveat_text = f" — {esc(ai_evidence_caveat)}" if ai_evidence_caveat else ""
        core_lines.append(
            f'<div class="meta" style="font-size:11px;">🤖 AI evidence check: {esc(confidence_display)}{caveat_text}</div>'
        )
    # Reuses evidence["hasCatalyst"] rather than re-deriving bool(news_items)
    # separately — evidence is already computed just above and, since the
    # classify_evidence fix, correctly accounts for a same-day broker event
    # too, not just news_items. Re-deriving has_news independently here
    # would silently reintroduce the exact inconsistency that was just fixed.
    flags = compute_opportunity_flags(change_pct, vol_ratio, above_ma20, rsi14, evidence["hasCatalyst"])
    if flags:
        flag_badges = " ".join(esc(label) for _fid, label, _reason in flags)
        core_lines.append(f'<div class="meta">🔎 {flag_badges}</div>')

    # --- Block: Entry / Exit Research (structured evidence panel) ---------------------------------------------
    # Deliberately evidence-based, never an instruction — see
    # compute_entry_exit_evidence's docstring. Each bullet is tagged with
    # its DATA_QUALITY_TAGS provenance so a reader can see at a glance
    # whether a line is a calculation, a broker's own opinion, or this
    # system's interpretation of the combination — never presented as if
    # all evidence carries the same weight or certainty.
    ee = compute_entry_exit_evidence(
        change_pct, change_pct_5d, above_ma20, ma_crossover, rsi14,
        vol_ratio, upside_pct, evidence, broker_momentum, latest_broker_event,
        ftse_relative, breakout_status,
    )

    def render_evidence_list(items):
        if not items:
            return '<span class="meta" style="font-size:11px;">None identified</span>'
        return "".join(
            f'<div class="meta" style="font-size:11px;">• {esc(text)} <span style="opacity:0.6;">[{DATA_QUALITY_TAGS.get(tag, tag)}]</span></div>'
            for text, tag in items
        )

    entry_exit_lines = []
    if any([ee["entry"]["supporting"], ee["entry"]["opposing"], ee["exit"]["supporting"], ee["exit"]["opposing"]]):
        entry_exit_lines.append(
            '<div class="meta" style="margin-top:6px;"><b style="font-size:12px;">🔍 ENTRY EVIDENCE</b></div>'
            '<div class="meta" style="font-size:11px;">Supporting further investigation:</div>'
            f'{render_evidence_list(ee["entry"]["supporting"])}'
            '<div class="meta" style="font-size:11px;margin-top:3px;">ENTRY RISKS / caution:</div>'
            f'{render_evidence_list(ee["entry"]["opposing"])}'
            '<div class="meta" style="margin-top:6px;"><b style="font-size:12px;">⚠️ EXIT / RISK EVIDENCE</b></div>'
            '<div class="meta" style="font-size:11px;">Supporting caution or reduction:</div>'
            f'{render_evidence_list(ee["exit"]["supporting"])}'
            '<div class="meta" style="font-size:11px;margin-top:3px;">Against exiting (position may still be sound):</div>'
            f'{render_evidence_list(ee["exit"]["opposing"])}'
        )
    if ee["holdWait"]:
        entry_exit_lines.append(
            '<div class="meta" style="margin-top:6px;"><b style="font-size:12px;">⏸️ HOLD / WAIT EVIDENCE</b></div>'
            f'{render_evidence_list(ee["holdWait"])}'
        )

    # --- Block: Contradictions (deeper, multi-pattern) ---------------------------------------------
    contradictions = detect_contradictions(
        change_pct, change_pct_5d, above_ma20, rsi14, vol_ratio,
        evidence, broker_momentum, upside_pct, ma200,
    )
    contradiction_lines = []
    if contradictions:
        contradiction_lines.append('<div class="meta" style="margin-top:6px;"><b style="font-size:12px;">⚡ CONTRADICTIONS</b></div>')
        for c in contradictions:
            missing_html = f' <span style="opacity:0.6;">(missing: {esc(", ".join(c["missingData"]))})</span>' if c["missingData"] else ""
            contradiction_lines.append(
                f'<div class="meta" style="font-size:11px;">• {esc(c["conflict"])}{missing_html}</div>'
            )

    # --- Block: Transparent Scorecard ---------------------------------------------
    scorecard = compute_research_scorecard(
        change_pct_5d, above_ma20, rsi14, ma_crossover, vol_ratio, change_pct,
        evidence, broker_momentum, breakout_status, ftse_relative, sector_context,
        upside_pct,
    )
    scorecard_lines = ['<div class="meta" style="margin-top:6px;"><b style="font-size:12px;">📋 RESEARCH SCORECARD</b></div>']
    subtotals = compute_scorecard_subtotals(scorecard["dimensions"])
    def _signed(n):
        return f'{"+" if n > 0 else ""}{n}'
    scorecard_lines.append(f'<div class="meta" style="font-size:12px;">Technical/Market Evidence: <span class="val">{_signed(subtotals["technicalMarket"])}</span></div>')
    scorecard_lines.append(f'<div class="meta" style="font-size:12px;">Research Evidence: <span class="val">{_signed(subtotals["researchEvidence"])}</span></div>')
    scorecard_lines.append(f'<div class="meta" style="font-size:12px;">Risk (caution): <span class="val">{_signed(subtotals["risk"])}</span></div>')
    for dim in SCORECARD_DIMENSIONS:
        score, reasons = scorecard["dimensions"][dim]
        sign = "+" if score > 0 else ""
        reason_str = f' — {"; ".join(reasons)}' if reasons else ' — no contributing signal'
        scorecard_lines.append(f'<div class="meta" style="font-size:11px;">{dim}: <span class="val">{sign}{score}</span>{esc(reason_str)}</div>')
    total_sign = "+" if scorecard["total"] > 0 else ""
    signal_quality = compute_signal_quality(scorecard["dimensions"], contradictions)
    SIGNAL_QUALITY_EXPLANATION = {
        "Strong": "scored dimensions agree in direction",
        "Mixed": "scored dimensions point in different directions",
        "Weak": "too few scored dimensions to judge agreement",
    }
    scorecard_lines.append(
        f'<div class="meta">TOTAL: <span class="val">{total_sign}{scorecard["total"]}</span> · '
        f'Confidence: <span class="val">{scorecard["confidence"]}</span> · '
        f'Signal Quality: <span class="val">{signal_quality}</span> '
        f'<span style="opacity:0.6;font-size:10px;">— {esc(SIGNAL_QUALITY_EXPLANATION.get(signal_quality, ""))}</span></div>'
        f'<div class="meta" style="opacity:0.6;font-size:10px;">Neither is a probability or prediction — Confidence reflects data completeness, Signal Quality reflects whether the scored dimensions agree.</div>'
    )
    extended_lines = extended_lines + contradiction_lines + entry_exit_lines + scorecard_lines

    if scorecard_summary_collector is not None:
        # Pure side-effect of the SAME scorecard/signal_quality/evidence
        # computation already performed above for this stock's own detail
        # lines — never a second call to compute_research_scorecard/
        # compute_signal_quality/classify_evidence. Originally built for
        # Strongest Agreeing Evidence (Phase 4); extended here (Phase 6)
        # with price and evidenceLabel, both already computed above, for
        # the daily "What Changed" snapshot — still zero duplicate
        # computation, just two more already-known facts captured.
        # Extended again (Phase 7B) with the full 8-dimension breakdown,
        # both subtotals, and DON'T CHASE state — all already computed
        # above for this stock's own display — so the prospective
        # evidence-history store can capture everything needed for a
        # genuinely honest future backtest of the News/Broker/AI
        # dimensions, without a single new calculation anywhere in this
        # function. Extended again for the Radar Summary: target/
        # upside/recommendation, the top news item, the first entry-
        # supporting (positive) and first contradiction-or-exit-
        # supporting (warning) fact, and DON'T CHASE's own reasons —
        # every single one already computed above in this exact
        # function call, none of it recomputed here.
        _top_news = news_items[0] if news_items else None
        scorecard_summary_collector.append({
            "ticker": ticker, "name": name, "total": scorecard["total"],
            "signalQuality": signal_quality, "confidence": scorecard["confidence"],
            "price": price, "evidenceLabel": evidence["label"],
            "dimensions": {dim: score for dim, (score, _reasons) in scorecard["dimensions"].items()},
            "technicalMarket": subtotals["technicalMarket"], "researchEvidence": subtotals["researchEvidence"],
            "risk": subtotals["risk"],
            "dontChase": dont_chase is not None,
            "dontChaseReasons": dont_chase["reasons"] if dont_chase else [],
            "target": target, "recommendation": recommendation, "upsidePct": upside_pct,
            "topNews": ({"title": _top_news.get("title"), "link": _top_news.get("link"),
                         "pubDate": _top_news.get("pubDate"), "source": _top_news.get("source")}
                        if _top_news else None),
            "mainPositiveEvidence": ee["entry"]["supporting"][0][0] if ee["entry"]["supporting"] else None,
            # Top 3 (not just the first) of the SAME already-computed
            # entry-supporting list — added for the compact Top Radar
            # card, which shows several evidence points at a glance
            # rather than just one. Zero new computation: entry["supporting"]
            # was already fully computed above by compute_entry_exit_evidence.
            "evidenceForTop": [text for text, _tag in ee["entry"]["supporting"][:3]],
            "mainWarning": contradictions[0]["conflict"] if contradictions else (ee["exit"]["supporting"][0][0] if ee["exit"]["supporting"] else None),
            "aiEvidenceConfidence": ai_evidence_confidence, "aiEvidenceCaveat": ai_evidence_caveat,
        })

    # Phase 5 chart — pure display, computed from the ALREADY-RETAINED
    # price_volume_series (see fetch_price_technicals's own docstring),
    # never a new fetch, never touching any calculation above this point.
    chart_svg = render_price_volume_sparkline(price_volume_series, currency_suffix)
    watchlist_chart_html = ""
    if chart_svg:
        if progressive_disclosure:
            # Screener/Moving Today: folds into the SAME shared "Full
            # Stock Intelligence" collapse built below — no new mechanism.
            extended_lines.append(f'<div class="meta" style="margin-top:4px;">{chart_svg}</div>')
        else:
            # Watchlist: its OWN narrow, separate collapse — a deliberate,
            # scoped exception for the chart specifically. Everything
            # else in a Watchlist entry stays exactly as visible as
            # before; this never touches core_lines/extended_lines'
            # existing concatenation below.
            watchlist_chart_html = (
                f'<details style="margin-top:4px;">'
                f'<summary style="cursor:pointer;color:#7fb3ff;font-size:12px;">▸ Show chart</summary>'
                f'<div class="meta" style="margin-top:4px;">{chart_svg}</div></details>'
            )

    if progressive_disclosure and extended_lines:
        # Native <details>/<summary> — zero JavaScript, zero new dependency,
        # standard HTML with built-in expand/collapse in every modern
        # browser. Collapsed by default so a Screener/Moving Today row
        # shows only the CORE line set at a glance; nothing in
        # extended_lines is removed or recalculated differently, only
        # WHERE it renders changes.
        extended_html = "".join(extended_lines)
        body = "".join(core_lines) + (
            f'<details style="margin-top:4px;">'
            f'<summary style="cursor:pointer;color:#7fb3ff;font-size:12px;font-weight:600;">▸ Full Stock Intelligence</summary>'
            f'{extended_html}</details>'
        )
    else:
        body = "".join(core_lines + extended_lines) + watchlist_chart_html

    attrs = []
    if anchor_id:
        attrs.append(f'id="{esc(anchor_id)}"')
    if css_class:
        attrs.append(f'class="{esc(css_class)}"')
    attr_str = (" " + " ".join(attrs)) if attrs else ""
    return f'<div{attr_str}>{header}{body}</div>'


def render_standalone_page(page_filename, title, heading_emoji_title, content_html, docs_dir):
    """
    Wraps a section's already-rendered content HTML (the exact same
    strings used on the main dashboard, not a re-render) in a minimal
    standalone page — same shared CSS, a clear back-to-dashboard link,
    and the same footer. Used for the 11 dedicated per-section pages
    (screener.html, heatmap.html, etc.) generated alongside index.html
    from the SAME poll run's data — never a separate fetch, never a
    separate poll.

    content_html should already contain everything below the page's own
    <h1> — including any "Data source"/"Retrieved"/freshness lines,
    since those already exist as part of the section's own rendered
    HTML on the main dashboard and are reused here unchanged.
    """
    page_html = f"""<!DOCTYPE html>
<html lang="en-GB"><head><meta charset="UTF-8"/><meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>{esc(title)} — UK Stock Watch</title>
{DASHBOARD_CSS}
</head>
<body>
<p style="margin:0 0 10px;"><a href="index.html" style="color:#7fb3ff;text-decoration:none;font-size:15px;">← Back to Dashboard</a></p>
<h1>{heading_emoji_title}</h1>
<main>
{content_html}
</main>
<p style="text-align:center;margin-top:18px;"><a href="index.html" style="color:#7fb3ff;text-decoration:none;font-size:14px;">← Back to Dashboard</a></p>
</body></html>"""
    os.makedirs(docs_dir, exist_ok=True)
    with open(os.path.join(docs_dir, page_filename), "w", encoding="utf-8") as f:
        f.write(page_html)
    return page_html


def render_dashboard(data, watchlist, latest_broker_events=None, events_by_ticker=None, prior_snapshot=None, backtest_results=None, radar_lifecycle=None):
    """
    latest_broker_events: optional dict of ticker -> latest non-superseded broker
    event within LATEST_BROKER_EVENT_MAX_AGE_DAYS (see get_latest_broker_event_per_
    ticker). events_by_ticker: optional dict of ticker -> full list of that
    ticker's events (see group_events_by_ticker), used for compute_broker_momentum
    (a multi-event trend, distinct from the single latest_broker_events entry).
    Both default to loading the events store fresh from disk if not provided —
    the defaults keep main()'s existing call site unchanged, while tests can
    inject specific data directly without touching the filesystem. Loaded from
    state/events.json as it stood at the START of this run (the broker-events
    collection step itself runs AFTER render_dashboard in main(), a pre-existing
    ordering untouched here) — a genuinely new rating change is reflected on the
    NEXT run's dashboard, not the same run that collected it. A roughly 5-minute
    lag on a signal that changes at most a few times a month is a minor, disclosed
    limitation, not a correctness bug.
    """
    if latest_broker_events is None or events_by_ticker is None:
        try:
            _events = load_events_store().get("events", [])
        except Exception:
            _events = []
        if latest_broker_events is None:
            try:
                latest_broker_events = get_latest_broker_event_per_ticker(_events)
            except Exception:
                latest_broker_events = {}  # never let a broker-events read failure break the whole dashboard
        if events_by_ticker is None:
            try:
                events_by_ticker = group_events_by_ticker(_events)
            except Exception:
                events_by_ticker = {}
    items_by_ticker = data.get("items", {})
    quotes = data.get("quotes", {})
    screener = data.get("screener", {})
    ftse100 = data.get("ftse100")
    screener_news = data.get("screenerNews", {})
    screener_news_recent = data.get("screenerNewsRecent", {})
    mover_news_fetch_attempts = data.get("moverNewsFetchAttempts", 0)
    mover_news_fetch_failures = data.get("moverNewsFetchFailures", 0)
    if mover_news_fetch_attempts == 0:
        mover_news_status = "not_checked"
    elif mover_news_fetch_failures >= mover_news_fetch_attempts:
        mover_news_status = "failed"
    else:
        mover_news_status = "ok"
    uptrend_stocks = data.get("uptrendStocks", [])
    big_movers = data.get("bigMovers", [])
    market_wide = data.get("marketWide", [])
    market_research = data.get("marketResearch", {})
    ftse_universe_status = data.get("ftseUniverseStatus", "not_checked")
    ftse_universe_source = data.get("ftseUniverseSource", "not_checked")
    ftse_universe_count = data.get("ftseUniverseCount", 0)
    last_poll_raw = data.get("lastPoll")
    if last_poll_raw:
        try:
            _last_poll_dt = datetime.strptime(last_poll_raw, "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
            last_poll = format_london_and_utc(_last_poll_dt)
            # ISO-8601 with an explicit "Z" — JS's `new Date(...)` parses this
            # unambiguously as UTC everywhere; a naive "YYYY-MM-DD HH:MM:SS" string
            # gets interpreted inconsistently (some engines treat it as local time),
            # which would silently corrupt every staleness calculation below.
            last_poll_iso_z = _last_poll_dt.strftime("%Y-%m-%dT%H:%M:%SZ")
        except Exception:
            last_poll = last_poll_raw  # fall back to raw string if format ever changes
            last_poll_iso_z = ""
    else:
        last_poll = "never"
        last_poll_iso_z = ""

    # Was THIS run's data actually fetched during UK market hours? Reuses the
    # exact same _is_uk_market_hours() already used for pipeline health —
    # not a separate calculation. None (unknown) when last_poll couldn't be
    # parsed; treated as "don't assert either way" rather than guessing.
    market_hours_at_generation = None
    if last_poll_raw:
        try:
            market_hours_at_generation = _is_uk_market_hours(_last_poll_dt)
        except Exception:
            market_hours_at_generation = None

    screener_status = data.get("screenerStatus", {})
    screener_source = data.get("screenerSource", {})
    screener_retrieved_at = data.get("screenerRetrievedAt")
    heatmap_retrieved_at = data.get("heatmapRetrievedAt")
    lse_coverage_report = data.get("lseCoverageReport", {})
    news_explorer = data.get("newsExplorer", {})

    def screener_empty_state_html(section_key):
        """
        Distinguishes the three genuinely different reasons a section can
        show zero rows — previously all three collapsed into the same
        generic "No data yet" text, making a real fetch failure
        indistinguishable from a quiet, fully-working day. Never invents a
        reason: when market-hours context itself couldn't be determined,
        it says so rather than guessing.
        """
        status = screener_status.get(section_key, "not_checked")
        if status == "failed":
            return '<tr><td colspan="3" class="meta" style="color:#f0997b;">⚠️ Fetch failed this run — this section could not be retrieved. Check again next cycle; if it persists, something needs investigating.</td></tr>'
        if status == "not_checked":
            return '<tr><td colspan="3" class="meta">Not checked this run.</td></tr>'
        # status == "ok": the fetch genuinely worked, it just found nothing —
        # note market-closed context when we can actually confirm it, rather
        # than assuming every empty result is because of that.
        if market_hours_at_generation is False:
            return '<tr><td colspan="3" class="meta">⚪ No qualifying movers this run — markets were closed at the time of this update, so this is expected.</td></tr>'
        if market_hours_at_generation is True:
            return '<tr><td colspan="3" class="meta">No qualifying movers found this run (checked successfully, markets were open).</td></tr>'
        return '<tr><td colspan="3" class="meta">No qualifying movers found this run (checked successfully).</td></tr>'

    all_items = []
    for ticker, its in items_by_ticker.items():
        all_items.extend(its)
    all_items.sort(key=item_sort_key, reverse=True)

    recent_items_by_ticker = data.get("recentItems", {})
    recent_market_wide = data.get("recentMarketWide", [])
    market_wide_alerts_status = data.get("marketWideAlertsStatus", "not_checked")
    news_fetch_attempts = data.get("newsFetchAttempts", 0)
    news_fetch_failures = data.get("newsFetchFailures", 0)
    if news_fetch_attempts == 0:
        news_fetch_status = "not_checked"
    elif news_fetch_failures >= news_fetch_attempts:
        news_fetch_status = "failed"
    else:
        news_fetch_status = "ok"  # includes partial failure — a real per-source outage
        # doesn't make every OTHER stock's successfully-fetched news untrustworthy

    def news_empty_state_html(status, recent_pool, label="news", render_fn=None):
        """
        Same four-way distinction as screener_empty_state_html, extended
        with a fifth behaviour specific to news: when the fetch genuinely
        worked but there's nothing from TODAY, fall back to showing the
        most recent AVAILABLE items instead of an empty section — each one
        still rendered through render_fn (defaults to item_div), which
        already shows its real pubDate, so a 2-day-old headline is never
        presented as if it were today's. This fallback pool is completely
        separate from what feeds the evidence/scorecard system (see
        passes_recency_filter_wide's docstring) — showing it here never
        changes what any "same-day catalyst" claim elsewhere on the page
        means.

        render_fn: how to render each fallback item — defaults to item_div
        (used by Market-wide Broker Alerts and the News & Broker Feed,
        where ticker is already embedded in each item's own dict). The
        Top Movers section instead uses screener_news_item(symbol, it),
        which takes the ticker as a SEPARATE argument — passed in here as
        a small wrapper rather than duplicating this whole function.
        """
        render_fn = render_fn or item_div
        if status == "failed":
            return f'<p class="meta" style="color:#f0997b;">⚠️ Fetch failed this run — {label} could not be retrieved. Check again next cycle; if it persists, something needs investigating.</p>'
        if status == "not_checked":
            return f'<p class="meta">Not checked this run.</p>'
        # status == "ok", nothing from today — try the wider recent pool first.
        if recent_pool:
            recent_sorted = sorted(recent_pool, key=item_sort_key, reverse=True)
            recent_html = "".join(render_fn(it) for it in recent_sorted[:20])
            context = ""
            if market_hours_at_generation is False:
                context = " Markets were closed at the time of this update, so no same-day items is expected."
            return (
                f'<p class="meta">No same-day {label} found this run (checked successfully).{context} '
                f'Showing the most recent available instead — each item below is dated, not from today:</p>{recent_html}'
            )
        if market_hours_at_generation is False:
            return f'<p class="meta">⚪ No {label} found this run — markets were closed at the time of this update, so this is expected.</p>'
        if market_hours_at_generation is True:
            return f'<p class="meta">No {label} found this run (checked successfully, markets were open).</p>'
        return f'<p class="meta">No {label} found this run (checked successfully).</p>'

    ftse_html = ""
    if ftse100:
        chg = ftse100.get("changePct") or 0
        cls = "up" if chg >= 0 else "down"
        arrow = "▲" if chg >= 0 else "▼"
        ftse_html = (
            f'<p style="font-size:14px;margin:0 0 10px;">FTSE 100: '
            f'<span class="{cls}">{ftse100.get("price",""):,.2f} {arrow}{abs(chg):.2f}%</span></p>'
        )

    watchlist_name_by_ticker = {stock["ticker"]: stock["name"] for stock in watchlist}
    ftse_change_pct_val = (ftse100.get("changePct") if ftse100 else None)

    # Combined pool of every enriched stock currently known this run (watchlist +
    # screener-ranked) — used ONLY for the deliberately-cautious sector-context
    # approximation (see compute_sector_relative_context's docstring on why this
    # is a small coincidental sample, never presented as the real sector).
    all_enriched_rows = [
        {"symbol_or_ticker": t, "sector": q.get("sector"), "changePct": q.get("changePct")}
        for t, q in quotes.items()
    ] + [
        {"symbol_or_ticker": row.get("symbol"), "sector": row.get("sector"), "changePct": row.get("changePct")}
        for section in ("volume", "gainers", "losers")
        for row in screener.get(section, [])
    ]

    # --- 📡 Radar Stocks -----------------------------------------------
    # A discovery section, NOT a recommendation system: every stock any
    # existing source has surfaced (Watchlist, Heat Map, Screener
    # Volume/Gainers/Losers), merged once per stock, showing exactly the
    # SAME evidence already computed and displayed elsewhere — never a
    # new score, never a verdict, never invented data.
    radar_discovery = discover_radar_stocks(watchlist, big_movers, screener, latest_broker_events)

    # Screener rows carry their own quote-shaped data (keyed by "symbol",
    # not present in the Watchlist-only `quotes` dict) — merged once here
    # so a Screener-only radar stock can still be rendered, reusing the
    # exact same fields Screener's own rendering already reads.
    screener_quotes_by_ticker = {
        row.get("symbol"): row
        for section in ("volume", "gainers", "losers")
        for row in screener.get(section, [])
        if row.get("symbol")
    }

    def radar_quote_for(ticker):
        return quotes.get(ticker) or screener_quotes_by_ticker.get(ticker) or {}

    _SOURCE_ORDER = ("Watchlist", "Heat Map", "LSE Volume", "LSE Gainers", "LSE Losers", "Broker Research")

    def radar_found_via_html(sources):
        by_label = {}
        for label, reason in sources:
            by_label.setdefault(label, []).append(reason)
        ordered_labels = [l for l in _SOURCE_ORDER if l in by_label] + [l for l in by_label if l not in _SOURCE_ORDER]
        found_via = " · ".join(esc(l) for l in ordered_labels)
        reasons = [r for label in ordered_labels for r in by_label[label] if r]
        why_line = f'<div class="meta">Why it appeared: {esc("; ".join(reasons))}</div>' if reasons else ""
        return f'<div class="meta"><b>Found via:</b> {found_via}</div>{why_line}'

    _STATUS_LABELS = {
        "NEW": "🆕 NEW", "ACTIVE": "🟢 ACTIVE", "AGING": "🟡 AGING",
        "STALE": "🟠 STALE", "CLEARED": "⚪ CLEARED",
    }

    def radar_lifecycle_html(entry):
        """
        Renders the exact Found/Last refreshed/Age/Status block the spec
        asks for, plus a MULTI-SOURCE callout when independent sources
        agree — built ENTIRELY from merge_radar_history's own output
        (computed once in main(), passed in read-only), never a second
        timestamp calculation here. Returns an empty string if no
        lifecycle entry exists for this ticker (radar_lifecycle not
        provided, or genuinely absent — never a fabricated timestamp).
        """
        if not entry:
            return ""
        try:
            found_dt = datetime.fromisoformat(entry["firstSeen"])
            refreshed_dt = datetime.fromisoformat(entry["lastSeen"])
        except (ValueError, TypeError, KeyError):
            return ""
        status = entry.get("status", "")
        status_label = _STATUS_LABELS.get(status, esc(status))
        age = entry.get("age")
        age_str = f"Age: {esc(age)}" if age else "Age: unavailable"
        source_count = len(entry.get("sourcesEverSeen", []))
        multi_source = (
            f'<div class="meta" style="color:#7fb3ff;font-weight:600;">MULTI-SOURCE: {source_count} sources</div>'
            if source_count >= 2 else ""
        )
        return (
            f'{multi_source}'
            f'<div class="meta">Found: {esc(format_london_and_utc(found_dt))}</div>'
            f'<div class="meta">Last refreshed: {esc(format_london_and_utc(refreshed_dt))} · {age_str} · '
            f'Status: <b>{status_label}</b></div>'
        )

    def radar_summary_table_row_html(disco, lifecycle_entry, summary, key):
        """
        ONE table row for the "Radar Summary" table — every field pulled
        directly from disco (discovery)/lifecycle_entry (Live Radar
        persistence)/summary (the SAME scorecard_summary_collector entry
        the detailed card below also uses) — zero new computation, zero
        new scoring. A quick-glance, at-a-glance table; the full,
        unmodified detailed card underneath (Full Radar Stocks evidence)
        remains the actual investigation, and Top Radar (a separate,
        deliberately compact card view) is untouched by this change.

        Column order (final, approved design): STOCK, SIGNAL, CONFIDENCE,
        PRICE/TARGET, WHY ON RADAR, EVIDENCE FOR, WARNINGS, NEWS/RESEARCH,
        FRESHNESS. Signal and Confidence are always separate cells, never
        merged. DON'T CHASE and any other warning are always BOTH shown
        together in Warnings — never one hiding the other.
        """
        ticker, name = disco["ticker"], disco["name"]
        price = summary.get("price") if summary else None
        anchor = f'<span id="radar-summary-{esc(key)}"></span>'
        ticker_cell = f'{anchor}<b>{esc(ticker)}</b><br/><span style="opacity:0.75;font-size:13px;">{esc(name)}</span>'

        _FRESHNESS_ICON = {"NEW": "🟢", "ACTIVE": "🟢", "AGING": "🟡", "STALE": "🔴", "CLEARED": "⚪"}
        if lifecycle_entry:
            try:
                status = lifecycle_entry.get("status", "")
                icon = _FRESHNESS_ICON.get(status, "⚪")
                age = lifecycle_entry.get("age")
                age_text = f"{esc(age)} ago" if age else "just now"
                refreshed_dt = datetime.fromisoformat(lifecycle_entry["lastSeen"])
                exact_ts = esc(format_london_and_utc(refreshed_dt))
                freshness_cell = (
                    f'{icon} <b>{esc(status)}</b><br/>'
                    f'<span style="opacity:0.85;font-size:13px;">{age_text}</span><br/>'
                    f'<span style="opacity:0.55;font-size:11px;">{exact_ts}</span>'
                )
                multi_badge = (
                    ' <span class="radar-multi" style="font-size:10px;">MULTI</span>'
                    if len(lifecycle_entry.get("sourcesEverSeen", [])) >= 2 else ""
                )
            except (ValueError, TypeError, KeyError):
                freshness_cell = "Unavailable"
                multi_badge = ""
        else:
            freshness_cell = "Unavailable"
            multi_badge = ""

        by_label = {}
        for label, _reason in disco["sources"]:
            by_label.setdefault(label, True)
        ordered_labels = [l for l in _SOURCE_ORDER if l in by_label] + [l for l in by_label if l not in _SOURCE_ORDER]
        found_via_cell = (
            "".join(f'<span class="source-pill">{esc(l)}</span>' for l in ordered_labels) + multi_badge
            if ordered_labels else '<span style="opacity:0.6;">Not available</span>'
        )

        if summary is None:
            price_cell_early = f'{price:.1f}p' if price is not None else "—"
            return (
                f'<tr><td>{ticker_cell}</td>'
                f'<td colspan="2" style="opacity:0.7;">Evidence unavailable for this stock right now</td>'
                f'<td>{price_cell_early}<br/><span style="opacity:0.7;">No current broker target available</span></td>'
                f'<td>{found_via_cell}</td><td>—</td><td>None flagged</td><td>—</td><td>{freshness_cell}</td></tr>'
            )

        sq, conf = summary["signalQuality"], summary["confidence"]
        _SIGNAL_CLASS = {"Strong": "signal-strong", "Mixed": "signal-mixed", "Weak": "signal-weak"}
        signal_cell = f'<span class="signal {_SIGNAL_CLASS.get(sq, "")}">{esc(sq).upper()}</span>'
        confidence_cell = f'<span class="confidence">{esc(conf).upper()}</span>'

        target, upside, rec = summary.get("target"), summary.get("upsidePct"), summary.get("recommendation")
        price_line = f'<b>{price:.1f}p</b>' if price is not None else '<span style="opacity:0.6;">Price unavailable</span>'
        if target:
            up_cls = "up" if (upside or 0) >= 0 else "down"
            rec_display = esc(rec) if (rec and rec.lower() != "none") else "none stated"
            price_target_cell = (
                f'{price_line}<br/>'
                f'<span style="font-weight:700;">{target:.1f}p</span>'
                + (f' <span class="{up_cls}">({"+" if (upside or 0) >= 0 else ""}{upside:.1f}%)</span>' if upside is not None else '')
                + f'<br/><span style="opacity:0.6;font-size:12px;">broker opinion: {rec_display}</span>'
            )
        else:
            price_target_cell = f'{price_line}<br/><span style="opacity:0.7;font-size:13px;">No current broker target available</span>'

        evidence_points = summary.get("evidenceForTop") or []
        evidence_cell = (
            "<ul class=\"evidence-list\">" + "".join(f"<li>{esc(p)}</li>" for p in evidence_points) + "</ul>"
            if evidence_points else '<span style="opacity:0.6;">None identified</span>'
        )

        # DON'T CHASE and any other warning are BOTH shown, always — one
        # never hides the other.
        warning_bits = []
        if summary.get("dontChase"):
            warning_bits.append(
                '<span class="dont-chase-badge">DON\'T CHASE</span><br/>'
                f'<span style="font-size:12px;opacity:0.85;">{esc("; ".join(summary.get("dontChaseReasons", [])))}</span>'
            )
        if summary.get("mainWarning"):
            warning_bits.append(f'<span style="color:#e8918a;">⚠ {esc(summary["mainWarning"])}</span>')
        warning_cell = "<br/>".join(warning_bits) if warning_bits else "None flagged"

        top_news = summary.get("topNews")
        news_bits = []
        if top_news:
            news_bits.append(
                f'<a href="{esc(top_news.get("link", "#"))}" target="_blank" style="color:#7fb3ff;">{esc(top_news["title"])}</a>'
                f'<br/><span style="opacity:0.6;font-size:12px;">{esc(top_news.get("source") or "?")}'
                f'{" · " + format_news_timestamp(top_news["pubDate"]) if top_news.get("pubDate") else ""}</span>'
            )
        else:
            news_bits.append('<span style="opacity:0.6;">No relevant recent news found</span>')
        ai_conf, ai_caveat = summary.get("aiEvidenceConfidence"), summary.get("aiEvidenceCaveat")
        if ai_conf:
            ai_label = AI_EVIDENCE_CONFIDENCE_LABELS.get(ai_conf, ai_conf)
            news_bits.append(
                f'<div style="margin-top:5px;font-size:12px;opacity:0.85;">🤖 <i>AI interpretation</i> '
                f'(not a source fact): {esc(ai_label)}{" — " + esc(ai_caveat) if ai_caveat else ""}</div>'
            )
        else:
            news_bits.append('<div style="margin-top:5px;font-size:12px;opacity:0.6;">🤖 AI interpretation: not available</div>')
        news_cell = "".join(news_bits)

        return (
            f'<tr>'
            f'<td>{ticker_cell}</td>'
            f'<td>{signal_cell}</td>'
            f'<td>{confidence_cell}</td>'
            f'<td>{price_target_cell}</td>'
            f'<td>{found_via_cell}</td>'
            f'<td>{evidence_cell}</td>'
            f'<td>{warning_cell}</td>'
            f'<td>{news_cell}</td>'
            f'<td>{freshness_cell}</td>'
            f'</tr>'
        )

    def radar_top_card_html(disco, lifecycle_entry, summary, key):
        """
        The GENUINELY compact Top Radar card — deliberately NOT a reuse
        of radar_summary_card_html's full grid layout. This is the
        fastest-possible scan: ticker/price/freshness/signal on one
        line, 3 short labeled blocks (Why On Radar / Evidence For /
        Warning), Target and News each one line, then a link down to
        the SAME stock's full card in Radar Summary (never duplicating
        that detailed content here). Every value still comes from the
        same summary/disco/lifecycle_entry inputs — zero new
        computation, zero new scoring, nothing here that isn't already
        computed and shown, in full, further down the page.
        """
        ticker, name = disco["ticker"], disco["name"]
        price = summary.get("price") if summary else None

        _FRESHNESS_ICON = {"NEW": "🟢", "ACTIVE": "🟢", "AGING": "🟡", "STALE": "🔴", "CLEARED": "⚪"}
        exact_ts_bit = ""
        if lifecycle_entry:
            status = lifecycle_entry.get("status", "")
            icon = _FRESHNESS_ICON.get(status, "⚪")
            age = lifecycle_entry.get("age")
            freshness_bit = f'{icon} {esc(status)} · {esc(age)} ago' if age else f'{icon} {esc(status)}'
            try:
                refreshed_dt = datetime.fromisoformat(lifecycle_entry["lastSeen"])
                exact_ts_bit = (
                    f'<div class="meta" style="font-size:13px;opacity:0.75;">'
                    f'{esc(format_london_and_utc(refreshed_dt))}</div>'
                )
            except (ValueError, TypeError, KeyError):
                pass
        else:
            freshness_bit = "Freshness unavailable"

        if summary is None:
            return (
                f'<div class="radar-compact"><div class="radar-compact-top">'
                f'<span class="radar-ticker" style="font-size:18px;">{esc(ticker)}</span> '
                f'<span style="color:#9aa0a6;">{esc(name)}</span></div>'
                f'<div class="meta" style="font-size:14px;">{freshness_bit}</div>{exact_ts_bit}'
                f'<div class="meta" style="font-size:14px;opacity:0.7;">Evidence unavailable for this stock right now.</div></div>'
            )

        sq, conf = summary["signalQuality"], summary["confidence"]
        top_line = (
            f'<span class="radar-ticker" style="font-size:18px;">{esc(ticker)}</span> '
            f'<span style="color:#9aa0a6;">{esc(name)}</span>'
        )
        stat_line = (
            f'<div class="meta" style="font-size:14px;">'
            + (f'<b>{price}p</b> · ' if price is not None else '')
            + f'{freshness_bit} · Signal: <b>{esc(sq)}</b> · Confidence: <b>{esc(conf)}</b></div>'
            + exact_ts_bit
        )

        by_label = {}
        for label, _reason in disco["sources"]:
            by_label.setdefault(label, True)
        ordered_labels = [l for l in _SOURCE_ORDER if l in by_label] + [l for l in by_label if l not in _SOURCE_ORDER]
        multi = ' <span class="radar-multi" style="font-size:10px;">MULTI</span>' if len(ordered_labels) >= 2 else ""
        why_line = f'<div class="meta" style="font-size:14px;"><b>Why:</b> {" / ".join(esc(l) for l in ordered_labels)}{multi}</div>'

        evidence_points = summary.get("evidenceForTop") or []
        evidence_line = (
            f'<div class="meta" style="font-size:14px;color:#5dcaa5;">✓ {esc("; ".join(evidence_points))}</div>'
            if evidence_points else
            '<div class="meta" style="font-size:14px;opacity:0.6;">No standout evidence identified</div>'
        )

        warning_bits = []
        if summary.get("mainWarning"):
            warning_bits.append(esc(summary["mainWarning"]))
        if summary.get("dontChase"):
            warning_bits.append("DON'T CHASE: " + esc("; ".join(summary.get("dontChaseReasons", []))))
        warning_line = (
            f'<div class="meta" style="font-size:14px;color:#e8918a;">⚠ {" · ".join(warning_bits)}</div>'
            if warning_bits else ""
        )

        target, upside, rec = summary.get("target"), summary.get("upsidePct"), summary.get("recommendation")
        if target:
            up_cls = "up" if (upside or 0) >= 0 else "down"
            target_line = (
                f'<div class="meta" style="font-size:14px;">🎯 <b>{target}p</b>'
                + (f' <span class="{up_cls}">({"+" if (upside or 0) >= 0 else ""}{upside:.1f}%)</span>' if upside is not None else '')
                + f' — <span style="opacity:0.7;">broker opinion: {esc(rec or "?")}</span></div>'
            )
        else:
            target_line = '<div class="meta" style="font-size:14px;">🎯 No current broker target available</div>'

        top_news = summary.get("topNews")
        news_line = (
            f'<div class="meta" style="font-size:14px;">📰 <a href="{esc(top_news.get("link", "#"))}" '
            f'target="_blank" style="color:#7fb3ff;">{esc(top_news["title"])}</a> '
            f'<span style="opacity:0.6;">({esc(top_news.get("source") or "?")}'
            f'{" · " + format_news_timestamp(top_news["pubDate"]) if top_news.get("pubDate") else ""})</span></div>'
            if top_news else '<div class="meta" style="font-size:14px;">📰 No relevant recent news found</div>'
        )

        return (
            f'<div class="radar-compact">{top_line}{stat_line}{why_line}{evidence_line}{warning_line}{target_line}{news_line}'
            f'<div style="margin-top:6px;"><a href="#radar-summary-{esc(key)}" style="color:#7fb3ff;font-size:14px;'
            f'text-decoration:none;">View full evidence ↓</a></div></div>'
        )

    radar_summary_collector = []  # a THROWAWAY collector, used only to sort this
    # section by existing Signal Quality/Confidence — never merged into the
    # real scorecard_summaries, so no stock is ever double-counted in
    # Strongest Agreeing Evidence, the daily snapshot, or evidence history
    # just because it also appears here.
    radar_entries = []
    radar_summary_cards = []
    radar_top_cards = []
    for key, disco in radar_discovery.items():
        ticker = disco["ticker"]
        q = radar_quote_for(ticker)
        name = disco["name"] or watchlist_name_by_ticker.get(ticker, "") or q.get("name") or ticker
        sector = q.get("sector")
        sector_context = compute_sector_relative_context(ticker, sector, all_enriched_rows)
        momentum = compute_broker_momentum(events_by_ticker.get(ticker, []) if events_by_ticker else [])
        news_items = items_by_ticker.get(ticker) or screener_news.get(ticker) or []
        before_len = len(radar_summary_collector)
        html_block = render_stock_research_html(
            ticker=ticker, name=name, price=q.get("price"), change_pct=q.get("changePct"), currency=q.get("currency"),
            volume=q.get("volume"), average_volume=q.get("averageVolume"),
            rsi14=q.get("rsi14"), ma20=q.get("ma20"), change_pct_5d=q.get("changePct5d"), above_ma20=q.get("aboveMA20"),
            target=q.get("targetMeanPrice"), recommendation=q.get("recommendationKey"),
            market_cap=q.get("marketCap"), wk_low=q.get("fiftyTwoWeekLow"), wk_high=q.get("fiftyTwoWeekHigh"), sector=sector,
            ftse_change_pct=ftse_change_pct_val, sector_context=sector_context,
            news_items=news_items,
            latest_broker_event=latest_broker_events.get(ticker) if latest_broker_events else None,
            broker_momentum=momentum, ma_crossover=q.get("maCrossover"),
            ma50=q.get("ma50"), ma200=q.get("ma200"), atr14=q.get("atr14"),
            support_resistance=q.get("supportResistance"), breakout_status=q.get("breakoutStatus"),
            ai_evidence_confidence=q.get("aiEvidenceConfidence"), ai_evidence_caveat=q.get("aiEvidenceCaveat"),
            include_header=True, show_stock_intelligence_label=True,
            progressive_disclosure=True, suppress_extended_market_cap=True,
            price_volume_series=q.get("priceVolumeSeries"),
            scorecard_summary_collector=radar_summary_collector,
        )
        # A sort key even when the underlying scorecard couldn't be computed
        # (e.g. a screener-only stock with genuinely too little data) — such
        # entries sort last, never dropped from the section entirely.
        summary = radar_summary_collector[before_len] if len(radar_summary_collector) > before_len else None
        _lifecycle_for_sort = radar_lifecycle.get(key) if radar_lifecycle else None
        # Sort priority exactly as specified: Signal Quality -> Confidence ->
        # multi-source discovery -> existing dashboard score -> freshness.
        # Every value reused from what's already computed above; no new
        # ranking system, just an ordering over existing facts.
        quality_rank = {"Strong": 0, "Mixed": 1, "Weak": 2}.get(summary["signalQuality"], 3) if summary else 3
        confidence_rank = {"High": 0, "Medium": 1, "Low": 2}.get(summary["confidence"], 3) if summary else 3
        multi_source_rank = 0 if (_lifecycle_for_sort and len(_lifecycle_for_sort.get("sourcesEverSeen", [])) >= 2) else 1
        score_rank = -abs(summary["total"]) if summary else 0
        _status_rank = {"NEW": 0, "ACTIVE": 0, "AGING": 1, "STALE": 2}.get(
            _lifecycle_for_sort.get("status") if _lifecycle_for_sort else None, 3)
        sort_key = (quality_rank, confidence_rank, multi_source_rank, score_rank, _status_rank)
        found_via_html = radar_found_via_html(disco["sources"])
        lifecycle_html = radar_lifecycle_html(radar_lifecycle.get(key)) if radar_lifecycle else ""
        # Explicit hierarchy separator: WHY IT APPEARED (found_via_html,
        # above) is deliberately distinct from CURRENT EVIDENCE (the
        # reused, unmodified html_block below) — a stock being on Radar
        # for a price move or a screener appearance must never be
        # conflated with the scoring system's own separate evidence
        # about it. The evidence block itself is untouched; only this
        # label is new.
        evidence_divider = '<div class="meta" style="margin-top:4px;font-weight:600;color:#7fb3ff;">Current Evidence:</div>'
        radar_entries.append((sort_key, key, f'{lifecycle_html}{found_via_html}{evidence_divider}{html_block}'))
        _lifecycle_for_this_stock = radar_lifecycle.get(key) if radar_lifecycle else None
        radar_summary_cards.append((sort_key, key, radar_summary_table_row_html(disco, _lifecycle_for_this_stock, summary, key)))
        radar_top_cards.append((sort_key, key, radar_top_card_html(disco, _lifecycle_for_this_stock, summary, key)))

    radar_entries.sort(key=lambda e: (e[0], e[1]))
    radar_stocks_html = "".join(e[2] for e in radar_entries) or '<span class="meta">Nothing on the radar right now.</span>'
    radar_summary_cards.sort(key=lambda e: (e[0], e[1]))
    radar_summary_html = "".join(e[2] for e in radar_summary_cards) or '<tr><td colspan="9" class="meta">Nothing on the radar right now.</td></tr>'

    # 🔥 TOP RADAR — the fastest possible scan, using radar_top_cards'
    # GENUINELY COMPACT, separately-rendered content (radar_top_card_html,
    # not a reuse of the full detailed card) — same underlying
    # disco/lifecycle/summary inputs, same Signal Quality/Confidence
    # ordering, but a deliberately different, shorter rendering. Full
    # detail remains exclusively in Radar Summary below via its own
    # "View full evidence ↓" link on each compact card.
    TOP_RADAR_COUNT = 8
    radar_top_cards.sort(key=lambda e: (e[0], e[1]))
    top_radar_html = (
        "".join(f'<div class="radar-compact-wrap">{card_html}</div>' for _sk, _k, card_html in radar_top_cards[:TOP_RADAR_COUNT])
        or '<span class="meta">Nothing on the radar right now.</span>'
    )

    # "At a Glance" — pure aggregation over radar_discovery (this run's
    # LIVE discoveries only, never a CLEARED/historical entry — matching
    # the same "stale info never shown as current" principle already
    # established elsewhere) and radar_lifecycle (computed once already,
    # never recomputed here). No new scoring — just counting existing,
    # already-computed statuses.
    _radar_status_counts = {"NEW": 0, "ACTIVE": 0, "AGING": 0, "STALE": 0}
    _radar_multi_source_count = 0
    for _key in radar_discovery:
        _entry = radar_lifecycle.get(_key) if radar_lifecycle else None
        if _entry:
            _status = _entry.get("status")
            if _status in _radar_status_counts:
                _radar_status_counts[_status] += 1
            if len(_entry.get("sourcesEverSeen", [])) >= 2:
                _radar_multi_source_count += 1
    _radar_total_count = len(radar_discovery)

    if _radar_total_count == 0:
        at_a_glance_html = '<span class="meta">Nothing on the radar right now.</span>'
    else:
        counts_line = (
            f'<div class="meta"><b>{_radar_total_count}</b> stock(s) detected · '
            f'<b>{_radar_status_counts["NEW"]}</b> NEW · <b>{_radar_status_counts["ACTIVE"]}</b> ACTIVE · '
            f'<b>{_radar_status_counts["AGING"]}</b> AGING · <b>{_radar_status_counts["STALE"]}</b> STALE · '
            f'<b>{_radar_multi_source_count}</b> multi-source</div>'
        )
        # Top 5 by the SAME existing sort (Signal Quality/Confidence-based,
        # already computed above for radar_summary_cards) — a one-line
        # jump-link list, not a new ranking.
        top_lines = []
        for _sort_key, _key, _card_html in radar_summary_cards[:5]:
            _disco = radar_discovery.get(_key, {})
            top_lines.append(
                f'<div class="meta" style="font-size:12px;">'
                f'<a href="#radar-summary-{esc(_key)}" style="color:#7fb3ff;text-decoration:none;">'
                f'<b>{esc(_disco.get("ticker", _key))}</b></a></div>'
            )
        at_a_glance_html = counts_line + (
            f'<div class="meta" style="margin-top:6px;font-size:12px;color:#9aa0a6;">Top current Radar stocks:</div>'
            + "".join(top_lines) if top_lines else ""
        )

    def quote_div(t, q):
        name = watchlist_name_by_ticker.get(t, "")
        sector = q.get("sector")
        sector_context = compute_sector_relative_context(t, sector, all_enriched_rows)
        momentum = compute_broker_momentum(events_by_ticker.get(t, []) if events_by_ticker else [])
        return render_stock_research_html(
            ticker=t, name=name, price=q.get("price"), change_pct=q.get("changePct"), currency=q.get("currency"),
            volume=q.get("volume"), average_volume=q.get("averageVolume"),
            rsi14=q.get("rsi14"), ma20=q.get("ma20"), change_pct_5d=q.get("changePct5d"), above_ma20=q.get("aboveMA20"),
            target=q.get("targetMeanPrice"), recommendation=q.get("recommendationKey"),
            market_cap=q.get("marketCap"), wk_low=q.get("fiftyTwoWeekLow"), wk_high=q.get("fiftyTwoWeekHigh"), sector=sector,
            ftse_change_pct=ftse_change_pct_val, sector_context=sector_context,
            news_items=items_by_ticker.get(t, []),
            latest_broker_event=latest_broker_events.get(t) if latest_broker_events else None,
            broker_momentum=momentum,
            anchor_id=f"stock-{t}", css_class="q", ma_crossover=q.get("maCrossover"),
            ma50=q.get("ma50"), ma200=q.get("ma200"), atr14=q.get("atr14"),
            support_resistance=q.get("supportResistance"), breakout_status=q.get("breakoutStatus"),
            show_stock_intelligence_label=True,
            ai_evidence_confidence=q.get("aiEvidenceConfidence"), ai_evidence_caveat=q.get("aiEvidenceCaveat"),
            scorecard_summary_collector=scorecard_summaries,
            price_volume_series=q.get("priceVolumeSeries"),
        )

    scorecard_summaries = []  # filled as a pure side-effect of the quote_div calls below
    quote_rows = "".join(quote_div(t, q) for t, q in quotes.items())

    # Strongest Agreeing Evidence — pure aggregation over scorecard_summaries,
    # which was populated ABOVE by the exact same render_stock_research_html
    # calls that already built quote_rows. Never a second scorecard/signal-
    # quality computation for any stock.
    strong_positive = sorted(
        (s for s in scorecard_summaries if s["signalQuality"] == "Strong" and s["total"] > 0),
        key=lambda s: -s["total"],
    )
    strong_negative = sorted(
        (s for s in scorecard_summaries if s["signalQuality"] == "Strong" and s["total"] < 0),
        key=lambda s: s["total"],
    )

    def _strongest_evidence_row(s):
        sign = "+" if s["total"] > 0 else ""
        return (
            f'<div class="quote-row"><a href="#stock-{esc(s["ticker"])}" style="color:#e8eaed;text-decoration:none;">'
            f'<b>{esc(s["ticker"])}</b> ({esc(s["name"])})</a> — TOTAL <span class="val">{sign}{s["total"]}</span> '
            f'(Strong · {esc(s["confidence"])} confidence)</div>'
        )

    strongest_positive_html = "".join(_strongest_evidence_row(s) for s in strong_positive) or '<span class="meta">None currently.</span>'
    strongest_negative_html = "".join(_strongest_evidence_row(s) for s in strong_negative) or '<span class="meta">None currently.</span>'

    # What Changed Since Last Snapshot — pure aggregation over
    # scorecard_summaries (the SAME data Strongest Agreeing Evidence just
    # used above) compared against prior_snapshot, which main() already
    # loaded from state/daily_snapshots.json before calling this function.
    # Never a new scorecard/evidence computation.
    if prior_snapshot is None:
        whats_changed_html = '<span class="meta">No prior snapshot available yet; comparison will appear from tomorrow\'s run.</span>'
        whats_changed_intro = "Comparison will appear once a prior day's snapshot exists."
    else:
        changes = compute_whats_changed(scorecard_summaries, prior_snapshot)
        since_label = format_since_label(prior_snapshot)
        whats_changed_intro = (
            f"Changed since {esc(since_label)}. Factual changes only — this does not imply "
            f"that any change is good or bad, or a signal to act."
        )
        if not changes:
            whats_changed_html = f'<span class="meta">No notable changes since {esc(since_label)}.</span>'
        else:
            rows = []
            for c in changes:
                parts = []
                if c["priceChangePct"] is not None:
                    cls = "up" if c["priceChangePct"] >= 0 else "down"
                    sign = "+" if c["priceChangePct"] >= 0 else ""
                    parts.append(f'price <span class="{cls}">{sign}{c["priceChangePct"]:.1f}%</span>')
                if c["totalFrom"] is not None and c["totalTo"] is not None and c["totalFrom"] != c["totalTo"]:
                    parts.append(f'TOTAL <span class="val">{c["totalFrom"]:+d} \u2192 {c["totalTo"]:+d}</span>')
                if c["signalQualityFrom"] is not None:
                    parts.append(f'Signal Quality: <span class="val">{esc(c["signalQualityFrom"])} \u2192 {esc(c["signalQualityTo"])}</span>')
                if c["evidenceLabelFrom"] is not None:
                    parts.append(f'Evidence: <span class="val">{esc(c["evidenceLabelFrom"])} \u2192 {esc(c["evidenceLabelTo"])}</span>')
                rows.append(
                    f'<div class="quote-row"><a href="#stock-{esc(c["ticker"])}" style="color:#e8eaed;text-decoration:none;">'
                    f'<b>{esc(c["ticker"])}</b> ({esc(c["name"])})</a> — {" · ".join(parts)}</div>'
                )
            whats_changed_html = "".join(rows)

    # Phase 7C: renders whatever load_backtest_results() produced,
    # loaded by main() and passed in here — this function never loads
    # the file itself, matching render_dashboard's existing "pure
    # function of its inputs" character elsewhere in this codebase.
    backtest_html = render_backtest_results_html(backtest_results)

    def item_div(it):
        # Defense in depth: only render a broker badge for genuine rating/target items,
        # even if something upstream ever slipped through — a "news" item mentioning a
        # bank's name is not the same as that bank issuing a rating.
        show_broker = it.get("broker") and it.get("category") in ("upgrade", "downgrade", "target", "target_raise", "target_cut", "initiation", "reiteration")
        broker_html = f'<span class="broker">{esc(it["broker"])}</span>' if show_broker else ""
        ticker_label = "LSE" if it.get("ticker") == "MARKET" else esc(it.get("ticker", ""))
        # Shows WHEN this item's classification was computed — separate from pubDate
        # (the article/event's own date), since a backfilled Yahoo history item can
        # be re-classified on a later run than when the actual rating happened.
        normalized_at_display = format_normalized_at(it.get("normalizedAt"))
        classified_html = f'<br/><span class="meta">Classified: {esc(normalized_at_display)}</span>' if normalized_at_display else ""
        # sourceLabel (e.g. "FT — Financial Times", "Reuters", "Broker Data",
        # "Unknown/Aggregated") — falls back to the raw source/hostname for
        # items persisted before this field existed, never blank.
        source_display = it.get("sourceLabel") or it.get("source", "")
        return (
            f'<div class="item"><span class="badge {it.get("category","news")}">{it.get("category","news").upper()}</span> '
            f'<b>{ticker_label}</b> '
            f'{broker_html} '
            f'<span class="meta">{esc(source_display)} · {format_news_timestamp(it.get("pubDate",""))}</span>{classified_html}<br/>'
            f'<a href="{esc(it.get("link","#"))}" target="_blank">{esc(it.get("title",""))}</a></div>'
        )

    item_rows = "".join(item_div(it) for it in all_items[:150])
    market_wide_rows = "".join(item_div(it) for it in market_wide[:60])

    all_recent_items = []
    for ticker, its in recent_items_by_ticker.items():
        all_recent_items.extend(its)
    # Exclude anything that's ALREADY shown in the same-day list (by link) —
    # the recent-fallback pool should only ever add genuinely NEW information
    # (older items not already visible), never duplicate what's already on screen.
    _same_day_links = {it.get("link") for it in all_items}
    all_recent_items = [it for it in all_recent_items if it.get("link") not in _same_day_links]

    _same_day_mw_links = {it.get("link") for it in market_wide}
    recent_market_wide_filtered = [it for it in recent_market_wide if it.get("link") not in _same_day_mw_links]

    def screener_table(rows, show_pct=True, section_key=None):
        def row(i, q):
            vol = q.get("volume")
            vol_str = f"{vol:,}" if isinstance(vol, (int, float)) else "?"
            chg = q.get("changePct") or 0
            chg_cls = "up" if chg >= 0 else "down"
            chg_str = f'{"▲" if chg >= 0 else "▼"}{abs(chg):.2f}%'
            last_col = chg_str if show_pct else vol_str
            last_cls = f' class="{chg_cls}"' if show_pct else ""
            symbol = q.get("symbol", "")
            name = esc(q.get("name") or symbol)

            # Screener-specific fundamentals NOT covered by the shared research
            # function (P/E, EPS, earnings/ex-div dates, dividend, insider %,
            # short interest, business summary) — kept here since they're a
            # discovery/vetting depth specific to this section, not duplicated
            # into the shared function (which would bloat the Watchlist with
            # fields that fit initial screening better than ongoing monitoring).
            earnings_date = format_epoch_date(q.get("nextEarningsDate"))
            ex_div_date = format_epoch_date(q.get("exDividendDate"))
            div_rate = q.get("dividendRate")
            div_yield = q.get("dividendYieldPct")
            calendar_html = ""
            if earnings_date:
                calendar_html += f'<br/><span class="meta">📅 next earnings: <span class="val">{earnings_date}</span></span>'
            if ex_div_date:
                calendar_html += f'<br/><span class="meta">💰 ex-dividend: <span class="val">{ex_div_date}</span></span>'
            if div_rate is not None or div_yield is not None:
                rate_str = f'<span class="val">{div_rate:.2f}</span>/share' if div_rate is not None else ''
                yield_str = f'<span class="val">{div_yield:.2f}%</span> yield' if div_yield is not None else ''
                joined = " · ".join(s for s in (rate_str, yield_str) if s)
                calendar_html += f'<br/><span class="meta">💷 dividend: {joined}</span>'
            pe = q.get("trailingPE")
            eps = q.get("trailingEps")
            fundamentals_parts = []
            if pe is not None:
                fundamentals_parts.append(f'P/E <span class="val">{pe:.1f}</span>')
            if eps is not None:
                fundamentals_parts.append(f'EPS <span class="val">{eps:.2f}</span>')
            if fundamentals_parts:
                calendar_html += f'<br/><span class="meta">📊 {" · ".join(fundamentals_parts)}</span>'
            scales_html = pe_scale(pe) + mktcap_scale(q.get("marketCap")) + eps_scale(eps)
            if scales_html:
                calendar_html += f'<div style="margin-top:4px;">{scales_html}</div>'
            insiders = q.get("heldPercentInsidersPct")
            if insiders is not None:
                calendar_html += f'<br/><span class="meta">🧑‍💼 insider ownership: <span class="val">{insiders:.1f}%</span></span>'
            short_pct = q.get("shortInterestPct")
            if short_pct is not None:
                calendar_html += f'<br/><span class="meta">📉 short interest: <span class="val">{short_pct:.2f}%</span> (FCA)</span>'
            industry = q.get("industry")
            sector = q.get("sector")
            if industry:
                calendar_html += f'<br/><span class="meta">🏷️ <span class="val">{esc(industry)}</span></span>'
            biz_summary = q.get("businessSummary")
            if biz_summary:
                calendar_html += f'<br/><span class="meta" style="display:block;max-width:520px;">ℹ️ {esc(biz_summary)}</span>'

            # Cross-link to the Watchlist's full Stock Intelligence view for
            # this exact stock, when it's also being watched — resolves the
            # SAME underlying stock identity across sections. No target exists
            # yet for a screener-only stock not on the watchlist, so no link
            # is rendered for those — never a dead link.
            symbol_html = esc(symbol)
            if symbol in watchlist_name_by_ticker:
                symbol_html = f'<a href="#stock-{esc(symbol)}" style="color:inherit;text-decoration:underline;">{esc(symbol)}</a>'

            # Everything else (news, target/upside, RSI/MA/MA50/MA200/
            # crossover/ATR/support-resistance/breakout, broker action +
            # momentum, evidence, contradictions, entry/exit/hold-wait,
            # scorecard) now comes from the SAME shared function the
            # Watchlist uses — one source of truth, no duplicate calculation
            # or rendering logic maintained in two places.
            news_for_symbol = screener_news.get(symbol) or []
            sector_context = compute_sector_relative_context(symbol, sector, all_enriched_rows)
            momentum = compute_broker_momentum(events_by_ticker.get(symbol, []) if events_by_ticker else [])
            research_html = render_stock_research_html(
                ticker=symbol, name=q.get("name") or symbol, price=q.get("price"), change_pct=chg, currency=q.get("currency"),
                volume=vol, average_volume=q.get("averageVolume"),
                rsi14=q.get("rsi14"), ma20=q.get("ma20"), change_pct_5d=q.get("changePct5d"), above_ma20=q.get("aboveMA20"),
                target=q.get("targetMeanPrice"), recommendation=q.get("recommendationKey"),
                market_cap=q.get("marketCap"), wk_low=q.get("fiftyTwoWeekLow"), wk_high=q.get("fiftyTwoWeekHigh"), sector=sector,
                ftse_change_pct=ftse_change_pct_val, sector_context=sector_context,
                news_items=news_for_symbol,
                latest_broker_event=latest_broker_events.get(symbol) if latest_broker_events else None,
                broker_momentum=momentum, ma_crossover=q.get("maCrossover"),
                ma50=q.get("ma50"), ma200=q.get("ma200"), atr14=q.get("atr14"),
                support_resistance=q.get("supportResistance"), breakout_status=q.get("breakoutStatus"),
                include_header=False, progressive_disclosure=True, suppress_extended_market_cap=True,
                price_volume_series=q.get("priceVolumeSeries"),
                # Extends this stock's own evidence capture (Strongest Agreeing
                # Evidence / daily snapshot / evidence history) to Screener-
                # discovered stocks too, not just Watchlist — the SAME
                # collector Watchlist already feeds, one shared list either way.
                scorecard_summary_collector=scorecard_summaries,
            )

            return (
                f'<tr><td>{i+1}</td><td><b style="font-size:14px;">{symbol_html}</b><br/>'
                f'<span class="meta">{name}</span>{calendar_html}{research_html}</td><td{last_cls}>{last_col}</td></tr>'
            )
        return "".join(row(i, q) for i, q in enumerate(rows)) or screener_empty_state_html(section_key)

    vol_rows = screener_table(screener.get("volume", []), show_pct=False, section_key="volume")
    gain_rows = screener_table(screener.get("gainers", []), section_key="gainers")
    lose_rows = screener_table(screener.get("losers", []), section_key="losers")

    def screener_news_item(symbol, it):
        broker_html = f'<span class="broker">{esc(it["broker"])}</span>' if it.get("broker") and it.get("category") in ("upgrade", "downgrade", "target", "target_raise", "target_cut", "initiation", "reiteration") else ""
        normalized_at_display = format_normalized_at(it.get("normalizedAt"))
        classified_html = f'<br/><span class="meta">Classified: {esc(normalized_at_display)}</span>' if normalized_at_display else ""
        source_display = it.get("sourceLabel") or it.get("source", "")
        return (
            f'<div class="item"><span class="badge {it.get("category","news")}">{it.get("category","news").upper()}</span> '
            f'<b>{esc(symbol)}</b> {broker_html} '
            f'<span class="meta">{esc(source_display)} · {format_news_timestamp(it.get("pubDate",""))}</span>{classified_html}<br/>'
            f'<a href="{esc(it.get("link","#"))}" target="_blank">{esc(it.get("title",""))}</a></div>'
        )

    screener_news_rows = "".join(
        screener_news_item(symbol, it)
        for symbol, items in screener_news.items()
        for it in items
    )

    all_recent_mover_news = []
    for symbol, items in screener_news_recent.items():
        all_recent_mover_news.extend(items)
    _same_day_mover_links = {it.get("link") for its in screener_news.values() for it in its}
    all_recent_mover_news = [it for it in all_recent_mover_news if it.get("link") not in _same_day_mover_links]

    uptrend_rows = "".join(
        f'<div class="quote-row"><b>{esc(s["symbol"])}</b> ({esc(s["name"])}) — '
        f'<span style="color:#2fbf71">▲{s["changePct5d"]:.1f}%</span> over 5 sessions</div>'
        for s in sorted(uptrend_stocks, key=lambda x: -x["changePct5d"])
    )

    # Every screener-ranked stock (Volume/Gainers/Losers) that has a real broker target
    # price attached — pulled together here as its own scannable list, in addition to
    # already showing inline under each screener row. ALSO includes watchlist stocks'
    # targets (via `quotes`), so this is genuinely every target the tool has, in one
    # place, not just the screener subset — deduped so a stock in both pools (e.g. KOO)
    # only appears once.
    all_screener_rows = (
        screener.get("volume", []) + screener.get("gainers", []) + screener.get("losers", [])
    )
    seen_target_symbols = set()
    target_price_rows = ""
    for q in all_screener_rows:
        symbol = q.get("symbol", "")
        target = q.get("targetMeanPrice")
        if not target or symbol in seen_target_symbols:
            continue
        seen_target_symbols.add(symbol)
        rec = q.get("recommendationKey")
        rec_html = f' · <span class="meta">{esc(rec)}</span>' if rec else ""
        target_price_rows += (
            f'<div class="quote-row"><b>{esc(symbol)}</b> ({esc(q.get("name") or symbol)}) — '
            f'🎯 target {target:.2f}{rec_html}</div>'
        )
    for stock in watchlist:
        ticker, name = stock["ticker"], stock["name"]
        yahoo_sym = yahoo_symbol(ticker)  # watchlist tickers are stored plain (e.g. "LLOY"),
                                           # screener symbols are Yahoo-format (e.g. "LLOY.L") —
                                           # normalize so the seen_target_symbols dedupe actually matches
        if yahoo_sym in seen_target_symbols:
            continue
        q = quotes.get(ticker, {})
        target = q.get("targetMeanPrice")
        if not target:
            continue
        seen_target_symbols.add(yahoo_sym)
        rec = q.get("recommendationKey")
        rec_html = f' · <span class="meta">{esc(rec)}</span>' if rec else ""
        target_price_rows += (
            f'<div class="quote-row"><b>{esc(ticker)}</b> ({esc(name)}) — '
            f'🎯 target {target:.2f}{rec_html}</div>'
        )

    # Upcoming Catalysts — same combined-pool source as Broker Target
    # Prices just above (quotes + screener), computed via the dedicated
    # collect_catalyst_events() so this rendering code stays purely
    # presentational, matching how every other section here separates
    # data-gathering from display.
    # News Explorer — genuine LSE regulatory/company announcements, no
    # Yahoo equivalent exists so an honest "unavailable" state is used
    # rather than a substitute. format_london_and_utc reused for
    # consistent BST/GMT-correct timestamps, same as everywhere else.
    def news_explorer_row(story):
        try:
            dt = datetime.fromisoformat(story["datetime"]) if story.get("datetime") else None
            when = format_london_and_utc(dt.replace(tzinfo=timezone.utc)) if dt else "—"
        except (ValueError, TypeError):
            when = esc(str(story.get("datetime") or "—"))
        price = f'{story["price"]:.2f}' if story.get("price") is not None else "—"
        if story.get("percentChange") is not None:
            pct = story["percentChange"]
            pct_cls = "up" if pct >= 0 else "down"
            pct_html = f'<span class="{pct_cls}">{"+" if pct >= 0 else ""}{pct:.2f}%</span>'
        else:
            pct_html = "—"
        headline_html = (
            f'<a href="{esc(story["url"])}" target="_blank" style="color:#7fb3ff;">{esc(story["headline"])}</a>'
            if story.get("url") else esc(story["headline"])
        )
        return (
            f'<tr><td>{headline_html}</td>'
            f'<td>{esc(story.get("companyName") or story.get("companyCode") or "—")}</td>'
            f'<td>{esc(story.get("source") or "—")}</td>'
            f'<td>{when}</td>'
            f'<td>{esc(story.get("rnsNumber") or "—")}</td>'
            f'<td>{price}</td>'
            f'<td>{pct_html}</td></tr>'
        )

    news_explorer_stories = news_explorer.get("stories", [])
    news_explorer_rows = "".join(news_explorer_row(s) for s in news_explorer_stories)

    catalyst_events = collect_catalyst_events(quotes, screener, watchlist)
    EVENT_TYPE_LABELS = {"earnings": "📊 Earnings", "ex_dividend": "💰 Ex-dividend"}
    catalyst_rows = "".join(
        f'<tr><td><b>{esc(e["ticker"])}</b> ({esc(e["name"])})</td>'
        f'<td>{esc(EVENT_TYPE_LABELS.get(e["eventType"], e["eventType"]))}</td>'
        f'<td>{esc(e["date"])}</td>'
        f'<td>{e["daysUntil"]} day{"s" if e["daysUntil"] != 1 else ""}</td>'
        f'<td>{esc(CATALYST_URGENCY_LABELS.get(e["urgency"], ""))}</td></tr>'
        for e in catalyst_events
    )

    def heatmap_cell(q):
        chg = q.get("changePct") or 0
        raw_symbol = q.get("symbol", "")
        symbol = esc(raw_symbol)
        name = esc(q.get("name") or symbol)
        # Colour by direction, intensity (lightness) by magnitude of the move —
        # bigger swings render darker/more saturated, capped at a 10% move.
        magnitude = min(abs(chg), 10) / 10
        lightness = 55 - (magnitude * 30)  # 55% (small move) down to 25% (big move)
        hue = 142 if chg >= 0 else 4  # green / red
        bg = f"hsl({hue}, 55%, {lightness:.0f}%)"
        # Cell SIZE now reflects market cap too, not just colour by % move — a tiny
        # illiquid stock and a multi-billion-pound company no longer look visually
        # identical. grid-column span (2 tiers: large-cap gets 2 columns, everything
        # else gets 1) rather than a continuous size — deterministic, and safe
        # within the existing fixed-column CSS grid, where a genuinely continuous
        # size would need a different (non-grid) layout entirely.
        mcap = q.get("marketCap") or 0
        span = 2 if mcap >= 5_000_000_000 else 1
        span_style = "grid-column:span 2;" if span == 2 else ""
        # Compact hover tooltip — a bit more than just the name, without
        # turning the Heat Map into a wall of text (that's what clicking
        # through to Stock Intelligence is for). Uses the browser's native
        # title attribute, so no extra JS/CSS machinery needed.
        tooltip_parts = [f"{name} ({symbol})", f"{'+' if chg >= 0 else ''}{chg:.1f}%"]
        vol_ratio_tip = compute_volume_ratio(q.get("volume"), q.get("averageVolume"))
        if vol_ratio_tip is not None:
            tooltip_parts.append(f"{vol_ratio_tip:.1f}x avg volume")
        mcap_str_tip = format_market_cap(mcap) if mcap else None
        if mcap_str_tip:
            tooltip_parts.append(f"mkt cap {mcap_str_tip}")
        if raw_symbol in watchlist_name_by_ticker:
            tooltip_parts.append("click for full research view")
        tooltip = esc(" · ".join(tooltip_parts))
        inner = (
            f'<div class="heat-symbol">{symbol}</div>'
            f'<div class="heat-pct">{"▲" if chg >= 0 else "▼"}{abs(chg):.1f}%</div>'
        )
        # Cross-link to the Watchlist's full research view for this exact stock,
        # when it's also being watched — resolves the same underlying stock
        # identity across sections. No target exists for a stock that's ONLY in
        # the screener/heat map (no separate deep-dive page built yet), so no
        # link is rendered for those — never a dead link.
        if raw_symbol in watchlist_name_by_ticker:
            return (
                f'<a class="heat-cell" href="#stock-{symbol}" '
                f'style="background:{bg};{span_style}display:block;text-decoration:none;color:inherit;" '
                f'title="{tooltip}">{inner}</a>'
            )
        return f'<div class="heat-cell" style="background:{bg};{span_style}" title="{tooltip}">{inner}</div>'

    # Grouped by sector for a scannable, clustered layout — sector groups
    # ordered by their own largest move (same "biggest moves surface
    # fastest" principle as the original flat sort, applied at the group
    # level too), with stocks WITHIN each sector still sorted by
    # magnitude (preserved from the pre-grouping sort order). A stock
    # with no sector data falls under an explicit "Other" heading rather
    # than being silently dropped. grid-column:1/-1 on the sector label
    # forces a clean new row in the existing CSS grid without needing a
    # different layout system. Capped at a modest total (24, a small
    # increase from the prior flat 20) so this stays a quick visual scan,
    # not a wall of every sector fully expanded.
    # Prefer the dedicated, broader LSE heatmap fetch (genuinely more of
    # the FTSE 100, not just the 20 stocks already surfaced as top
    # gainers/losers) — falls back to the narrower gainers+losers pool,
    # unchanged from before, when the dedicated heatmap fetch wasn't
    # available this run. Never silently mixes the two within one pool.
    _heatmap_instruments = data.get("heatmapInstruments")
    if _heatmap_instruments:
        heatmap_pool = list(_heatmap_instruments)
    else:
        heatmap_pool = (screener.get("gainers", []) + screener.get("losers", []))
    heatmap_pool.sort(key=lambda q: abs(q.get("changePct") or 0), reverse=True)
    heatmap_pool = heatmap_pool[:24]

    sector_groups = {}
    for q in heatmap_pool:
        sector_groups.setdefault(q.get("sector") or "Other", []).append(q)
    ordered_sectors = sorted(
        sector_groups.items(),
        key=lambda kv: max(abs(s.get("changePct") or 0) for s in kv[1]),
        reverse=True,
    )
    heatmap_cell_parts = []
    for sector_name, stocks_in_sector in ordered_sectors:
        heatmap_cell_parts.append(
            f'<div style="grid-column:1/-1;font-size:12px;font-weight:700;color:#9aa0a6;margin:8px 0 2px;">{esc(sector_name)}</div>'
        )
        heatmap_cell_parts.extend(heatmap_cell(q) for q in stocks_in_sector)
    heatmap_cells = "".join(heatmap_cell_parts)

    def heatmap_empty_state_html():
        """
        Heat Map draws from the SAME gainers+losers fetch as the Screener —
        same status source, same three-way distinction, not a separate
        calculation. Both failed -> failure message; either genuinely
        working -> the normal "nothing to show" framing (a real move
        showing up in gainers or losers alone is enough to populate this,
        so this branch only fires when both are actually empty).
        """
        gainers_status = screener_status.get("gainers", "not_checked")
        losers_status = screener_status.get("losers", "not_checked")
        if gainers_status == "failed" and losers_status == "failed":
            return '<span class="meta" style="color:#f0997b;">⚠️ Fetch failed this run — the heat map could not be retrieved. Check again next cycle.</span>'
        if gainers_status == "not_checked" and losers_status == "not_checked":
            return '<span class="meta">Not checked this run.</span>'
        if market_hours_at_generation is False:
            return '<span class="meta">⚪ No significant movers this run — markets were closed at the time of this update, so this is expected.</span>'
        if market_hours_at_generation is True:
            return '<span class="meta">No significant movers found this run (checked successfully, markets were open).</span>'
        return '<span class="meta">No significant movers found this run (checked successfully).</span>'

    # Explicit, visible data-source line — Volume/Gainers/Losers are either
    # ALL genuinely from LSE's own first-party endpoint, or ALL from the
    # Yahoo fallback (never a silent per-row mix of the two) — this makes
    # which one happened this run impossible to miss.
    _screener_sources_seen = set(screener_source.values()) if screener_source else set()
    _screener_retrieved_html = (
        f' · Retrieved: {esc(format_london_and_utc(datetime.fromisoformat(screener_retrieved_at)))}'
        if screener_retrieved_at else ""
    )
    if not _screener_sources_seen or _screener_sources_seen == {"not_checked"}:
        screener_source_line = '<p class="status-warn">ℹ️ Data source: not checked this run</p>'
    elif _screener_sources_seen == {"LSE"}:
        screener_source_line = f'<p class="status-ok">✅ Data source: LSE (London Stock Exchange, first-party){_screener_retrieved_html}</p>'
    elif _screener_sources_seen == {"Yahoo (fallback)"}:
        screener_source_line = '<p class="status-warn">⚠️ Data source: Yahoo Finance (fallback — LSE unavailable this run)</p>'
    else:
        screener_source_line = f'<p class="status-warn">⚠️ Data source: mixed this run ({dict(screener_source)})</p>'

    # The universe/coverage line — its MEANING depends entirely on which
    # source is actually active this run, never shown as a stale,
    # unrelated description of a mechanism that isn't even in use. When
    # LSE is the source, this describes LSE's own genuine FTSE 100
    # coverage (real matched/unmatched counts from the coverage check,
    # never assumed) — the OLD "FTSE 100 + FTSE 250 (350)" line
    # describes the Yahoo-fallback's own filtering scope specifically,
    # and is only meaningful/shown when Yahoo is actually what's running
    # this cycle.
    if _screener_sources_seen == {"LSE"}:
        cov = lse_coverage_report or {}
        if cov.get("status") == "ok":
            returned = cov.get("lseInstrumentsReturned", 0)
            matched, unmatched = cov.get("matched", 0), cov.get("unmatched", 0)
            # The right question is "of the rows this widget actually
            # returned, how many are genuine FTSE 100 members?" — NOT
            # "does this widget's row count approach the full ~100
            # constituent universe". risersFallersVolume is a top-movers
            # widget by design (its own name says so) and routinely
            # returns far fewer than 100 rows on a quiet day; that is
            # not itself a problem, so the warning threshold is based on
            # the match rate WITHIN what was returned, never against the
            # full universe size.
            match_rate_ok = returned == 0 or (matched / returned) >= 0.9
            cov_class = "status-ok" if match_rate_ok else "status-warn"
            cov_icon = "✅" if cov_class == "status-ok" else "⚠️"
            universe_status_line = (
                f'<p class="{cov_class}">{cov_icon} Universe: FTSE 100 (LSE first-party) — '
                f'{returned} row(s) returned by this widget (a top-movers list, not a full '
                f'constituent dump), {matched} matched to FTSE 100, {unmatched} unmatched '
                f'(out of {cov.get("ftse100Expected", 0)} total FTSE 100 constituents)</p>'
            )
        else:
            universe_status_line = (
                '<p class="status-warn">ℹ️ Universe: FTSE 100 (LSE first-party) — '
                'coverage not verified this run</p>'
            )
    else:
        # Explicit, visible universe/source status — never let a drop from full FTSE
        # 100+250 coverage down to FTSE-100-only, a stale cache, or fully unrestricted
        # scoring stay invisible on the actual page just because the workflow log said so.
        # Only relevant/shown when Yahoo is genuinely the active fallback source.
        _universe_status_html = {
            "healthy": ('status-ok', f'✅ Universe (Yahoo fallback): FTSE 100 + FTSE 250 ({ftse_universe_count} constituents)'),
            "degraded_ftse100_only": ('status-warn', f'⚠️ Universe (Yahoo fallback): FTSE 100 ONLY — FTSE 250 source unavailable, no fallback data exists yet ({ftse_universe_count} constituents, screener coverage reduced)'),
            "stale_cache": ('status-warn', f'⚠️ Universe (Yahoo fallback): using last known-good cached list (source={esc_safe(ftse_universe_source)}, {ftse_universe_count} constituents) — today\'s refresh failed'),
            "unavailable": ('status-bad', '🛑 Universe (Yahoo fallback): UNAVAILABLE — screener showing the whole unrestricted LSE, no FTSE 100/250 filtering applied this run'),
            "not_checked": ('status-warn', 'ℹ️ Universe: not checked this run (hourly-only cycle)'),
            "unknown": ('status-warn', f'⚠️ Universe (Yahoo fallback): unrecognized status (source={esc_safe(ftse_universe_source)})'),
        }
        _status_class, _status_text = _universe_status_html.get(ftse_universe_status, _universe_status_html["unknown"])
        universe_status_line = f'<p class="{_status_class}">{_status_text}</p>'

    def research_card(stock):
        ticker, name = stock["ticker"], stock["name"]
        entry = market_research.get(ticker)

        # Free, always-on facts layer — same data already gathered for News Feed/
        # Watchlist, just pulled together per-stock. No API key needed for this part.
        quote = quotes.get(ticker, {})
        facts_parts = []
        target = quote.get("targetMeanPrice")
        rec = quote.get("recommendationKey")
        if target:
            facts_parts.append(f'🎯 Broker consensus target: <span class="val">{target}</span>' + (f' · {esc(rec)}' if rec else ''))
        recent_items = (items_by_ticker.get(ticker) or [])[:5]
        headlines_html = ""
        if recent_items:
            headlines_html = '<ul style="margin:6px 0 0;padding-left:18px;font-size:12px;line-height:1.7;">' + "".join(
                f'<li><span class="meta">{format_news_timestamp(it.get("pubDate",""))} — {esc(it.get("source",""))}</span><br/>'
                f'<a href="{esc(it.get("link","#"))}" target="_blank" style="color:#e8eaed;">{esc(it.get("title",""))}</a></li>'
                for it in recent_items
            ) + '</ul>'
        facts_html = ""
        if facts_parts:
            facts_html += f'<p class="meta" style="margin:6px 0 0;">{" · ".join(facts_parts)}</p>'
        facts_html += headlines_html
        if not facts_html:
            facts_html = '<p class="meta" style="margin:6px 0 0;">No recent news/broker data gathered for this stock yet.</p>'

        # Optional AI-written paragraph on top, only if a key is configured and this
        # stock's write-up has actually been generated.
        ai_html = ""
        if entry and entry.get("text"):
            ai_html = (
                f'<p style="margin:10px 0 0;font-size:13px;line-height:1.6;border-top:1px solid #2a2e37;padding-top:8px;">'
                f'🤖 {esc(entry["text"])}</p>'
                f'<span class="meta">AI summary generated {esc(entry.get("generatedAt","?"))} UTC from the facts above — not fresh web research, not a recommendation.</span>'
            )

        return (
            f'<div class="item"><b>{esc(ticker)}</b> <span class="meta">({esc(name)})</span>'
            f'{facts_html}{ai_html}</div>'
        )

    research_rows = "".join(research_card(s) for s in watchlist)

    # "Moving Today" now shows the same full research picture as the
    # Watchlist/Screener — the SAME shared function, SAME source of truth —
    # rather than just a bare ticker and %. Looks up each mover's ticker in
    # `quotes` (the fully-enriched dict at render time) rather than using
    # `big_movers` for anything beyond identifying WHICH tickers qualify —
    # big_movers itself is a snapshot taken at the moment of the initial
    # lightweight quote fetch, before enrichment; quotes is what's current
    # by the time this renders.
    def mover_div(m):
        ticker = m.get("ticker", "")
        q = quotes.get(ticker, m)  # fall back to the snapshot if genuinely absent from quotes
        name = watchlist_name_by_ticker.get(ticker, "")
        sector = q.get("sector")
        sector_context = compute_sector_relative_context(ticker, sector, all_enriched_rows)
        momentum = compute_broker_momentum(events_by_ticker.get(ticker, []) if events_by_ticker else [])
        return render_stock_research_html(
            ticker=ticker, name=name, price=q.get("price"), change_pct=q.get("changePct"), currency=q.get("currency"),
            volume=q.get("volume"), average_volume=q.get("averageVolume"),
            rsi14=q.get("rsi14"), ma20=q.get("ma20"), change_pct_5d=q.get("changePct5d"), above_ma20=q.get("aboveMA20"),
            target=q.get("targetMeanPrice"), recommendation=q.get("recommendationKey"),
            market_cap=q.get("marketCap"), wk_low=q.get("fiftyTwoWeekLow"), wk_high=q.get("fiftyTwoWeekHigh"), sector=sector,
            ftse_change_pct=ftse_change_pct_val, sector_context=sector_context,
            news_items=items_by_ticker.get(ticker, []),
            latest_broker_event=latest_broker_events.get(ticker) if latest_broker_events else None,
            broker_momentum=momentum, ma_crossover=q.get("maCrossover"),
            ma50=q.get("ma50"), ma200=q.get("ma200"), atr14=q.get("atr14"),
            support_resistance=q.get("supportResistance"), breakout_status=q.get("breakoutStatus"),
            css_class="q", progressive_disclosure=True,
            price_volume_series=q.get("priceVolumeSeries"),
        )

    mover_rows = "".join(mover_div(m) for m in big_movers)

    # --- ⚠️ Warnings / Contradictions — pure aggregation over scorecard_summaries
    # (ALREADY populated above by quote_div/screener_table's own render_stock_
    # research_html calls) — never a new computation, never a second pass over
    # any stock's evidence. Only stocks with a genuine warning/contradiction or
    # a firing DON'T CHASE are listed; nothing here is invented.
    _warning_rows = []
    for _s in scorecard_summaries:
        _parts = []
        if _s.get("mainWarning"):
            _parts.append(f'⚠ {esc(_s["mainWarning"])}')
        if _s.get("dontChase"):
            _parts.append(f'⏳ DON\'T CHASE: {esc("; ".join(_s.get("dontChaseReasons", [])))}')
        if _parts:
            _warning_rows.append(
                f'<div class="quote-row"><a href="#radar-summary-{esc(_s["ticker"])}" style="color:#e8eaed;text-decoration:none;">'
                f'<b>{esc(_s["ticker"])}</b></a> — {" · ".join(_parts)}</div>'
            )
    warnings_html = "".join(_warning_rows) or '<span class="meta">No warnings or contradictions flagged right now.</span>'

    # --- ℹ️ Data / Freshness — pure aggregation over status values ALREADY
    # computed above (universe status, news/mover-news fetch status,
    # market-wide alerts status, screener status, last poll) — never a new
    # freshness calculation, just gathering what's already known into one
    # place for a quick scan.
    _freshness_rows = [
        f'<div class="meta">Last successful poll: <b>{esc(str(last_poll))}</b></div>',
        f'<div class="meta">FTSE universe: <b>{esc(ftse_universe_status)}</b> ({esc(ftse_universe_source)}, {ftse_universe_count} constituents)</div>',
        f'<div class="meta">Watchlist news fetch: <b>{esc(news_fetch_status)}</b></div>',
        f'<div class="meta">Mover news fetch: <b>{esc(mover_news_status)}</b></div>',
        f'<div class="meta">Market-wide broker alerts: <b>{esc(market_wide_alerts_status)}</b></div>',
        f'<div class="meta">Screener — Volume: <b>{esc(screener_status.get("volume", "not_checked"))}</b> · '
        f'Gainers: <b>{esc(screener_status.get("gainers", "not_checked"))}</b> · '
        f'Losers: <b>{esc(screener_status.get("losers", "not_checked"))}</b></div>',
    ]
    data_freshness_html = "".join(_freshness_rows)

    html = f"""<!DOCTYPE html>
<html lang="en-GB"><head><meta charset="UTF-8"/><meta name="viewport" content="width=device-width, initial-scale=1"/>
<meta http-equiv="refresh" content="300">
<!-- Reload interval matches the actual server-side poll cadence (~5 min) —
     was 90s, which reloaded far more often than the underlying data could
     genuinely change, and is also a real, documented accessibility concern
     (WCAG 2.2.1: unannounced auto-refreshing content can disrupt someone
     using a screen reader or with limited motor control mid-read). The
     client-side freshness ticker just below already tells the reader
     exactly how stale the page is without needing a reload to do it. -->
<title>UK Stock Watch</title>
{DASHBOARD_CSS}
</head>
<body>
<h1>UK Stock Watch — Live Feed</h1>
<p id="pipeline-status" class="status-neutral" style="text-align:left;font-size:13px;margin:0 0 10px;">🕐 Last successful poll: {esc(str(last_poll))} — checking freshness…</p>
<script>
(function() {{
  var lastPollIso = {json.dumps(last_poll_iso_z)};
  var el = document.getElementById('pipeline-status');
  if (!el) return;
  var GREEN_MIN = {PIPELINE_GREEN_THRESHOLD_MIN};
  var AMBER_MIN = {PIPELINE_AMBER_THRESHOLD_MIN};
  var STALE_OOH_MIN = {PIPELINE_OUT_OF_HOURS_STALE_THRESHOLD_MIN};
  var OPEN_MIN = {UK_MARKET_OPEN_MINUTES};
  var CLOSE_MIN = {UK_MARKET_CLOSE_MINUTES};

  function isUkMarketHours(now) {{
    // Intl with an explicit IANA zone handles the GMT/BST switch automatically,
    // the same way Python's zoneinfo does server-side — never a fixed UTC offset.
    var parts = new Intl.DateTimeFormat('en-GB', {{
      timeZone: 'Europe/London', hour: 'numeric', minute: 'numeric',
      hour12: false, weekday: 'short'
    }}).formatToParts(now);
    var weekday, hour, minute;
    parts.forEach(function(p) {{
      if (p.type === 'weekday') weekday = p.value;
      if (p.type === 'hour') hour = parseInt(p.value, 10);
      if (p.type === 'minute') minute = parseInt(p.value, 10);
    }});
    if (weekday === 'Sat' || weekday === 'Sun') return false;
    var minutesSinceMidnight = hour * 60 + minute;
    return minutesSinceMidnight >= OPEN_MIN && minutesSinceMidnight < CLOSE_MIN;
  }}

  function formatLondonAndUtc(dt) {{
    // Client-side equivalent of the server-side format_london_and_utc() —
    // every OTHER timestamp on this page shows London time first (with
    // the correct BST/GMT label) then UTC in parens; this freshness banner
    // was the one place still showing UTC only. Uses Intl with an explicit
    // Europe/London zone (timeZoneName:'short' gives the correct BST/GMT
    // abbreviation), so the DST switch is handled automatically here too,
    // never a fixed UTC+1 assumption. Seconds are kept (unlike the
    // server-side minute-only format) since this banner is specifically
    // about freshness, where finer precision is more useful, not less.
    var parts = new Intl.DateTimeFormat('en-GB', {{
      timeZone: 'Europe/London', weekday: 'short', day: '2-digit', month: 'short', year: 'numeric',
      hour: '2-digit', minute: '2-digit', second: '2-digit', hour12: false, timeZoneName: 'short'
    }}).formatToParts(dt);
    var get = function(type) {{
      var p = parts.find(function(x) {{ return x.type === type; }});
      return p ? p.value : '';
    }};
    var londonStr = get('weekday') + ' ' + get('day') + ' ' + get('month') + ' ' + get('year') + ', ' +
                    get('hour') + ':' + get('minute') + ':' + get('second') + ' ' + get('timeZoneName');
    var utcStr = dt.toISOString().substr(11, 8) + ' UTC';
    return londonStr + ' (' + utcStr + ')';
  }}

  function update() {{
    if (!lastPollIso) {{
      el.textContent = '⚪ No successful poll has been recorded yet.';
      el.className = 'status-neutral';
      return;
    }}
    var now = new Date();
    var lastPoll = new Date(lastPollIso);
    var ageMinutes = (now - lastPoll) / 60000;
    var marketHours = isUkMarketHours(now);
    var label = formatLondonAndUtc(lastPoll) + ' (' + ageMinutes.toFixed(0) + ' min ago)';
    var cls, text;
    if (marketHours) {{
      if (ageMinutes <= GREEN_MIN) {{ cls = 'status-ok'; text = '✅ Up to date — last poll ' + label; }}
      else if (ageMinutes <= AMBER_MIN) {{ cls = 'status-warn'; text = '⚠️ Delayed — last poll ' + label; }}
      else {{ cls = 'status-bad'; text = '🛑 Overdue — last poll ' + label; }}
    }} else {{
      if (ageMinutes <= STALE_OOH_MIN) {{ cls = 'status-neutral'; text = '⚪ Markets closed — last poll ' + label; }}
      else {{ cls = 'status-warn'; text = '⚠️ Markets closed, but no poll in ' + (ageMinutes / 60).toFixed(1) + 'h — worth checking'; }}
    }}
    el.className = cls;
    el.textContent = text + ' — this page checks its own freshness every 30s.';
  }}
  update();
  setInterval(update, 30000);
}})();
</script>
{ftse_html}
<p class="disclaimer">LSE-listed stocks only. Informational only — not investment advice, not a guarantee of any outcome.</p>

<nav aria-label="Section navigation" style="margin:14px 0;padding:12px;background:#161920;border-radius:6px;font-size:16px;line-height:2.4;position:sticky;top:0;z-index:100;box-shadow:0 2px 8px rgba(0,0,0,0.4);">
<b style="color:#9aa0a6;margin-right:8px;">Quick Navigation:</b>
<a href="radar.html#top-radar" style="color:#7fb3ff;margin-right:14px;text-decoration:none;">🔥 Top Radar</a>
<a href="radar.html#radar-now" style="color:#7fb3ff;margin-right:14px;text-decoration:none;">📡 Radar Now</a>
<a href="radar.html#radar-stocks" style="color:#7fb3ff;margin-right:14px;text-decoration:none;">🛰️ Radar Stocks</a>
<a href="radar.html#radar-summary" style="color:#7fb3ff;margin-right:14px;text-decoration:none;">📋 Radar Summary</a>
<a href="strongest-evidence.html" style="color:#7fb3ff;margin-right:14px;text-decoration:none;">🏆 Strongest Evidence</a>
<a href="screener.html" style="color:#7fb3ff;margin-right:14px;text-decoration:none;">📊 Screener</a>
<a href="heatmap.html" style="color:#7fb3ff;margin-right:14px;text-decoration:none;">🗺️ Heat Map</a>
<a href="gainers.html" style="color:#7fb3ff;margin-right:14px;text-decoration:none;">🟢 Gainers</a>
<a href="losers.html" style="color:#7fb3ff;margin-right:14px;text-decoration:none;">🔴 Losers</a>
<a href="volume.html" style="color:#7fb3ff;margin-right:14px;text-decoration:none;">📊 Volume</a>
<a href="watchlist.html" style="color:#7fb3ff;margin-right:14px;text-decoration:none;">👀 Watchlist</a>
<a href="news-explorer.html" style="color:#7fb3ff;margin-right:14px;text-decoration:none;">📰 News Explorer</a>
<a href="news-feed.html" style="color:#7fb3ff;margin-right:14px;text-decoration:none;">📰 News / Evidence</a>
<a href="catalysts.html" style="color:#7fb3ff;margin-right:14px;text-decoration:none;">📅 Catalysts</a>
<a href="warnings.html" style="color:#7fb3ff;margin-right:14px;text-decoration:none;">⚠️ Warnings</a>
<a href="data-freshness.html" style="color:#7fb3ff;text-decoration:none;white-space:nowrap;">ℹ️ Data&nbsp;/&nbsp;Freshness</a>
<details class="nav-more" style="margin-top:4px;">
<summary style="cursor:pointer;color:#5a6072;font-size:13px;list-style:none;">More navigation ▾</summary>
<div style="margin-top:4px;font-size:13px;line-height:2;">
<a href="more-data.html#mover-news" style="color:#5a80a8;margin-right:12px;text-decoration:none;">Mover News</a>
<a href="more-data.html#uptrend" style="color:#5a80a8;margin-right:12px;text-decoration:none;">5-Day Uptrend</a>
<a href="more-data.html#targets" style="color:#5a80a8;margin-right:12px;text-decoration:none;">Target Prices</a>
<a href="more-data.html#movers-today" style="color:#5a80a8;margin-right:12px;text-decoration:none;">Moving Today</a>
<a href="more-data.html#whats-changed" style="color:#5a80a8;margin-right:12px;text-decoration:none;">What Changed</a>
<a href="more-data.html#market-research" style="color:#5a80a8;margin-right:12px;text-decoration:none;">Market Research</a>
<a href="more-data.html#broker-alerts" style="color:#5a80a8;margin-right:12px;text-decoration:none;">Broker Alerts</a>
<a href="more-data.html#backtest" style="color:#5a80a8;text-decoration:none;">Signal Backtest</a>
</div>
</details>
</nav>

<main>
<div id="radar-now" style="background:#161920;border-radius:6px;padding:12px 14px;margin-bottom:14px;border-left:3px solid #7fb3ff;">
<b style="color:#7fb3ff;">📡 RADAR NOW</b>
{at_a_glance_html}
</div>

<h2 id="top-radar">🔥 Top Radar</h2>
<p class="meta">The strongest current Radar evidence, ranked by the same existing Signal Quality/Confidence ordering used throughout — not a new score, not a recommendation. Click any ticker's evidence below to jump straight to it, or scroll to Radar Summary for every currently detected stock.</p>
{top_radar_html}
<p style="text-align:center;margin:6px 0 18px;"><a href="#radar-now" style="color:#7fb3ff;text-decoration:none;font-size:13px;">↑ Back to top</a></p>

<h2 id="radar-stocks">📡 Radar Stocks</h2>
<p class="meta">Every stock any existing source has surfaced — Watchlist, Heat Map, and LSE Screener — merged once per stock. A discovery list, not a recommendation: this shows the same evidence already computed elsewhere, ranked by existing Signal Quality so the clearest-agreeing evidence appears first, without ranking, scoring, or calling anything a "buy" or "opportunity."</p>
<div class="disclaimer" style="margin-bottom:10px;"><b>Discovery sources (find NEW stocks):</b> Watchlist, Heat Map, LSE Screener (Volume/Gainers/Losers), and market-wide Broker Research (rating/target changes across the whole LSE, not just watchlist stocks). <b>Evidence sources (enrich stocks already discovered, but cannot discover a stock on their own):</b> Google/Yahoo/Reuters/Bloomberg/FT news feeds and the existing AI Evidence Review — these are per-ticker lookups, not a whole-market scan, so they can only add evidence to a stock some other source already surfaced. <b>Not available, shown honestly rather than invented:</b> genuine insider buy/sell transaction data (only static insider ownership % exists, which is never presented as trading activity) and any web source beyond the news feeds listed above.</div>

<h3 id="radar-summary" style="margin-top:14px;">📋 Radar Summary — all current radar stocks at a glance</h3>
<p class="meta">One row per stock — scan Signal, Confidence, target, discovery source, evidence, warnings and freshness in seconds. Technical/Market/Research/Risk sub-scores remain in the full evidence cards below.</p>
<div class="radar-table-wrap">
<table class="radar-table">
<thead><tr>
<th>Stock</th><th>Signal</th><th>Confidence</th><th>Price / Target</th><th>Why On Radar</th>
<th>Evidence For</th><th>Warnings</th><th>News / Research</th><th>Freshness</th>
</tr></thead>
<tbody>
{radar_summary_html}
</tbody>
</table>
</div>

<h3 style="margin-top:18px;">🔬 Full Radar Stocks evidence</h3>
{radar_stocks_html}

<h2 id="heatmap">🗺️ Heat Map (top movers, by size of move)</h2>
{('<p class="status-ok">✅ Data source: LSE (London Stock Exchange, first-party)'
  + (' · Retrieved: ' + esc(format_london_and_utc(datetime.fromisoformat(heatmap_retrieved_at))) if heatmap_retrieved_at else '')
  + '</p>') if _heatmap_instruments else '<p class="status-warn">⚠️ Data source: derived from Screener Gainers/Losers (dedicated LSE heatmap unavailable this run)</p>'}
<div class="heatmap-grid">{heatmap_cells or heatmap_empty_state_html()}</div>

<h2 id="screener">📊 LSE Screener (Volume / Gainers / Losers)</h2>
{universe_status_line}
{screener_source_line}
<div class="screener-grid">
  <div><h3 id="volume">Top Volume</h3><table><tr><th>#</th><th>Symbol</th><th>Volume</th></tr>{vol_rows}</table></div>
  <div><h3 id="gainers">Top Gainers</h3><table><tr><th>#</th><th>Symbol</th><th>Chg%</th></tr>{gain_rows}</table></div>
  <div><h3 id="losers">Top Losers</h3><table><tr><th>#</th><th>Symbol</th><th>Chg%</th></tr>{lose_rows}</table></div>
</div>

<h2 id="mover-news">📰 News on Today's Top Movers</h2>
<p class="meta">Real, dated-today news for any stock currently in Volume/Gainers/Losers or the Heat Map — not limited to your watchlist. Matched by company-name/headline keyword matching and classified by rule-based pattern detection (upgrade/downgrade/target-change etc.) — not AI-generated. See Market Research below for the one section that does use an AI-written summary.</p>
<div>{screener_news_rows or news_empty_state_html(mover_news_status, all_recent_mover_news, "news for today's ranked stocks", render_fn=lambda it: screener_news_item(it.get("ticker", ""), it))}</div>

<h2 id="uptrend">📈 5-Day Uptrend ({UPTREND_5DAY_THRESHOLD_PCT:.0f}%+, screener + watchlist)</h2>
<p class="meta">Real closing-price history over the last 5 trading days — a fact about the past, not a forecast of what happens next.</p>
<div class="quotes">{uptrend_rows or '<span class="meta">Nothing has met the 5-day threshold right now</span>'}</div>

<h2 id="targets">🎯 Broker Target Prices</h2>
<p class="meta">Real, already-published broker consensus targets from Yahoo's aggregation — not generated by this tool. Covers both your watchlist and today's screener-ranked stocks (Volume/Gainers/Losers).</p>
<div class="quotes">{target_price_rows or '<span class="meta">No target price data available yet.</span>'}</div>

<h2 id="catalysts">🗓️ Upcoming Catalysts</h2>
<p class="meta">Real, already-published earnings and ex-dividend dates from Yahoo's calendar data — informational only, not a suggestion to act around any of these dates. Covers both your watchlist and today's screener-ranked stocks.</p>
<table><tr><th>Stock</th><th>Event</th><th>Date</th><th>Days until</th><th>Timing</th></tr>{catalyst_rows or '<tr><td colspan="5" class="meta">No known upcoming earnings or ex-dividend dates right now.</td></tr>'}</table>

<h2 id="movers-today">🔥 Already Moving Today (watchlist, ±{BIG_MOVER_THRESHOLD_PCT:.0f}%+)</h2>
<p class="meta">A fact about what already happened today — not a forecast of what happens next.</p>
<div class="quotes">{mover_rows or '<span class="meta">Nothing past the threshold right now</span>'}</div>

<h2 id="strongest-evidence">🏆 Strongest Agreeing Evidence</h2>
<p class="meta">Stocks where the scored evidence dimensions agree most clearly right now — not a ranking of what to buy or sell, and not a prediction of future performance. See each stock's full Research Scorecard below for the underlying facts.</p>
<h3>Strongest agreeing-positive evidence</h3>
<div class="quotes">{strongest_positive_html}</div>
<h3>Strongest agreeing-negative evidence</h3>
<div class="quotes">{strongest_negative_html}</div>

<h2 id="whats-changed">📅 What Changed Since Last Snapshot</h2>
<p class="meta">{whats_changed_intro}</p>
<div class="quotes">{whats_changed_html}</div>

<h2 id="watchlist">👀 Your Watchlist</h2>
<div class="quotes">{quote_rows or '<span class="meta">No quotes yet</span>'}</div>

<h2 id="market-research">🔎 Market Research</h2>
<p class="meta">Real broker targets, recent news, and consensus ratings, already gathered by this tool — free, always live, no AI or API cost involved. If an ANTHROPIC_API_KEY is configured, a short AI-written summary appears too (🤖), synthesised only from these same facts — never fresh web research, never a recommendation. Always cross-check anything here against primary sources before acting on it.</p>
{research_rows or '<p class="meta">No watchlist stocks to show yet.</p>'}

<h2 id="broker-alerts">⬆⬇🎯 Market-wide Broker Alerts (all LSE, not just watchlist)</h2>
<p class="meta">Upgrades/downgrades from anywhere on the LSE, not limited to your watchlist below.</p>
{market_wide_rows or news_empty_state_html(market_wide_alerts_status, recent_market_wide_filtered, "market-wide alerts")}

<h2 id="news-explorer">📰 News Explorer (LSE regulatory &amp; company announcements)</h2>
<p class="meta">Genuine LSE first-party data — regulatory news and company announcements across the whole market, not limited to your watchlist. No equivalent exists via any other source used on this page, so this section shows an honest "unavailable" state rather than a substitute when the LSE source can't be reached.</p>
{('<p class="status-ok">✅ Data source: London Stock Exchange (first-party) · Retrieved: '
  + esc(format_london_and_utc(datetime.fromisoformat(news_explorer['retrievedAt']))) + ' · '
  + str(news_explorer.get('totalElements', 0)) + ' total result(s), showing '
  + str(len(news_explorer_stories)) + '</p>')
 if news_explorer.get('status') == 'ok' else
 ('<p class="status-warn">⚠️ News Explorer unavailable this run'
  + (' (' + esc(news_explorer['error']) + ')' if news_explorer.get('error') else '') + '</p>')}
{f'<table><tr><th>Headline</th><th>Company</th><th>Source</th><th>Date / Time</th><th>RNS Number</th><th>Price</th><th>Change %</th></tr>{news_explorer_rows}</table>' if news_explorer_stories else '<span class="meta">No stories available this run.</span>'}

<h2 id="news-feed">📰 News &amp; Broker Feed (watchlist)</h2>
{item_rows or news_empty_state_html(news_fetch_status, all_recent_items, "news")}

<h2 id="warnings">⚠️ Warnings / Contradictions</h2>
<p class="meta">Every stock currently on Radar with a flagged contradiction or an active DON'T CHASE warning, gathered here from the same evidence shown on each stock's own card — nothing new calculated.</p>
<div class="quotes">{warnings_html}</div>

<h2 id="data-freshness">ℹ️ Data / Freshness</h2>
<p class="meta">A quick summary of how current each data source is right now, gathered from the same status checks already shown throughout this page.</p>
{data_freshness_html}

<h2 id="backtest">📊 Signal Backtest (Technical Signals Only)</h2>
{backtest_html}

<p class="lastpoll">Last checked: {esc(str(last_poll))}</p>
<p style="text-align:center;margin-top:10px;"><a href="#radar-now" style="color:#7fb3ff;text-decoration:none;font-size:14px;">↑ Back to top</a></p>
</main>
</body></html>"""
    os.makedirs(DOCS_DIR, exist_ok=True)
    with open(os.path.join(DOCS_DIR, DOCS_FILENAME), "w", encoding="utf-8") as f:
        f.write(html)

    # Dedicated standalone pages — built from the EXACT SAME already-computed
    # section variables used just above for index.html, within this same
    # function call. No re-fetch, no second poll, no separate data path:
    # whatever index.html shows for a section is byte-identical to what its
    # dedicated page shows, because both come from the same strings.
    render_standalone_page("radar.html", "Radar", "📡 Radar", f"""
<h2 id="radar-now">📡 Radar Now</h2>
<div style="background:#161920;border-radius:6px;padding:12px 14px;margin-bottom:14px;border-left:3px solid #7fb3ff;">
{at_a_glance_html}
</div>

<h2 id="top-radar">🔥 Top Radar</h2>
<p class="meta">The strongest current Radar evidence, ranked by the same existing Signal Quality/Confidence ordering used throughout — not a new score, not a recommendation. Click any ticker's evidence below to jump straight to it, or scroll to Radar Summary for every currently detected stock.</p>
{top_radar_html}

<h2 id="radar-stocks">📡 Radar Stocks</h2>
<p class="meta">Every stock any existing source has surfaced — Watchlist, Heat Map, and LSE Screener — merged once per stock. A discovery list, not a recommendation: this shows the same evidence already computed elsewhere, ranked by existing Signal Quality so the clearest-agreeing evidence appears first, without ranking, scoring, or calling anything a "buy" or "opportunity."</p>
<div class="disclaimer" style="margin-bottom:10px;"><b>Discovery sources (find NEW stocks):</b> Watchlist, Heat Map, LSE Screener (Volume/Gainers/Losers), and market-wide Broker Research (rating/target changes across the whole LSE, not just watchlist stocks). <b>Evidence sources (enrich stocks already discovered, but cannot discover a stock on their own):</b> Google/Yahoo/Reuters/Bloomberg/FT news feeds and the existing AI Evidence Review — these are per-ticker lookups, not a whole-market scan, so they can only add evidence to a stock some other source already surfaced. <b>Not available, shown honestly rather than invented:</b> genuine insider buy/sell transaction data (only static insider ownership % exists, which is never presented as trading activity) and any web source beyond the news feeds listed above.</div>

<h3 id="radar-summary" style="margin-top:14px;">📋 Radar Summary — all current radar stocks at a glance</h3>
<p class="meta">One row per stock — scan Signal, Confidence, target, discovery source, evidence, warnings and freshness in seconds. Technical/Market/Research/Risk sub-scores remain in the full evidence cards below.</p>
<div class="radar-table-wrap">
<table class="radar-table">
<thead><tr>
<th>Stock</th><th>Signal</th><th>Confidence</th><th>Price / Target</th><th>Why On Radar</th>
<th>Evidence For</th><th>Warnings</th><th>News / Research</th><th>Freshness</th>
</tr></thead>
<tbody>
{radar_summary_html}
</tbody>
</table>
</div>

<h3 style="margin-top:18px;">🔬 Full Radar Stocks evidence</h3>
{radar_stocks_html}
""", DOCS_DIR)

    render_standalone_page("strongest-evidence.html", "Strongest Evidence", "🏆 Strongest Agreeing Evidence", f"""
<p class="meta">Stocks where the scored evidence dimensions agree most clearly right now — not a ranking of what to buy or sell, and not a prediction of future performance. See each stock's full Research Scorecard on the main dashboard for the underlying facts.</p>
<h3>Strongest agreeing-positive evidence</h3>
<div class="quotes">{strongest_positive_html}</div>
<h3>Strongest agreeing-negative evidence</h3>
<div class="quotes">{strongest_negative_html}</div>
""", DOCS_DIR)

    render_standalone_page("screener.html", "Screener", "📊 LSE Screener Overview", f"""
<p class="meta">A summary view of the LSE Risers/Fallers/Volume widget. For the full focused list of each category, use the dedicated Top Gainers, Top Losers, or Top Volume pages linked from Quick Navigation.</p>
{universe_status_line}
{screener_source_line}
<div class="screener-grid">
  <div><h2>Top Volume <a href="volume.html" style="color:#7fb3ff;font-size:13px;text-decoration:none;">(full page →)</a></h2><table><tr><th>#</th><th>Symbol</th><th>Volume</th></tr>{vol_rows}</table></div>
  <div><h2>Top Gainers <a href="gainers.html" style="color:#7fb3ff;font-size:13px;text-decoration:none;">(full page →)</a></h2><table><tr><th>#</th><th>Symbol</th><th>Chg%</th></tr>{gain_rows}</table></div>
  <div><h2>Top Losers <a href="losers.html" style="color:#7fb3ff;font-size:13px;text-decoration:none;">(full page →)</a></h2><table><tr><th>#</th><th>Symbol</th><th>Chg%</th></tr>{lose_rows}</table></div>
</div>
""", DOCS_DIR)

    BROKER_ENRICHMENT_NOTE = ('<p class="meta">Price, % change, and volume above are LSE first-party data. '
                               'Where a 🎯 target price appears within a stock\'s expanded evidence, that figure '
                               'is separate Yahoo Finance broker-consensus enrichment, not LSE data.</p>')

    render_standalone_page("gainers.html", "Top Gainers", "🟢 FTSE 100 Top Gainers", f"""
<p class="meta">The stocks with the largest positive price movement today, from the LSE Risers/Fallers/Volume widget.</p>
{universe_status_line}
{screener_source_line}
{BROKER_ENRICHMENT_NOTE}
<table><tr><th>#</th><th>Symbol</th><th>Chg%</th></tr>{gain_rows}</table>
""", DOCS_DIR)

    render_standalone_page("losers.html", "Top Losers", "🔴 FTSE 100 Top Losers", f"""
<p class="meta">The stocks with the largest negative price movement today, from the LSE Risers/Fallers/Volume widget.</p>
{universe_status_line}
{screener_source_line}
{BROKER_ENRICHMENT_NOTE}
<table><tr><th>#</th><th>Symbol</th><th>Chg%</th></tr>{lose_rows}</table>
""", DOCS_DIR)

    render_standalone_page("volume.html", "Top Volume", "📊 FTSE 100 Top Volume", f"""
<p class="meta">The most heavily traded stocks by volume today, from the LSE Risers/Fallers/Volume widget.</p>
{universe_status_line}
{screener_source_line}
{BROKER_ENRICHMENT_NOTE}
<table><tr><th>#</th><th>Symbol</th><th>Volume</th></tr>{vol_rows}</table>
""", DOCS_DIR)

    render_standalone_page("heatmap.html", "Heat Map", "🗺️ Heat Map (top movers, by size of move)", f"""
{('<p class="status-ok">✅ Data source: LSE (London Stock Exchange, first-party)'
  + (' · Retrieved: ' + esc(format_london_and_utc(datetime.fromisoformat(heatmap_retrieved_at))) if heatmap_retrieved_at else '')
  + '</p>') if _heatmap_instruments else '<p class="status-warn">⚠️ Data source: derived from Screener Gainers/Losers (dedicated LSE heatmap unavailable this run)</p>'}
<div class="heatmap-grid">{heatmap_cells or heatmap_empty_state_html()}</div>
""", DOCS_DIR)

    render_standalone_page("watchlist.html", "Watchlist", "👀 Your Watchlist", f"""
<div class="quotes">{quote_rows or '<span class="meta">No quotes yet</span>'}</div>
""", DOCS_DIR)

    render_standalone_page("news-explorer.html", "News Explorer", "📰 News Explorer (LSE regulatory &amp; company announcements)", f"""
<p class="meta">Genuine LSE first-party data — regulatory news and company announcements across the whole market, not limited to your watchlist. No equivalent exists via any other source used on this page, so this section shows an honest "unavailable" state rather than a substitute when the LSE source can't be reached.</p>
{('<p class="status-ok">✅ Data source: London Stock Exchange (first-party) · Retrieved: '
  + esc(format_london_and_utc(datetime.fromisoformat(news_explorer['retrievedAt']))) + ' · '
  + str(news_explorer.get('totalElements', 0)) + ' total result(s), showing '
  + str(len(news_explorer_stories)) + '</p>')
 if news_explorer.get('status') == 'ok' else
 ('<p class="status-warn">⚠️ News Explorer unavailable this run'
  + (' (' + esc(news_explorer['error']) + ')' if news_explorer.get('error') else '') + '</p>')}
{f'<table><tr><th>Headline</th><th>Company</th><th>Source</th><th>Date / Time</th><th>RNS Number</th><th>Price</th><th>Change %</th></tr>{news_explorer_rows}</table>' if news_explorer_stories else '<span class="meta">No stories available this run.</span>'}
""", DOCS_DIR)

    render_standalone_page("news-feed.html", "News / Evidence", "📰 News &amp; Broker Feed (watchlist)", f"""
{item_rows or news_empty_state_html(news_fetch_status, all_recent_items, "news")}
""", DOCS_DIR)

    render_standalone_page("catalysts.html", "Catalysts", "🗓️ Upcoming Catalysts", f"""
<p class="meta">Real, already-published earnings and ex-dividend dates from Yahoo's calendar data — informational only, not a suggestion to act around any of these dates. Covers both your watchlist and today's screener-ranked stocks.</p>
<table><tr><th>Stock</th><th>Event</th><th>Date</th><th>Days until</th><th>Timing</th></tr>{catalyst_rows or '<tr><td colspan="5" class="meta">No known upcoming earnings or ex-dividend dates right now.</td></tr>'}</table>
""", DOCS_DIR)

    render_standalone_page("warnings.html", "Warnings", "⚠️ Warnings / Contradictions", f"""
<p class="meta">Every stock currently on Radar with a flagged contradiction or an active DON'T CHASE warning, gathered here from the same evidence shown on each stock's own card — nothing new calculated.</p>
<div class="quotes">{warnings_html}</div>
""", DOCS_DIR)

    render_standalone_page("data-freshness.html", "Data / Freshness", "ℹ️ Data / Freshness", f"""
<p class="meta">A quick summary of how current each data source is right now, gathered from the same status checks already shown on the main dashboard.</p>
{data_freshness_html}
""", DOCS_DIR)

    render_standalone_page("more-data.html", "More Data", "📚 More Data", f"""
<h2 id="mover-news">📰 News on Today's Top Movers</h2>
<p class="meta">Real, dated-today news for any stock currently in Volume/Gainers/Losers or the Heat Map — not limited to your watchlist. Matched by company-name/headline keyword matching and classified by rule-based pattern detection (upgrade/downgrade/target-change etc.) — not AI-generated. See Market Research below for the one section that does use an AI-written summary.</p>
<div>{screener_news_rows or news_empty_state_html(mover_news_status, all_recent_mover_news, "news for today's ranked stocks", render_fn=lambda it: screener_news_item(it.get("ticker", ""), it))}</div>

<h2 id="uptrend">📈 5-Day Uptrend ({UPTREND_5DAY_THRESHOLD_PCT:.0f}%+, screener + watchlist)</h2>
<p class="meta">Real closing-price history over the last 5 trading days — a fact about the past, not a forecast of what happens next.</p>
<div class="quotes">{uptrend_rows or '<span class="meta">Nothing has met the 5-day threshold right now</span>'}</div>

<h2 id="targets">🎯 Broker Target Prices</h2>
<p class="meta">Real, already-published broker consensus targets from Yahoo's aggregation — not generated by this tool. Covers both your watchlist and today's screener-ranked stocks (Volume/Gainers/Losers).</p>
<div class="quotes">{target_price_rows or '<span class="meta">No target price data available yet.</span>'}</div>

<h2 id="movers-today">🔥 Already Moving Today (watchlist, ±{BIG_MOVER_THRESHOLD_PCT:.0f}%+)</h2>
<p class="meta">A fact about what already happened today — not a forecast of what happens next.</p>
<div class="quotes">{mover_rows or '<span class="meta">Nothing past the threshold right now</span>'}</div>

<h2 id="whats-changed">📅 What Changed Since Last Snapshot</h2>
<p class="meta">{whats_changed_intro}</p>
<div class="quotes">{whats_changed_html}</div>

<h2 id="market-research">🔎 Market Research</h2>
<p class="meta">Real broker targets, recent news, and consensus ratings, already gathered by this tool — free, always live, no AI or API cost involved. If an ANTHROPIC_API_KEY is configured, a short AI-written summary appears too (🤖), synthesised only from these same facts — never fresh web research, never a recommendation. Always cross-check anything here against primary sources before acting on it.</p>
{research_rows or '<p class="meta">No watchlist stocks to show yet.</p>'}

<h2 id="broker-alerts">⬆⬇🎯 Market-wide Broker Alerts (all LSE, not just watchlist)</h2>
<p class="meta">Upgrades/downgrades from anywhere on the LSE, not limited to your watchlist below.</p>
{market_wide_rows or news_empty_state_html(market_wide_alerts_status, recent_market_wide_filtered, "market-wide alerts")}

<h2 id="backtest">📊 Signal Backtest (Technical Signals Only)</h2>
{backtest_html}
""", DOCS_DIR)

    # Attaches each stock's Radar Stocks discovery sources (Watchlist,
    # Heat Map, LSE Volume/Gainers/Losers) onto its scorecard_summaries
    # entry, so Phase 7B's evidence-history persistence can save WHY a
    # stock was on the radar alongside its evidence — a future audit
    # trail, never touching the rendered HTML above (already written to
    # disk by this point) and never adding a new scorecard/evidence
    # computation of its own.
    for s in scorecard_summaries:
        disco = radar_discovery.get(bare_ticker(s["ticker"]))
        s["discoveredVia"] = [label for label, _reason in disco["sources"]] if disco else []

    # Additive: exposes the per-Watchlist-stock scorecard summary (price,
    # TOTAL, Signal Quality, Confidence, Evidence label) built as a pure
    # side-effect above — every existing caller that ignores this return
    # value (there was none before this) continues to work exactly as
    # before. Used by main() to persist the daily "What Changed" snapshot
    # without any duplicate scorecard/evidence computation.
    return scorecard_summaries


HEARTBEAT_INTERVAL_SECONDS = 6 * 3600  # confirmation ping every 6h, not every 5 min — a signal, not spam


AI_DIGEST_INTERVAL_SECONDS = 6 * 3600  # throttled like the heartbeat — a paid API call, not free

# CallMeBot's free WhatsApp tier allows ~16 messages per 240 minutes (~4/hour sustained).
# Sending the screener report every 5-minute cycle would blow through that in half an hour
# regardless of anything else — throttle the WhatsApp SEND specifically (the dashboard/
# data itself still refreshes every cycle; only the push notification is rationed).
SCREENER_MESSAGE_INTERVAL_SECONDS = 60 * 60  # once per hour
MAX_ALERTS_PER_RUN = 5  # cap a burst of alerts in one run; rest are still on the dashboard
# A stock is flagged as "getting attention today" purely on mention COUNT — how many
# already-published, real items exist about it today. This is not a signal about where
# the price might go; it's a fact about today's coverage volume, same category as the
# screener's volume ranking.
ATTENTION_MENTION_THRESHOLD = 3


def format_uptrend_message(uptrend_stocks):
    """Purely descriptive: stocks that have genuinely risen UPTREND_5DAY_THRESHOLD_PCT%+
    over the last 5 real trading days, computed from actual closing prices. This is a
    fact about the past — explicitly not a prediction the rise continues."""
    if not uptrend_stocks:
        return None
    flagged = sorted(uptrend_stocks, key=lambda x: -x["changePct5d"])
    lines = [
        f"📈 5-DAY UPTREND ({UPTREND_5DAY_THRESHOLD_PCT:.0f}%+, screener + watchlist)  🕐 {now_stamp()}",
        "Real closing-price history over the last 5 trading days — not a forecast of what happens next.",
    ]
    for s in flagged[:15]:
        lines.append(f'{s["symbol"]} ({s["name"]}) — ▲{s["changePct5d"]:.1f}% over 5 sessions')
    return "\n".join(lines)


def format_screener_news_message(screener_news):
    """Real, dated-today news for stocks currently ranked in Volume/Gainers/Losers —
    not limited to the watchlist. Capped per message to stay a reasonable length."""
    lines = []
    for symbol, items in screener_news.items():
        for it in items[:2]:  # at most 2 headlines per stock in the message (full list is on the dashboard)
            lines.append(f'{symbol}: {it["title"]}')
    if not lines:
        return None
    header = f"📰 NEWS ON TODAY'S TOP MOVERS  🕐 {now_stamp()}"
    return header + "\n" + "\n".join(lines[:15])  # overall cap so one message can't run away


def format_attention_message(mention_counts):
    """Purely descriptive: which watchlist stocks have unusually high real news/broker
    mention counts today. Never says why that might matter for price — just the count."""
    flagged = sorted(
        ((t, d) for t, d in mention_counts.items() if d["count"] >= ATTENTION_MENTION_THRESHOLD),
        key=lambda x: -x[1]["count"],
    )
    if not flagged:
        return None
    lines = [f"📢 GETTING ATTENTION TODAY (watchlist, {ATTENTION_MENTION_THRESHOLD}+ mentions)  🕐 {now_stamp()}"]
    lines.append("A count of today's coverage — not a signal about where price might go.")
    for ticker, d in flagged:
        lines.append(f"{ticker} ({d['name']}) — {d['count']} mentions today")
    return "\n".join(lines)
AI_DIGEST_MODEL = "claude-haiku-4-5-20251001"  # cheap/fast model, appropriate for a short summary

AI_DIGEST_SYSTEM_PROMPT = """You are a strictly factual summarizer for a UK stock market news digest sent to one person's WhatsApp.

CRITICAL RULES — violating any of these makes your output unusable:
- NEVER recommend buying, selling, or holding any stock, in any form or phrasing.
- NEVER predict future price movements or say what "might" or "could" happen next.
- NEVER use directive/advisory language: "should", "consider", "opportunity", "worth watching for a play", "good time to", etc.
- ONLY describe what has already happened: which broker rated which stock how, what news broke, what moved and by how much, per the data given to you.
- Attribute every claim to its source (the broker name, or "news reports") — never state something as your own conclusion.
- Do not add analysis, speculation, or connect-the-dots reasoning between separate facts.
- Keep it under 120 words, plain prose, no markdown headers.

If you cannot summarize something factually without it reading like advice, state the raw fact more plainly instead of omitting it or softening it into a suggestion."""

# Safety net beyond the system prompt — if the model drifts into advice-shaped language
# despite the instructions, this blocks the message from ever being sent rather than
# trusting the prompt alone.
FORBIDDEN_DIGEST_PATTERNS = [
    r"\byou should\b", r"\bconsider (buying|selling|holding)\b", r"\brecommend(s|ed|ing)?\b",
    r"\bworth (a )?(buy|look|watch)\b", r"\bgood (time|opportunity)\b", r"\bi (suggest|advise)\b",
    r"\bmight (rise|fall|climb|drop)\b", r"\bcould (rise|fall|climb|drop|see)\b",
    r"\bexpect(ed)? to (rise|fall|climb|drop)\b",
]


def generate_ai_digest(context_text):
    """
    Optional, opt-in, paid: summarizes already-gathered facts into a short WhatsApp
    digest. Returns None (silently, no error) if no API key is configured — the rest
    of the tool works identically either way. Never asked to judge or predict, only to
    restate what was already found; output is scanned before sending as a second layer
    of protection against advice-shaped language slipping through.
    """
    api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        return None
    body = json.dumps({
        "model": AI_DIGEST_MODEL,
        "max_tokens": 300,
        "system": AI_DIGEST_SYSTEM_PROMPT,
        "messages": [{"role": "user", "content": f"Today's gathered data:\n{context_text}\n\nWrite the factual summary."}],
    }).encode("utf-8")
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=body,
        headers={"x-api-key": api_key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
        method="POST",
    )
    try:
        resp_data = json.loads(urllib.request.urlopen(req, timeout=30).read())
        text = "".join(b.get("text", "") for b in resp_data.get("content", []) if b.get("type") == "text").strip()
        if not text:
            return None
        if any(re.search(pat, text, re.IGNORECASE) for pat in FORBIDDEN_DIGEST_PATTERNS):
            print("  ! AI digest blocked: output matched an advice-shaped pattern, not sent.", file=sys.stderr)
            return None
        return text
    except urllib.error.HTTPError as e:
        try:
            error_body = e.read().decode("utf-8", errors="replace")
        except Exception:
            error_body = "(couldn't read error body)"
        print(f"  ! AI digest generation failed: HTTP {e.code} — {error_body[:300]}", file=sys.stderr)
        return None
    except Exception as e:
        print(f"  ! AI digest generation failed: {e}", file=sys.stderr)
        return None


AI_EVIDENCE_REVIEW_MODEL = AI_DIGEST_MODEL  # same cheap/fast tier — this is a bounded
# classification task (pick one of 6 labels), not open-ended generation; no reason to
# use a larger/more expensive model for it.

# The ONLY allowed values — anything else in the AI's response is treated as
# malformed and the whole response is discarded. Never partially trusted.
AI_EVIDENCE_CONFIDENCE_VALUES = {"strong", "weak", "irrelevant", "not_same_day", "ambiguous", "opinion_speculation"}

AI_EVIDENCE_CONFIDENCE_LABELS = {
    "strong": "✅ Confirmed",
    "weak": "⚠️ Weak/uncertain",
    "irrelevant": "⚠️ Irrelevant",
    "not_same_day": "⚠️ Not genuinely same-day",
    "ambiguous": "⚠️ Ambiguous",
    "opinion_speculation": "⚠️ Opinion/speculation, not a factual report",
}

AI_EVIDENCE_REVIEW_SYSTEM_PROMPT = """You are reviewing evidence already used by a deterministic stock research system. You are NOT generating new evidence, NOT making a recommendation, and NOT predicting anything.

You will be given: a ticker, a company name, today's price move, the deterministic system's classification (Supported or Conflicting), and the specific same-day news/broker item(s) it was based on.

Your ONLY task: assess whether those specific item(s), taken at face value, genuinely and clearly explain today's price move as a factual matter.

CRITICAL RULES — violating any of these makes your output unusable:
- NEVER recommend buying, selling, or holding any stock, in any form.
- NEVER predict future price movement.
- NEVER use directive/advisory language: "should", "consider", "opportunity", "worth watching", "good time to", etc.
- You are not deciding whether the classification is "right" — the deterministic system's label is fixed and will be shown regardless of your answer. You are only assessing how CLEARLY the underlying material actually explains the move, and whether it is genuinely about this specific company and genuinely from today.

Respond with ONLY a JSON object, nothing else, no markdown, no explanation outside the JSON:
{"confidence": "<one of: strong, weak, irrelevant, not_same_day, ambiguous, opinion_speculation>", "caveat": "<a short plain-English note, under 200 characters, or empty string if confidence is strong>"}

Meaning of each confidence value:
- strong: the item(s) clearly and directly explain the price move
- weak: the item(s) are genuinely about this company but don't clearly explain WHY the price moved today
- irrelevant: the item(s) do not appear to genuinely be about this specific company (e.g. a name collision with a different company)
- not_same_day: the item(s) do not appear to genuinely be from today, despite being provided as same-day material
- ambiguous: there is genuine uncertainty, or the item could reasonably be read as supporting the move or not
- opinion_speculation: the item is commentary/opinion/speculation rather than a factual report of something that happened"""


def review_evidence_with_ai(ticker, company_name, change_pct, evidence_label, same_day_items):
    """
    Optional, opt-in, paid, downgrade-only review of an ALREADY-COMPUTED
    deterministic evidence classification — never a new evidence source,
    never able to change what's shown as the actual Evidence: label.
    Returns None (silently) if no API key is configured, if evidence_label
    isn't "supported"/"conflicting" (nothing directional to review), if the
    API call fails/times out, or if the response is malformed or contains
    anything resembling advice/recommendation language — in every one of
    those cases the dashboard behaves exactly as it does without this
    feature. Returns {"aiEvidenceConfidence": ..., "aiEvidenceCaveat": ...,
    "aiEvidenceReviewed": True} only when the response passed every check.

    same_day_items MUST already be the same-day-filtered, relevance-
    filtered pool (the exact same items_by_ticker[ticker] that
    classify_evidence itself was given) — never the wider recent-fallback
    pool. This is what "AI can only review evidence the deterministic
    system already identified" and "never present recent-fallback news as
    same-day evidence" mean in practice: the input itself is already
    constrained to genuine same-day material before the AI ever sees it.
    """
    if evidence_label not in ("supported", "conflicting"):
        return None  # nothing directional for the AI to review
    api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        return None
    if not same_day_items:
        return None  # nothing to review — shouldn't normally happen if label is supported/conflicting, but never guess

    items_text = "\n".join(
        f"- \"{it.get('title', '')}\" (category: {it.get('category', 'news')}, published: {it.get('pubDate', '?')})"
        for it in same_day_items[:5]  # cap — this is a review of the evidentiary basis, not a full digest
    )
    user_content = (
        f"Ticker: {ticker} ({company_name})\n"
        f"Today's price move: {change_pct:+.1f}%\n"
        f"Deterministic classification: {evidence_label}\n"
        f"Same-day item(s) this classification was based on:\n{items_text}\n\n"
        f"Assess these specific item(s) per your instructions and respond with the JSON object only."
    )
    body = json.dumps({
        "model": AI_EVIDENCE_REVIEW_MODEL,
        "max_tokens": 200,
        "system": AI_EVIDENCE_REVIEW_SYSTEM_PROMPT,
        "messages": [{"role": "user", "content": user_content}],
    }).encode("utf-8")
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=body,
        headers={"x-api-key": api_key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
        method="POST",
    )
    try:
        resp_data = json.loads(urllib.request.urlopen(req, timeout=30).read())
        text = "".join(b.get("text", "") for b in resp_data.get("content", []) if b.get("type") == "text").strip()
        if not text:
            return None
        parsed = json.loads(text)
        confidence = parsed.get("confidence")
        caveat = parsed.get("caveat", "")
        if confidence not in AI_EVIDENCE_CONFIDENCE_VALUES:
            print(f"  ! AI evidence review discarded: '{confidence}' is not an allowed value.", file=sys.stderr)
            return None
        if not isinstance(caveat, str) or len(caveat) > 300:
            print("  ! AI evidence review discarded: caveat malformed or too long.", file=sys.stderr)
            return None
        # Same forbidden-pattern scan already used for the digest — reused, not
        # duplicated, so both features stay behind the identical safety bar.
        if any(re.search(pat, caveat, re.IGNORECASE) for pat in FORBIDDEN_DIGEST_PATTERNS):
            print("  ! AI evidence review discarded: caveat matched an advice-shaped pattern.", file=sys.stderr)
            return None
        return {"aiEvidenceConfidence": confidence, "aiEvidenceCaveat": caveat, "aiEvidenceReviewed": True}
    except json.JSONDecodeError:
        print("  ! AI evidence review discarded: response was not valid JSON.", file=sys.stderr)
        return None
    except urllib.error.HTTPError as e:
        try:
            error_body = e.read().decode("utf-8", errors="replace")
        except Exception:
            error_body = "(couldn't read error body)"
        print(f"  ! AI evidence review failed: HTTP {e.code} — {error_body[:300]}", file=sys.stderr)
        return None
    except Exception as e:
        print(f"  ! AI evidence review failed: {e}", file=sys.stderr)
        return None


def build_digest_context(screener, market_wide_enriched, big_movers, watchlist_alerts):
    """Plain-text summary of this run's facts, handed to the AI as its only input —
    it never sees anything beyond what's already been gathered and shown elsewhere."""
    lines = []
    if screener.get("gainers"):
        lines.append("Top gainers: " + ", ".join(f"{q['symbol']} +{q.get('changePct',0):.1f}%" for q in screener["gainers"][:5]))
    if screener.get("losers"):
        lines.append("Top losers: " + ", ".join(f"{q['symbol']} {q.get('changePct',0):.1f}%" for q in screener["losers"][:5]))
    if market_wide_enriched:
        lines.append("Market-wide broker calls: " + "; ".join(
            f'{it.get("broker","?")} {it["category"]} on {it["title"][:80]}' for it in market_wide_enriched[:8]
        ))
    if big_movers:
        lines.append("Watchlist stocks already moving today: " + ", ".join(
            f"{m['ticker']} {m.get('changePct',0):+.1f}%" for m in big_movers
        ))
    if watchlist_alerts:
        lines.append("Watchlist news/broker alerts: " + "; ".join(
            f'{a.get("ticker","")} {a["category"]}: {a["title"][:80]}' for a in watchlist_alerts[:8]
        ))
    return "\n".join(lines) if lines else "No notable items this cycle."


def maybe_send_ai_digest(seen_state, screener, market_wide_enriched, big_movers, watchlist_alerts):
    if not os.environ.get("ANTHROPIC_API_KEY", "").strip():
        return False  # feature is off — no key configured, no cost, no message
    last_digest = seen_state.get("lastAiDigest", 0)
    if time.time() - last_digest < AI_DIGEST_INTERVAL_SECONDS:
        return False
    context = build_digest_context(screener, market_wide_enriched, big_movers, watchlist_alerts)
    digest = generate_ai_digest(context)
    if digest:
        msg = f"🤖 AI DAILY DIGEST (factual summary only, not advice)  🕐 {now_stamp()}\n\n{digest}"
        print(msg)
        send_webhook(msg)
    seen_state["lastAiDigest"] = time.time()
    return True


def maybe_send_heartbeat(seen_state, watchlist, alert_count, mover_count):
    last_heartbeat = seen_state.get("lastHeartbeat", 0)
    now = time.time()
    if now - last_heartbeat < HEARTBEAT_INTERVAL_SECONDS:
        return False
    ts = datetime.now(timezone.utc).strftime("%H:%M UTC")
    msg = (
        f"✅ UK Stock Watch — still checking. {ts}, watching {len(watchlist)} stocks. "
        f"{alert_count} broker alert(s) and {mover_count} big mover(s) since last heartbeat."
    )
    print(msg)
    send_webhook(msg)
    seen_state["lastHeartbeat"] = now
    return True


# =========================================================================
# MARKET RESEARCH — per-stock factual synthesis of ALREADY-GATHERED data
# =========================================================================
# IMPORTANT — what this is and isn't:
# This does NOT do fresh web research (no search capability exists in this
# script). It synthesizes facts poll.py has ALREADY collected this run —
# recent news/broker items, price, target price, consensus label — into a
# short factual write-up per stock. Same safety pattern as the existing AI
# digest: strict factual-only system prompt, output scanned against
# FORBIDDEN_DIGEST_PATTERNS before use, opt-in (requires ANTHROPIC_API_KEY).
#
# "Live" here means: refreshed automatically on a rolling basis, always
# current within ~1 day, WITHOUT refreshing all watchlist stocks every run —
# doing that would mean one paid API call per stock every 5 minutes, which
# doesn't scale to a 40+ stock watchlist. A capped number of the STALEST
# entries get refreshed each run, cycling through the whole list over time.

MARKET_RESEARCH_REFRESH_SECONDS = 24 * 3600  # each stock's write-up refreshes at most once/day
MARKET_RESEARCH_MAX_PER_RUN = 3  # cost/runtime cap — NOT a technical limit, a deliberate budget

MARKET_RESEARCH_SYSTEM_PROMPT = """You are writing a short "Market Research" note for one UK-listed stock, shown on a personal dashboard.

CRITICAL RULES — violating any of these makes your output unusable:
- Base EVERY sentence ONLY on the facts provided below. Do not draw on any outside knowledge of this company beyond what's given — if you know more about it from training, ignore that; use only the supplied facts.
- NEVER recommend buying, selling, or holding, in any form or phrasing.
- NEVER predict future price movements or say what "might" or "could" happen next.
- NEVER use directive/advisory language: "should", "consider", "opportunity", "worth watching", "good time to", etc.
- If broker ratings or targets are in the data, state them as attributed facts ("Broker X rates it Y with a target of Z") — never as your own conclusion or endorsement.
- If the provided facts are thin or contradictory, say so plainly rather than filling gaps with speculation.
- Do not explain what the "strong_buy"/"buy"/"hold" label conceptually means — that reasoning isn't in the data you're given, so don't invent it.
- Keep it under 100 words, plain prose, no markdown headers.

If you cannot write something factual from the given data, write "Not enough recent data gathered for a research note yet" instead of inventing content."""


def build_stock_research_context(ticker, name, quote, items):
    lines = [f"Stock: {name} ({ticker})"]
    if quote:
        price = quote.get("price")
        chg = quote.get("changePct")
        if price is not None:
            lines.append(f"Current price: {price} ({chg:+.2f}% today)" if chg is not None else f"Current price: {price}")
        target = quote.get("targetMeanPrice")
        rec = quote.get("recommendationKey")
        if target:
            lines.append(f"Broker consensus target price: {target}" + (f", consensus rating: {rec}" if rec else ""))
    if items:
        lines.append("Recent news/broker items (most recent first):")
        for it in items[:8]:
            broker_bit = f" [{it['broker']}]" if it.get("broker") else ""
            lines.append(f"- ({it.get('category','news')}){broker_bit} {it.get('title','')} — {it.get('source','')}, {it.get('pubDate','')}")
    return "\n".join(lines)


def generate_stock_research(ticker, name, quote, items):
    api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        return None
    context = build_stock_research_context(ticker, name, quote, items)
    body = json.dumps({
        "model": AI_DIGEST_MODEL,
        "max_tokens": 250,
        "system": MARKET_RESEARCH_SYSTEM_PROMPT,
        "messages": [{"role": "user", "content": f"Facts gathered for this stock:\n{context}\n\nWrite the research note."}],
    }).encode("utf-8")
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=body,
        headers={"x-api-key": api_key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
        method="POST",
    )
    try:
        resp_data = json.loads(urllib.request.urlopen(req, timeout=30).read())
        text = "".join(b.get("text", "") for b in resp_data.get("content", []) if b.get("type") == "text").strip()
        if not text:
            return None
        if any(re.search(pat, text, re.IGNORECASE) for pat in FORBIDDEN_DIGEST_PATTERNS):
            print(f"  ! market research blocked for {ticker}: output matched an advice-shaped pattern", file=sys.stderr)
            return None
        return text
    except urllib.error.HTTPError as e:
        try:
            error_body = e.read().decode("utf-8", errors="replace")
        except Exception:
            error_body = "(couldn't read error body)"
        print(f"  ! market research generation failed for {ticker}: HTTP {e.code} — {error_body[:300]}", file=sys.stderr)
        return None
    except Exception as e:
        print(f"  ! market research generation failed for {ticker}: {e}", file=sys.stderr)
        return None


def update_market_research(watchlist, quotes, items_by_ticker, existing_research):
    """
    Refreshes MARKET_RESEARCH_MAX_PER_RUN of the stalest watchlist entries per run.
    Returns the updated research dict (unchanged if no API key configured — the
    dashboard section still renders, just shows nothing generated yet).
    """
    if not os.environ.get("ANTHROPIC_API_KEY", "").strip():
        print("  market research: skipped, no ANTHROPIC_API_KEY configured", file=sys.stderr)
        return existing_research  # feature off — no key, no cost, section shows "not enabled"

    now = time.time()

    def staleness(ticker):
        entry = existing_research.get(ticker)
        if not entry or not entry.get("generatedAtEpoch"):
            return float("inf")  # never generated — highest priority
        return now - entry["generatedAtEpoch"]

    candidates = sorted(watchlist, key=lambda s: -staleness(s["ticker"]))
    due = [s for s in candidates if staleness(s["ticker"]) >= MARKET_RESEARCH_REFRESH_SECONDS]
    refresh_list = due[:MARKET_RESEARCH_MAX_PER_RUN]
    print(f"  market research: {len(due)} stock(s) due, refreshing {len(refresh_list)} this run")

    for stock in refresh_list:
        ticker, name = stock["ticker"], stock["name"]
        text = generate_stock_research(ticker, name, quotes.get(ticker), items_by_ticker.get(ticker, []))
        if text:
            existing_research[ticker] = {
                "text": text,
                "generatedAtEpoch": now,
                "generatedAt": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
            }
            print(f"  market research refreshed: {ticker}")
        time.sleep(1)  # be polite between successive API calls in the same run

    return existing_research


def main():
    watchlist = load_json(WATCHLIST_FILE, [])
    if not watchlist:
        print("watchlist.json is empty — nothing to do.")
        return

    seen_state = load_json(SEEN_FILE, {"seen": []})
    seen = set(seen_state.get("seen", []))
    global _recent_send_times
    _recent_send_times = list(seen_state.get("recentSendTimes", []))
    data = load_json(DATA_FILE, {"items": {}, "quotes": {}, "lastPoll": None})
    items_by_ticker = data.get("items", {})
    recent_items_by_ticker = data.get("recentItems", {})
    quotes = data.get("quotes", {})
    market_research = data.get("marketResearch", {})
    news_fetch_attempts = 0
    news_fetch_failures = 0

    yahoo_crumb = get_yahoo_crumb()
    print(f"Yahoo auth crumb: {'obtained' if yahoo_crumb else 'FAILED — screener/analyst-history will likely 401'}")

    new_alerts = []
    mention_counts = {}  # ticker -> {"name": ..., "count": ...} — today's mention volume, purely descriptive
    big_movers = []
    market_wide_enriched = []
    market_wide_enriched_wide = []
    screener = {}
    screener_status = {"volume": "not_checked", "gainers": "not_checked", "losers": "not_checked"}
    screener_source = {"volume": "not_checked", "gainers": "not_checked", "losers": "not_checked"}
    screener_lse_result = None
    heatmap_instruments = None
    heatmap_source = "not_checked"
    heatmap_lse_result = None
    news_explorer_result = {"status": "not_checked", "source": "LSE", "retrievedAt": None,
                             "stories": [], "totalElements": None, "totalPages": None, "error": None}
    lse_coverage_report = {"status": "not_checked"}
    ratings_fetch_failed = None  # None = not attempted this run (SKIP_MARKET_WIDE), True/False once it is
    ft_fetch_failed = None
    ft_items_pool = []
    market_wide_fetch_failed = None
    ratings_items = []
    screener_news = {}
    screener_news_recent = {}
    mover_news_fetch_attempts = 0
    mover_news_fetch_failures = 0
    uptrend_stocks = []
    screener_targets = {}
    _bare_to_screener_key = {}
    ftse_universe_names = None
    ftse_universe_source = "not_checked"  # SKIP_MARKET_WIDE runs never touch FTSE universe status at all
    ftse_universe_status = "not_checked"
    ftse100 = fetch_ftse100()

    if not SKIP_MARKET_WIDE:
        ratings_items, ratings_fetch_failed = fetch_feed(ANALYST_RATINGS_FEED_URL)
        market_wide_items, market_wide_fetch_failed = fetch_feed(market_wide_broker_news_url())

        # FT — fetched ONCE per run (both are general feeds, not per-company
        # searches — FT offers no company-specific search endpoint), then
        # text-matched against each watchlist/screener company below, the
        # same way general_news_url-sourced items already are. Deduplicated
        # between the two FT feeds specifically (a story can legitimately
        # appear in both Markets and the International homepage).
        ft_markets_items, ft_markets_failed = fetch_feed(FT_MARKETS_URL)
        ft_international_items, ft_international_failed = fetch_feed(FT_INTERNATIONAL_URL)
        ft_fetch_failed = ft_markets_failed and ft_international_failed  # only a genuine failure if BOTH broke
        _ft_seen_links = set()
        ft_items_pool = []
        for it in ft_markets_items + ft_international_items:
            if it["link"] in _ft_seen_links:
                continue
            _ft_seen_links.add(it["link"])
            ft_items_pool.append(it)

        # Restrict the screener to genuine FTSE 100/250 constituents — Yahoo's raw LSE
        # screener includes the whole market (AIM micro-caps and all), which is why
        # names like "Forgent", "Tower Resources", "Premier African Minerals" were
        # showing up under a section that's supposed to represent significant, broadly
        # tracked UK stocks. ftse_universe_names is None only when BOTH sources failed
        # AND no cache exists at all — in that specific case only, screener is left
        # unrestricted rather than incorrectly filtering out everything.
        #
        # raw_count=60: micro-caps often dominate a RAW top-10-by-volume/change ranking
        # (a penny stock's tiny price makes its share COUNT huge for a small pound
        # value) — fetching only 10 candidates and then FTSE-filtering them could
        # easily leave zero genuine FTSE 350 names. Over-fetching a larger pool first,
        # THEN filtering, THEN trimming to the final display count of 10 mirrors the
        # exact same over-fetch-then-filter pattern fetch_gb_screener already uses
        # internally for its own liquidity filter.
        ftse_universe_names, ftse_universe_source = load_ftse_universe()
        ftse_universe_status = ftse_universe_status_label(ftse_universe_source)
        # ONE run_id shared by every LSE call this poll run makes — added
        # specifically so a genuine duplicate fetch (same run_id, same
        # label, twice) is distinguishable from output that only LOOKS
        # duplicated due to buffered/reordered log lines. See _lse_log's
        # own docstring for the full reasoning.
        lse_run_id = new_lse_run_id()
        _lse_log(lse_run_id, "LSE poll sequence: START (Screener -> Heatmap -> News Explorer)")
        screener, screener_status, screener_source, screener_lse_result = fetch_lse_screener_primary(
            raw_count=60 if ftse_universe_names is not None else 10, display_count=10, run_id=lse_run_id)
        # A genuine, logged-every-run coverage check — not an assumption
        # made once and never re-verified. Only meaningful when LSE was
        # actually the source this run; skipped (not fabricated) when we
        # fell back to Yahoo, since there's nothing LSE-specific to check.
        # Reuses screener_lse_result's OWN already-fetched instruments —
        # this used to make a second, entirely redundant POST to the
        # same LSE endpoint; that redundancy has been removed, both to
        # cut needless load and to reduce the number of rapid
        # back-to-back requests this run makes to the same first-party
        # endpoint (a real, avoidable rate-limiting risk factor).
        if screener_source.get("gainers") == "LSE" and screener_lse_result is not None:
            try:
                lse_coverage_report = check_lse_ftse100_coverage(instruments=screener_lse_result["instruments"])
            except Exception as e:
                print(f"  ! LSE coverage check itself failed unexpectedly: {e}", file=sys.stderr)
                lse_coverage_report = {"status": "failed", "error": str(e)}
        else:
            lse_coverage_report = {"status": "skipped", "reason": "LSE was not the source this run"}

        # A short, deliberate pause between separate LSE components/refresh
        # calls within the same run. The pattern actually observed live
        # (Screener's own call succeeding while Heatmap's and News
        # Explorer's subsequent calls failed) is consistent with
        # rate-limiting from rapid back-to-back requests to the same
        # first-party endpoint — this doesn't attempt to defeat any
        # anti-bot mechanism, it simply behaves less like a burst of
        # automated requests, which is a legitimate, ordinary thing for
        # any well-behaved client to do.
        time.sleep(2)

        # Dedicated LSE heatmap fetch — a genuinely broader FTSE 100 pool
        # than just the top-10 gainers/losers already in screener (which
        # is what the heatmap grid fell back to before this). Uses the
        # SAME proven fetch_lse_ftse100_market_data, just the "heatmap"
        # tab specifically. Falls back to None (not a fabricated list) on
        # failure — render_dashboard already knows how to build the
        # heatmap pool from screener[gainers]/[losers] alone when this
        # is absent, so nothing breaks, it's just a narrower pool.
        heatmap_lse_result = fetch_lse_ftse100_market_data("heatmap", run_id=lse_run_id)
        if heatmap_lse_result["status"] == "ok":
            heatmap_instruments = heatmap_lse_result["instruments"]
            heatmap_source = "LSE"
            print(f"  > LSE heatmap: {len(heatmap_instruments)} instruments for the full-coverage heat map pool")
        else:
            heatmap_instruments = None
            heatmap_source = "unavailable"
            # Loud, unmistakable — this is exactly the failure mode that
            # was previously silent-ish (a single "!" line easy to miss
            # in a long log) while the live dashboard quietly fell back.
            print(f"!!! LSE HEATMAP FETCH FAILED this run: {heatmap_lse_result['error']} "
                  f"— heat map will use the narrower gainers/losers pool instead", file=sys.stderr)

        time.sleep(2)

        # News Explorer — genuinely no Yahoo equivalent exists for this
        # (regulatory/company announcement data), so failure here means
        # an honest "unavailable" state, never a substitute silently
        # presented as if it were the same thing.
        news_explorer_result = fetch_lse_news_explorer(run_id=lse_run_id)
        if news_explorer_result["status"] != "ok":
            print(f"!!! LSE NEWS EXPLORER FETCH FAILED this run: {news_explorer_result['error']}", file=sys.stderr)

        _lse_log(lse_run_id, f"LSE poll sequence: END — Screener={screener_source.get('gainers')}, "
                              f"Heatmap={heatmap_source}, News Explorer={news_explorer_result['status']}")

        # The FTSE-universe name filter below exists specifically to correct
        # Yahoo's broader GB-region screener down to genuine FTSE 100/250
        # constituents. When the LSE-primary fetch succeeded, the data is
        # ALREADY scoped to genuine FTSE 100 constituents by construction
        # (it's LSE's own official FTSE 100 page data) — running the
        # Yahoo-oriented name-matching filter over it would be unnecessary
        # and risks incorrectly dropping genuine rows whose LSE-sourced
        # name field doesn't happen to match the cached Yahoo-derived name
        # list. Only apply this filter when we've actually fallen back to
        # Yahoo for this run.
        used_yahoo_fallback = any(v == "Yahoo (fallback)" for v in screener_source.values())
        if ftse_universe_names is not None and used_yahoo_fallback:
            counts_before = {k: len(v) for k, v in screener.items()}
            for section in ("volume", "gainers", "losers"):
                filtered = [
                    row for row in screener.get(section, [])
                    if clean_company_name(row.get("name", "")).lower() in ftse_universe_names
                ]
                screener[section] = filtered[:10]
            counts_after = {k: len(v) for k, v in screener.items()}
            # The degraded (FTSE 100 only, no FTSE 250 at all) and stale-cache states are
            # deliberately loud here — this is exactly the "silently appears healthy while
            # coverage has quietly dropped" failure mode being guarded against. A plain
            # "healthy" status logs at the normal level; anything else gets an unmistakable
            # "!!!" marker so it can't blend into routine per-run output.
            marker = "" if ftse_universe_status == "healthy" else "!!! "
            print(f"{marker}FTSE universe filter (status={ftse_universe_status}, source={ftse_universe_source}, "
                  f"{len(ftse_universe_names)} names): "
                  f"volume {counts_before['volume']}->{counts_after['volume']}, "
                  f"gainers {counts_before['gainers']}->{counts_after['gainers']}, "
                  f"losers {counts_before['losers']}->{counts_after['losers']}")
        else:
            print(f"!!! FTSE universe UNAVAILABLE this run (status={ftse_universe_status}) — "
                  f"screener left completely unrestricted (whole LSE, no FTSE filtering applied)", file=sys.stderr)

        # News for every stock ranked in Volume/Gainers/Losers, not just the watchlist —
        # deduped by symbol (a stock can appear in more than one list), one query each,
        # staggered to avoid the burst-triggered 503s seen earlier in this project.
        #
        # Heat Map is now ALSO included, since it can genuinely surface a stock
        # that isn't in Volume/Gainers/Losers at all (confirmed: it's a broader,
        # separately-fetched ~100-instrument pool, not merely derived from the
        # other three). Deduped against symbols already covered above — a stock
        # in both pools is fetched once, not twice.
        #
        # Heat Map alone can be up to ~100 instruments, and news-fetching has no
        # batch API (confirmed: neither Yahoo function this project uses supports
        # multiple symbols per request) — genuinely adding one request per
        # Heat-Map-only stock would multiply this section's request count by
        # roughly 4x every single run. Capped instead to the top N Heat-Map-only
        # movers by |% change| (the same "biggest moves surface fastest"
        # principle already used for the Heat Map's own cell ordering), so this
        # stays a modest, bounded addition rather than an unmeasured explosion.
        ranked_stocks = {}
        for section in ("volume", "gainers", "losers"):
            for row in screener.get(section, []):
                ranked_stocks[row["symbol"]] = row.get("name", row["symbol"])
        gvl_symbol_count = len(ranked_stocks)

        # Broker-target enrichment (Yahoo) — a SEPARATE step from LSE market
        # data, restoring the targetMeanPrice/targetHighPrice/targetLowPrice/
        # analyst-count fields the LSE migration lost (LSE's own market-data
        # response has no broker-consensus fields at all). Scoped to
        # Volume/Gainers/Losers only — the SAME scope this feature already
        # had before the LSE migration (never included Heat Map) — and
        # deduped against watchlist symbols, which already get their own
        # target price via the existing per-ticker quotes lookup elsewhere.
        # This NEVER touches price/changePct/volume — those stay exactly as
        # LSE supplied them; only the broker-consensus fields are attached.
        watchlist_tickers_set = {stock["ticker"] for stock in watchlist}
        broker_enrichment_symbols = [
            s for s in ranked_stocks if s not in watchlist_tickers_set
        ] if screener_source.get("gainers") == "LSE" else []
        broker_enrichment_count = 0
        broker_target_by_symbol = {}
        for symbol in broker_enrichment_symbols:
            enrichment = fetch_yahoo_broker_target(symbol)
            broker_enrichment_count += 1
            if enrichment:
                broker_target_by_symbol[symbol] = enrichment
        if broker_enrichment_symbols:
            print(f"  > Broker-target enrichment (Yahoo): {broker_enrichment_count} request(s) for "
                  f"{len(broker_enrichment_symbols)} Volume/Gainers/Losers symbol(s) not already on "
                  f"the watchlist, {len(broker_target_by_symbol)} genuine target(s) found")
        for section in ("volume", "gainers", "losers"):
            for row in screener.get(section, []):
                enrichment = broker_target_by_symbol.get(row["symbol"])
                if enrichment:
                    # Attached alongside, never replacing, the LSE-sourced
                    # price/changePct/volume fields already on this row.
                    row.update(enrichment)
                    row["brokerTargetSource"] = "Yahoo Finance (enrichment)"

        HEATMAP_NEWS_CAP = 15
        heatmap_only_candidates = [
            row for row in (heatmap_instruments or [])
            if row.get("symbol") and row["symbol"] not in ranked_stocks
        ]
        heatmap_only_candidates.sort(key=lambda r: abs(r.get("changePct") or 0), reverse=True)
        heatmap_added = 0
        for row in heatmap_only_candidates[:HEATMAP_NEWS_CAP]:
            ranked_stocks[row["symbol"]] = row.get("name", row["symbol"])
            heatmap_added += 1

        print(f"Fetching news for {len(ranked_stocks)} screener-ranked stocks "
              f"({gvl_symbol_count} volume/gainers/losers + {heatmap_added} Heat-Map-only, "
              f"capped at {HEATMAP_NEWS_CAP})...")
        for symbol, name in ranked_stocks.items():
            # Search using a cleaned name (see clean_company_name docstring) — the
            # dashboard still displays the full "name" as-is, only the search query
            # uses the cleaned version, since that's what actually matches real news.
            cleaned_name = clean_company_name(name)
            items, mover_news_failed = fetch_feed(general_news_url(cleaned_name))
            mover_news_fetch_attempts += 1
            mover_news_fetch_failures += 1 if mover_news_failed else 0
            # Google matches a keyword search against full article content, not just
            # the headline — confirmed live: this pool surfaced items unrelated to the
            # searched stock. Requiring the (cleaned) company name to actually appear
            # in the item's own title before tagging it with this stock's ticker closes
            # that gap — same fix applied to the watchlist's own per-stock news loop.
            items = [("g", it) for it in items if cleaned_name.lower() in it.get("title", "").lower()]
            # FT — matched against the shared FT pool fetched once above, same
            # title-must-contain-company-name requirement as the Google-News
            # path (FT has no company-specific search either).
            ft_matched_mover = [("ft", it) for it in ft_items_pool if cleaned_name.lower() in it.get("title", "").lower()]
            items = items + ft_matched_mover
            items = [(src, it) for src, it in items if passes_relevance_filter(it.get("title", ""), cleaned_name, src)]
            # Built ONCE against the wider age ceiling, same restructuring
            # pattern as the other two news sections — the existing
            # same-day screener_news (which feeds evidence/scorecard via
            # render_stock_research_html's news_items parameter) is DERIVED
            # from this, guaranteeing byte-identical existing behaviour,
            # while a separate recent-fallback pool becomes available
            # purely for the "News on Today's Top Movers" DISPLAY, never
            # passed into evidence.
            items_wide = [(src, it) for src, it in items if passes_recency_filter_wide(it.get("pubDate"))]
            items_today = [(src, it) for src, it in items_wide if is_today_in_london(it.get("pubDate"))]
            now_iso_sc = datetime.now(timezone.utc).isoformat()

            def _enrich_mover_items(raw_items):
                out = []
                for fetch_source, it in raw_items[:5]:  # cap per-stock to keep dashboard/message size sane
                    category = classify(it["title"])
                    broker = detect_broker(it["title"]) if category in ("upgrade", "downgrade", "target", "target_raise", "target_cut", "initiation", "reiteration") else None
                    source_type, source_label = classify_source_type(fetch_source, it.get("link"), it.get("title"))
                    out.append({**it, "ticker": symbol, "company": name, "category": category, "broker": broker,
                                "detectedAt": now_iso_sc, "normalizedAt": now_iso_sc,
                                "normalizedAction": CATEGORY_TO_NORMALIZED_ACTION.get(category),
                                "fetchSource": fetch_source, "sourceType": source_type, "sourceLabel": source_label})
                return out

            if items_today:
                screener_news[symbol] = dedupe_near_duplicate_headlines(_enrich_mover_items(items_today))

            # Recent-fallback pool: merge with whatever was persisted for this
            # symbol, revalidate against the wide recency+relevance rule
            # (same principle as the other two sections — a stored item that
            # would no longer qualify today, whether by age or a relevance
            # rule change, doesn't linger indefinitely), dedupe, cap.
            existing_recent_sc = revalidate_stored_news_items(
                screener_news_recent.get(symbol, []), name, date_filter_fn=passes_recency_filter_wide,
            )
            merged_recent_sc = _enrich_mover_items(items_wide) + existing_recent_sc
            seen_links_sc = set()
            deduped_recent_sc = []
            for it in merged_recent_sc:
                if it["link"] in seen_links_sc:
                    continue
                seen_links_sc.add(it["link"])
                deduped_recent_sc.append(it)
            deduped_recent_sc = dedupe_near_duplicate_headlines(deduped_recent_sc)
            deduped_recent_sc.sort(key=item_sort_key, reverse=True)
            if deduped_recent_sc:
                screener_news_recent[symbol] = deduped_recent_sc[:5]

            time.sleep(1)  # stagger — this is the change most likely to trip Google's rate limiting if rushed

        # 5-day uptrend: real closing-price history for the same deduped stock set (no
        # extra tickers beyond what's already being fetched news for) plus the watchlist.
        # Deduped by BARE ticker (stripping the screener's ".L" suffix for comparison
        # only) — screener symbols and watchlist tickers refer to the same stock but
        # were never compared as equal before, so a stock present in BOTH pools (e.g.
        # KOO.L from the screener and KOO from the watchlist) was silently checked and
        # displayed TWICE. Whichever pool is seen first keeps its own key format
        # unchanged (screener_news's own lookups elsewhere depend on the raw ".L" key,
        # so that dict itself is never touched) — this only prevents the same real
        # stock from entering uptrend_targets under two different-looking keys.
        uptrend_stocks = []
        uptrend_targets = {}
        _seen_bare_tickers = set()
        for symbol, name in ranked_stocks.items():
            bare = bare_ticker(symbol)
            if bare not in _seen_bare_tickers:
                uptrend_targets[symbol] = name
                _seen_bare_tickers.add(bare)
        for stock in watchlist:
            bare = bare_ticker(stock["ticker"])
            if bare not in _seen_bare_tickers:
                uptrend_targets[stock["ticker"]] = stock["name"]
                _seen_bare_tickers.add(bare)
        # Cross-reference index: bare ticker -> whichever key form actually
        # WON the dedup above and is genuinely present in screener_targets.
        # Fixes a real, confirmed bug (GLEN, screener-ranked AND watchlisted
        # this run): the merge loop below looks up screener_targets by the
        # WATCHLIST's own ticker form ("GLEN"), but when a stock is BOTH
        # screener-ranked and watchlisted, the screener's form ("GLEN.L")
        # is what actually won the dedup and is the only key present —
        # "GLEN" was never added, so the lookup silently found nothing and
        # skipped the whole merge, even though the technicals data was
        # genuinely fetched (just under a different key). Built once here,
        # after uptrend_targets is finalized, reused by that merge below.
        _bare_to_screener_key = {bare_ticker(k): k for k in uptrend_targets}
        print(f"Checking price technicals, targets, earnings/dividend dates for {len(uptrend_targets)} stocks...")
        short_interest_map = fetch_short_interest()  # one fetch per run, not per ticker
        screener_targets = {}  # symbol -> {targetMeanPrice, recommendationKey, nextEarningsDate, exDividendDate, rsi14, ma20, aboveMA20}
        for symbol, name in uptrend_targets.items():
            hist = fetch_price_technicals(symbol)
            if hist and hist.get("changePct5d") is not None and hist["changePct5d"] >= UPTREND_5DAY_THRESHOLD_PCT:
                uptrend_stocks.append({"symbol": symbol, "name": name, **hist})
            time.sleep(0.3)  # lighter stagger — this hits Yahoo, not Google, different rate-limit budget

            # Real broker price targets + earnings/dividend calendar dates — all from
            # Yahoo's own published data, in the SAME request as before (calendarEvents
            # was added to the existing modules list, not a new call).
            analyst = fetch_yahoo_analyst(symbol)
            if analyst:
                entry = {}
                if analyst.get("targetMeanPrice"):
                    entry["targetMeanPrice"] = analyst["targetMeanPrice"]
                    entry["recommendationKey"] = analyst.get("recommendationKey")
                if analyst.get("nextEarningsDate"):
                    entry["nextEarningsDate"] = analyst["nextEarningsDate"]
                if analyst.get("exDividendDate"):
                    entry["exDividendDate"] = analyst["exDividendDate"]
                if analyst.get("dividendRate") is not None:
                    entry["dividendRate"] = analyst["dividendRate"]
                if analyst.get("dividendYieldPct") is not None:
                    entry["dividendYieldPct"] = analyst["dividendYieldPct"]
                if analyst.get("trailingPE") is not None:
                    entry["trailingPE"] = analyst["trailingPE"]
                if analyst.get("trailingEps") is not None:
                    entry["trailingEps"] = analyst["trailingEps"]
                if analyst.get("marketCap") is not None:
                    entry["marketCap"] = analyst["marketCap"]
                if analyst.get("averageVolume") is not None:
                    entry["averageVolume"] = analyst["averageVolume"]
                if analyst.get("fiftyTwoWeekLow") is not None:
                    entry["fiftyTwoWeekLow"] = analyst["fiftyTwoWeekLow"]
                if analyst.get("fiftyTwoWeekHigh") is not None:
                    entry["fiftyTwoWeekHigh"] = analyst["fiftyTwoWeekHigh"]
                if analyst.get("heldPercentInsidersPct") is not None:
                    entry["heldPercentInsidersPct"] = analyst["heldPercentInsidersPct"]
                if analyst.get("sector"):
                    entry["sector"] = analyst["sector"]
                if analyst.get("industry"):
                    entry["industry"] = analyst["industry"]
                if analyst.get("businessSummary"):
                    entry["businessSummary"] = analyst["businessSummary"]
                if hist:
                    entry["rsi14"] = hist.get("rsi14")
                    entry["ma20"] = hist.get("ma20")
                    entry["aboveMA20"] = hist.get("aboveMA20")
                    # Named distinctly from "volume" (screener rows already have their own,
                    # more precise regularMarketVolume-derived "volume" field from
                    # fetch_gb_screener) — this exists purely so watchlist rows, which have
                    # no other volume source, can get one via the merge below without ever
                    # risking overwriting a screener row's existing value.
                    entry["latestVolume"] = hist.get("latestVolume")
                    entry["changePct5d"] = hist.get("changePct5d")
                    entry["ma50"] = hist.get("ma50")
                    entry["ma200"] = hist.get("ma200")
                    entry["maCrossover"] = hist.get("maCrossover")
                    entry["atr14"] = hist.get("atr14")
                    entry["supportResistance"] = hist.get("supportResistance")
                    entry["breakoutStatus"] = hist.get("breakoutStatus")
                    entry["priceVolumeSeries"] = hist.get("priceVolumeSeries")
                si = match_short_interest(name, short_interest_map)
                if si:
                    entry["shortInterestPct"] = si["pct"]
                    entry["shortInterestDate"] = si["position_date"]
                if entry:
                    screener_targets[symbol] = entry
            time.sleep(0.3)

        # Attach the target-price data directly onto each screener row so it displays
        # with the row itself — no separate lookup needed on the dashboard/message side.
        for section in ("volume", "gainers", "losers"):
            for row in screener.get(section, []):
                extra = screener_targets.get(row["symbol"])
                if extra:
                    row.update(extra)  # attach whichever fields were found: target, recommendation, earnings/ex-div dates, RSI/MA

        # NOTE: the enrichment merge from screener_targets into `quotes` for
        # watchlist tickers happens AFTER the per-stock watchlist loop below,
        # not here — see the comment there for why (this merge must run
        # against THIS run's freshly-fetched quotes, not the persisted ones
        # still sitting in `quotes` at this point in the function).

        # Market-wide broker alerts: covers ALL LSE companies for upgrade/downgrade news,
        # not just the watchlist — this is what "all LSE companies" actually needs, without
        # trying to poll ~1,900 individual tickers (which would take hours per cycle and
        # get rate-limited). Combines the unfiltered ratings feed + the market-wide search.
        # general_market_items (investing.com's general "Stock Market News" feed) was
        # removed from this pool entirely, and its fetch removed too — despite being on
        # the uk.investing.com subdomain, real data proved it's not actually UK-scoped
        # (surfaced US broker actions on Workday, Ulta Beauty, Elastic, Autodesk as if
        # they were "all LSE" alerts). Not reused elsewhere either, so keeping the fetch
        # around would just be a wasted request for data that's no longer used anywhere.
        #
        # ratings_items (ANALYST_RATINGS_FEED_URL) turned out to have the EXACT SAME
        # problem, confirmed live on the deployed dashboard: "DA Davidson raises Workday
        # stock price target", "...raises Ulta Beauty...", "...raises Elastic..." — all
        # US stocks — were appearing in this "all LSE" section. The "uk." subdomain does
        # not mean UK-only content; it's a global analyst-ratings feed. There is no free
        # source that gives a full LSE company universe to check against (see project
        # history — short-interest work hit the identical wall), so items from this feed
        # are only kept when their subject company resolves to a ticker from THIS run's
        # own watchlist/screener pool — the same resolve_ticker_by_substring() machinery
        # already built and tested for the broker-events pipeline, including excluding
        # the detected broker's own name as a candidate subject. This is a real,
        # documented trade-off: it narrows Investing.com-sourced market-wide coverage to
        # companies already in the pool (fewer items, but every one confirmed genuinely
        # LSE-relevant) rather than the previous unfiltered/unverified global feed.
        # market_wide_items (Google News) is NOT put through this same gate — its query
        # is already scoped to "(LSE OR London Stock Exchange)" text, and requiring
        # pool-membership there would defeat the point of covering LSE companies beyond
        # the ~50-80 stock pool, which is what genuinely makes this section market-WIDE.
        mw_ticker_lookup = build_name_ticker_lookup(
            watchlist, screener.get("volume", []) + screener.get("gainers", []) + screener.get("losers", [])
        )
        ratings_resolved = {}  # link -> (ticker, company), only for CONFIRMED pool matches
        confirmed_ratings_items = []
        for it in ratings_items:
            rb = detect_broker(it["title"])
            r_ticker, r_company = resolve_ticker_by_substring(it["title"], mw_ticker_lookup, exclude_name=rb)
            if r_ticker:
                confirmed_ratings_items.append(it)
                ratings_resolved[it["link"]] = (r_ticker, r_company)
        market_wide_pool = [("ratings", it) for it in confirmed_ratings_items] + [("g", it) for it in market_wide_items]
        # Built ONCE against the WIDER age ceiling (not same-day), then the
        # existing same-day subset is DERIVED from it below — guarantees
        # market_wide_enriched (which feeds alerts/seen-tracking/AI-digest,
        # unchanged) ends up with the EXACT same items as before this
        # change, while a new wider pool becomes available alongside it for
        # the "most recent available, not from today" display fallback —
        # without re-running the classification/enrichment logic twice.
        market_wide_pool = [(src, it) for src, it in market_wide_pool if passes_recency_filter_wide(it.get("pubDate"))]
        now_iso_mw = datetime.now(timezone.utc).isoformat()
        market_wide_enriched_wide = []
        for mw_fetch_source, it in market_wide_pool:
            category = classify(it["title"])
            # "target" = a broker raising/cutting their price target — a genuine, already-
            # published broker action, same category of fact as an upgrade/downgrade, so it
            # belongs in the same alert stream rather than being silently dropped.
            if category not in ("upgrade", "downgrade", "target", "target_raise", "target_cut", "initiation", "reiteration"):
                continue
            resolved = ratings_resolved.get(it["link"])
            source_type, source_label = classify_source_type(mw_fetch_source, it.get("link"), it.get("title"))
            market_wide_enriched_wide.append({
                **it,
                "ticker": resolved[0] if resolved else "MARKET",
                "company": resolved[1] if resolved else "",
                "category": category,
                "broker": detect_broker(it["title"]),
                "detectedAt": now_iso_mw,
                "normalizedAt": now_iso_mw,
                "normalizedAction": CATEGORY_TO_NORMALIZED_ACTION.get(category),
                "sourceType": source_type,
                "sourceLabel": source_label,
                "fetchSource": mw_fetch_source,
            })
        market_wide_dedup_wide = {}
        for it in market_wide_enriched_wide:
            market_wide_dedup_wide[it["link"]] = it
        market_wide_enriched_wide = list(market_wide_dedup_wide.values())
        # The EXISTING same-day pool — identical items to before this change,
        # just now derived from the wider set rather than separately filtered.
        market_wide_enriched = [it for it in market_wide_enriched_wide if is_today_in_london(it.get("pubDate"))]
        # Near-duplicate dedup must stay WITHIN each ticker — this pool is
        # flat across many different companies, and comparing headlines
        # ACROSS tickers would risk two unrelated companies' genuinely
        # different alerts being wrongly treated as duplicates just for
        # sharing common rating-change wording. Grouped by ticker first,
        # deduped within each group, then flattened back.
        _mw_by_ticker = {}
        for it in market_wide_enriched:
            _mw_by_ticker.setdefault(it.get("ticker"), []).append(it)
        market_wide_enriched = [
            it for tkr_items in _mw_by_ticker.values()
            for it in dedupe_near_duplicate_headlines(tkr_items)
        ]

        for it in market_wide_enriched:
            if it["link"] in seen:
                continue
            seen.add(it["link"])
            new_alerts.append(it)
    else:
        print("SKIP_MARKET_WIDE set — skipping market-wide search/screener/heatmap (covered by the other job).")

    for stock in watchlist:
        ticker, name = stock["ticker"], stock["name"]
        print(f"Polling {ticker} ({name})...")

        g_items, g_failed = fetch_feed(google_news_url(name))
        y_items, y_failed = fetch_feed(yahoo_news_url(ticker))
        rb_items, rb_failed = fetch_feed(reuters_bloomberg_url(name))
        news_fetch_attempts += 3
        news_fetch_failures += sum([g_failed, y_failed, rb_failed])
        _cleaned_name = clean_company_name(name)
        # ONE shared relevance gate (passes_relevance_filter) now applied to every
        # source, including y_items — investigated directly, not assumed exempt: Yahoo's
        # per-ticker feed has no query-level scoping of its own, and real live examples
        # (an ABDN-tagged bond-coupon item, a SHEL-tagged Allianz/AA takeover item —
        # neither mentioning the company at all) confirmed items were reaching the
        # dashboard without ever being relevance-checked. matched_ratings now also uses
        # the cleaned name (previously used the raw, uncleaned name — inconsistent with
        # every other source here).
        g_items = [it for it in g_items if passes_relevance_filter(it.get("title", ""), _cleaned_name, "g")]
        y_items = [it for it in y_items if passes_relevance_filter(it.get("title", ""), _cleaned_name, "y")]
        rb_items = [it for it in rb_items if passes_relevance_filter(it.get("title", ""), _cleaned_name, "rb")]
        matched_ratings = [it for it in ratings_items if passes_relevance_filter(it.get("title", ""), _cleaned_name, "ratings")]
        # FT — matched against the ONE shared FT pool fetched above (not a
        # per-ticker fetch), same title-must-contain-the-company-name
        # requirement as general_news_url's own post-filter (FT offers no
        # company-specific search, so its general feed needs the same
        # tightening), THEN the same shared relevance gate as every other
        # source — no separate, looser rule for FT.
        ft_matched = [it for it in ft_items_pool if _cleaned_name.lower() in it.get("title", "").lower()]
        ft_matched = [it for it in ft_matched if passes_relevance_filter(it.get("title", ""), _cleaned_name, "ft")]
        combined_tagged_all = [("g", it) for it in g_items] + [("y", it) for it in y_items] \
            + [("rb", it) for it in rb_items] + [("ratings", it) for it in matched_ratings] \
            + [("ft", it) for it in ft_matched]
        # Built ONCE against the wider age ceiling, same pattern as the
        # market-wide restructuring above — the existing same-day subset is
        # DERIVED from it, guaranteeing byte-identical behaviour for
        # anything already consuming `combined_tagged` (evidence/scorecard
        # included), while a wider pool becomes available alongside it.
        combined_tagged_wide = [(src, it) for src, it in combined_tagged_all if passes_recency_filter_wide(it.get("pubDate"))]
        combined_tagged = [(src, it) for src, it in combined_tagged_wide if is_today_in_london(it.get("pubDate"))]
        # Purely a count of real, already-published items mentioning this stock today
        # (deduped by link) — a fact about today's coverage volume, not a prediction of
        # anything. NEWS_SAME_LONDON_DAY_ONLY already restricts `combined` to today.
        mention_links = {it["link"] for _, it in combined_tagged}
        mention_counts[ticker] = {"name": name, "count": len(mention_links)}

        quote = fetch_yahoo_quote(ticker)
        if quote:
            quotes[ticker] = quote
            # Purely descriptive: this stock has already moved a lot today.
            # Not a prediction of further movement, just a factual flag.
            if abs(quote.get("changePct") or 0) >= BIG_MOVER_THRESHOLD_PCT:
                big_movers.append({**quote, "ticker": ticker})

        analyst = fetch_yahoo_analyst(ticker)
        analyst_items = analyst_history_to_items(ticker, analyst)
        if analyst and analyst.get("targetMeanPrice") and ticker in quotes:
            quotes[ticker]["targetMeanPrice"] = analyst["targetMeanPrice"]
            quotes[ticker]["recommendationKey"] = analyst.get("recommendationKey")

        now_iso = datetime.now(timezone.utc).isoformat()

        def _enrich_tagged(tagged_pairs):
            out = []
            for fetch_source, it in tagged_pairs:
                category = classify(it["title"])
                # Only tag a broker when the item is actually a rating/target call —
                # otherwise a story that merely mentions a bank's name (e.g. a
                # personnel/legal story about "Barclays") gets mislabeled as if that
                # bank issued the rating.
                broker = detect_broker(it["title"]) if category in ("upgrade", "downgrade", "target", "target_raise", "target_cut", "initiation", "reiteration") else None
                source_type, source_label = classify_source_type(fetch_source, it.get("link"), it.get("title"))
                out.append({
                    **it,
                    "ticker": ticker,
                    "company": name,
                    "category": category,
                    "broker": broker,
                    "detectedAt": now_iso,
                    "normalizedAt": now_iso,
                    "normalizedAction": CATEGORY_TO_NORMALIZED_ACTION.get(category),
                    # Recorded so a LATER run can re-apply the correct, source-aware
                    # relevance policy when this item is carried forward — see
                    # revalidate_stored_news_items().
                    "fetchSource": fetch_source,
                    "sourceType": source_type,
                    "sourceLabel": source_label,
                })
            return out

        enriched = _enrich_tagged(combined_tagged)
        enriched_wide = _enrich_tagged(combined_tagged_wide)
        # Analyst history items already carry structured category/broker/pubDate —
        # merge as-is rather than re-running keyword classification on them. Not a text
        # search result at all (Yahoo's quoteSummary API returns these keyed directly to
        # this exact ticker), so relevance checking doesn't apply — tagged "analyst" so
        # carry-forward revalidation knows to exempt these too, not just this run.
        for it in analyst_items:
            enriched.append({**it, "ticker": ticker, "company": name, "detectedAt": now_iso, "fetchSource": "analyst", "sourceType": "broker_data", "sourceLabel": "Broker Data"})
            enriched_wide.append({**it, "ticker": ticker, "company": name, "detectedAt": now_iso, "fetchSource": "analyst", "sourceType": "broker_data", "sourceLabel": "Broker Data"})


        # Re-validate PERSISTED items against TODAY's date filter AND today's relevance
        # rule — not date alone. A stored item that would fail the relevance check if
        # fetched fresh today doesn't get a free pass just because it arrived on an
        # earlier run: confirmed live, an ABDN item about a bond coupon schedule and a
        # SHEL item about an unrelated Allianz/AA takeover — neither mentioning the
        # company at all — persisted for hours because only date was ever re-checked.
        # This check is purely a function of each stored item's OWN recorded title/
        # company/fetchSource — never affected by whether THIS run's fresh fetch of any
        # source succeeded or failed, so a temporarily-unavailable source can never cause
        # a genuinely relevant, already-stored item to be discarded.
        existing = revalidate_stored_news_items(items_by_ticker.get(ticker, []), name)
        merged = enriched + existing
        seen_links = set()
        deduped = []
        for it in merged:
            if it["link"] in seen_links:
                continue
            seen_links.add(it["link"])
            deduped.append(it)
        # Near-duplicate dedup: the SAME story covered independently by
        # different outlets with different links (exact-link dedup above
        # only catches identical URLs) - keeps the most authoritative
        # source's version when two headlines are genuinely near-identical.
        deduped = dedupe_near_duplicate_headlines(deduped)
        deduped.sort(key=item_sort_key, reverse=True)
        items_by_ticker[ticker] = deduped[:MAX_ITEMS_PER_TICKER]

        # Wider-window ("most recent available, not necessarily today")
        # parallel pool for this ticker — display-only, NEVER fed into
        # evidence/scorecard/contradiction logic (that stays exclusively on
        # items_by_ticker above). Same merge-with-persisted-and-revalidate
        # pattern, just against the wider recency rule.
        existing_recent = revalidate_stored_news_items(
            recent_items_by_ticker.get(ticker, []), name, date_filter_fn=passes_recency_filter_wide,
        )
        merged_recent = enriched_wide + existing_recent
        seen_links_recent = set()
        deduped_recent = []
        for it in merged_recent:
            if it["link"] in seen_links_recent:
                continue
            seen_links_recent.add(it["link"])
            deduped_recent.append(it)
        deduped_recent = dedupe_near_duplicate_headlines(deduped_recent)
        deduped_recent.sort(key=item_sort_key, reverse=True)
        recent_items_by_ticker[ticker] = deduped_recent[:MAX_ITEMS_PER_TICKER]

        for it in enriched:
            if it["link"] in seen:
                continue
            seen.add(it["link"])
            # Analyst-history items are backfilled (up to 15 past ratings) so first-run
            # dashboard context isn't empty — but only alert on genuinely recent ones,
            # not old history that happens to be new to our "seen" set. For your
            # watchlist specifically, event-category news (mergers, trading updates,
            # profit warnings, results) is genuinely price-moving and worth pushing —
            # unlike the market-wide stream, this is scoped to stocks you actually track.
            if it["category"] in ("upgrade", "downgrade", "target", "target_raise", "target_cut", "initiation", "director_dealing", "event") and it.get("_recent", True):
                new_alerts.append(it)

        time.sleep(0.5)  # be polite to free sources

    # Merge the screener_targets enrichment (RSI/MA20/MA50/MA200/crossover/ATR/
    # support-resistance/breakout/market cap/average volume/52-week range/
    # sector — everything fetched in the shared uptrend_targets loop above)
    # into the watchlist's `quotes` dict. THIS MUST RUN HERE — after the
    # per-stock loop above has finished setting quotes[ticker] = quote for
    # every watchlist ticker — not earlier. An earlier version of this merge
    # ran BEFORE that per-stock loop, against the persisted `quotes` from the
    # previous run's data.json; since quotes[ticker] = quote a few lines
    # above is a FULL replacement (not an update), that earlier ordering
    # meant every field this merge added was silently wiped out moments
    # later for any ticker that got a fresh quote this run — which is the
    # normal case. Caught by checking main()'s actual execution order end to
    # end, not just by testing render_dashboard() in isolation against
    # hand-constructed data (which is exactly why it wasn't caught earlier —
    # isolated rendering tests never exercise this ordering at all).
    # targetMeanPrice/recommendationKey are left as already set by the
    # watchlist's own earlier fetch_yahoo_analyst call (still needed there
    # for analyst_items/news-feed history, a separate purpose from this
    # merge) rather than overwritten, though both ultimately come from the
    # same underlying figure.
    for stock in watchlist:
        ticker = stock["ticker"]
        if ticker not in quotes:
            continue
        # Falls back to the bare-ticker cross-reference when the watchlist's
        # own ticker form isn't a key in screener_targets — see
        # _bare_to_screener_key's own comment for exactly why this is
        # needed (a screener-ranked-AND-watchlisted stock like GLEN).
        extra = screener_targets.get(ticker) or screener_targets.get(_bare_to_screener_key.get(bare_ticker(ticker), ""))
        if not extra:
            continue
        for key, value in extra.items():
            if key in ("targetMeanPrice", "recommendationKey"):
                continue  # already set by the watchlist's own earlier fetch; don't disturb
            if key == "latestVolume":
                # Named distinctly upstream specifically so it can be safely mapped
                # to "volume" here — fetch_yahoo_quote (the watchlist's own price
                # source) never returns volume at all, so there's nothing to collide
                # with; a screener row's own "volume" field is never touched by this
                # loop at all, since this only ever writes into `quotes`.
                if value is not None:
                    quotes[ticker]["volume"] = value
                continue
            quotes[ticker][key] = value

    # AI evidence-quality review (Phase 1, opt-in via ANTHROPIC_API_KEY) —
    # runs AFTER the merge above so quotes[ticker] has the SAME fields
    # (averageVolume, etc) that render_dashboard itself will use, computing
    # evidence via the identical classify_evidence() call with identical
    # inputs — never a second, potentially-different evidence computation.
    # Watchlist only for this initial conservative rollout (not Screener
    # rows) — a deliberate scoping choice, not an oversight; extending to
    # Screener would reuse the exact same fields once this is proven out.
    if os.environ.get("ANTHROPIC_API_KEY", "").strip():
        try:
            _ai_events = load_events_store().get("events", [])
        except Exception:
            _ai_events = []
        _ai_latest_events = get_latest_broker_event_per_ticker(_ai_events)
        for stock in watchlist:
            ticker = stock["ticker"]
            if ticker not in quotes:
                continue
            q = quotes[ticker]
            same_day_items = items_by_ticker.get(ticker, [])
            vol_ratio = compute_volume_ratio(q.get("volume"), q.get("averageVolume"))
            evidence = classify_evidence(
                q.get("changePct"), vol_ratio, same_day_items,
                latest_broker_event=_ai_latest_events.get(ticker),
            )
            ai_review = review_evidence_with_ai(
                ticker, stock["name"], q.get("changePct") or 0, evidence["label"], same_day_items,
            )
            if ai_review:
                quotes[ticker]["aiEvidenceConfidence"] = ai_review["aiEvidenceConfidence"]
                quotes[ticker]["aiEvidenceCaveat"] = ai_review["aiEvidenceCaveat"]
                quotes[ticker]["aiEvidenceReviewed"] = True

    # Merge this run's market-wide items with previously stored ones (same pattern as
    # the per-ticker feed) so the dashboard shows recent history, not just this cycle.
    # Existing/persisted items are re-validated against TODAY's filter too — same fix
    # and same reasoning as the per-ticker merge above (see its comment).
    existing_market_wide = [it for it in data.get("marketWide", []) if passes_news_filters(it.get("pubDate"))]
    merged_market_wide = market_wide_enriched + existing_market_wide
    seen_links_mw = set()
    deduped_market_wide = []
    for it in merged_market_wide:
        if it["link"] in seen_links_mw:
            continue
        seen_links_mw.add(it["link"])
        deduped_market_wide.append(it)
    deduped_market_wide.sort(key=item_sort_key, reverse=True)
    deduped_market_wide = deduped_market_wide[:MAX_ITEMS_PER_TICKER]

    # Wider-window ("most recent available, not necessarily today") parallel
    # pool — display-only, NEVER fed into evidence/scorecard/contradiction
    # logic, which stays exclusively on the same-day deduped_market_wide
    # above. Same merge-with-persisted pattern, just against
    # RECENT_NEWS_FALLBACK_MAX_AGE_DAYS instead of same-day-only.
    existing_market_wide_recent = [it for it in data.get("recentMarketWide", []) if passes_recency_filter_wide(it.get("pubDate"))]
    merged_market_wide_recent = market_wide_enriched_wide + existing_market_wide_recent
    seen_links_mw_recent = set()
    deduped_market_wide_recent = []
    for it in merged_market_wide_recent:
        if it["link"] in seen_links_mw_recent:
            continue
        seen_links_mw_recent.add(it["link"])
        deduped_market_wide_recent.append(it)
    deduped_market_wide_recent.sort(key=item_sort_key, reverse=True)
    deduped_market_wide_recent = deduped_market_wide_recent[:MAX_ITEMS_PER_TICKER]

    # Refresh a capped batch of the stalest Market Research write-ups (see function
    # docstring for why this doesn't refresh everyone every run) using this run's
    # freshly-gathered items_by_ticker/quotes as the factual source.
    market_research = update_market_research(watchlist, quotes, items_by_ticker, market_research)

    _last_poll_now = datetime.now(timezone.utc)
    _last_poll_str = _last_poll_now.strftime("%Y-%m-%d %H:%M:%S")
    data = {
        "items": items_by_ticker,
        "recentItems": recent_items_by_ticker,
        "quotes": quotes,
        "screener": screener,
        "screenerStatus": screener_status,
        "screenerSource": screener_source,
        "screenerRetrievedAt": screener_lse_result["retrievedAt"] if screener_lse_result else None,
        "lseCoverageReport": lse_coverage_report,
        "heatmapInstruments": heatmap_instruments,
        "heatmapSource": heatmap_source,
        "heatmapRetrievedAt": heatmap_lse_result["retrievedAt"] if heatmap_lse_result and heatmap_lse_result.get("status") == "ok" else None,
        "newsExplorer": news_explorer_result,
        "ftse100": ftse100,
        "screenerNews": screener_news,
        "screenerNewsRecent": screener_news_recent,
        "moverNewsFetchAttempts": mover_news_fetch_attempts,
        "moverNewsFetchFailures": mover_news_fetch_failures,
        "uptrendStocks": uptrend_stocks,
        "bigMovers": big_movers,
        "marketWide": deduped_market_wide,
        "recentMarketWide": deduped_market_wide_recent,
        # Per-section fetch health, same "did it actually work" distinction
        # as screenerStatus — None means genuinely not attempted this run
        # (SKIP_MARKET_WIDE), never conflated with a real failure.
        "marketWideAlertsStatus": (
            "not_checked" if ratings_fetch_failed is None and market_wide_fetch_failed is None
            else "failed" if (ratings_fetch_failed and market_wide_fetch_failed)
            else "ok"
        ),
        # FT's own fetch status — separate from the per-ticker news
        # aggregate below, since FT is ONE shared fetch for the whole run
        # (not per-ticker), matching how marketWideAlertsStatus is tracked.
        "ftFetchStatus": (
            "not_checked" if ft_fetch_failed is None
            else "failed" if ft_fetch_failed
            else "ok"
        ),
        # Aggregate across every per-ticker news fetch this run (3 feeds ×
        # every watchlist stock) — a per-ticker granular breakdown would be
        # a much larger change for limited extra value; this answers the
        # real question ("is something actually broken right now") without
        # needing to inspect each of potentially dozens of individual fetches.
        "newsFetchAttempts": news_fetch_attempts,
        "newsFetchFailures": news_fetch_failures,
        "marketResearch": market_research,
        "lastPoll": _last_poll_str,
        # Explicit, persisted universe/source status — surfaced on the dashboard itself
        # (not just in the workflow log) specifically so a drop from full FTSE 100+250
        # coverage down to FTSE-100-only, a stale cache, or fully unrestricted can never
        # be silently invisible to someone just looking at the live page.
        "ftseUniverseStatus": ftse_universe_status,
        "ftseUniverseSource": ftse_universe_source,
        "ftseUniverseCount": len(ftse_universe_names) if ftse_universe_names is not None else 0,
        # Pipeline (write-time) freshness — deliberately separate from feed-level health
        # above. See compute_pipeline_health()'s docstring: this is a snapshot evaluated
        # NOW, at write time — it is not proof of current health once the file ages. The
        # dashboard's own client-side script is what determines CURRENT freshness.
        "pipelineHealth": compute_pipeline_health(_last_poll_str, _last_poll_now),
    }
    save_json(DATA_FILE, data)

    # --- Daily snapshot lookup for "What Changed" (Phase 6) — wrapped so
    # ANY failure here (corrupt state/daily_snapshots.json, unexpected
    # exception) can NEVER stop the dashboard from rendering. A missing/
    # unreadable prior snapshot just means render_dashboard shows the
    # honest "no prior snapshot yet" message, same as first-run.
    prior_snapshot = None
    try:
        _now_for_snapshot = datetime.now(timezone.utc)
        _today_london = _now_for_snapshot.astimezone(LONDON_TZ).strftime("%Y-%m-%d")
        _snapshots_store = load_daily_snapshots()
        prior_snapshot = find_prior_snapshot(_snapshots_store, _today_london)
    except DailySnapshotsCorruptError as e:
        print(f"  ! daily snapshots store CORRUPT — 'What Changed' will show as unavailable this run: {e}", file=sys.stderr)
    except Exception as e:
        print(f"  ! daily snapshots lookup failed — 'What Changed' will show as unavailable this run: {e}", file=sys.stderr)

    # Phase 7C: loads whatever backtest results were most recently saved
    # by a SEPARATE backtest run (see --backtest below) — never computed
    # here, since a full historical backtest is far heavier than this
    # 5-minute poll cycle. A missing/corrupt results file just means the
    # dashboard shows its own honest "no backtest run yet" state.
    backtest_results = None
    try:
        backtest_results = load_backtest_results()
    except BacktestResultsCorruptError as e:
        print(f"  ! backtest results file CORRUPT — Signal Backtest section will show as unavailable this run: {e}", file=sys.stderr)
    except Exception as e:
        print(f"  ! backtest results lookup failed — Signal Backtest section will show as unavailable this run: {e}", file=sys.stderr)

    # Live Radar: load the persistent cross-run history, merge THIS run's
    # fresh discovery into it (never mutating the loaded copy — see
    # merge_radar_history's own docstring), and hold the updated store to
    # save AFTER rendering — same load-before/save-after pattern as every
    # other persistence path here. A corrupt history file means the
    # lifecycle info (Found/Age/Status) is unavailable this run, but never
    # blocks the rest of the dashboard from rendering.
    radar_lifecycle = None
    updated_radar_history = None
    try:
        _radar_now_iso = datetime.now(timezone.utc).isoformat()
        _radar_history = load_radar_history()
        # Mirrors render_dashboard's own default-loading logic for
        # latest_broker_events (see that function's docstring) — main()
        # doesn't otherwise have this available before render_dashboard
        # runs, and it's a pure read of the already-persisted events
        # store, no side effects, safe to load here too.
        try:
            _radar_broker_events = get_latest_broker_event_per_ticker(load_events_store().get("events", []))
        except Exception:
            _radar_broker_events = {}
        _radar_discovery = discover_radar_stocks(
            watchlist, data.get("bigMovers", []), data.get("screener", {}), _radar_broker_events,
        )
        updated_radar_history, radar_lifecycle = merge_radar_history(_radar_history, _radar_discovery, _radar_now_iso)
    except RadarHistoryCorruptError as e:
        print(f"  ! radar history store CORRUPT — Radar Stocks will show without Found/Age/Status this run: {e}", file=sys.stderr)
    except Exception as e:
        print(f"  ! radar history lookup failed — Radar Stocks will show without Found/Age/Status this run: {e}", file=sys.stderr)

    scorecard_summaries = render_dashboard(
        data, watchlist, prior_snapshot=prior_snapshot, backtest_results=backtest_results,
        radar_lifecycle=radar_lifecycle,
    )

    # Saved only once rendering has succeeded — a render failure never
    # persists a lifecycle update whose corresponding HTML was never
    # actually produced.
    if updated_radar_history is not None:
        try:
            save_radar_history(updated_radar_history)
        except Exception as e:
            print(f"  ! radar history save FAILED — previous file left untouched: {e}", file=sys.stderr)

    # Captures TODAY's snapshot from the SAME scorecard_summaries
    # render_dashboard just returned — never a second scorecard/evidence
    # computation. Wrapped with the same fail-safe discipline as the
    # broker-events block just below: a failure here can never stop the
    # rest of this poll cycle.
    try:
        snapshot_result = capture_daily_snapshot(scorecard_summaries or [])
        if not snapshot_result.get("written"):
            print(f"Daily snapshot: not written this run (reason={snapshot_result.get('reason')})")
    except Exception as e:
        print(f"  ! Daily snapshot capture failed entirely — state/daily_snapshots.json left untouched: {e}", file=sys.stderr)

    # Phase 7B: prospective evidence-history capture — from the SAME
    # scorecard_summaries, same fail-safe wrapping, entirely separate
    # from the Phase 6 snapshot above (different file, richer schema,
    # different purpose — building toward a future honest backtest of
    # the News/Broker/AI dimensions).
    try:
        evidence_history_result = capture_daily_evidence_snapshot(scorecard_summaries or [])
        if not evidence_history_result.get("written"):
            print(f"Evidence history: not written this run (reason={evidence_history_result.get('reason')})")
    except Exception as e:
        print(f"  ! Evidence history capture failed entirely — state/evidence_history.json left untouched: {e}", file=sys.stderr)

    # --- Broker-events pipeline (additive, isolated, separate from state/
    # data.json above) -------------------------------------------------
    # Wrapped so ANY failure here — Yahoo down, Investing.com down, a
    # parsing bug, a corrupt state/events.json — can NEVER stop the rest
    # of this poll cycle. Everything below (screener messages, alerts,
    # heartbeat, digest) still runs exactly as before regardless of what
    # happens in this block. state/events.json is created automatically
    # on first run if it doesn't exist yet; on later runs, existing
    # events are enriched or added to, never duplicated or wiped.
    try:
        events_result = collect_and_persist_broker_events(
            watchlist,
            screener_rows=screener.get("volume", []) + screener.get("gainers", []) + screener.get("losers", []),
        )
        if not events_result.get("written"):
            print(f"Broker events: not written this run (reason={events_result.get('reason')})")
    except Exception as e:
        print(f"  ! Broker events collection failed entirely — state/events.json left untouched: {e}", file=sys.stderr)

    screener_last_sent = seen_state.get("lastScreenerMessage", 0)
    screener_interval = SCREENER_MESSAGE_INTERVAL_SECONDS if _current_template() == "callmebot" else 0
    if time.time() - screener_last_sent >= screener_interval:
        screener_messages = format_screener_sections(screener)
        for i, msg in enumerate(screener_messages):
            print(msg)
            send_webhook(msg)
            if i < len(screener_messages) - 1:
                time.sleep(2)  # small gap so rapid-fire messages don't collapse/drop
        if screener_messages:
            seen_state["lastScreenerMessage"] = time.time()
    else:
        print("Screener WhatsApp send throttled (sent within the last hour) — dashboard/data still updated above.")

    screener_news_msg = format_screener_news_message(screener_news)
    if screener_news_msg:
        print(screener_news_msg)
        send_webhook(screener_news_msg)

    uptrend_msg = format_uptrend_message(uptrend_stocks)
    if uptrend_msg:
        print(uptrend_msg)
        send_webhook(uptrend_msg)

    attention_msg = format_attention_message(mention_counts)
    if attention_msg:
        print(attention_msg)
        send_webhook(attention_msg)

    if big_movers:
        lines = [f"🔥 ALREADY MOVING TODAY (±{BIG_MOVER_THRESHOLD_PCT:.0f}%+, watchlist)  🕐 {now_stamp()} — facts about today, not a forecast:"]
        for m in big_movers:
            arrow = "▲" if (m.get("changePct") or 0) >= 0 else "▼"
            lines.append(f'{m["ticker"]} {arrow}{abs(m.get("changePct") or 0):.2f}%')
        movers_msg = "\n".join(lines)
        print(movers_msg)
        send_webhook(movers_msg)

    ALERT_LABELS = {"upgrade": "⬆ UPGRADE", "downgrade": "⬇ DOWNGRADE", "target": "🎯 PRICE TARGET", "target_raise": "🎯⬆ TARGET RAISED", "target_cut": "🎯⬇ TARGET CUT", "initiation": "🆕 INITIATED", "reiteration": "🔁 REITERATED", "director_dealing": "🧑‍💼 DIRECTOR DEALING", "event": "📰 NEWS"}
    alert_cap = MAX_ALERTS_PER_RUN if _current_template() == "callmebot" else len(new_alerts)
    alerts_to_send = new_alerts[:alert_cap]
    skipped_count = len(new_alerts) - len(alerts_to_send)
    for alert in alerts_to_send:
        label = ALERT_LABELS.get(alert["category"], "📰 NEWS")
        broker_tag = f' ({alert["broker"]})' if alert["broker"] else ""
        ticker_bit = "LSE (market-wide)" if alert["ticker"] == "MARKET" else alert["ticker"]
        article_date = alert.get("pubDate") or ""
        msg = f'{label}: {ticker_bit}{broker_tag}\n{alert["title"]}\n{alert["link"]}\n📅 {article_date}  🕐 seen {now_stamp()}'
        print(f"Alert: {msg}")
        send_webhook(msg)
        time.sleep(2)  # CallMeBot can silently drop messages sent in rapid succession
                        # even while still returning 200 — space every send out.
    if skipped_count > 0:
        skip_msg = f"ℹ️ {skipped_count} more alert(s) this run — see the full list on the dashboard (capped to avoid WhatsApp rate limits)."
        print(skip_msg)
        send_webhook(skip_msg)

    print(f"Done. {len(new_alerts)} new alert(s) found, {len(alerts_to_send)} sent to WhatsApp.")
    # Explicit, immediately-flushed overall completion marker — not
    # LSE-specific, but ties off the same chronology gap: confirms the
    # whole poll run reached its natural end, not just that individual
    # steps along the way logged successfully.
    print(f"[{datetime.now(timezone.utc).isoformat()}] POLL RUN COMPLETE", file=sys.stderr, flush=True)

    if new_alerts:
        time.sleep(2)  # gap before heartbeat/digest, same reasoning as between alerts

    heartbeat_sent = maybe_send_heartbeat(seen_state, watchlist, len(new_alerts), len(big_movers))
    digest_sent = maybe_send_ai_digest(seen_state, screener, market_wide_enriched, big_movers, new_alerts)

    # Always persist here (not just when heartbeat/digest fired) — this is the only place
    # lastScreenerMessage's updated value actually gets written to disk.
    save_json(SEEN_FILE, {
        "seen": list(seen)[-MAX_SEEN:],
        "lastHeartbeat": seen_state.get("lastHeartbeat", 0),
        "lastAiDigest": seen_state.get("lastAiDigest", 0),
        "lastScreenerMessage": seen_state.get("lastScreenerMessage", 0),
        "recentSendTimes": _recent_send_times,  # carries the rate-limit budget forward to the next run
    })



# =========================================================================
# BROKER-EVENTS PIPELINE (steps 1-3) — target extraction, ticker
# resolution, normalization, and cross-source deduplication/merging.
#
# SEPARATE, ADDITIVE subsystem — nothing here is called from main(), the
# dashboard, or the existing news/alert pipeline yet. Pure functions
# operating on in-memory lists only; no filesystem I/O. This is what makes
# idempotency achievable once a persistent store is added in a later step:
# calling this pipeline twice on the same raw input always produces the
# same output, since every ID is deterministic (hash-based or built from
# the data itself, never a random/incrementing value or a timestamp of
# when the pipeline ran).
#
# Pipeline shape, exactly as specified:
#   raw_events -> normalized_events -> candidate_matches -> merged_events
# Nothing is ever discarded: every input event ends up in exactly one of
# merged_events / conflicts / unmatched_events in the final output.
# =========================================================================

import hashlib

# -------------------------------------------------------------------
# STEP 1 — Investing.com target-price extraction
# -------------------------------------------------------------------
# Only extracts a target when the headline gives EXPLICIT numbers with an
# unambiguous currency marker (£/GBP prefix, or p/pence suffix). A bare
# number with neither marker is never guessed at — real UK financial
# headlines always mark pence explicitly ("450p") or pounds explicitly
# ("£4.50"); a marker-less number is genuinely ambiguous and is left as
# no-extraction rather than assumed.

_MONEY_TOKEN_RE = r"(?:GBP\s*)?(?:£\s*(\d{1,3}(?:,\d{3})*(?:\.\d+)?)|(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*p(?:ence)?\b)"
# group(1) = pounds-marked number (£ prefix), group(2) = pence-marked number (p suffix)
# — exactly one is populated per match, since the alternation is exclusive.


def _parse_money_match(m):
    """Given a regex match against _MONEY_TOKEN_RE, returns (value_in_pounds, currency)."""
    pounds_str, pence_str = m.group(1), m.group(2)
    if pounds_str is not None:
        return float(pounds_str.replace(",", "")), "GBP"
    if pence_str is not None:
        return float(pence_str.replace(",", "")) / 100.0, "GBP"
    return None, None  # should not happen given the alternation, but never guess


def extract_target_from_headline(title):
    """
    Returns (old_target, new_target, currency) — all None if no explicit,
    unambiguous target-price change is stated in the headline. old_target
    may be None even when new_target is found (e.g. "cuts price target to
    £3.80" states only the new value) — that's a valid partial extraction,
    not a failure; old_target simply stays unknown, never invented.

    Only searches within a window around the word "target" (or "price
    target"/"pt") — avoids accidentally grabbing an unrelated number
    elsewhere in a long headline (e.g. a percentage, a share count, a date).
    """
    t = title.lower()
    target_idx = None
    for kw in ("price target", "target price", " pt ", "target"):
        idx = t.find(kw)
        if idx != -1:
            target_idx = idx
            break
    if target_idx is None:
        return None, None, None

    window_start = max(0, target_idx - 15)
    window_end = min(len(title), target_idx + 70)
    window = title[window_start:window_end]

    money_re = re.compile(_MONEY_TOKEN_RE, re.IGNORECASE)
    matches = list(money_re.finditer(window))
    if not matches:
        return None, None, None

    if len(matches) == 1:
        value, currency = _parse_money_match(matches[0])
        return None, value, currency

    if len(matches) == 2:
        v1, c1 = _parse_money_match(matches[0])
        v2, c2 = _parse_money_match(matches[1])
        if v1 is None or v2 is None:
            return None, None, None
        currency = c1 or c2
        # "from X to Y" => old=X, new=Y. Anything else ("to Y from X",
        # "raises ... to Y from X") => the token right after "from" is old,
        # the other is new. Checked via the word immediately preceding the
        # FIRST token in reading order.
        before_first = window[:matches[0].start()].strip().split()
        word_before_first = before_first[-1].lower() if before_first else ""
        if word_before_first == "from":
            return v1, v2, currency
        return v2, v1, currency

    # 3+ money-like tokens in the window is ambiguous — a malformed or
    # unusually-structured headline. Never guess which pair is the target.
    return None, None, None


def compute_target_change_pct(old_target, new_target):
    """(new - old) / old * 100 — only when both values are present and
    old_target is non-zero. Never estimated otherwise."""
    if old_target is None or new_target is None or old_target == 0:
        return None
    return (new_target - old_target) / old_target * 100.0


# -------------------------------------------------------------------
# STEP 2 — Name -> LSE ticker resolution
# -------------------------------------------------------------------

def build_name_ticker_lookup(watchlist, screener_rows):
    """
    Builds a company-name -> (ticker, display_name) lookup from data
    ALREADY available in this run (watchlist + screener pool) — never
    invented, never fetched from a separate source. Keys are lowercased
    and cleaned via the existing clean_company_name() (strips "ORD
    10P"-style jargon), so lookups tolerate the same naming noise Yahoo's
    screener returns. display_name is the SAME cleaned name, kept
    alongside the ticker so a successful match can populate the event's
    "company" field from the identical source that resolved the ticker —
    never a separately-invented value.
    """
    lookup = {}
    for stock in watchlist:
        cleaned = clean_company_name(stock["name"]).strip()
        key = cleaned.lower()
        if key:
            lookup[key] = (stock["ticker"].upper(), cleaned)
    for row in screener_rows:
        cleaned = clean_company_name(row.get("name", "")).strip()
        key = cleaned.lower()
        symbol = row.get("symbol", "")
        ticker = symbol.upper().rsplit(".L", 1)[0] if symbol.upper().endswith(".L") else symbol.upper()
        if key and ticker:
            lookup.setdefault(key, (ticker, cleaned))  # watchlist takes priority if both define the same name
    return lookup


def resolve_ticker_by_substring(headline, lookup, exclude_name=None):
    """
    The Investing.com RSS has no ticker field, only a headline — so this
    checks whether any KNOWN company name (from the lookup, i.e. from this
    run's own watchlist/screener) appears as a substring of the headline.
    A company outside that pool genuinely can't be resolved from data this
    run has, and stays (None, None) rather than guessed.

    Returns (ticker, company_name) — company_name comes from the SAME
    matched lookup entry that produced the ticker, never invented
    separately.

    exclude_name: typically the broker ALREADY detected in this same
    headline. Any lookup entry whose cleaned name equals it is skipped as
    a match candidate — fixes a confirmed real bug: "Barclays cuts
    Pearson rating" was resolving ticker=BARC, because "Barclays" is both
    the ACTING BROKER in this headline and a company on the watchlist.
    The broker performing an action is never the SUBJECT of that action.
    """
    if not headline:
        return None, None
    t_lower = headline.lower()
    exclude_lower = exclude_name.strip().lower() if exclude_name else None
    for name, (ticker, display_name) in lookup.items():
        if not name or name not in t_lower:
            continue
        if exclude_lower and name == exclude_lower:
            continue
        return ticker, display_name
    return None, None


# -------------------------------------------------------------------
# STEP 3 — Normalization + cross-source deduplication/merging
# -------------------------------------------------------------------

CONFIDENCE_SINGLE_STRUCTURED = "SINGLE_SOURCE_STRUCTURED"
CONFIDENCE_SINGLE_PARSED = "SINGLE_SOURCE_PARSED"
CONFIDENCE_MERGED_HIGH = "MERGED_HIGH"
CONFIDENCE_MERGED_PARTIAL = "MERGED_PARTIAL"
CONFIDENCE_CONFLICT = "CONFLICT"


def make_source_event_id(source, **fields):
    """Deterministic per-source ID — identical inputs always produce the
    identical ID, which is what makes re-polling idempotent at this layer."""
    raw = source + "|" + "|".join(f"{k}={fields[k]}" for k in sorted(fields))
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:16]


def make_canonical_key(ticker, broker_canonical, date_str):
    """
    Cross-source canonical key — identifies the underlying broker EVENT
    (this broker, this company, this day), never the amount of
    information currently known about it. Deliberately excludes rating
    and target values entirely: those are DATA on the event, discovered
    incrementally as sources report them, not part of the event's
    identity. An event first seen via Yahoo (rating only) and later
    enriched with a target from Investing.com must keep the exact same
    ID throughout — baking rating/target into the key would make that
    impossible, which was the bug in the original design.

    Ticker is mandatory: without it, two events can't be safely confirmed
    as the same company, so no canonical key is produced.
    """
    if not ticker:
        return None
    return f"LSE|{ticker.upper()}|{broker_canonical.upper()}|{date_str}"


def make_conflict_key(ticker, broker_canonical, date_str, old_rating, new_rating, old_target, new_target):
    """
    Used ONLY to disambiguate a fragment that actively CONFLICTS with
    whatever is already recorded under the canonical (ticker|broker|date)
    key for that slot. A conflicting fragment must never collide with —
    and must never overwrite — the canonical "confirmed" record, so it
    gets its own distinct, deterministic ID built from the specific
    rating/target values that make it different. Two runs reporting the
    exact same conflicting fragment produce the same conflict key, so
    repeats are still idempotent even in the conflict case.
    """
    base = make_canonical_key(ticker, broker_canonical, date_str)
    if base is None:
        return None
    old_r = (old_rating or "NA").strip().upper()
    new_r = (new_rating or "NA").strip().upper()
    old_t = str(round(old_target * 100)) if old_target is not None else "NA"
    new_t = str(round(new_target * 100)) if new_target is not None else "NA"
    return f"{base}|CONFLICT|{old_r}>{new_r}|{old_t}>{new_t}"


def _canonical_broker(name):
    """Resolves a broker name to its canonical form from the existing
    BROKER_NAMES list (e.g. so 'Barclays' from two sources always matches)."""
    if not name:
        return None
    for b in BROKER_NAMES:
        if b.lower() == name.lower() or b.lower() in name.lower():
            return b
    return name


def normalize_yahoo_event(ticker, company, firm, from_grade, to_grade, action_code, epoch, link):
    """
    One Yahoo upgradeDowngradeHistory entry -> a normalized event. Always
    SINGLE_SOURCE_STRUCTURED confidence — Yahoo never carries target
    history (verified in the forensic pass), so old_target/new_target are
    always None here; a later merge with an Investing.com event may fill
    them in.
    """
    dt_london = datetime.fromtimestamp(epoch, tz=timezone.utc).astimezone(LONDON_TZ)
    date_str = dt_london.strftime("%Y-%m-%d")
    action = normalize_action_from_grades(from_grade, to_grade, action_code)
    return {
        "source": "yahoo",
        "source_event_id": make_source_event_id("yahoo", ticker=ticker, firm=firm, epoch=epoch),
        "source_url": link,
        "timestamp": dt_london.astimezone(timezone.utc).isoformat(),
        "date": date_str,
        "company": company,
        "ticker": ticker.upper(),
        "exchange": "LSE",
        "broker": _canonical_broker(firm),
        "action": action,
        "old_rating": from_grade or None,
        "new_rating": to_grade,
        "old_rating_bucket": normalize_rating_bucket(from_grade),
        "new_rating_bucket": normalize_rating_bucket(to_grade),
        "old_target": None,
        "new_target": None,
        "target_currency": None,
        "target_change_pct": None,
        "confidence": CONFIDENCE_SINGLE_STRUCTURED,
    }


def normalize_investing_event(title, link, pub_date, ticker_lookup):
    """
    One Investing.com RSS item -> a normalized event. Broker/action come
    from the existing, already-tested classify()/detect_broker()/
    normalize_headline_action(). Target comes from Step 1's regex
    extractor. Ticker AND company come together from the SAME resolved
    lookup match (Step 2), with the detected broker excluded as a
    candidate so the broker can never be mistaken for the subject
    company. new_rating comes from the narrow explicit-destination-rating
    extractor — only populated when the headline states one outright.
    """
    category = classify(title)
    broker = detect_broker(title)
    old_target, new_target, currency = extract_target_from_headline(title)
    new_rating_text = extract_new_rating_from_headline(title)
    dt = _parse_pub_date(pub_date)
    if dt is None:
        date_str, timestamp = None, None
    else:
        dt_london = dt.astimezone(LONDON_TZ)
        date_str = dt_london.strftime("%Y-%m-%d")
        timestamp = dt.astimezone(timezone.utc).isoformat()

    ticker, matched_company = resolve_ticker_by_substring(title, ticker_lookup, exclude_name=broker)

    return {
        "source": "investing_com",
        "source_event_id": make_source_event_id("investing_com", link=link),
        "source_url": link,
        "timestamp": timestamp,
        "date": date_str,
        "company": matched_company,  # from the SAME match that resolved ticker — never invented separately
        "ticker": ticker,  # may be None — genuinely unresolvable for out-of-pool companies
        "exchange": "LSE" if ticker else None,
        "broker": _canonical_broker(broker) if broker else None,
        "action": normalize_headline_action(title) or category.upper(),
        "old_rating": None,  # headlines don't reliably state an explicit OLD rating structurally
        "new_rating": new_rating_text,  # only when explicitly stated, e.g. "upgrade to Neutral"
        "old_rating_bucket": None,
        "new_rating_bucket": normalize_rating_bucket(new_rating_text) if new_rating_text else None,
        "old_target": old_target,
        "new_target": new_target,
        "target_currency": currency,
        "target_change_pct": compute_target_change_pct(old_target, new_target),
        "confidence": CONFIDENCE_SINGLE_PARSED,
    }


# Actions that structurally imply pre-existing broker coverage (the source
# data explicitly stated an old rating, or the action is a change from one).
# INITIATION structurally means the OPPOSITE: no prior coverage existed.
# The two stories are semantically contradictory for the same broker/
# company/day even when the specific rating TEXT doesn't happen to overlap
# (e.g. an initiation fragment with no target vs an upgrade fragment with
# no target — same-day fields never literally disagree, but the underlying
# claims still can't both be true).
_PRIOR_COVERAGE_ACTIONS = {"UPGRADE", "DOWNGRADE", "REITERATION", "RATING_CHANGE"}


def _action_types_conflict(action1, action2):
    """True if the two actions tell mutually exclusive stories about
    whether prior coverage existed, regardless of what the rating/target
    fields themselves say."""
    if action1 == "INITIATION" and action2 in _PRIOR_COVERAGE_ACTIONS:
        return True
    if action2 == "INITIATION" and action1 in _PRIOR_COVERAGE_ACTIONS:
        return True
    return False


def compute_evidence_fingerprint(e):
    """
    PURELY DESCRIPTIVE metadata — a snapshot of what's currently known
    about an event, stored on the record for human/debugging visibility
    only. Deliberately NEVER used for identity (event_id) or for deciding
    whether two fragments should merge — a literal fingerprint-equality
    matching rule was considered and rejected, since two genuinely
    complementary fragments of the SAME real event (e.g. Yahoo's
    rating-only report and Investing.com's target-only report) produce
    different fingerprints by design, and requiring equality would have
    prevented exactly the enrichment this pipeline exists to do.
    """
    action = e.get("action") or "NA"
    old_r = (e.get("old_rating") or "NA").upper()
    new_r = (e.get("new_rating") or "NA").upper()
    old_t = f"{e['old_target']:.2f}" if e.get("old_target") is not None else "NA"
    new_t = f"{e['new_target']:.2f}" if e.get("new_target") is not None else "NA"
    return f"{action}|{old_r}>{new_r}|{old_t}>{new_t}"


def _events_conflict(e1, e2):
    """
    True if two same-ticker/same-broker/same-day events actively
    contradict each other and must NOT be merged. Compatible (mergeable):
    one side has no rating info and the other does (a target-only fragment
    naturally complementing a rating-only fragment); one has no target and
    the other does. Conflicting: both state a rating transition and they
    differ; both state a target value and it differs; OR the two actions
    structurally imply mutually exclusive stories (initiation vs. a
    change to pre-existing coverage) even when the rating/target fields
    themselves don't literally overlap — see _action_types_conflict.
    """
    if _action_types_conflict(e1.get("action"), e2.get("action")):
        return True
    if e1["new_rating"] and e2["new_rating"]:
        if (e1.get("old_rating") or "").lower() != (e2.get("old_rating") or "").lower():
            return True
        if e1["new_rating"].lower() != e2["new_rating"].lower():
            return True
    if e1["new_target"] is not None and e2["new_target"] is not None:
        if round(e1["new_target"], 2) != round(e2["new_target"], 2):
            return True
    if e1["old_target"] is not None and e2["old_target"] is not None:
        if round(e1["old_target"], 2) != round(e2["old_target"], 2):
            return True
    return False


def _merge_pair(e1, e2, canonical_key):
    """Combines two compatible (non-conflicting) same-event records.
    Never overwrites — fills gaps, always keeps both sources' IDs/URLs."""
    def pick(a, b):
        return a if a is not None else b
    timestamps = [t for t in (e1.get("timestamp"), e2.get("timestamp")) if t]
    merged = {
        "event_id": canonical_key,
        "timestamp": min(timestamps) if timestamps else None,
        "date": e1.get("date") or e2.get("date"),
        "company": pick(e1.get("company"), e2.get("company")),
        "ticker": e1.get("ticker") or e2.get("ticker"),
        "exchange": "LSE",
        "broker": e1.get("broker") or e2.get("broker"),
        "action": e1.get("action") if e1.get("action") not in (None, "NO_CHANGE") else e2.get("action"),
        "old_rating": pick(e1.get("old_rating"), e2.get("old_rating")),
        "new_rating": pick(e1.get("new_rating"), e2.get("new_rating")),
        "old_rating_bucket": pick(e1.get("old_rating_bucket"), e2.get("old_rating_bucket")),
        "new_rating_bucket": pick(e1.get("new_rating_bucket"), e2.get("new_rating_bucket")),
        "old_target": pick(e1.get("old_target"), e2.get("old_target")),
        "new_target": pick(e1.get("new_target"), e2.get("new_target")),
        "target_currency": pick(e1.get("target_currency"), e2.get("target_currency")),
        "source": sorted(set([e1["source"], e2["source"]])),
        "source_url": [e1["source_url"], e2["source_url"]],
        "source_event_ids": [e1["source_event_id"], e2["source_event_id"]],
        # source_refs: correctly-PAIRED (source, source_event_id) entries —
        # NOTE "source" above is sorted(set(...)), independently of
        # source_url/source_event_ids' e1-then-e2 order, so those three
        # lists are NOT safely zippable back together. source_refs exists
        # specifically to give supersession-matching (added in a later
        # stage) an unambiguous pairing to rely on, without touching or
        # reordering any of the three existing fields above.
        "source_refs": [
            {"source": e1["source"], "source_event_id": e1["source_event_id"]},
            {"source": e2["source"], "source_event_id": e2["source_event_id"]},
        ],
        "confidence": CONFIDENCE_MERGED_HIGH,
    }
    merged["target_change_pct"] = compute_target_change_pct(merged["old_target"], merged["new_target"])
    both_sides_gave_target = e1.get("new_target") is not None and e2.get("new_target") is not None
    if not both_sides_gave_target:
        # Only one side contributed target data — genuinely a partial
        # merge, not the strong two-source agreement MERGED_HIGH implies.
        merged["confidence"] = CONFIDENCE_MERGED_PARTIAL
    merged["evidence_fingerprint"] = compute_evidence_fingerprint(merged)  # descriptive only, see docstring
    return merged


def finalize_unmatched_event(e, is_conflict_side=False):
    """
    Gives a normalized event its own deterministic event_id.

    is_conflict_side=False (the default — a genuinely standalone
    single-source event, no conflicting sibling encountered): uses the
    clean canonical key (ticker|broker|date) when all three are known, so
    a later compatible fragment from another source can find and enrich
    this exact record. Falls back to the source-specific ID when ticker/
    broker/date aren't all known (never invented, never guessed).

    is_conflict_side=True (this event actively conflicts with another
    fragment sharing the same ticker/broker/day): uses make_conflict_key
    instead — deliberately NOT the clean canonical key, so a disputed
    fragment can never collide with or silently overwrite the confirmed
    canonical record for that slot.
    """
    canonical = None
    if e.get("ticker") and e.get("broker") and e.get("date"):
        if is_conflict_side:
            canonical = make_conflict_key(
                e["ticker"], e["broker"], e["date"],
                e.get("old_rating"), e.get("new_rating"),
                e.get("old_target"), e.get("new_target"),
            )
        else:
            canonical = make_canonical_key(e["ticker"], e["broker"], e["date"])
    out = dict(e)
    out["event_id"] = canonical or e["source_event_id"]
    out["source"] = [e["source"]]
    out["source_url"] = [e["source_url"]]
    out["source_event_ids"] = [e["source_event_id"]]
    out["source_refs"] = [{"source": e["source"], "source_event_id": e["source_event_id"]}]
    out["evidence_fingerprint"] = compute_evidence_fingerprint(out)  # descriptive only, see docstring
    return out


def _fold_compatible_fragment(acc, fragment):
    """
    Folds a raw normalized fragment into an accumulator that may itself be
    either a raw fragment (first fold in a cluster) or an already-folded
    (list-shaped source/source_url/source_event_ids) partial result from
    earlier folds. Used ONLY inside deduplicate_and_merge's WITHIN-A-SINGLE-
    RUN clustering — never for cross-run store reconciliation, which is
    enrich_stored_event's job. Caller guarantees `fragment` is compatible
    with `acc` (i.e. _events_conflict already checked) before calling this.
    Returns the updated accumulator, always in finalized (list-shaped)
    form, ready for another fold or final output.
    """
    def pick(a, b):
        return a if a is not None else b

    acc_source = acc.get("source")
    acc_is_finalized = isinstance(acc_source, list)
    acc_sources = list(acc_source) if acc_is_finalized else [acc_source]
    acc_source_urls = list(acc.get("source_url")) if acc_is_finalized else [acc.get("source_url")]
    acc_source_ids = list(acc.get("source_event_ids")) if acc_is_finalized else [acc.get("source_event_id")]
    acc_source_refs = list(acc.get("source_refs")) if acc_is_finalized else [{"source": acc.get("source"), "source_event_id": acc.get("source_event_id")}]
    acc_had_own_target = acc.get("new_target") is not None  # BEFORE this fold — needed for the confidence rule below

    merged = dict(acc)
    merged["old_rating"] = pick(acc.get("old_rating"), fragment.get("old_rating"))
    merged["new_rating"] = pick(acc.get("new_rating"), fragment.get("new_rating"))
    merged["old_rating_bucket"] = pick(acc.get("old_rating_bucket"), fragment.get("old_rating_bucket"))
    merged["new_rating_bucket"] = pick(acc.get("new_rating_bucket"), fragment.get("new_rating_bucket"))
    merged["old_target"] = pick(acc.get("old_target"), fragment.get("old_target"))
    merged["new_target"] = pick(acc.get("new_target"), fragment.get("new_target"))
    merged["target_currency"] = pick(acc.get("target_currency"), fragment.get("target_currency"))
    merged["target_change_pct"] = compute_target_change_pct(merged["old_target"], merged["new_target"])
    merged["ticker"] = acc.get("ticker") or fragment.get("ticker")
    merged["broker"] = acc.get("broker") or fragment.get("broker")
    merged["date"] = acc.get("date") or fragment.get("date")
    merged["company"] = pick(acc.get("company"), fragment.get("company"))
    if not merged.get("action") or merged.get("action") == "NO_CHANGE":
        merged["action"] = fragment.get("action")

    if fragment["source"] not in acc_sources:
        acc_sources.append(fragment["source"])
    merged["source"] = sorted(set(acc_sources))
    merged["source_url"] = acc_source_urls + [fragment["source_url"]]
    merged["source_event_ids"] = acc_source_ids + [fragment["source_event_id"]]
    merged["source_refs"] = acc_source_refs + [{"source": fragment["source"], "source_event_id": fragment["source_event_id"]}]

    # MERGED_HIGH requires genuine independent agreement — the accumulator
    # already had its OWN target before this fold, AND this fragment
    # brings its own target too (they must have agreed, since a conflict
    # would already have excluded this fragment from the cluster). If only
    # one side ever contributed a target, it's MERGED_PARTIAL — filled a
    # gap, not independently corroborated. Same distinction enrich_stored_event
    # uses, kept consistent deliberately.
    fragment_had_own_target = fragment.get("new_target") is not None
    if len(merged["source"]) > 1:
        merged["confidence"] = CONFIDENCE_MERGED_HIGH if (acc_had_own_target and fragment_had_own_target) else CONFIDENCE_MERGED_PARTIAL

    timestamps = [t for t in (acc.get("timestamp"), fragment.get("timestamp")) if t]
    merged["timestamp"] = min(timestamps) if timestamps else None
    return merged


def deduplicate_and_merge(normalized_events):
    """
    normalized_events -> candidate_matches -> merged_events, per the
    required pipeline shape. Returns (merged_events, conflicts,
    unmatched_events) — every input event ends up in exactly one of these
    three, never silently dropped.

    Matching evidence, exactly as specified: ticker + canonical broker +
    calendar day are REQUIRED to even consider two events a candidate
    match — same ticker+broker+day alone is NOT assumed to mean the same
    event; the actual rating/target/action transitions are then checked
    for compatibility (_events_conflict) before merging.

    CLUSTERING IS ORDER-INDEPENDENT BY DESIGN. An earlier version merged
    fragments greedily in arrival order (first compatible pair found), which
    was proven — by direct test — to produce DIFFERENT results depending on
    the input list's order when 3+ fragments share a slot and compatibility
    isn't transitive (fragment X compatible with Y, Y compatible with Z, but
    X actively conflicts with Z — entirely possible when X/Z each state a
    DIFFERENT rating and Y states none). In one observed ordering, greedy
    pairing even produced two DIFFERENT merged records under the IDENTICAL
    event_id — a genuine collision. Fixed here by computing the FULL
    pairwise compatibility matrix for the group first: a fragment only
    joins the shared "confirmed" cluster if it's compatible with EVERY
    other fragment in that cluster, not just its first-encountered match.
    Any fragment that conflicts with at least one other group member is
    excluded from the cluster entirely and recorded as its own separate
    conflict-side record instead — this is deterministic given the same
    input SET, regardless of the order that set is provided in.
    """
    groups = {}
    unmatched = []
    for e in normalized_events:
        if not (e.get("ticker") and e.get("broker") and e.get("date")):
            unmatched.append(finalize_unmatched_event(e))
            continue
        key = (e["ticker"], e["broker"], e["date"])
        groups.setdefault(key, []).append(e)

    merged_events = []
    conflicts = []
    for key, group in groups.items():
        n = len(group)
        if n == 1:
            unmatched.append(finalize_unmatched_event(group[0]))
            continue

        # Full pairwise compatibility matrix — computed once, independent
        # of any arrival order, since it only depends on the SET of
        # fragments present in this group.
        conflicted_indices = set()
        for i in range(n):
            for j in range(i + 1, n):
                if _events_conflict(group[i], group[j]):
                    conflicts.append({
                        "reason": "rating_or_target_mismatch",
                        "ticker": key[0], "broker": key[1], "date": key[2],
                        "events": [group[i], group[j]],
                    })
                    conflicted_indices.add(i)
                    conflicted_indices.add(j)

        # A fragment qualifies for the ONE shared "confirmed" cluster only
        # if it never conflicted with ANY other group member — not merely
        # its first-encountered pairing. This is what makes the result
        # independent of input order: the qualifying set is the same
        # regardless of which order fragments were compared in.
        cluster_indices = [i for i in range(n) if i not in conflicted_indices]

        if len(cluster_indices) >= 2:
            acc = group[cluster_indices[0]]
            for idx in cluster_indices[1:]:
                acc = _fold_compatible_fragment(acc, group[idx])
            canonical_key = make_canonical_key(acc["ticker"], acc["broker"], acc["date"])
            acc = dict(acc)
            acc["event_id"] = canonical_key
            acc["evidence_fingerprint"] = compute_evidence_fingerprint(acc)
            merged_events.append(acc)
        elif len(cluster_indices) == 1:
            unmatched.append(finalize_unmatched_event(group[cluster_indices[0]]))
        # else (0 qualify): every fragment in this group conflicted with at
        # least one other — nothing to merge; all fall through to the
        # conflict-side finalization below.

        # Every fragment that conflicted with at least one other group
        # member is finalized individually (conflict-suffixed ID) — once
        # each, even if it was part of MULTIPLE conflicting pairs.
        for i in conflicted_indices:
            unmatched.append(finalize_unmatched_event(group[i], is_conflict_side=True))

    return merged_events, conflicts, unmatched




# =========================================================================
# PERSISTENT EVENT STORE — state/events.json
#
# SEPARATE from state/data.json entirely — different file, different
# functions, never touches the existing data.json load/save calls. This
# layer is additive: nothing here is called from main() yet (verified by
# the same "not wired in" check used for the pipeline itself).
#
# Design principles, per spec:
# - Append-only: existing entries are NEVER modified or removed by normal
#   operation. Only explicit user action (not built here) would ever
#   remove an event.
# - Idempotent: running the exact same poll cycle twice produces byte-
#   identical file content the second time (0 events added, same sorted
#   order) — this is verified directly in the tests below.
# - Atomic writes: temp file + os.replace(), so a crash or interruption
#   mid-write can never leave a truncated/corrupt events.json in place —
#   either the old file survives intact, or the new one lands complete.
# - A CORRUPT existing file is never silently overwritten. If the file
#   can't be parsed as the expected {version, events} shape, persistence
#   is skipped for that cycle entirely and the corrupt file is left
#   exactly as found — overwriting it with a "fresh" store would destroy
#   whatever real history it might still contain in partially-recoverable
#   form, which is a worse outcome than just skipping a cycle.
# =========================================================================

import tempfile

EVENTS_STATE_FILE = os.path.join(STATE_DIR, "events.json")
EVENTS_STORE_VERSION = 1


# --- Daily snapshots (Phase 6: "What Changed") --------------------------
# Same file/directory (state/), same reliability the events store has
# already proven across many real GitHub Actions runs (broker momentum's
# 90-day rolling history could not work at all if state/ weren't
# persisting reliably between runs — this is direct, already-verified
# evidence, not an assumption).
DAILY_SNAPSHOTS_FILE = os.path.join(STATE_DIR, "daily_snapshots.json")
DAILY_SNAPSHOTS_VERSION = 1
DAILY_SNAPSHOTS_RETENTION_DAYS = 30
WHATS_CHANGED_PRICE_THRESHOLD_PCT = 5.0
# Every field this feature actually reads from a prior-snapshot stock
# record — used to reject a partially-written or older-format record
# before it can produce a misleading delta. Values MAY legitimately be
# None (missing data at capture time); the KEYS must all be present.
REQUIRED_SNAPSHOT_STOCK_FIELDS = {"ticker", "price", "total", "signalQuality", "evidenceLabel"}


class DailySnapshotsCorruptError(Exception):
    """Same discipline as EventsStoreCorruptError: a file that EXISTS but
    fails to parse or doesn't match the expected shape must never be
    silently treated as empty — that would destroy real history. Only a
    genuinely MISSING file is a legitimate fresh-start case."""
    pass


def load_daily_snapshots(path=None):
    """Loads the daily snapshots store. Mirrors load_events_store exactly:
    a MISSING file is the only case that legitimately produces a fresh
    empty store; a file that EXISTS but fails to parse or match the
    expected shape raises DailySnapshotsCorruptError instead of silently
    returning empty."""
    path = path or DAILY_SNAPSHOTS_FILE
    if not os.path.exists(path):
        return {"version": DAILY_SNAPSHOTS_VERSION, "snapshots": []}
    with open(path, "r", encoding="utf-8") as f:
        raw = f.read()
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        raise DailySnapshotsCorruptError(f"{path} contains invalid JSON: {e}") from e
    if not isinstance(data, dict) or not isinstance(data.get("snapshots"), list):
        raise DailySnapshotsCorruptError(f"{path} does not match the expected {{version, snapshots}} shape")
    return data


def is_valid_snapshot_stock_entry(entry):
    """A prior-snapshot stock record is only used for comparison when it
    has every field this feature actually reads — guards against a
    partially-written or older-format record silently producing a
    misleading or wrong delta."""
    return isinstance(entry, dict) and REQUIRED_SNAPSHOT_STOCK_FIELDS.issubset(entry.keys())


def find_prior_snapshot(store, before_date):
    """
    Returns the most recent snapshot with a date STRICTLY BEFORE
    before_date (today's London date), or None if none exists yet.
    Never assumes "yesterday" specifically — this is what correctly
    skips weekends/bank holidays by construction: whatever the latest
    dated entry before today actually is, that's what gets returned and
    later labelled with its REAL date.
    """
    candidates = [s for s in store.get("snapshots", []) if s.get("date") and s["date"] < before_date]
    if not candidates:
        return None
    return max(candidates, key=lambda s: s["date"])


def capture_daily_snapshot(scorecard_summaries, now=None, path=None):
    """
    Appends today's snapshot, built entirely from scorecard_summaries —
    the SAME per-stock data render_dashboard already produces as a side
    effect of its existing scorecard/signal-quality/evidence computation
    (see that computation's own comments) — never a new calculation.

    Deduplicated by London calendar date: if today's date already has a
    snapshot (a later run on the same day), this is a no-op. On a
    corrupt existing store, this deliberately does NOT write — same
    fail-safe-by-skipping discipline as collect_and_persist_broker_events,
    so a parsing bug can never silently wipe real history. Retention:
    keeps at most DAILY_SNAPSHOTS_RETENTION_DAYS most recent snapshots.
    """
    path = path or DAILY_SNAPSHOTS_FILE
    now = now or datetime.now(timezone.utc)
    today_london = now.astimezone(LONDON_TZ).strftime("%Y-%m-%d")

    try:
        store = load_daily_snapshots(path)
    except DailySnapshotsCorruptError as e:
        print(f"  ! daily snapshots store CORRUPT — skipping persistence this cycle to avoid data loss: {e}", file=sys.stderr)
        return {"written": False, "reason": "existing_store_corrupt"}

    if any(s.get("date") == today_london for s in store["snapshots"]):
        return {"written": False, "reason": "already_captured_today"}

    stocks = {}
    for s in scorecard_summaries:
        ticker = s.get("ticker")
        if not ticker:
            continue
        stocks[ticker] = {
            "ticker": ticker, "price": s.get("price"), "total": s.get("total"),
            "signalQuality": s.get("signalQuality"), "evidenceLabel": s.get("evidenceLabel"),
        }

    store["snapshots"].append({
        "date": today_london,
        "capturedAt": now.astimezone(timezone.utc).isoformat(),
        "stocks": stocks,
    })
    store["snapshots"].sort(key=lambda s: s["date"])
    store["snapshots"] = store["snapshots"][-DAILY_SNAPSHOTS_RETENTION_DAYS:]

    try:
        atomic_write_json(path, store)
    except Exception as e:
        print(f"  ! daily snapshots write FAILED — previous file left untouched: {e}", file=sys.stderr)
        return {"written": False, "reason": "write_failed"}

    return {"written": True, "date": today_london, "stockCount": len(stocks)}


def compute_whats_changed(scorecard_summaries, prior_snapshot):
    """
    Returns a list of per-stock changes for stocks with a MATERIAL
    change — a price move past WHATS_CHANGED_PRICE_THRESHOLD_PCT, OR a
    genuine Signal Quality or Evidence label CATEGORY change (always
    surfaced regardless of price magnitude — a category changing is
    inherently notable, not a matter of degree). Purely factual deltas;
    never implies a change is good or bad.

    Returns [] both when there's no prior snapshot at all AND when
    nothing material changed — the caller distinguishes those two cases
    using prior_snapshot itself (None vs not-None), not this return
    value.
    """
    if prior_snapshot is None:
        return []
    prior_stocks = prior_snapshot.get("stocks", {})
    changes = []
    for s in scorecard_summaries:
        ticker = s.get("ticker")
        prior = prior_stocks.get(ticker)
        if not is_valid_snapshot_stock_entry(prior):
            continue  # no valid prior data for this specific stock — skip it, never guess
        price_change_pct = None
        if prior.get("price") and s.get("price") is not None:
            price_change_pct = (s["price"] - prior["price"]) / prior["price"] * 100

        signal_changed = prior.get("signalQuality") != s.get("signalQuality")
        evidence_changed = prior.get("evidenceLabel") != s.get("evidenceLabel")
        price_material = price_change_pct is not None and abs(price_change_pct) >= WHATS_CHANGED_PRICE_THRESHOLD_PCT

        if not (price_material or signal_changed or evidence_changed):
            continue

        changes.append({
            "ticker": ticker, "name": s.get("name"),
            "priceChangePct": price_change_pct if price_material else None,
            "totalFrom": prior.get("total"), "totalTo": s.get("total"),
            "signalQualityFrom": prior.get("signalQuality") if signal_changed else None,
            "signalQualityTo": s.get("signalQuality") if signal_changed else None,
            "evidenceLabelFrom": prior.get("evidenceLabel") if evidence_changed else None,
            "evidenceLabelTo": s.get("evidenceLabel") if evidence_changed else None,
        })
    return changes


def format_since_label(prior_snapshot, now=None):
    """'Fri 28 Aug (3 days ago)' — the REAL date of the prior snapshot,
    never a hardcoded 'yesterday' that would be quietly wrong across a
    weekend or bank holiday."""
    now = now or datetime.now(timezone.utc)
    today_london = now.astimezone(LONDON_TZ).date()
    prior_date = datetime.strptime(prior_snapshot["date"], "%Y-%m-%d").date()
    days_ago = (today_london - prior_date).days
    day_word = "day" if days_ago == 1 else "days"
    return f"{prior_date.strftime('%a %d %b')} ({days_ago} {day_word} ago)"


# --- Phase 7B: Prospective evidence-history store -----------------------
# A SEPARATE, additive store from daily_snapshots.json (Phase 6) — that
# store's schema is fixed and working in production for the "What
# Changed" feature; this is a NEW store with a RICHER schema (the full
# 8-dimension breakdown, both subtotals, RISK, DON'T CHASE state) built
# specifically so a genuinely honest backtest of the News/Broker/AI
# Evidence dimensions becomes possible in the future — built from what
# ACTUALLY happened, going forward from whenever this is deployed, never
# reconstructed or guessed for any date before that. Same file/directory
# reliability already proven by events.json and daily_snapshots.json
# across many real GitHub Actions runs.
EVIDENCE_HISTORY_FILE = os.path.join(STATE_DIR, "evidence_history.json")
EVIDENCE_HISTORY_VERSION = 1
EVIDENCE_HISTORY_RETENTION_DAYS = 730  # ~2 years — deliberately much longer
# than daily_snapshots.json's 30 days, since THAT store exists only to
# support a short "what changed recently" comparison, while THIS store
# exists specifically to accumulate enough history for a future backtest.


class EvidenceHistoryCorruptError(Exception):
    """Same discipline as EventsStoreCorruptError/DailySnapshotsCorruptError:
    a file that EXISTS but fails to parse or match the expected shape must
    never be silently treated as empty — only a genuinely MISSING file is
    a legitimate fresh-start case."""
    pass


def load_evidence_history(path=None):
    """Mirrors load_daily_snapshots/load_events_store exactly: a MISSING
    file is the only case that legitimately produces a fresh empty store;
    a file that EXISTS but fails to parse or match the expected shape
    raises EvidenceHistoryCorruptError instead of silently returning
    empty."""
    path = path or EVIDENCE_HISTORY_FILE
    if not os.path.exists(path):
        return {"version": EVIDENCE_HISTORY_VERSION, "days": []}
    with open(path, "r", encoding="utf-8") as f:
        raw = f.read()
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        raise EvidenceHistoryCorruptError(f"{path} contains invalid JSON: {e}") from e
    if not isinstance(data, dict) or not isinstance(data.get("days"), list):
        raise EvidenceHistoryCorruptError(f"{path} does not match the expected {{version, days}} shape")
    return data


def capture_daily_evidence_snapshot(evidence_summaries, now=None, path=None):
    """
    Appends today's FULL scorecard state for every watchlist AND
    Radar-Stocks-discovered stock — all 8 dimensions, both subtotals,
    RISK, Signal Quality, Confidence, Evidence label, DON'T CHASE
    state, and discoveredVia (which Radar Stocks sources found this
    stock — Watchlist, Heat Map, LSE Volume/Gainers/Losers, or a
    combination) — built entirely from evidence_summaries, the SAME
    per-stock data render_dashboard already produces as a side effect
    of its existing scorecard computation (see that computation's own
    comments) — never a new calculation.

    Deduplicated by London calendar date, same discipline as
    capture_daily_snapshot (Phase 6): a later run on the same day is a
    no-op. On a corrupt existing store, this deliberately does NOT
    write — same fail-safe-by-skipping discipline as every other
    persistence path in this project. Retention: keeps at most
    EVIDENCE_HISTORY_RETENTION_DAYS most recent days.
    """
    path = path or EVIDENCE_HISTORY_FILE
    now = now or datetime.now(timezone.utc)
    today_london = now.astimezone(LONDON_TZ).strftime("%Y-%m-%d")

    try:
        store = load_evidence_history(path)
    except EvidenceHistoryCorruptError as e:
        print(f"  ! evidence history store CORRUPT — skipping persistence this cycle to avoid data loss: {e}", file=sys.stderr)
        return {"written": False, "reason": "existing_store_corrupt"}

    if any(d.get("date") == today_london for d in store["days"]):
        return {"written": False, "reason": "already_captured_today"}

    stocks = {}
    for s in evidence_summaries:
        ticker = s.get("ticker")
        if not ticker:
            continue
        stocks[ticker] = {
            "ticker": ticker, "price": s.get("price"), "total": s.get("total"),
            "signalQuality": s.get("signalQuality"), "confidence": s.get("confidence"),
            "evidenceLabel": s.get("evidenceLabel"), "dimensions": s.get("dimensions"),
            "technicalMarket": s.get("technicalMarket"), "researchEvidence": s.get("researchEvidence"),
            "risk": s.get("risk"), "dontChase": s.get("dontChase"),
            "discoveredVia": s.get("discoveredVia") or [],
        }

    store["days"].append({
        "date": today_london,
        "capturedAt": now.astimezone(timezone.utc).isoformat(),
        "stocks": stocks,
    })
    store["days"].sort(key=lambda d: d["date"])
    store["days"] = store["days"][-EVIDENCE_HISTORY_RETENTION_DAYS:]

    try:
        atomic_write_json(path, store)
    except Exception as e:
        print(f"  ! evidence history write FAILED — previous file left untouched: {e}", file=sys.stderr)
        return {"written": False, "reason": "write_failed"}

    return {"written": True, "date": today_london, "stockCount": len(stocks)}


class EventsStoreCorruptError(Exception):
    """Raised when state/events.json exists but doesn't parse as valid
    JSON, or doesn't match the expected {version, events} shape. The
    caller must NOT write a fresh store in response to this — see the
    module docstring above."""
    pass


def load_events_store(path=None):
    """
    Loads the events store. A MISSING file is the only case that
    legitimately produces a fresh empty store (first run ever) — a file
    that EXISTS but fails to parse raises EventsStoreCorruptError instead
    of silently returning empty, since those are very different
    situations: one has no history yet, the other has history that must
    not be destroyed.
    """
    path = path or EVENTS_STATE_FILE
    if not os.path.exists(path):
        return {"version": EVENTS_STORE_VERSION, "events": []}
    with open(path, "r", encoding="utf-8") as f:
        raw = f.read()
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        raise EventsStoreCorruptError(f"{path} contains invalid JSON: {e}") from e
    if not isinstance(data, dict) or not isinstance(data.get("events"), list):
        raise EventsStoreCorruptError(f"{path} does not match the expected {{version, events}} shape")
    return data


# Not "recent" beyond this — the broker-events pipeline itself keeps full
# history regardless, this is purely about what counts as a "NEW" change
# worth surfacing prominently on the research view specifically.
LATEST_BROKER_EVENT_MAX_AGE_DAYS = 30


def get_latest_broker_event_per_ticker(events, max_age_days=LATEST_BROKER_EVENT_MAX_AGE_DAYS, now_utc=None):
    """
    Given the flat event list from load_events_store()["events"], returns
    a dict: ticker -> the single most recent, NON-SUPERSEDED event for
    that ticker, only if it falls within max_age_days of now_utc.

    Pure function of its inputs — reads the ALREADY-COLLECTED events
    store, never triggers a new fetch and never touches the collection
    pipeline itself. Never invents or infers a change: a ticker with no
    qualifying event is simply absent from the result, and the caller
    must render nothing for it rather than a fabricated "no recent
    change" line dressed up as a real data point.

    An event with `superseded_by` set is skipped — a newer record has
    already replaced it, so it's no longer "the latest" for that ticker
    even if its own timestamp would otherwise qualify.
    """
    now_utc = now_utc or datetime.now(timezone.utc)
    latest_by_ticker = {}
    for e in events:
        if e.get("superseded_by"):
            continue
        ticker = e.get("ticker")
        ts = e.get("timestamp")
        if not ticker or not ts:
            continue
        try:
            event_dt = datetime.fromisoformat(ts)
        except (ValueError, TypeError):
            continue
        if event_dt.tzinfo is None:
            event_dt = event_dt.replace(tzinfo=timezone.utc)
        if (now_utc - event_dt).days > max_age_days:
            continue
        existing = latest_by_ticker.get(ticker)
        if existing is None or event_dt > existing[1]:
            latest_by_ticker[ticker] = (e, event_dt)
    return {ticker: e for ticker, (e, _dt) in latest_by_ticker.items()}


def group_events_by_ticker(events):
    """
    Plain grouping — ticker -> list of that ticker's events, in whatever
    order they appear in the source list. No filtering (date, superseded)
    happens here; that's each consumer's own job (compute_broker_momentum
    filters superseded + lookback window itself). Kept as a separate, tiny
    function so the SAME grouped structure can feed both the latest-event
    lookup and the multi-event momentum calculation without re-scanning
    the full flat event list twice.
    """
    grouped = {}
    for e in (events or []):
        ticker = e.get("ticker")
        if not ticker:
            continue
        grouped.setdefault(ticker, []).append(e)
    return grouped


def atomic_write_json(path, data):
    """
    Writes JSON atomically: content goes to a temp file in the SAME
    directory first (so the final os.replace is a same-filesystem rename,
    guaranteed atomic on POSIX and Windows), then replaces the target in
    one step. An interrupted write leaves the ORIGINAL file completely
    intact — there is no window where a half-written file sits at the
    real path.
    """
    dir_name = os.path.dirname(path) or "."
    os.makedirs(dir_name, exist_ok=True)
    fd, tmp_path = tempfile.mkstemp(dir=dir_name, prefix=".events_tmp_", suffix=".json")
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        os.replace(tmp_path, path)
    except Exception:
        try:
            os.remove(tmp_path)
        except OSError:
            pass
        raise


def merge_events_into_store(store, new_events):
    """
    Append-only, idempotent merge. An event whose event_id is already
    present is left COMPLETELY untouched — the SIMPLE variant, kept for
    cases where pure append-only behavior is genuinely what's wanted.

    Superseded as the entry point used by collect_and_persist_broker_events
    by reconcile_events_with_store() below, which additionally allows an
    existing record to be ENRICHED by a later-arriving compatible source —
    see that function's docstring for why the plain append-only version
    couldn't do this (event_id previously encoded rating/target, which is
    now fixed at the make_canonical_key() level, but this function's own
    "never touch an existing entry" behavior is still exactly right for
    contexts where enrichment isn't wanted).

    Returns (updated_store, added_count). added_count == 0 on a repeated/
    duplicate run is the CORRECT outcome, not a failure.
    """
    existing_ids = {e["event_id"] for e in store["events"]}
    added = 0
    for e in new_events:
        if e["event_id"] in existing_ids:
            continue
        store["events"].append(e)
        existing_ids.add(e["event_id"])
        added += 1
    store["events"].sort(key=lambda e: e["event_id"])
    return store, added


def _event_conflicts_with_stored(candidate, stored):
    """
    Same compatibility test as _events_conflict, but comparing a freshly
    normalized candidate fragment against an existing STORED event's
    CURRENT fields (which may already reflect earlier enrichment from a
    previous run, not just its original raw values). Includes the same
    action-type contradiction check (see _action_types_conflict) — an
    initiation reported on a later run against an already-stored
    upgrade/downgrade/reiteration for the same slot is a genuine conflict
    even if no rating/target field literally overlaps.
    """
    if _action_types_conflict(candidate.get("action"), stored.get("action")):
        return True
    if candidate.get("new_rating") and stored.get("new_rating"):
        if (candidate.get("old_rating") or "").lower() != (stored.get("old_rating") or "").lower():
            return True
        if candidate["new_rating"].lower() != stored["new_rating"].lower():
            return True
    if candidate.get("new_target") is not None and stored.get("new_target") is not None:
        if round(candidate["new_target"], 2) != round(stored["new_target"], 2):
            return True
    if candidate.get("old_target") is not None and stored.get("old_target") is not None:
        if round(candidate["old_target"], 2) != round(stored["old_target"], 2):
            return True
    return False


# Actions that carry little/no real information about what actually
# happened — a classifier fallback, not a genuine finding. Anything else
# (UPGRADE, DOWNGRADE, INITIATION, REITERATION, TARGET_RAISE, TARGET_CUT,
# RATING_CHANGE) is "meaningful": a real, specific classification.
_WEAK_ACTIONS = {None, "", "NEWS", "NO_CHANGE"}


def _is_weak_action(action):
    return action in _WEAK_ACTIONS


def enrich_stored_event(stored, candidate, now_iso):
    """
    Fills gaps in an existing stored event using a compatible candidate
    fragment. Returns (new_event_dict, changed_bool) — never mutates the
    original `stored` dict in place.

    `candidate` here is always an already-FINALIZED event (from
    finalize_unmatched_event or _merge_pair), meaning its "source",
    "source_url", "source_event_ids", and "source_refs" fields are already
    lists — NOT the raw normalized event's plain-string fields. Every list
    field here is combined by concatenation/union, never by wrapping an
    already-list value in another list.

    IMPORTANT: whether the candidate's source_event_id(s) were already
    present in the stored record is NOT used to short-circuit BEFORE
    comparing fields — an earlier version did this, which meant a
    same-source-event re-arriving with IMPROVED normalized data (e.g. a
    parser fix now extracting a rating that used to be null) was silently
    discarded, since the "already seen" check fired before any field
    comparison ever happened. Fixed: fields are always computed first: a
    genuine no-op is now determined by comparing the RESULT against what
    was already stored, not by source_event_id novelty alone.
    Note: rating/target/company fields only ever FILL a null (pick()) —
    never overwrite one non-null value with a different non-null value.

    `action` is the one field handled DIFFERENTLY from the null-fill
    pattern above, because a classifier ALWAYS produces some action value
    (never null) — "NEWS"/"NO_CHANGE" are the classifier's OWN way of
    saying "couldn't tell", not a genuine absence of information the way
    a null rating is. So: a WEAK stored action (NEWS/NO_CHANGE/absent)
    CAN be corrected to a MEANINGFUL one from the candidate (this is what
    lets a parser-classification fix like "NEWS" -> "DOWNGRADE" actually
    take effect on reprocessing). But a stored action that's ALREADY
    meaningful is never silently overwritten by a DIFFERENT candidate
    action, even if that candidate action is also meaningful — two
    meaningful-but-different classifications for the same event_id is a
    genuine disagreement, and resolving that is _event_conflicts_with_stored's
    job (checked by the caller before this function is ever reached), not
    a silent overwrite here.

    Otherwise: rating/target/currency fields are filled ONLY where
    currently null, source/source_url/source_event_ids/source_refs are
    unioned (each new entry added at most once — never duplicated, even
    when the candidate carries a source_event_id already present),
    confidence is recalculated from the resulting source count and target
    completeness, last_seen is bumped to now only when something actually
    changed, and first_seen/event_id are always preserved exactly as they
    were.
    """
    candidate_source_ids = candidate.get("source_event_ids") or [candidate.get("source_event_id")]
    stored_source_ids = stored.get("source_event_ids", [])
    new_source_ids = [sid for sid in candidate_source_ids if sid not in stored_source_ids]

    def pick(a, b):
        return a if a is not None else b

    enriched = dict(stored)
    enriched["old_rating"] = pick(stored.get("old_rating"), candidate.get("old_rating"))
    enriched["new_rating"] = pick(stored.get("new_rating"), candidate.get("new_rating"))
    enriched["old_rating_bucket"] = pick(stored.get("old_rating_bucket"), candidate.get("old_rating_bucket"))
    enriched["new_rating_bucket"] = pick(stored.get("new_rating_bucket"), candidate.get("new_rating_bucket"))
    enriched["old_target"] = pick(stored.get("old_target"), candidate.get("old_target"))
    enriched["new_target"] = pick(stored.get("new_target"), candidate.get("new_target"))
    enriched["target_currency"] = pick(stored.get("target_currency"), candidate.get("target_currency"))
    enriched["target_change_pct"] = compute_target_change_pct(enriched["old_target"], enriched["new_target"])
    if enriched.get("company") is None:
        enriched["company"] = candidate.get("company")
    if _is_weak_action(enriched.get("action")) and not _is_weak_action(candidate.get("action")):
        enriched["action"] = candidate.get("action")
    # else: stored action is either already meaningful (never blindly
    # overwritten by a different candidate action — see docstring), or the
    # candidate brings nothing better than what's already stored.

    # Only add entries that are genuinely NEW — guards against duplicating
    # source_url/source_refs when the candidate's source_event_id(s) were
    # already present (previously this path was unreachable due to the
    # early-return above, which masked the fact these appends had no such
    # guard; now that field-improvement can flow through even for an
    # already-seen source_event_id, the guard is required to avoid
    # introducing duplicate entries here).
    sources = list(stored.get("source", []))
    for s in candidate.get("source", []):
        if s not in sources:
            sources.append(s)
    enriched["source"] = sorted(set(sources))

    stored_refs = list(stored.get("source_refs", []))
    stored_ref_keys = {(r["source"], r["source_event_id"]) for r in stored_refs}
    new_refs = [r for r in candidate.get("source_refs", []) if (r["source"], r["source_event_id"]) not in stored_ref_keys]
    enriched["source_refs"] = stored_refs + new_refs

    if new_source_ids:
        # New evidence: append its url/id alongside the existing ones.
        candidate_urls_by_id = dict(zip(candidate_source_ids, candidate.get("source_url", [])))
        new_urls = [candidate_urls_by_id[sid] for sid in new_source_ids if sid in candidate_urls_by_id]
        enriched["source_url"] = list(stored.get("source_url", [])) + new_urls
        enriched["source_event_ids"] = list(stored_source_ids) + new_source_ids
    # else: every source_event_id the candidate carries was already known —
    # source_url/source_event_ids stay exactly as stored, unchanged.

    fields_changed = any(
        enriched.get(f) != stored.get(f)
        for f in ("old_rating", "new_rating", "old_rating_bucket", "new_rating_bucket",
                   "old_target", "new_target", "target_currency", "target_change_pct",
                   "company", "action", "source", "source_event_ids")
    )
    if not fields_changed:
        return stored, False  # genuine no-op — nothing new, not even a field improved

    # MERGED_HIGH requires genuine independent agreement: the incoming
    # candidate AND the already-stored record must EACH have reported
    # their own target value (if they disagreed, _event_conflicts_with_stored
    # would already have routed this to the conflict path, never reaching
    # here) — that's a real second confirmation, not just a gap being
    # filled. If only one side ever had target data at all, it's a
    # MERGED_PARTIAL: useful combined information, but not independently
    # corroborated by two sources. This mirrors the exact same distinction
    # _merge_pair uses for a same-run merge — kept consistent deliberately.
    candidate_had_target = candidate.get("new_target") is not None
    stored_had_target_before = stored.get("new_target") is not None
    if len(enriched["source"]) > 1:
        if candidate_had_target and stored_had_target_before:
            enriched["confidence"] = CONFIDENCE_MERGED_HIGH
        else:
            enriched["confidence"] = CONFIDENCE_MERGED_PARTIAL
    # else: still single-source overall (e.g. an already-seen source improved
    # one of its own fields) — left as-is defensively rather than asserted,
    # since confidence display should never crash the pipeline over an edge case.

    enriched["first_seen"] = stored.get("first_seen", now_iso)  # explicitly preserved, never changed
    enriched["last_seen"] = now_iso
    enriched["evidence_fingerprint"] = compute_evidence_fingerprint(enriched)  # recomputed to reflect new fields
    return enriched, True


def _find_active_superseded_candidate(events_by_id, source, source_event_id):
    """
    Scans currently ACTIVE (non-superseded) records for one whose
    source_refs contains this exact (source, source_event_id) pair.
    Restricted to SINGLE-SOURCE stored records only — a conservative,
    explicit scope limit: disentangling which specific source contributed
    which field on an already-multi-source-enriched record, if that
    record's data needs correcting, is a materially harder problem this
    implementation deliberately doesn't attempt yet. A multi-source
    record is simply never returned here, so it falls through to the
    ordinary event_id-based enrich/conflict path unchanged.

    Returns the matching record, or None. A record with superseded_by
    already set is never matched again — it has already been superseded
    once, and the CURRENT active representative for that source item is
    whichever record superseded it, not the stale one itself.
    """
    for rec in events_by_id.values():
        if rec.get("superseded_by"):
            continue
        if len(rec.get("source", [])) != 1:
            continue
        for ref in rec.get("source_refs", []):
            if ref["source"] == source and ref["source_event_id"] == source_event_id:
                return rec
    return None


def reconcile_events_with_store(store, candidate_events, now_iso=None):
    """
    The enrichment-aware entry point used by collect_and_persist_broker_events.
    For each candidate event from THIS run (already carrying its event_id
    from the normalize/merge stage):

    0. SUPERSESSION CHECK (single-source candidates only): if this exact
       (source, source_event_id) pair was already recorded under a
       DIFFERENT event_id on some other ACTIVE single-source record —
       meaning our own parser/classification logic changed and this same
       source article now derives a different event_id than it used to
       (e.g. a ticker-resolution bug fix) — that OLD record is marked
       `superseded_by: <this candidate's event_id>`. It is NEVER deleted,
       NEVER overwritten, and its first_seen/event_id/other fields are
       completely untouched — only the one new `superseded_by` key is
       added. The candidate itself then proceeds through the NORMAL steps
       below using its own event_id, same as any other candidate. This is
       genuinely different from steps 1-3: those are about NEW EVIDENCE
       about the same real-world event; this is about OUR OWN
       classification of the SAME source article changing.

    1. event_id not yet in the store -> appended as a brand-new record
       (first_seen = last_seen = now).
    2. event_id already in the store, and the candidate is COMPATIBLE
       with what's stored -> enriched in place (gaps filled, sources
       unioned, confidence recalculated, last_seen bumped ONLY if a field
       actually changed) via enrich_stored_event() — this also now
       correctly handles a same-source-event returning with IMPROVED
       data (previously silently discarded; see enrich_stored_event's
       docstring for that fix).
    3. event_id already in the store, but the candidate ACTIVELY
       CONFLICTS with what's stored -> the stored record is NEVER
       touched. The candidate is re-keyed via make_conflict_key() and
       appended as its OWN separate record instead (or matched against
       an existing conflict-side record with that same disambiguated
       key, if this exact conflict was already seen before — so repeats
       of a conflicting fragment are still idempotent, not re-appended).

    Returns (updated_store, stats) where stats = {"added", "enriched",
    "conflicts_recorded", "superseded"}.
    """
    now_iso = now_iso or datetime.now(timezone.utc).isoformat()
    events_by_id = {e["event_id"]: e for e in store["events"]}
    added = 0
    enriched_count = 0
    conflicts_recorded = 0
    superseded_count = 0

    for cand in candidate_events:
        eid = cand["event_id"]

        cand_refs = cand.get("source_refs", [])
        if len(cand_refs) == 1:
            ref = cand_refs[0]
            stale = _find_active_superseded_candidate(events_by_id, ref["source"], ref["source_event_id"])
            if stale is not None and stale["event_id"] != eid:
                events_by_id[stale["event_id"]] = dict(stale, superseded_by=eid)
                superseded_count += 1
            # If stale is not None and stale["event_id"] == eid, this is the
            # SAME slot as before — nothing special here, falls through to
            # the ordinary logic below exactly as always.

        if eid not in events_by_id:
            new_record = dict(cand)
            new_record.setdefault("first_seen", now_iso)
            new_record["last_seen"] = now_iso
            events_by_id[eid] = new_record
            added += 1
            continue

        stored = events_by_id[eid]
        if _event_conflicts_with_stored(cand, stored):
            conflict_key = make_conflict_key(
                cand.get("ticker"), cand.get("broker"), cand.get("date"),
                cand.get("old_rating"), cand.get("new_rating"),
                cand.get("old_target"), cand.get("new_target"),
            )
            if conflict_key and conflict_key not in events_by_id:
                new_record = dict(cand)
                new_record["event_id"] = conflict_key
                new_record.setdefault("first_seen", now_iso)
                new_record["last_seen"] = now_iso
                events_by_id[conflict_key] = new_record
                added += 1
                conflicts_recorded += 1
            # else: this exact conflicting fragment was already recorded
            # under its own conflict key — genuine no-op, still idempotent.
            continue

        new_stored, changed = enrich_stored_event(stored, cand, now_iso)
        events_by_id[eid] = new_stored
        if changed:
            enriched_count += 1

    updated_events = sorted(events_by_id.values(), key=lambda e: e["event_id"])
    store["events"] = updated_events
    return store, {"added": added, "enriched": enriched_count, "conflicts_recorded": conflicts_recorded, "superseded": superseded_count}


def fetch_yahoo_broker_events(watchlist):
    """
    Fetches and normalizes Yahoo upgradeDowngradeHistory events for every
    watchlist ticker. Fails soft per-ticker AND overall — one ticker's
    fetch failing never blocks the rest (same pattern as every other
    fetch_* in this file), and if Yahoo is down entirely this returns []
    rather than raising, so the caller can proceed with Investing.com
    alone.
    """
    events = []
    for stock in watchlist:
        ticker, name = stock["ticker"], stock["name"]
        try:
            analyst = fetch_yahoo_analyst(ticker)
            history_count = len((analyst or {}).get("history", []))
            print(f"Broker events: Yahoo {ticker} — {history_count} analyst-history record(s) returned")
            if not analyst:
                continue
            for h in analyst.get("history", [])[:15]:
                firm = h.get("firm", "")
                to_grade = h.get("toGrade", "")
                epoch = h.get("epochGradeDate")
                if not firm or not to_grade or not epoch:
                    continue
                link = f"https://finance.yahoo.com/quote/{yahoo_symbol(ticker)}/analysis#{firm}-{epoch}".replace(" ", "-")
                events.append(normalize_yahoo_event(
                    ticker=ticker, company=name, firm=firm,
                    from_grade=h.get("fromGrade"), to_grade=to_grade,
                    action_code=h.get("action", ""), epoch=epoch, link=link,
                ))
        except Exception as e:
            print(f"  ! events pipeline: yahoo fetch failed for {ticker}: {e}", file=sys.stderr)
            continue
    return events


def fetch_investing_broker_events(ticker_lookup):
    """Fetches and normalizes Investing.com analyst-ratings RSS events.
    Fails soft — returns [] on any error, never raises, so the caller can
    proceed with Yahoo's events alone."""
    try:
        items, had_error = fetch_feed(ANALYST_RATINGS_FEED_URL)
        if had_error:
            return []
        return [
            normalize_investing_event(
                title=it["title"], link=it["link"], pub_date=it.get("pubDate"),
                ticker_lookup=ticker_lookup,
            )
            for it in items
        ]
    except Exception as e:
        print(f"  ! events pipeline: investing.com fetch failed: {e}", file=sys.stderr)
        return []


def collect_and_persist_broker_events(watchlist, screener_rows=None, dry_run=False, path=None):
    """
    ONE full poll-cycle pass: fetch both sources -> normalize -> dedupe/
    merge (within this run) -> load existing store -> RECONCILE against
    history (new records appended, compatible existing records enriched,
    conflicting fragments kept separate) -> write atomically. Called from
    main() as an additive, isolated step, wrapped so any failure here can
    never stop the rest of the poll cycle — see main()'s own try/except
    around this call.

    Failure handling, exactly as specified:
    - Yahoo fails -> proceeds using Investing.com's events only.
    - Investing.com fails -> proceeds using Yahoo's events only.
    - Both fail -> new_events is empty; the existing store is loaded and
      re-saved (0 added, 0 enriched) rather than skipped — safe because
      reconcile_events_with_store only ever appends or fills gaps, so
      re-saving unchanged content can never truncate or lose data.
    - Existing store is CORRUPT -> persistence is skipped entirely this
      cycle; the corrupt file is left exactly as-is, loudly logged.
    - Write itself fails (disk full, permissions, etc) -> reported, the
      PREVIOUS file on disk is untouched (atomic_write_json guarantees
      this structurally, not just by convention).

    Conflicting pairs discovered WITHIN this run (from deduplicate_and_merge)
    are finalized as individually conflict-keyed records before reaching
    the store — never merged into one, never discarded. A candidate that
    conflicts with something ALREADY IN THE STORE (discovered across runs,
    e.g. a later day's report disagreeing with an earlier one) is handled
    by reconcile_events_with_store itself, the same way.

    dry_run=True runs the full pipeline and returns the result WITHOUT
    writing anything to disk.
    """
    path = path or EVENTS_STATE_FILE
    screener_rows = screener_rows or []

    print("Broker events: collection started")
    yahoo_events = fetch_yahoo_broker_events(watchlist)
    print(f"Broker events: Yahoo returned {len(yahoo_events)} event(s)")
    ticker_lookup = build_name_ticker_lookup(watchlist, screener_rows)
    investing_events = fetch_investing_broker_events(ticker_lookup)
    print(f"Broker events: Investing.com returned {len(investing_events)} event(s)")

    all_normalized = yahoo_events + investing_events
    print(f"Broker events: {len(all_normalized)} normalized candidate(s) before dedup/merge")
    merged, conflicts, unmatched = deduplicate_and_merge(all_normalized)

    # `unmatched` already contains every conflict-side fragment, correctly
    # finalized with is_conflict_side=True and deduplicated (deduplicate_and_merge
    # handles this internally now — see its docstring). `conflicts` itself
    # is kept only as pair-level diagnostic/logging info, not re-finalized
    # here — doing so used to create duplicate conflict records.
    new_events = merged + unmatched

    try:
        store = load_events_store(path)
    except EventsStoreCorruptError as e:
        print(f"  ! events store CORRUPT — skipping persistence this cycle to avoid data loss: {e}", file=sys.stderr)
        return {
            "written": False, "reason": "existing_store_corrupt",
            "new_events_found": len(new_events), "added": 0, "enriched": 0,
            "conflicts": conflicts, "store": None,
        }

    updated_store, stats = reconcile_events_with_store(store, new_events)
    added, enriched_count = stats["added"], stats["enriched"]

    if dry_run:
        return {
            "written": False, "reason": "dry_run",
            "new_events_found": len(new_events), "added": added, "enriched": enriched_count,
            "conflicts": conflicts, "store": updated_store,
        }

    try:
        atomic_write_json(path, updated_store)
    except Exception as e:
        print(f"  ! events store write FAILED — previous file left untouched: {e}", file=sys.stderr)
        return {
            "written": False, "reason": "write_failed",
            "new_events_found": len(new_events), "added": 0, "enriched": 0,
            "conflicts": conflicts, "store": None,
        }

    print(
        f"Broker events: added={added} enriched={enriched_count} conflicts={len(conflicts)} "
        f"total_in_store={len(updated_store['events'])}"
    )
    return {
        "written": True, "reason": "ok",
        "new_events_found": len(new_events), "added": added, "enriched": enriched_count,
        "conflicts": conflicts, "store": updated_store,
    }



import statistics


# =========================================================================
# Phase 7A: Technical Backtesting Engine
# =========================================================================
#
# Scope, exactly as specified: tests the EXISTING scoring logic already
# live in this dashboard, using historical price/volume data reconstructed
# walk-forward (never using future data to compute a past day's signal).
# Reuses the SAME pure functions the live dashboard uses
# (compute_rsi, compute_atr, compute_support_resistance,
# compute_breakout_status, compute_research_scorecard,
# compute_scorecard_subtotals, compute_signal_quality,
# compute_dont_chase_warning) rather than reimplementing any of this
# logic a second time — this IS "test the existing logic, don't optimise
# it" in concrete form.
#
# Explicitly NOT reconstructed: NEWS, BROKER, and the target-proximity/
# conflicting-evidence components of RISK — there is no reliable
# historical source for point-in-time broker ratings or news relevance.
# This is achieved by passing NEUTRAL inputs (evidence label "no_signal",
# broker_momentum=None, upside_pct=None) into the SAME
# compute_research_scorecard() the live dashboard calls — confirmed by
# direct test that this produces NEWS=0, BROKER=0, and a RISK score
# reflecting ONLY the RSI-overextension component, with zero new/
# duplicate scoring logic required.

BACKTEST_FORWARD_HORIZONS = [5, 20, 60]  # trading days
BACKTEST_MIN_SAMPLE_SIZE = 30  # n below this -> "insufficient data", never a claim
BACKTEST_AVERAGE_VOLUME_WINDOW = 20  # trailing trading days — a specific,
# documented methodological choice: historical chart data has no
# point-in-time "average volume" field the way a live quote does, so this
# defines it explicitly as a trailing 20-day average rather than silently
# assuming a number.

# Pre-registered signal list — FIXED before any evaluation, per the
# explicit instruction not to cherry-pick, add, or remove hypotheses based
# on results. Each is a small, specific, individually named condition
# reusing the live dashboard's own existing thresholds/functions.
BACKTEST_SIGNAL_IDS = [
    "technical_subtotal_strong_positive",  # Technical/Market subtotal >= +4
    "technical_subtotal_strong_negative",  # Technical/Market subtotal <= -4
    "technical_signal_quality_strong_positive",  # technical-only Signal Quality: Strong, positive direction
    "technical_signal_quality_strong_negative",  # technical-only Signal Quality: Strong, negative direction
    "rsi_overbought",  # RSI(14) >= OVEREXTENDED_RSI_THRESHOLD (70)
    "rsi_oversold",  # RSI(14) <= 30 (the existing MOMENTUM dimension's own "weak" threshold)
    "bullish_ma_crossover",  # MA20 crosses above MA50 — a FRESH crossover, not merely "currently bullish"
    "bearish_ma_crossover",  # MA20 crosses below MA50 — a FRESH crossover
    "high_volume_breakout",  # breakout above 20-day high AND volume >= HIGH_VOLUME_RATIO_THRESHOLD
    "dont_chase_triggered",  # compute_dont_chase_warning fires
]
BACKTEST_SUBTOTAL_STRONG_THRESHOLD = 4  # named, adjustable — matches the
# magnitude already used informally elsewhere in this project as "a
# clearly decisive score," not a new invented number


def reconstruct_technical_state(closes, volumes, highs, lows, index):
    """
    Reconstructs exactly the technical inputs the live dashboard computes
    for TODAY, but as of `index` in a historical series — using ONLY
    closes[:index+1] / volumes[:index+1] / highs[:index+1] / lows[:index+1]
    at every step. This is THE walk-forward discipline: never pass the
    full series to any calculation, always the trailing slice up to and
    including the evaluation point, so a signal computed for day T can
    never see day T+1 or later.

    Reuses compute_rsi / compute_atr / compute_support_resistance /
    compute_breakout_status directly — the exact same functions
    fetch_price_technicals uses for the live dashboard — never a second,
    parallel implementation of any of this arithmetic.

    Returns None if there isn't enough trailing history at this index for
    even the most basic calculation (5-day change) — never a partially
    fabricated state.
    """
    trailing_closes = closes[:index + 1]
    trailing_volumes = volumes[:index + 1]
    trailing_highs = highs[:index + 1]
    trailing_lows = lows[:index + 1]
    if len(trailing_closes) < 6:
        return None
    latest_close = trailing_closes[-1]
    five_days_ago = trailing_closes[-6]
    change_pct_5d = (latest_close - five_days_ago) / five_days_ago * 100 if five_days_ago else None
    change_pct = None
    if len(trailing_closes) >= 2 and trailing_closes[-2]:
        change_pct = (latest_close - trailing_closes[-2]) / trailing_closes[-2] * 100

    rsi14 = compute_rsi(trailing_closes, 14)
    ma20 = sum(trailing_closes[-20:]) / len(trailing_closes[-20:]) if len(trailing_closes) >= 20 else None
    ma50 = sum(trailing_closes[-50:]) / 50 if len(trailing_closes) >= 50 else None
    ma_crossover = None
    if ma20 is not None and ma50 is not None:
        ma_crossover = "bullish" if ma20 > ma50 else ("bearish" if ma20 < ma50 else "flat")
    above_ma20 = (latest_close > ma20) if ma20 is not None else None

    paired = list(zip(trailing_closes, trailing_volumes, trailing_highs, trailing_lows))
    atr14 = compute_atr(paired, 14)
    support_resistance = compute_support_resistance(trailing_highs, trailing_lows)
    breakout_status = compute_breakout_status(latest_close, trailing_highs[:-1], trailing_lows[:-1])

    avg_volume = None
    if len(trailing_volumes) >= BACKTEST_AVERAGE_VOLUME_WINDOW:
        window_vols = [v for v in trailing_volumes[-BACKTEST_AVERAGE_VOLUME_WINDOW:] if v is not None]
        if len(window_vols) == BACKTEST_AVERAGE_VOLUME_WINDOW:
            avg_volume = sum(window_vols) / BACKTEST_AVERAGE_VOLUME_WINDOW
    latest_volume = trailing_volumes[-1]
    volume_ratio = compute_volume_ratio(latest_volume, avg_volume)

    return {
        "close": latest_close, "changePct": change_pct, "changePct5d": change_pct_5d,
        "rsi14": rsi14, "ma20": ma20, "ma50": ma50, "maCrossover": ma_crossover,
        "aboveMA20": above_ma20, "atr14": atr14, "supportResistance": support_resistance,
        "breakoutStatus": breakout_status, "volumeRatio": volume_ratio,
    }


def compute_technical_only_scorecard(state, ftse_relative=None):
    """
    Calls compute_research_scorecard/compute_scorecard_subtotals/
    compute_signal_quality DIRECTLY — the exact live functions — with
    NEUTRAL inputs for the non-reconstructable dimensions (evidence
    label "no_signal", broker_momentum=None, upside_pct=None,
    sector_context=None), so NEWS and BROKER always resolve to 0 and
    RISK reflects ONLY its RSI-overextension component. Confirmed by a
    direct test that this produces exactly that result — not assumed.
    """
    neutral_evidence = {"label": "no_signal", "hasCatalyst": False, "catalystDirection": None, "volumeConfirms": None}
    scorecard = compute_research_scorecard(
        state["changePct5d"], state["aboveMA20"], state["rsi14"], state["maCrossover"],
        state["volumeRatio"], state["changePct"], neutral_evidence, None,
        state["breakoutStatus"], ftse_relative, None, None,
    )
    subtotals = compute_scorecard_subtotals(scorecard["dimensions"])
    signal_quality = compute_signal_quality(scorecard["dimensions"], [])  # no contradiction detection reconstructed — technical-only dims can't produce the news/broker-based contradiction types
    return {"dimensions": scorecard["dimensions"], "technicalMarket": subtotals["technicalMarket"], "signalQuality": signal_quality}


def _signal_fired(signal_id, state, prev_state, tech_scorecard, dont_chase):
    """
    Evaluates ONE pre-registered signal ID against the current
    reconstructed state (and, for crossover-type signals, the
    immediately preceding day's state — needed to detect a FRESH
    crossover rather than "currently in a long-running bullish/bearish
    state", which would otherwise make an episode span months and be a
    near-meaningless single "event"). Every threshold reused here is the
    SAME named constant the live dashboard already uses
    (OVEREXTENDED_RSI_THRESHOLD, HIGH_VOLUME_RATIO_THRESHOLD,
    BACKTEST_SUBTOTAL_STRONG_THRESHOLD) — never a new number invented
    just for this function.
    """
    if state is None:
        return False
    if signal_id == "technical_subtotal_strong_positive":
        return tech_scorecard["technicalMarket"] >= BACKTEST_SUBTOTAL_STRONG_THRESHOLD
    if signal_id == "technical_subtotal_strong_negative":
        return tech_scorecard["technicalMarket"] <= -BACKTEST_SUBTOTAL_STRONG_THRESHOLD
    if signal_id == "technical_signal_quality_strong_positive":
        return tech_scorecard["signalQuality"] == "Strong" and tech_scorecard["technicalMarket"] > 0
    if signal_id == "technical_signal_quality_strong_negative":
        return tech_scorecard["signalQuality"] == "Strong" and tech_scorecard["technicalMarket"] < 0
    if signal_id == "rsi_overbought":
        return state["rsi14"] is not None and state["rsi14"] >= OVEREXTENDED_RSI_THRESHOLD
    if signal_id == "rsi_oversold":
        return state["rsi14"] is not None and state["rsi14"] < 30
    if signal_id == "bullish_ma_crossover":
        return (state["maCrossover"] == "bullish" and prev_state is not None
                and prev_state["maCrossover"] is not None and prev_state["maCrossover"] != "bullish")
    if signal_id == "bearish_ma_crossover":
        return (state["maCrossover"] == "bearish" and prev_state is not None
                and prev_state["maCrossover"] is not None and prev_state["maCrossover"] != "bearish")
    if signal_id == "high_volume_breakout":
        return (state["breakoutStatus"] == "breakout" and state["volumeRatio"] is not None
                and state["volumeRatio"] >= HIGH_VOLUME_RATIO_THRESHOLD)
    if signal_id == "dont_chase_triggered":
        return dont_chase is not None
    raise ValueError(f"Unknown signal id: {signal_id}")  # a typo here must fail loudly, never silently evaluate nothing


def detect_signal_events(closes, volumes, highs, lows, signal_id):
    """
    Walks the FULL historical series day by day, reconstructing the
    technical state at each index using ONLY data through that index
    (see reconstruct_technical_state's own walk-forward discipline),
    and records the FIRST day of every consecutive run where the given
    signal fires as one "episode" — collapsing a persistent condition
    (e.g. RSI >= 70 for five consecutive days) into a single event at
    its start, rather than five separate (and non-independent)
    observations. This IS the required episode/de-duplication rule.

    Crossover-type signals are inherently single-day by their own
    "fresh change" definition and don't need this collapsing, but
    sharing one pass keeps the logic in one place rather than two
    parallel walks.

    Returns a list of event indices into the `closes` array.
    """
    n = len(closes)
    events = []
    prev_state = None
    was_firing = False
    for i in range(n):
        state = reconstruct_technical_state(closes, volumes, highs, lows, i)
        if state is None:
            prev_state = None
            was_firing = False
            continue
        tech_scorecard = compute_technical_only_scorecard(state)
        dont_chase = compute_dont_chase_warning(state["changePct5d"], state["rsi14"], state["volumeRatio"])
        firing = _signal_fired(signal_id, state, prev_state, tech_scorecard, dont_chase)
        if firing and not was_firing:
            events.append(i)
        was_firing = firing
        prev_state = state
    return events


def forward_return(closes, event_index, horizon_days):
    """
    (close[event_index + horizon_days] - close[event_index]) /
    close[event_index] — a raw price return, explicitly NOT adjusted for
    dividends (see the module-level note on this), and explicitly NOT
    net of any transaction cost or slippage — reported as a plain fact
    about what the price actually did, nothing assumed on top of it.
    Returns None when there isn't enough FUTURE data yet to compute this
    horizon (expected and normal near the end of the available history —
    never approximated from a shorter window).
    """
    target_index = event_index + horizon_days
    if target_index >= len(closes):
        return None
    start_price = closes[event_index]
    end_price = closes[target_index]
    if not start_price:
        return None
    return (end_price - start_price) / start_price * 100


import math


def _standard_normal_cdf(x):
    return 0.5 * (1 + math.erf(x / math.sqrt(2)))


def _two_tailed_p_value(mean, stdev, n):
    """
    Two-tailed p-value for a one-sample test of mean != 0, using a
    NORMAL approximation rather than a true t-distribution. This is a
    standard, well-justified simplification given this project is
    stdlib-only (no scipy) — and BACKTEST_MIN_SAMPLE_SIZE is already 30,
    exactly the sample size at which the t-distribution and normal
    distribution converge closely enough that this approximation is
    defensible, not a shortcut that meaningfully changes conclusions at
    the sample sizes this backtest actually reports on.
    """
    if n < 2 or stdev is None or stdev == 0:
        return None
    standard_error = stdev / math.sqrt(n)
    z = mean / standard_error
    return 2 * (1 - _standard_normal_cdf(abs(z)))


def summarize_signal_horizon(returns, benchmark_returns, baseline_returns):
    """
    Statistical summary for ONE signal at ONE forward horizon. `returns`
    are the signal's own forward returns at every episode where it
    fired; `benchmark_returns` are the FTSE 100's forward returns over
    the SAME event windows; `baseline_returns` are the stock's own
    UNCONDITIONAL forward returns at this horizon (sampled across every
    trading day, not just signal-firing days) — both comparators
    required, matching "does this signal add anything beyond a
    generally-drifting market or a generally-drifting stock."
    """
    n = len(returns)
    if n == 0:
        return {"n": 0, "insufficientSample": True, "meanReturn": None, "medianReturn": None,
                "winRate": None, "stdev": None, "pValue": None, "vsBenchmark": None, "vsBaseline": None}
    mean_ret = statistics.mean(returns)
    median_ret = statistics.median(returns)
    win_rate = sum(1 for r in returns if r > 0) / n
    stdev = statistics.stdev(returns) if n > 1 else None
    p_value = _two_tailed_p_value(mean_ret, stdev, n)
    benchmark_mean = statistics.mean(benchmark_returns) if benchmark_returns else None
    baseline_mean = statistics.mean(baseline_returns) if baseline_returns else None
    return {
        "n": n, "meanReturn": mean_ret, "medianReturn": median_ret, "winRate": win_rate,
        "stdev": stdev, "pValue": p_value,
        "vsBenchmark": (mean_ret - benchmark_mean) if benchmark_mean is not None else None,
        "vsBaseline": (mean_ret - baseline_mean) if baseline_mean is not None else None,
        "insufficientSample": n < BACKTEST_MIN_SAMPLE_SIZE,
    }


BACKTEST_HISTORY_RANGE = "1y"  # matches fetch_price_technicals's own proven
# range exactly — NOT verified today to be extendable to 2y/5y/max; that
# needs a real, live check against Yahoo's actual API before ever being
# widened, since I cannot make that live call myself (see this module's
# own capability notes). Widening this constant is a one-line change
# once verified — deliberately NOT done speculatively here.


def fetch_backtest_history(symbol):
    """
    Fetches one symbol's historical daily close/volume/high/low for
    backtesting — reuses http_get and the exact same Yahoo chart
    endpoint fetch_price_technicals already uses successfully, just
    keeping the RAW series instead of collapsing it to today's
    scalars. Returns {"closes": [...], "volumes": [...], "highs": [...],
    "lows": [...], "dates": [...]} or None on any failure — same
    fail-safe-by-returning-None contract as every other fetch function
    in this file, never a partial/fabricated series.
    """
    url = f"https://query1.finance.yahoo.com/v8/finance/chart/{urllib.parse.quote(symbol)}?interval=1d&range={BACKTEST_HISTORY_RANGE}"
    try:
        data = json.loads(http_get(url))
        result = (data.get("chart") or {}).get("result") or [None]
        if not result[0]:
            return None
        timestamps = result[0].get("timestamp") or []
        quote0 = (result[0].get("indicators", {}).get("quote", [{}])[0] or {})
        raw_closes = quote0.get("close") or []
        raw_volumes = quote0.get("volume") or []
        raw_highs = quote0.get("high") or []
        raw_lows = quote0.get("low") or []
        paired = [
            (t, c, v, h, l) for t, c, v, h, l in zip(timestamps, raw_closes, raw_volumes, raw_highs, raw_lows)
            if c is not None
        ]
        if len(paired) < 30:
            return None  # not enough real history to be worth including
        dates = [datetime.fromtimestamp(t, tz=timezone.utc).astimezone(LONDON_TZ).strftime("%Y-%m-%d") for t, *_ in paired]
        return {
            "dates": dates,
            "closes": [p[1] for p in paired], "volumes": [p[2] for p in paired],
            "highs": [p[3] for p in paired], "lows": [p[4] for p in paired],
        }
    except Exception as e:
        print(f"  ! backtest history fetch failed: {symbol} ({e})", file=sys.stderr)
        return None


def run_backtest_cli(watchlist):
    """
    The actual entry point for `python poll.py --backtest` — fetches
    real historical data for every watchlist stock plus the FTSE 100
    index, runs run_technical_backtest(), and saves the result via
    save_backtest_results(). This is DELIBERATELY separate from the
    normal 5-minute poll cycle (main()) — a full historical backtest
    across many stocks is far heavier than a live poll, and has no
    business running on that cadence.

    Must be run somewhere with genuine network access to Yahoo Finance
    — this sandbox's own network access does not include that domain,
    so this function's real-data path has been written and reviewed
    but not personally executed against live data; see this module's
    capability notes.
    """
    print(f"Backtest: fetching historical data for {len(watchlist)} watchlist stock(s) + FTSE 100...")
    stock_histories = {}
    all_dates = None
    for stock in watchlist:
        ticker = stock["ticker"]
        hist = fetch_backtest_history(yahoo_symbol(ticker))
        if hist is None:
            print(f"  ! skipping {ticker}: history fetch failed or too short")
            continue
        stock_histories[ticker] = hist
        if all_dates is None or len(hist["dates"]) < len(all_dates):
            all_dates = hist["dates"]  # shortest series sets the shared alignment window
        time.sleep(0.3)

    if not stock_histories:
        print("Backtest: no usable stock histories fetched — aborting.", file=sys.stderr)
        return

    ftse_hist = fetch_backtest_history("%5EFTSE")
    if ftse_hist is None:
        print("Backtest: FTSE 100 history fetch failed — cannot compute a benchmark, aborting.", file=sys.stderr)
        return

    # Align every series onto the SAME trailing window length (the
    # shortest available series) — see run_technical_backtest's own
    # documented precondition that all series must be index-aligned to
    # one shared calendar. This is a simple, honest alignment by
    # trailing-length only; it does not attempt to reconcile genuine
    # calendar gaps (a stock's own trading halt, for instance) beyond
    # this — a real limitation, disclosed here rather than silently
    # assumed away.
    min_len = min(len(h["closes"]) for h in stock_histories.values())
    min_len = min(min_len, len(ftse_hist["closes"]))
    for ticker in stock_histories:
        for key in ("closes", "volumes", "highs", "lows", "dates"):
            stock_histories[ticker][key] = stock_histories[ticker][key][-min_len:]
    ftse_closes = ftse_hist["closes"][-min_len:]
    dates = stock_histories[next(iter(stock_histories))]["dates"]

    in_sample_cutoff_index = int(min_len * 0.7)  # 70/30 in-sample/out-of-sample split
    stock_histories_for_engine = {t: {k: v for k, v in h.items() if k != "dates"} for t, h in stock_histories.items()}
    results = run_technical_backtest(stock_histories_for_engine, ftse_closes, in_sample_cutoff_index)

    metadata = {
        "runAt": datetime.now(timezone.utc).isoformat(),
        "historyStartDate": dates[0], "historyEndDate": dates[-1],
        "stockCount": len(stock_histories), "inSampleCutoffDate": dates[in_sample_cutoff_index],
    }
    save_backtest_results(results, metadata)
    print(f"Backtest: complete. {len(stock_histories)} stocks, {dates[0]} to {dates[-1]}, "
          f"in-sample cutoff {dates[in_sample_cutoff_index]}. Results saved to {BACKTEST_RESULTS_FILE}.")


def run_technical_backtest(stock_histories, ftse_closes, in_sample_cutoff_index):
    """
    Full Phase 7A orchestration. Preconditions the caller must satisfy
    (documented here since this module cannot itself fetch or align
    real market data — see the module's own capability notes):
    - `stock_histories`: {ticker: {"closes": [...], "volumes": [...],
      "highs": [...], "lows": [...]}} — one list per field, all the SAME
      length for a given ticker, in chronological order.
    - `ftse_closes`: a list of FTSE 100 closes, INDEX-ALIGNED to the
      SAME trading-day calendar as every stock series (i.e. index i in
      ftse_closes is the same calendar day as index i in every stock's
      closes). Real-world stocks can have small gaps (suspensions,
      listing dates); aligning series onto one shared calendar is the
      caller's responsibility before this function is used with real
      data.
    - `in_sample_cutoff_index`: the index splitting in-sample
      (development) from out-of-sample (held-out) — the held-out period
      must never have influenced the signal definitions above, which it
      cannot, since BACKTEST_SIGNAL_IDS is fixed in this module's own
      source, not derived from any data.

    Returns a nested results structure: for each signal, for each of
    "inSample"/"outOfSample", for each forward horizon, a
    summarize_signal_horizon() result — plus the Bonferroni-corrected
    significance threshold applied across the whole batch.
    """
    total_tests = len(BACKTEST_SIGNAL_IDS) * len(BACKTEST_FORWARD_HORIZONS) * 2  # *2 for in-sample + out-of-sample, each a separate test
    bonferroni_alpha = 0.05 / total_tests

    # Unconditional baseline forward returns per horizon, sampled once
    # across every trading day of every stock — computed independently
    # of any signal, used as every signal's baseline comparator.
    baseline_returns_by_period_horizon = {"inSample": {h: [] for h in BACKTEST_FORWARD_HORIZONS},
                                            "outOfSample": {h: [] for h in BACKTEST_FORWARD_HORIZONS}}
    for ticker, hist in stock_histories.items():
        closes = hist["closes"]
        for i in range(len(closes)):
            period = "inSample" if i < in_sample_cutoff_index else "outOfSample"
            for h in BACKTEST_FORWARD_HORIZONS:
                r = forward_return(closes, i, h)
                if r is not None:
                    baseline_returns_by_period_horizon[period][h].append(r)

    results = {"signals": {}, "bonferroniAlpha": bonferroni_alpha, "totalTests": total_tests}
    for signal_id in BACKTEST_SIGNAL_IDS:
        period_results = {"inSample": {}, "outOfSample": {}}
        events_by_period = {"inSample": [], "outOfSample": []}
        for ticker, hist in stock_histories.items():
            closes, volumes, highs, lows = hist["closes"], hist["volumes"], hist["highs"], hist["lows"]
            event_indices = detect_signal_events(closes, volumes, highs, lows, signal_id)
            for idx in event_indices:
                period = "inSample" if idx < in_sample_cutoff_index else "outOfSample"
                events_by_period[period].append((ticker, idx))
        for period in ("inSample", "outOfSample"):
            horizon_results = {}
            for h in BACKTEST_FORWARD_HORIZONS:
                sig_returns, bench_returns = [], []
                for ticker, idx in events_by_period[period]:
                    r = forward_return(stock_histories[ticker]["closes"], idx, h)
                    if r is None:
                        continue
                    sig_returns.append(r)
                    fr_bench = forward_return(ftse_closes, idx, h)
                    if fr_bench is not None:
                        bench_returns.append(fr_bench)
                horizon_results[h] = summarize_signal_horizon(
                    sig_returns, bench_returns, baseline_returns_by_period_horizon[period][h],
                )
            period_results[period] = horizon_results
        results["signals"][signal_id] = {
            "eventCount": {"inSample": len(events_by_period["inSample"]), "outOfSample": len(events_by_period["outOfSample"])},
            "horizons": period_results,
        }
    return results


# --- Phase 7C: Backtest results persistence + dashboard reporting -------
BACKTEST_RESULTS_FILE = os.path.join(STATE_DIR, "backtest_results.json")


class BacktestResultsCorruptError(Exception):
    """Same discipline as every other state file in this project — a file
    that EXISTS but fails to parse must never be silently treated as
    absent."""
    pass


def load_backtest_results(path=None):
    """
    Returns the most recently saved backtest results, or None if no
    backtest has ever been run — the dashboard's own honest "not run
    yet" state, never a fabricated placeholder. Raises
    BacktestResultsCorruptError on a file that exists but fails to
    parse, same discipline as every other state file here — even though
    this specific file isn't accumulating irreplaceable history the way
    events.json/evidence_history.json are (a corrupt results file just
    means re-running the backtest, not losing anything), silently
    treating corruption as "no backtest run yet" would misrepresent the
    actual situation to whoever reads the dashboard.
    """
    path = path or BACKTEST_RESULTS_FILE
    if not os.path.exists(path):
        return None
    with open(path, "r", encoding="utf-8") as f:
        raw = f.read()
    try:
        return json.loads(raw)
    except json.JSONDecodeError as e:
        raise BacktestResultsCorruptError(f"{path} contains invalid JSON: {e}") from e


def save_backtest_results(results, run_metadata, path=None):
    """
    Persists a completed backtest run. `run_metadata` should include at
    minimum: runAt (ISO timestamp), historyStartDate, historyEndDate,
    stockCount, inSampleCutoffDate — the concrete facts needed to know
    WHAT was actually tested, not just the statistical output, so the
    dashboard can honestly state its own scope rather than imply a
    single canonical "the" backtest.
    """
    path = path or BACKTEST_RESULTS_FILE
    payload = {"metadata": run_metadata, "results": results}
    atomic_write_json(path, payload)
    return {"written": True}


def _format_pct(x, decimals=2):
    return f"{'+' if x >= 0 else ''}{x:.{decimals}f}%" if x is not None else "—"


def render_backtest_results_html(loaded):
    """
    Renders the Phase 7C dashboard section from whatever
    load_backtest_results() returned. Deliberately verbose about
    limitations — every one of survivorship bias, data revisions, the
    missing News/Broker/AI reconstruction, pre-cost/pre-slippage
    returns, sample-size limits, correlation-vs-causation, and
    out-of-sample limits is stated EXPLICITLY and PROMINENTLY, not
    buried in a single footnote, because a reader skimming straight to
    a "beat the benchmark" number is exactly the failure mode this
    section exists to prevent.
    """
    limitations_html = (
        '<div class="disclaimer" style="margin-bottom:10px;">'
        '<b>Read this before the numbers below:</b><ul style="margin:6px 0 0 18px;padding:0;">'
        '<li><b>Survivorship bias</b> — this universe reflects TODAY\'s FTSE 100/250 membership only; '
        'stocks that were delisted, acquired, or relegated during the tested period are invisible to this test.</li>'
        '<li><b>No historical News/Broker/AI Evidence data</b> — this backtest covers ONLY the Technical/Market '
        'dimensions and the RSI-overextension component of RISK. It says nothing about whether NEWS, BROKER, '
        'or AI Evidence Review signals have ever worked.</li>'
        '<li><b>Pre-cost, pre-slippage</b> — returns shown are raw price moves, not what a real trade would net.</li>'
        '<li><b>Price returns, not total return</b> — not adjusted for dividends.</li>'
        '<li><b>Data revisions</b> — historical prices can occasionally be revised by the data source after the fact; '
        'this cannot be fully corrected for with a free data source.</li>'
        '<li><b>Small samples are flagged, not hidden</b> — any result built from fewer than '
        f'{BACKTEST_MIN_SAMPLE_SIZE} events is marked "insufficient data," not reported as a finding.</li>'
        '<li><b>Multiple-testing corrected</b> — with many signals and horizons tested together, apparent '
        '"significance" is checked against a Bonferroni-corrected threshold, not the usual 0.05.</li>'
        '<li><b>Correlation, not causation</b> — even a result that clears every bar above only shows a historical '
        'association, never a guarantee, explanation, or prediction of future performance.</li>'
        '</ul></div>'
    )
    if loaded is None:
        return limitations_html + '<span class="meta">No backtest has been run yet.</span>'

    metadata = loaded.get("metadata", {})
    results = loaded.get("results", {})
    meta_line = (
        f'<p class="meta">Tested {esc(str(metadata.get("stockCount", "?")))} stock(s), '
        f'{esc(str(metadata.get("historyStartDate", "?")))} to {esc(str(metadata.get("historyEndDate", "?")))} '
        f'· in-sample/out-of-sample split at {esc(str(metadata.get("inSampleCutoffDate", "?")))} '
        f'· run {esc(str(metadata.get("runAt", "?")))}</p>'
    )
    bonferroni_alpha = results.get("bonferroniAlpha")
    rows = []
    for signal_id, signal_data in results.get("signals", {}).items():
        for period in ("inSample", "outOfSample"):
            period_label = "In-sample" if period == "inSample" else "Out-of-sample"
            for horizon in BACKTEST_FORWARD_HORIZONS:
                r = signal_data.get("horizons", {}).get(period, {}).get(str(horizon)) or signal_data.get("horizons", {}).get(period, {}).get(horizon)
                if r is None:
                    continue
                if r.get("insufficientSample"):
                    rows.append(
                        f'<tr><td>{esc(signal_id)}</td><td>{period_label}</td><td>{horizon}d</td>'
                        f'<td colspan="6" class="meta">Insufficient data (n={r.get("n", 0)} &lt; {BACKTEST_MIN_SAMPLE_SIZE})</td></tr>'
                    )
                    continue
                significant = (bonferroni_alpha is not None and r.get("pValue") is not None
                               and r["pValue"] < bonferroni_alpha)
                rows.append(
                    f'<tr><td>{esc(signal_id)}</td><td>{period_label}</td><td>{horizon}d</td>'
                    f'<td>{r["n"]}</td><td>{_format_pct(r["meanReturn"])}</td><td>{_format_pct(r["medianReturn"])}</td>'
                    f'<td>{r["winRate"]*100:.0f}%</td><td>{_format_pct(r["vsBenchmark"])}</td>'
                    f'<td>{"✓ significant" if significant else "—"}</td></tr>'
                )
    table = (
        '<table><tr><th>Signal</th><th>Period</th><th>Horizon</th><th>N</th><th>Mean</th><th>Median</th>'
        '<th>Win rate</th><th>vs FTSE</th><th>After correction</th></tr>' + "".join(rows) + '</table>'
    ) if rows else '<span class="meta">No signal events found in the tested history.</span>'
    return limitations_html + meta_line + table


if __name__ == "__main__":
    if "--backtest" in sys.argv:
        # Phase 7A/7C entry point — deliberately separate from the
        # normal 5-minute poll cycle. Must be run somewhere with real
        # network access to Yahoo Finance (this repo's own GitHub
        # Actions runner already has this, proven by every successful
        # Poller run) — see run_backtest_cli's own docstring.
        _watchlist = load_json(WATCHLIST_FILE, [])
        if not _watchlist:
            print("watchlist.json is empty — nothing to backtest.", file=sys.stderr)
            sys.exit(1)
        run_backtest_cli(_watchlist)
        sys.exit(0)
    try:
        main()
    except Exception as e:
        # A totally unhandled crash (vs the per-source try/excepts already inside main)
        # should still tell you something's wrong, not just go silent.
        err_msg = f"⚠️ UK Stock Watch poller crashed: {type(e).__name__}: {e}\nWill retry next scheduled run."
        print(err_msg, file=sys.stderr)
        try:
            send_webhook(err_msg, bypass_market_hours_gate=True)
        except Exception:
            pass  # if even the failure notification fails, there's nothing more we can do here
        raise  # re-raise so the GitHub Actions run shows red/failed in the Actions tab too
